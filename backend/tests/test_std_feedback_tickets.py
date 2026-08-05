from __future__ import annotations

import os
import unittest
import uuid
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

os.environ.setdefault("DATABASE_URL", "postgresql+asyncpg://postgres:postgres@localhost/primex_test")
os.environ.setdefault("JWT_SECRET", "test-secret")

import httpx
from fastapi import HTTPException
from openpyxl import load_workbook

from app.api.routers.external_tickets import (
    _external_tickets_workbook,
    _search_condition,
    external_ticket_task_options,
    sync_external_tickets_now,
)
from app.models.enums import UserRole
from app.models.std_feedback_ticket import StdFeedbackTicket
from app.services.std_feedback_task_creation import (
    assignee_task_title,
    create_ticket_task_bundle,
    default_bundle_description,
    default_bundle_title,
    task_type_fields,
    user_initials,
)
from app.services.std_feedback_tickets import (
    StdFeedbackClient,
    _cursor_from_page,
    _needs_detail,
    _upsert_std_ticket,
    is_external_ticket_payload,
    sync_std_feedback_tickets,
    ticket_comments,
    ticket_files,
)


class _ListResult:
    def __init__(self, values):
        self.values = list(values)

    def scalars(self):
        return self

    def all(self):
        return list(self.values)

    def scalar_one(self):
        return self.values[0]

    def scalar_one_or_none(self):
        return self.values[0] if self.values else None


class _FakeDb:
    def __init__(self, select_batches=None):
        self.select_batches = list(select_batches or [])
        self.added = []
        self.commit_count = 0
        self.rollback_count = 0

    async def execute(self, _statement, *_args, **_kwargs):
        return _ListResult(self.select_batches.pop(0) if self.select_batches else [])

    def add(self, value):
        self.added.append(value)

    async def flush(self):
        for value in self.added:
            if hasattr(value, "id") and getattr(value, "id", None) is None:
                value.id = uuid.uuid4()

    async def commit(self):
        self.commit_count += 1

    async def rollback(self):
        self.rollback_count += 1


class _FakeStdClient:
    def __init__(self, pages, details=None, failure: Exception | None = None):
        self.pages = list(pages)
        self.details = details or {}
        self.failure = failure
        self.list_params = []
        self.detail_calls = []

    async def list_tickets(self, params):
        self.list_params.append(dict(params))
        if self.failure:
            raise self.failure
        return self.pages.pop(0)

    async def get_ticket(self, external_id):
        self.detail_calls.append(external_id)
        return self.details[external_id]


def _ticket_payload(external_id: str, email: str, updated_at: str = "2026-08-03T10:30:00Z"):
    return {
        "id": external_id,
        "issue_number": 1034219,
        "related_ticket_number": "ORD-55",
        "title": "Wrong product data",
        "description": "The order contains an incorrect field.",
        "creator": {"id": "creator-1", "username": "client", "email": email},
        "status": "open",
        "category": "data",
        "priority": "high",
        "created_at": "2026-08-03T09:00:00Z",
        "updated_at": updated_at,
        "comments": [{"id": "c1", "body": "Please fix"}],
        "files": [{"id": "f1", "filename": "proof.png"}],
    }


class TestStdFeedbackClient(unittest.IsolatedAsyncioTestCase):
    async def test_server_to_server_authorization_header(self) -> None:
        seen = {}

        async def handler(request: httpx.Request) -> httpx.Response:
            seen["authorization"] = request.headers.get("authorization")
            return httpx.Response(200, json={"status": "ok"})

        client = StdFeedbackClient(
            token="server-secret",
            base_url="https://std.example.test/api",
            transport=httpx.MockTransport(handler),
        )
        try:
            self.assertEqual((await client.health())["status"], "ok")
        finally:
            await client.close()
        self.assertEqual(seen["authorization"], "Bearer server-secret")

    def test_external_domain_filter_is_backend_owned(self) -> None:
        self.assertTrue(is_external_ticket_payload(_ticket_payload("1", "person@staudmoebel.de")))
        self.assertTrue(is_external_ticket_payload(_ticket_payload("2", "PERSON@STAUDMOEBEL.DE")))
        self.assertFalse(is_external_ticket_payload(_ticket_payload("3", "person@primexeu.com")))

    def test_cursor_prefers_api_next_cursor_and_can_fallback_to_last_row(self) -> None:
        rows = [_ticket_payload("a", "a@staudmoebel.de", "2026-08-03T10:00:00Z")]
        data = {
            "tickets": rows,
            "pagination": {"next_cursor": {"after_updated_at": "2026-08-03T11:00:00Z", "after_id": "b"}},
        }
        self.assertEqual(_cursor_from_page(data, rows), ("2026-08-03T11:00:00Z", "b"))
        self.assertEqual(_cursor_from_page({"tickets": rows}, rows), ("2026-08-03T10:00:00Z", "a"))

    def test_detail_is_loaded_only_for_new_or_changed_ticket(self) -> None:
        summary = _ticket_payload("a", "a@staudmoebel.de")
        existing = StdFeedbackTicket(
            external_id="a",
            source_updated_at=datetime(2026, 8, 3, 10, 30, tzinfo=timezone.utc),
        )
        self.assertFalse(_needs_detail(existing, summary))
        summary["updated_at"] = "2026-08-03T11:00:00Z"
        self.assertTrue(_needs_detail(existing, summary))
        self.assertTrue(_needs_detail(None, summary))


class TestStdFeedbackSync(unittest.IsolatedAsyncioTestCase):
    async def test_initial_sync_paginates_filters_and_updates_cursor(self) -> None:
        external_a = _ticket_payload("a", "a@staudmoebel.de", "2026-08-03T10:00:00Z")
        internal = _ticket_payload("internal", "staff@primexeu.com", "2026-08-03T10:05:00Z")
        external_b = _ticket_payload("b", "b@staudmoebel.de", "2026-08-03T11:00:00Z")
        pages = [
            {
                "tickets": [external_a, internal],
                "pagination": {
                    "has_more": True,
                    "next_cursor": {"after_updated_at": "2026-08-03T10:05:00Z", "after_id": "internal"},
                },
            },
            {"tickets": [external_b], "pagination": {"has_more": False}},
        ]
        client = _FakeStdClient(pages, {"a": external_a, "internal": internal, "b": external_b})
        state = SimpleNamespace(after_updated_at=None, after_id=None, last_sync_error=None, last_successful_sync_at=None)
        db = _FakeDb()
        upsert = AsyncMock()
        with (
            patch("app.services.std_feedback_tickets._get_sync_state", new=AsyncMock(return_value=state)),
            patch("app.services.std_feedback_tickets._existing_by_external_ids", new=AsyncMock(return_value={})),
            patch("app.services.std_feedback_tickets._upsert_std_ticket", new=upsert),
        ):
            result = await sync_std_feedback_tickets(db, client=client)

        self.assertTrue(result["ok"])
        self.assertTrue(result["initial_sync"])
        self.assertEqual(result["pages"], 2)
        self.assertEqual(result["synced"], 2)
        self.assertEqual(upsert.await_count, 2)
        self.assertEqual(client.list_params[1]["after_id"], "internal")
        self.assertEqual(state.after_id, "b")

    async def test_incremental_sync_starts_from_saved_cursor(self) -> None:
        state = SimpleNamespace(
            after_updated_at=datetime(2026, 8, 3, 10, 30, tzinfo=timezone.utc),
            after_id="saved-id",
            last_sync_error=None,
            last_successful_sync_at=None,
        )
        client = _FakeStdClient([{"tickets": [], "pagination": {"has_more": False}}])
        db = _FakeDb()
        with (
            patch("app.services.std_feedback_tickets._get_sync_state", new=AsyncMock(return_value=state)),
            patch("app.services.std_feedback_tickets._existing_by_external_ids", new=AsyncMock(return_value={})),
        ):
            result = await sync_std_feedback_tickets(db, client=client)
        self.assertTrue(result["ok"])
        self.assertFalse(result["initial_sync"])
        self.assertEqual(client.list_params[0]["after_id"], "saved-id")
        self.assertIn("2026-08-03T10:30:00", client.list_params[0]["after_updated_at"])

    async def test_failure_does_not_advance_cursor(self) -> None:
        original_time = datetime(2026, 8, 3, 10, 30, tzinfo=timezone.utc)
        state = SimpleNamespace(
            after_updated_at=original_time,
            after_id="saved-id",
            last_sync_error=None,
            last_successful_sync_at=None,
        )
        db = _FakeDb()
        client = _FakeStdClient([], failure=httpx.ConnectError("offline"))
        with patch("app.services.std_feedback_tickets._get_sync_state", new=AsyncMock(return_value=state)):
            result = await sync_std_feedback_tickets(db, client=client)
        self.assertFalse(result["ok"])
        self.assertEqual(state.after_updated_at, original_time)
        self.assertEqual(state.after_id, "saved-id")
        self.assertIn("ConnectError", state.last_sync_error)
        self.assertEqual(db.rollback_count, 1)

    async def test_upsert_reuses_existing_unique_ticket(self) -> None:
        existing = StdFeedbackTicket(external_id="a")
        db = _FakeDb()
        payload = _ticket_payload("a", "a@staudmoebel.de")
        first = await _upsert_std_ticket(db, payload, existing=existing)
        second = await _upsert_std_ticket(db, payload, existing=existing)
        self.assertIs(first, existing)
        self.assertIs(second, existing)
        self.assertEqual(db.added, [])
        self.assertEqual(existing.comment_count, 1)
        self.assertEqual(existing.file_count, 1)
        self.assertTrue(existing.is_external)


class TestStdFeedbackWorkflow(unittest.IsolatedAsyncioTestCase):
    async def test_staff_from_any_department_can_access_std_task_options(self) -> None:
        std_project = SimpleNamespace(id=uuid.uuid4(), title="STD", department_id=uuid.uuid4())
        staff = SimpleNamespace(
            id=uuid.uuid4(),
            role=UserRole.STAFF,
            department_id=uuid.uuid4(),
            full_name="Staff User",
            username="staff",
            email="staff@example.com",
            is_active=True,
        )
        db = _FakeDb([[std_project], [staff]])

        result = await external_ticket_task_options(db=db, user=staff)

        self.assertEqual([project.id for project in result.projects], [std_project.id])

    def test_assignee_initials_are_added_to_task_title(self) -> None:
        user = SimpleNamespace(full_name="Laurent Hoxha", username="laurent", email="laurent@example.com")
        self.assertEqual(user_initials(user), "LH")
        self.assertEqual(
            assignee_task_title("STD - 3 TIK EXT PËR RREGULLIM", user),
            "LH: STD - 3 TIK EXT PËR RREGULLIM",
        )

    def test_all_real_task_types_map_to_task_fields(self) -> None:
        self.assertEqual(task_type_fields("NORMAL")["priority"], "NORMAL")
        self.assertEqual(task_type_fields("HIGH")["priority"], "HIGH")
        self.assertTrue(task_type_fields("1H")["is_1h_report"])
        self.assertTrue(task_type_fields("R1")["is_r1"])
        self.assertTrue(task_type_fields("PERSONAL")["is_personal"])
        self.assertTrue(task_type_fields("BLLOK")["is_bllok"])

    async def test_one_h_bundle_keeps_ticket_list_only_in_note(self) -> None:
        ticket = StdFeedbackTicket(
            id=uuid.uuid4(),
            external_id="ticket-a",
            issue_number=1038116,
            title="Issue: 1038116",
            is_external=True,
        )
        project = SimpleNamespace(
            id=uuid.uuid4(),
            title="STD",
            department_id=uuid.uuid4(),
            current_phase="DEVELOPMENT",
        )
        assignee = SimpleNamespace(
            id=uuid.uuid4(),
            full_name="Laurent Hoxha",
            username="laurent",
            email="laurent@example.com",
            department_id=project.department_id,
        )
        db = _FakeDb([[ticket], [project], [assignee]])

        result = await create_ticket_task_bundle(
            db,
            ticket_ids=[ticket.id],
            project_id=project.id,
            assignee_ids=[assignee.id],
            actor_user_id=uuid.uuid4(),
            title="STD - 1 TIK EXT PËR RREGULLIM",
            description=None,
            review_note=None,
            priority="1H",
            start_date=None,
            due_date=None,
        )

        self.assertTrue(result.created)
        self.assertEqual(result.tasks[0].title, "LH: STD - 1 TIK EXT PËR RREGULLIM")
        self.assertIsNone(result.tasks[0].description)
        self.assertEqual(result.tasks[0].priority, "NORMAL")
        self.assertTrue(result.tasks[0].is_1h_report)
        self.assertIn("LH: STD - 1 TIK EXT PËR RREGULLIM", result.note.content)
        self.assertIn("1. 1038116", result.note.content)
        self.assertNotIn("#1038116", result.note.content)

    def test_task_content_contains_selected_ticket_references(self) -> None:
        tickets = [
            StdFeedbackTicket(external_id="a", issue_number=101, order_ticket_number="ORD-1", title="First"),
            StdFeedbackTicket(external_id="b", issue_number=102, title="Second"),
        ]
        self.assertEqual(default_bundle_title(tickets), "STD - 2 TIK EXT PËR RREGULLIM")
        description = default_bundle_description(tickets)
        self.assertIn("1. ORD-1", description)
        self.assertIn("2. 102", description)
        self.assertNotIn("#101", description)
        self.assertNotIn("STD 2.0", description)
        self.assertIn("STD External", description)

    async def test_repeating_conversion_returns_existing_tasks_without_duplicates(self) -> None:
        note_id = uuid.uuid4()
        ticket = StdFeedbackTicket(
            id=uuid.uuid4(), external_id="a", is_external=True, ga_note_id=note_id, task_id=uuid.uuid4()
        )
        note = SimpleNamespace(id=note_id)
        task = SimpleNamespace(id=ticket.task_id, ga_note_origin_id=note_id, is_active=True)
        db = _FakeDb([[ticket], [note], [task]])
        result = await create_ticket_task_bundle(
            db,
            ticket_ids=[ticket.id],
            project_id=uuid.uuid4(),
            assignee_ids=[uuid.uuid4()],
            actor_user_id=uuid.uuid4(),
            title=None,
            description=None,
            review_note=None,
            priority="NORMAL",
            start_date=None,
            due_date=None,
        )
        self.assertFalse(result.created)
        self.assertEqual(result.tasks, [task])
        self.assertEqual(db.added, [])

    async def test_sync_route_is_admin_only(self) -> None:
        user = SimpleNamespace(role=UserRole.STAFF)
        with self.assertRaises(HTTPException) as raised:
            await sync_external_tickets_now(db=AsyncMock(), user=user)
        self.assertEqual(raised.exception.status_code, 403)

    def test_search_covers_required_ticket_and_reporter_fields(self) -> None:
        sql = str(_search_condition("needle"))
        for column in (
            "issue_number",
            "order_ticket_number",
            "title",
            "description",
            "reporter_username",
            "reporter_email",
        ):
            self.assertIn(column, sql)

    def test_detail_helpers_keep_comments_and_files(self) -> None:
        ticket = StdFeedbackTicket(
            external_id="a",
            raw={"comments": [{"id": "c1"}], "attachments": [{"id": "f1"}]},
        )
        self.assertEqual(ticket_comments(ticket), [{"id": "c1"}])
        self.assertEqual(ticket_files(ticket), [{"id": "f1"}])

    def test_excel_export_contains_full_ticket_information_and_excel_dates(self) -> None:
        ticket = StdFeedbackTicket(
            id=uuid.uuid4(),
            external_id="std-266",
            issue_number=266,
            order_ticket_number="1038116",
            title="Issue: 1038116",
            description="Wrong customer number",
            affected_fields=["customer_number"],
            category="Data Issue",
            priority="High",
            status="Open",
            dashboard_area="Orders",
            reporter_username="Laurent Hoxha",
            reporter_email="laurent@example.com",
            comment_count=1,
            file_count=1,
            reported_at=datetime(2026, 8, 5, 8, 30, tzinfo=timezone.utc),
            source_updated_at=datetime(2026, 8, 5, 9, 15, tzinfo=timezone.utc),
            review_status="PENDING",
            raw={
                "comments": [{"author": "STD User", "body": "Please fix"}],
                "files": [{"filename": "proof.png"}],
            },
        )

        workbook = load_workbook(_external_tickets_workbook([ticket]))
        worksheet = workbook["STD Tickets"]
        headers = [cell.value for cell in worksheet[1]]

        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(worksheet.auto_filter.ref, "A1:AC2")
        self.assertIn("Problem / Description", headers)
        self.assertIn("Reporter Email", headers)
        self.assertIn("Comments", headers)
        self.assertIn("Review Decision", headers)
        self.assertEqual(worksheet.cell(2, headers.index("Order Ticket #") + 1).value, "1038116")
        self.assertIn("Please fix", worksheet.cell(2, headers.index("Comments") + 1).value)
        self.assertEqual(worksheet.cell(2, headers.index("Attachments") + 1).value, "proof.png")
        created_cell = worksheet.cell(2, headers.index("Created At (UTC)") + 1)
        self.assertIsInstance(created_cell.value, datetime)
        self.assertEqual(created_cell.number_format, "dd.mm.yyyy hh:mm")

    def test_frontend_never_contains_the_std_bearer_token_setting(self) -> None:
        frontend = Path(__file__).resolve().parents[2] / "frontend" / "src"
        sources = "\n".join(path.read_text(encoding="utf-8") for path in frontend.rglob("*.tsx"))
        self.assertNotIn("STD_FEEDBACK_API_TOKEN", sources)
        self.assertNotIn("STD_PRIMEFLOW_API_TOKEN", sources)


if __name__ == "__main__":
    unittest.main()
