from io import BytesIO
from datetime import date

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

from app.services.tomorrow_print_report import (
    _comment_user_initials,
    _comments_table_html,
    _dated_meetings_html,
    _excel_table_attachment,
    _html_table,
    _meeting_rows,
    _one_h_checklists_html,
    _png_table_attachment,
    _task_rows,
    build_today_print_report,
    build_tomorrow_print_report,
    ensure_required_shtypi_recipient,
)


def test_required_shtypi_recipients_are_always_in_to_without_duplicates() -> None:
    assert ensure_required_shtypi_recipient({
        "to": ["ga@primexeu.com"], "cc": ["info@primexeu.com"], "bcc": []
    }) == {
        "to": [
            "ga@primexeu.com",
            "130primex.eu@gmail.com",
            "313primex.eu@gmail.com",
            "131primex.eu@gmail.com",
            "info@primexeu.com",
        ],
        "cc": [],
        "bcc": [],
    }
    assert ensure_required_shtypi_recipient({
        "to": ["ga@primexeu.com", "313PRIMEX.EU@GMAIL.COM"],
        "cc": ["130PRIMEX.EU@GMAIL.COM"],
        "bcc": ["INFO@PRIMEXEU.COM"],
    }) == {
        "to": [
            "ga@primexeu.com",
            "313PRIMEX.EU@GMAIL.COM",
            "130primex.eu@gmail.com",
            "131primex.eu@gmail.com",
            "info@primexeu.com",
        ],
        "cc": [],
        "bcc": [],
    }


def test_staff_comment_users_keep_fixed_order_then_pcm_weekly_plan_order() -> None:
    payload = {
        "departments": [
            {"id": "pcm", "code": "PCM", "name": "Project Content Manager"},
            {"id": "dev", "code": "DEV", "name": "Development"},
        ],
        "users": [
            {"full_name": "Zana Meta", "department_id": "pcm", "is_active": True,
             "weekly_planner_sort_order": 2},
            {"full_name": "Bora Kola", "department_id": "pcm", "is_active": True,
             "weekly_planner_sort_order": 0},
            {"full_name": "Inactive User", "department_id": "pcm", "is_active": False,
             "weekly_planner_sort_order": 1},
            {"full_name": "Dev User", "department_id": "dev", "is_active": True,
             "weekly_planner_sort_order": 0},
            {"full_name": "Elsa Hoxha", "department_id": "pcm", "is_active": True,
             "weekly_planner_sort_order": 1},
        ],
    }

    assert _comment_user_initials(payload) == ["AT", "EF", "RA", "EH", "LH", "FG", "BK", "ZM"]


def test_staff_comments_use_compact_write_in_lines_in_all_formats() -> None:
    initials = ["AT", "RA", "EF", "EH", "LH", "FG", "BK"]
    html = _comments_table_html(initials)
    assert 'data-user-comments-lines="true"' in html
    assert html.count('data-user-comment-line="true"') == 2
    assert html.count('data-comment-department=') == 2
    assert '<strong>DEV:</strong>' in html
    assert '<strong>PX:</strong>' in html
    assert '<strong>GD:</strong>' not in html
    assert '<strong>PCM:</strong>' not in html
    assert html.index("<strong>AT:</strong>") < html.index("<strong>EF:</strong>")
    assert html.index("<strong>EF:</strong>") < html.index("<strong>RA:</strong>")
    assert html.index("<strong>AT:</strong>") < html.index("<strong>BK:</strong>")
    assert html.count("data-user-comment=") == len(initials)
    assert html.count('width="100%"') >= 3
    assert "border-bottom:1px solid #111827" in html
    assert ">INC<" not in html
    assert ">KOM<" not in html

    _, workbook_bytes, _ = _excel_table_attachment(
        [], [], date(2026, 8, 14), include_meetings=False, comment_initials=initials
    )
    sheet = load_workbook(BytesIO(workbook_bytes)).active
    title_cells = [cell for row in sheet.iter_rows() for cell in row if cell.value == "KOMENTE PER STAF"]
    assert len(title_cells) == 1
    title_row = title_cells[0].row
    assert sheet.cell(title_row + 1, 1).value.startswith("DEV: AT: ____________________")
    assert "EF: ____________________" in sheet.cell(title_row + 1, 1).value
    assert "LH: ____________________" in sheet.cell(title_row + 1, 1).value
    assert sheet.cell(title_row + 2, 1).value.startswith("PX: FG: ____________________")
    assert "BK: ____________________" in sheet.cell(title_row + 2, 1).value
    assert "GD:" not in sheet.cell(title_row + 2, 1).value
    assert "PCM:" not in sheet.cell(title_row + 2, 1).value
    values = [str(cell.value or "") for row in sheet.iter_rows() for cell in row]
    assert "INC" not in values
    assert "KOM" not in values

    _, png, _ = _png_table_attachment([], date(2026, 8, 14), initials)
    assert png.startswith(b"\x89PNG\r\n\x1a\n")


def test_html_table_keeps_grid_styles_inline_for_email_clients() -> None:
    report_html = _html_table(
        [("P", [{"title": "DM/GA: A task"}, {"title": "ER: A task"}], True)]
    )

    assert '<table role="presentation" width="100%" border="1"' in report_html
    assert 'style="width:100%;border-collapse:collapse;table-layout:fixed' in report_html
    assert 'style="border:1px solid #000;padding:5px' in report_html
    assert '<col width="2.5%"><col width="10.5%"><col width="14.5%" span="6">' in report_html
    assert ">LLOJI DHE SLOTI</th>" in report_html
    assert '<th colspan="6"' in report_html
    assert ">TASKS</th>" in report_html
    assert ">TASK 1</th>" not in report_html
    assert ">TASK 6</th>" not in report_html
    assert 'bgcolor="#D8B4FE"' in report_html
    assert report_html.count('background-color:#D8B4FE') == 1
    assert '<style>' not in report_html


def test_task_grid_uses_light_dividers_inside_a_slot_and_bold_slot_labels() -> None:
    report_html = _html_table(
        [("BLL", [{"title": f"Task {index}"} for index in range(7)], False)]
    )

    assert "border-top:2px solid #111827" in report_html
    assert "border-top:1px solid #cbd5e1" in report_html
    assert "border-bottom:2px solid #111827" in report_html
    assert "border:3px solid #111827" in report_html
    assert "border-top:3px solid #111827;border-bottom:3px solid #111827" in report_html
    assert 'style="border:1px solid #000;padding:5px;vertical-align:top;text-align:left;overflow-wrap:anywhere;word-break:break-word;font-weight:700;border-top:2px solid #111827;border-bottom:2px solid #111827">BLL</th>' in report_html


def test_excel_task_grid_has_thick_header_outer_frame_and_category_edges() -> None:
    _, content, _ = _excel_table_attachment(
        [
            ("1H 10:00", [{"title": f"Task {index}"} for index in range(7)], False),
            ("1H 11:00", [{"title": "Next task"}], False),
        ],
        [],
        date(2026, 8, 14),
        include_meetings=False,
    )
    sheet = load_workbook(BytesIO(content)).active

    # Row 5 is the task header; rows 6-7 are one multi-line category; row 8 is the next category.
    assert sheet["A5"].border.top.style == "medium"
    assert sheet["A5"].border.bottom.style == "medium"
    assert sheet["A5"].border.left.style == "medium"
    assert sheet["H5"].border.right.style == "medium"
    assert sheet["C6"].border.top.style == "medium"
    assert sheet["C6"].border.bottom.style == "thin"
    assert sheet["C7"].border.top.style == "thin"
    assert sheet["C7"].border.bottom.style == "medium"
    assert sheet["C8"].border.top.style == "medium"
    assert sheet["C8"].border.bottom.style == "medium"
    assert sheet["A8"].border.left.style == "medium"
    assert sheet["H8"].border.right.style == "medium"


def test_one_h_checklists_render_side_by_side_before_the_task_grid() -> None:
    checklists_html = _one_h_checklists_html()

    assert 'data-one-h-checklist-columns="true"' in checklists_html
    assert 'width="50%"' in checklists_html
    assert "PYETJET PER 1H - BORD" in checklists_html
    assert "STAFF - HAPAT PER 1H" in checklists_html
    assert "Slotin paraprak/aktual" in checklists_html
    assert 'data-board-checklist-columns="true"' in checklists_html
    assert "7. Done? / Strikes?" in checklists_html
    assert "8. Notes te reja? Data? AM/PM? Kujt?" in checklists_html
    assert "9. BZ Notes" in checklists_html
    assert "Secili i lexon vet para BZ me GA" in checklists_html
    assert "Share screen side by side DET/REZULTATIN" in checklists_html
    assert "4. BZ Det nga Stafi per GA" in checklists_html
    assert "Komunikimi GA temas Det nga Stafi/ KA email" in checklists_html
    assert checklists_html.count('data-compact-checklist-row="true"') == 2
    assert 'font-size:20px;font-weight:900' in checklists_html

    _, content, _ = _excel_table_attachment([], [], date(2026, 8, 14))
    sheet = load_workbook(BytesIO(content)).active
    assert sheet["A3"].value == "STAFF - HAPAT PER 1H"
    assert sheet["E3"].value == "PYETJET PER 1H - BORD"
    assert sheet["A4"].value.startswith("1. Hap doc dhe det / 2. Share screen")
    assert "4. BZ Det nga Stafi per GA (Komunikimi GA temas Det nga Stafi/ KA email)" in sheet["A4"].value
    assert sheet["E4"].value.startswith("1. Slotin paraprak/aktual / 2. A ke filluar")
    assert "7. Done? / Strikes?" in sheet["E4"].value
    assert "8. Notes te reja? Data? AM/PM? Kujt?" in sheet["E4"].value
    assert "9. BZ Notes (Secili i lexon vet para BZ me GA)" in sheet["E4"].value
    assert sheet["B5"].value == "LLOJI DHE SLOTI"
    assert sheet["C5"].value == "TASKS"
    assert "C5:H5" in {str(cell_range) for cell_range in sheet.merged_cells.ranges}


def test_email_table_removes_added_and_done_editor_markers() -> None:
    report_html = _html_table(
        [("1H 10:00", [{"title": "LH: [[added]]New[[/added]] [[done]]completed[[/done]] task"}], False)]
    )

    assert "[[added]]" not in report_html
    assert "[[/added]]" not in report_html
    assert "[[done]]" not in report_html
    assert "[[/done]]" not in report_html
    assert "New completed task" in report_html


def test_downloadable_email_table_uses_clean_task_text() -> None:
    filename, content, mime_type = _excel_table_attachment(
        [("1H 10:00", [{"title": "LH: [[added]]New task[[/added]]"}], False)],
        [("TAK EXT", [{"title": "[[added]]Meeting[[/added]]", "time": "10:00"}], False)],
        date(2026, 8, 14),
    )

    workbook = load_workbook(BytesIO(content))
    values = [str(cell.value or "") for row in workbook.active.iter_rows() for cell in row]
    assert filename == "1H_SHTYPI_2026-08-14.xlsx"
    assert mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert any("New task" in value for value in values)
    assert any("Meeting 10:00" in value for value in values)
    assert not any("[[added]]" in value or "[[/added]]" in value for value in values)


def test_task_status_colours_apply_to_email_cells_but_ga_personal_stays_purple() -> None:
    report_html = _html_table(
        [
            (
                "1H 10:00",
                [
                    {"title": "TODO task", "status": "TODO"},
                    {"title": "Progress task", "status": "IN_PROGRESS"},
                    {"title": "Waiting task", "status": "WAITING_CONFIRMATION"},
                    {"title": "Done task", "status": "DONE"},
                ],
                False,
            ),
            ("P", [{"title": "GA: Personal task", "status": "TODO"}], True),
        ]
    )

    assert 'bgcolor="#FFC4ED"' in report_html
    assert 'bgcolor="#FFFF00"' in report_html
    assert 'bgcolor="#FFEDD5"' in report_html
    assert 'bgcolor="#C4FDC4"' in report_html
    assert 'bgcolor="#D8B4FE"' in report_html


def test_task_cards_show_their_am_pm_period_in_email_and_excel() -> None:
    tasks = [
        {"title": "Morning task", "finishPeriod": "AM"},
        {"title": "Afternoon task", "finish_period": "pm"},
        {"title": "Unscheduled period"},
    ]

    report_html = _html_table([("1H 10:00", tasks, False)])

    assert report_html.count('data-task-badge="finish-period"') == 3
    assert 'data-task-badge="finish-period" style=' in report_html
    assert '>AM</span>' in report_html
    assert '>PM</span>' in report_html
    assert '>AM/PM</span>' in report_html
    assert "background-color:#E0F2FE" in report_html
    assert "border:1px solid #BAE6FD" in report_html
    assert "color:#0369A1" in report_html
    assert "border-radius:999px" in report_html

    _, content, _ = _excel_table_attachment(
        [("1H 10:00", tasks, False)], [], date(2026, 8, 14)
    )
    sheet = load_workbook(BytesIO(content)).active
    assert "[AM]\n" in sheet["C6"].value
    assert "[PM]\n" in sheet["D6"].value
    assert "[AM/PM]\n" in sheet["E6"].value

    _, png, _ = _png_table_attachment(
        [("1H 10:00", tasks, False)], date(2026, 8, 14)
    )
    assert png.startswith(b"\x89PNG\r\n\x1a\n")


def test_0800_and_am_pm_badges_keep_their_distinct_designs_together() -> None:
    report_html = _html_table(
        [(
            "1H NO SLOT",
            [{"title": "08:00 Combined task", "finishPeriod": None, "status": "DONE"}],
            False,
        )],
        report_date=date(2026, 8, 14),
    )

    assert report_html.count('data-task-badge="finish-period"') == 1
    assert report_html.count('data-task-badge="08:00"') == 1
    assert "background-color:#E0F2FE;border:1px solid #BAE6FD;color:#0369A1" in report_html
    assert "background-color:#DC2626;border:1px solid #B91C1C;color:#FFFFFF" in report_html


def test_em_title_without_time_is_treated_as_0800() -> None:
    report_html = _html_table(
        [("1H NO SLOT", [{"title": "EF: EM TEST", "status": "TODO"}], False)],
        report_date=date(2026, 8, 14),
    )

    assert 'data-task-badge="08:00"' in report_html


def test_deadline_and_0800_tasks_are_highlighted_in_email_and_excel() -> None:
    tasks = [
        {
            "title": "Deadline task",
            "status": "TODO",
            "is_deadline_important": True,
            "due_date": "2026-08-14T16:00:00+02:00",
        },
        {"title": "Morning task", "status": "TODO", "due_date": "2026-08-14T08:00:00+02:00"},
    ]
    report_html = _html_table([("DEADLINE / 08:00", tasks, False)], report_date=date(2026, 8, 14))

    assert 'bgcolor="#DC2626"' in report_html
    assert 'bgcolor="#FFC4ED"' in report_html
    assert "border:2px solid #DC2626" in report_html
    assert 'data-task-badge="08:00"' in report_html
    assert report_html.count('data-task-badge="due-date"') == 1
    assert 'data-badge-position="bottom-right"' in report_html
    assert 'data-due-today="true"' in report_html
    assert ">SOT</span>" in report_html
    assert ">14.08.2026</span>" not in report_html
    assert "DUE TODAY" not in report_html
    assert "background-color:#EFF6FF" in report_html
    assert "border:1px solid #93C5FD" in report_html
    assert "color:#1D4ED8" in report_html
    assert "background-color:#DC2626" in report_html
    assert "border:1px solid #B91C1C" in report_html
    assert "color:#FFFFFF" in report_html
    assert "font-weight:900" in report_html
    assert 'height="100%"' in report_html
    assert 'valign="bottom" align="right"' in report_html
    assert "position:absolute" not in report_html

    _, content, _ = _excel_table_attachment([("DEADLINE / 08:00", tasks, False)], [], date(2026, 8, 14))
    sheet = load_workbook(BytesIO(content)).active
    assert sheet["C6"].fill.fgColor.rgb.endswith("DC2626")
    assert "[14.08.2026]" in sheet["C6"].value
    assert "DUE" not in sheet["C6"].value
    assert "[08:00]" in sheet["D6"].value
    assert sheet["D6"].border.left.color.rgb.endswith("DC2626")


def test_ga_personal_purple_overrides_deadline_red_in_email_and_excel() -> None:
    ga_deadline = {
        "title": "EF/GA: Personal deadline",
        "status": "TODO",
        "is_deadline_important": True,
        "due_date": "2026-08-14",
    }
    other_deadline = {
        "title": "ER/KA: Personal deadline",
        "status": "TODO",
        "is_deadline_important": True,
        "due_date": "2026-08-14",
    }
    rows = [("P: GA", [ga_deadline], True), ("P: KA", [other_deadline], True)]

    report_html = _html_table(rows, report_date=date(2026, 8, 14))
    assert 'bgcolor="#D8B4FE"' in report_html
    assert 'bgcolor="#DC2626"' in report_html
    assert report_html.count('data-task-badge="due-date"') == 2
    assert report_html.count(">SOT</span>") == 2

    _, content, _ = _excel_table_attachment(rows, [], date(2026, 8, 14))
    sheet = load_workbook(BytesIO(content)).active
    assert sheet["C6"].fill.fgColor.rgb.endswith("D8B4FE")
    assert sheet["C7"].fill.fgColor.rgb.endswith("DC2626")


def test_done_task_stays_green_even_when_it_is_a_deadline() -> None:
    task = {
        "title": "EH: 08:00 DERGO EMAIL STD",
        "status": "DONE",
        "isDone": True,
        "is_deadline_important": True,
        "due_date": "2026-09-03",
    }
    rows = [("1H 10:00", [task], False)]

    report_html = _html_table(rows, report_date=date(2026, 9, 3))
    assert 'bgcolor="#C4FDC4"' in report_html
    assert 'bgcolor="#DC2626"' not in report_html
    assert "border:2px solid #DC2626" in report_html

    _, content, _ = _excel_table_attachment(rows, [], date(2026, 9, 3))
    sheet = load_workbook(BytesIO(content), rich_text=True).active
    assert sheet["C6"].fill.fgColor.rgb.endswith("C4FDC4")


def test_wfc_title_token_is_red_and_uses_white_highlight_on_red_deadline_cards() -> None:
    rows = [
        ("1H 10:00", [{"title": "EF: PF: WFC TEST P", "status": "TODO"}], False),
        (
            "1H 11:00",
            [{"title": "EF: PF: WFC DEADLINE", "is_deadline_important": True, "due_date": "2026-08-14"}],
            False,
        ),
        (
            "P: GA",
            [{"title": "EF/GA: WFC PERSONAL", "is_deadline_important": True, "due_date": "2026-08-14"}],
            True,
        ),
    ]

    report_html = _html_table(rows, report_date=date(2026, 8, 14))
    assert report_html.count('data-task-token="wfc"') == 3
    assert report_html.count("color:#DC2626;font-weight:800;") == 3
    assert report_html.count("background-color:#FFFFFF;border-radius:2px;padding:0 2px;") == 1
    assert 'bgcolor="#DC2626"' in report_html
    assert 'bgcolor="#D8B4FE"' in report_html

    _, content, _ = _excel_table_attachment(rows, [], date(2026, 8, 14))
    sheet = load_workbook(BytesIO(content), rich_text=True).active
    assert isinstance(sheet["C6"].value, CellRichText)
    assert isinstance(sheet["C7"].value, CellRichText)
    assert isinstance(sheet["C8"].value, CellRichText)
    normal_wfc = next(block for block in sheet["C6"].value if getattr(block, "text", "") == "WFC")
    red_card_wfc = next(block for block in sheet["C7"].value if getattr(block, "text", "") == "WFC")
    purple_card_wfc = next(block for block in sheet["C8"].value if getattr(block, "text", "") == "WFC")
    assert normal_wfc.font.color.rgb == "FFDC2626"
    assert red_card_wfc.font.color.rgb == "FFFFFF00"
    assert purple_card_wfc.font.color.rgb == "FFDC2626"


def test_future_deadline_uses_plain_white_date_text_on_the_red_cell() -> None:
    report_html = _html_table(
        [
            (
                "1H 10:00",
                [{"title": "Future deadline", "is_deadline_important": True, "due_date": "2026-08-15"}],
                False,
            )
        ],
        report_date=date(2026, 8, 14),
    )

    assert 'data-due-today="false"' in report_html
    assert ">15.08.2026</span>" in report_html
    assert "DUE" not in report_html
    assert 'bgcolor="#DC2626"' in report_html
    assert "border:0;background-color:transparent;color:#FFFFFF" in report_html


def test_overdue_deadline_uses_white_date_text_on_the_red_cell() -> None:
    report_html = _html_table(
        [(
            "1H 10:00",
            [{"title": "Overdue", "is_deadline_important": True, "due_date": "2026-08-13"}],
            False,
        )],
        report_date=date(2026, 8, 14),
    )

    assert 'bgcolor="#DC2626"' in report_html
    assert ">13.08.2026</span>" in report_html
    assert "border:0;background-color:transparent;color:#FFFFFF" in report_html
    assert "DUE" not in report_html


def test_deadline_and_0800_tasks_have_a_dedicated_printed_row() -> None:
    rows = _task_rows(
        {
            "important": [
                {"title": "Deadline task", "date": "2026-08-14", "is_deadline_important": True},
                {"title": "08:00 task", "date": "2026-08-14"},
            ]
        },
        date(2026, 8, 14),
    )

    important_row = next(row for row in rows if row[0] == "DEADLINE / 08:00")
    assert [item["title"] for item in important_row[1]] == ["Deadline task", "08:00 task"]


def test_personal_tasks_are_split_exclusively_into_ga_ka_and_px_rows() -> None:
    rows = _task_rows(
        {
            "personal": [
                {"title": "EF/GA: WFC", "date": "2026-08-14"},
                {"title": "GA/KA: GA wins", "date": "2026-08-14"},
                {"title": "ER: KA: Teams", "date": "2026-08-14"},
                {"title": "EP/ESH: Personal PX", "date": "2026-08-14"},
                {"title": "Personal task without initials", "date": "2026-08-14"},
            ]
        },
        date(2026, 8, 14),
    )
    personal_rows = [row for row in rows if row[2]]

    assert [row[0] for row in personal_rows] == [
        "P: GA\n08:15 / 13:15",
        "P: KA\n08:30 / 13:15",
        "P: PX\n08:45 / 14:00",
    ]
    assert {item["title"] for item in personal_rows[0][1]} == {"EF/GA: WFC", "GA/KA: GA wins"}
    assert [item["title"] for item in personal_rows[1][1]] == ["ER: KA: Teams"]
    assert {item["title"] for item in personal_rows[2][1]} == {
        "EP/ESH: Personal PX",
        "Personal task without initials",
    }
    assert sum(len(row[1]) for row in personal_rows) == 5
    report_html = _html_table(personal_rows)
    assert "font-size:10px" in report_html
    assert "P: GA" in report_html
    assert "P: KA" in report_html
    assert "P: PX" in report_html
    assert 'P: GA<br><span style="font-size:13px;line-height:1.2;font-weight:800;white-space:nowrap">08:15 / 13:15</span>' in report_html
    assert 'P: KA<br><span style="font-size:13px;line-height:1.2;font-weight:800;white-space:nowrap">08:30 / 13:15</span>' in report_html
    assert 'P: PX<br><span style="font-size:13px;line-height:1.2;font-weight:800;white-space:nowrap">08:45 / 14:00</span>' in report_html


def test_blocked_row_label_uses_full_afternoon_interval_and_keeps_report_time() -> None:
    rows = _task_rows({}, date(2026, 8, 14))
    blocked_row = next(row for row in rows if row[0].startswith("BLL"))

    assert blocked_row[0] == "BLL\n14:30 - 16:00\nRAP 16:10"


def test_excel_status_colours_and_done_overrides_ga_personal() -> None:
    _, content, _ = _excel_table_attachment(
        [
            ("1H 10:00", [{"title": "TODO task", "status": "TODO"}, {"title": "Progress", "status": "IN_PROGRESS"}], False),
            ("P", [{"title": "GA: Personal", "status": "DONE"}], True),
        ],
        [],
        date(2026, 8, 14),
    )

    sheet = load_workbook(BytesIO(content)).active
    assert sheet["C6"].fill.fgColor.rgb.endswith("FFC4ED")
    assert sheet["D6"].fill.fgColor.rgb.endswith("FFFF00")
    assert sheet["C7"].fill.fgColor.rgb.endswith("C4FDC4")


def test_done_tasks_are_last_within_each_printed_row() -> None:
    rows = _task_rows(
        {
            "oneH": [
                {"title": "Done first alphabetically", "date": "2026-08-14", "status": "DONE", "oneHReportSlot": "10:00"},
                {"title": "Todo later alphabetically", "date": "2026-08-14", "status": "TODO", "oneHReportSlot": "10:00"},
            ]
        },
        date(2026, 8, 14),
    )

    first_slot_items = rows[0][1]
    assert [item["title"] for item in first_slot_items] == ["Todo later alphabetically", "Done first alphabetically"]


def test_non_daily_or_weekly_meetings_get_blue_borders_in_email_and_excel() -> None:
    meetings = [
        ("TAK EXT", [
            {"title": "One-off", "time": "10:00", "recurrence_type": "none"},
            {"title": "Weekly", "time": "11:00", "recurrence_type": "weekly"},
        ], False)
    ]

    report_html = _html_table(meetings, meeting=True)
    assert "border:2px solid #2563EB" in report_html
    assert report_html.count("border:2px solid #2563EB") == 1

    _, content, _ = _excel_table_attachment([], meetings, date(2026, 8, 14))
    sheet = load_workbook(BytesIO(content)).active
    assert sheet["C8"].border.left.color.rgb.endswith("2563EB")
    assert sheet["D8"].border.left.style == "thin"


def test_email_meetings_use_grouped_today_tomorrow_columns() -> None:
    sections = [
        (
            date(2026, 8, 25),
            "SOT",
            [
                ("TAK EXT", [{"title": "Today one-off", "time": "10:00", "recurrence_type": "none"}], False),
                ("TAK INT", [{"title": "Today early internal", "time": "8:15", "recurrence_type": "weekly"}], False),
            ],
        ),
        (
            date(2026, 8, 26),
            "NESER",
            [("TAK EXT", [
                {"title": "Tomorrow weekly", "time": "11:00", "recurrence_type": "weekly"},
                {"title": "Tomorrow second", "time": "13:00", "recurrence_type": "weekly"},
            ], False)],
        ),
    ]

    report_html = _dated_meetings_html(sections)

    assert 'data-side-by-side-meetings="true"' in report_html
    assert "border:4px solid #111827" in report_html
    assert report_html.count("border:3px solid #111827") >= 6
    assert "TAKIMET SOT - 25.08.2026" in report_html
    assert "TAKIMET NESER - 26.08.2026" in report_html
    assert report_html.count(">LLOJI</th>") == 2
    assert report_html.count(">KOHA</th>") == 2
    assert report_html.count(">TAKIMET</th>") == 2
    assert report_html.count('data-meeting-time="true"') == 6
    assert ">10:00</td>" in report_html
    assert "Today one-off 10:00" not in report_html
    assert "border-left:4px solid #2563EB" in report_html
    assert report_html.count("border:2px solid #2563EB") == 2
    assert report_html.count('data-meeting-row="true"') == 3
    assert report_html.count('rowspan="2"') == 2
    assert report_html.index("Today one-off") < report_html.index("Today early internal")
    assert report_html.index("Tomorrow weekly") < report_html.index("Tomorrow second")

    _, png, _ = _png_table_attachment([], date(2026, 8, 25), meeting_sections=sections)
    assert png.startswith(b"\x89PNG\r\n\x1a\n")

    _, content, _ = _excel_table_attachment(
        [], [], date(2026, 8, 25), meeting_sections=sections
    )
    values = [str(cell.value or "") for row in load_workbook(BytesIO(content)).active.iter_rows() for cell in row]
    assert "KOHA" in values
    assert "8:15" in values
    assert "10:00" in values
    assert "1. Today one-off" in values
    assert "2. Today early internal" in values


def test_meetings_are_ordered_chronologically_even_without_leading_zero() -> None:
    rows = _meeting_rows(
        {
            "external": [
                {"title": "Late", "date": "2026-08-25", "time": "13:00"},
                {"title": "Early", "date": "2026-08-25", "time": "8:15"},
                {"title": "Middle", "date": "2026-08-25", "time": "10:00"},
            ]
        },
        date(2026, 8, 25),
    )

    assert [item["title"] for item in rows[0][1]] == ["Early", "Middle", "Late"]


def test_today_and_tomorrow_reports_separate_two_days_of_meetings(monkeypatch) -> None:
    class FakeClient:
        def __init__(self, *args, **kwargs) -> None:
            pass

        async def common_view(self, target_date: date) -> dict:
            return {
                "items": {
                    "oneH": [
                        {"title": "GA: Today task", "date": target_date.isoformat(), "oneHReportSlot": "10:00"},
                        {"title": "GA: Tomorrow task", "date": "2026-08-25", "oneHReportSlot": "10:00"},
                    ],
                    "external": [
                        {"title": "Today meeting", "date": "2026-08-24", "time": "10:00"},
                        {"title": "Tomorrow meeting", "date": "2026-08-25", "time": "11:00"},
                        {"title": "Following meeting", "date": "2026-08-26", "time": "12:00"},
                    ],
                }
            }

    monkeypatch.setenv("PRIMEFLOW_API_BASE_URL", "http://primeflow.test")
    monkeypatch.setattr("app.services.tomorrow_print_report.PrimeFlowClient", FakeClient)

    import asyncio

    report = asyncio.run(build_today_print_report(date(2026, 8, 24), include_attachment=True))

    assert report["target_date"] == "2026-08-24"
    assert report["subject"] == "1H SHTYPI  SOT— 24.08.2026"
    assert "1H SHTYPI  SOT— 24.08.2026" in report["html"]
    assert "Today task" in report["html"]
    assert "Tomorrow task" not in report["html"]
    assert "Today meeting" in report["html"]
    assert "Tomorrow meeting" in report["html"]
    assert "Following meeting" not in report["html"]
    assert "SOT - 24.08.2026" in report["html"]
    assert "NESER - 25.08.2026" in report["html"]
    assert 'data-side-by-side-meetings="true"' in report["html"]
    assert report["html"].index("Today meeting") < report["html"].index("Tomorrow meeting")
    assert 'data-today-print-report="true"' in report["content_html"]
    assert [attachment[2] for attachment in report["attachments"]] == [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "image/png",
    ]
    assert report["attachments"][1][1].startswith(b"\x89PNG\r\n\x1a\n")
    workbook = load_workbook(BytesIO(report["attachments"][0][1]))
    values = [str(cell.value or "") for row in workbook.active.iter_rows() for cell in row]
    assert "TAKIMET SOT - 24.08.2026" in values
    assert "TAKIMET NESER - 25.08.2026" in values
    assert any("Today meeting" in value for value in values)
    assert any("Tomorrow meeting" in value for value in values)

    tomorrow = asyncio.run(build_tomorrow_print_report(date(2026, 8, 24), include_attachment=True))
    assert tomorrow["target_date"] == "2026-08-25"
    assert tomorrow["subject"] == "1H SHTYPI  NESER — 25.08.2026"
    assert "1H SHTYPI  NESER — 25.08.2026" in tomorrow["html"]
    assert "Today meeting" not in tomorrow["html"]
    assert "Tomorrow meeting" in tomorrow["html"]
    assert "Following meeting" in tomorrow["html"]
    assert "NESER - 25.08.2026" in tomorrow["html"]
    assert "PAS NESER - 26.08.2026" in tomorrow["html"]
    assert "DITA PAS NESER" not in tomorrow["html"]
