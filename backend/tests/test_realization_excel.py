import io
import unittest

from openpyxl import load_workbook

from app.services.realization_calculator import (
    QUESTION_LABELS,
    REPORT_QUESTION_SECTIONS,
    build_live_questions,
)
from app.services.realization_excel import build_realization_workbook


class TestRealizationExcel(unittest.TestCase):
    def test_live_questions_cover_the_full_reference_catalog(self) -> None:
        questions = build_live_questions(
            {
                "weekly_planned_count": 5,
                "weekly_completed_count": 2,
                "weekly_additional_count": 1,
                "weekly_fast_task_count": 1,
                "daily_planned_count": 2,
                "daily_completed_count": 1,
                "counters": {"in_progress_count": 1, "no_progress_count": 0},
                "tasks": [],
                "observations": [],
                "daily_timeline": [],
                "attendance": [],
            }
        )
        expected_keys = [
            key
            for _section_title, question_keys in REPORT_QUESTION_SECTIONS
            for key in question_keys
        ]
        self.assertEqual([question["key"] for question in questions], expected_keys)
        self.assertEqual(len(questions), 15)
        by_key = {question["key"]: question for question in questions}
        self.assertEqual(by_key["task_status"]["source_status"], "AUTO")
        self.assertEqual(
            by_key["helped_colleague"]["source_status"],
            "AUTO_NEEDS_CONFIRMATION",
        )
        self.assertIsNone(by_key["helped_colleague"]["auto_value"])

    def test_export_has_department_evidence_guide_and_weekly_bonus(self) -> None:
        payload = build_realization_workbook(
            week_start="2026-08-03",
            week_end="2026-08-07",
            departments=[
                {
                    "name": "Development",
                    "status": "CALCULATED",
                    "people": [
                        {
                            "user_name": "Test Person",
                            "planned_count": 5,
                            "completed_on_time_count": 4,
                            "completed_late_count": 0,
                            "additional_count": 1,
                            "suggested_level": "A",
                            "suggested_symbol": "+",
                            "final_level": None,
                            "final_symbol": None,
                            "facts_json": {
                                "weekly_progress_percent": 80,
                                "questions": [
                                    {
                                        "key": "task_status",
                                        "label": "Statusi i detyrave",
                                        "auto_value": {"planned": 5, "completed": 4},
                                        "final_value": None,
                                        "source_status": "AUTO",
                                        "evidence_ids": ["task-1"],
                                        "explanation": "",
                                    }
                                ],
                                "tasks": [
                                    {
                                        "task_id": "task-1",
                                        "match_key": "task-1",
                                        "title": "Task",
                                        "source_type": "system",
                                        "classification": "completed_on_time",
                                    }
                                ],
                                "observations": [],
                            },
                        }
                    ],
                }
            ],
        )
        workbook = load_workbook(io.BytesIO(payload), data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            ["Përmbledhje", "Development", "Evidenca", "Udhëzuesi"],
        )
        values = " ".join(
            str(cell.value or "")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        ).lower()
        # The weekly € bonus (matching the manual grading guide's per-level
        # table) is a deliberate inclusion; the monthly "PAGA" section is not.
        self.assertIn("bonusi javor", values)
        self.assertNotIn("pagë bazë", values)
        self.assertNotIn("bonus mujor", values)

        development = workbook["Development"]
        development_values = " ".join(
            str(cell.value or "") for row in development.iter_rows() for cell in row
        ).lower()
        self.assertIn("40", development_values)
        self.assertIn("totali i javës", development_values)
        self.assertIn("nënshkrimet", development_values)

    def test_live_export_is_populated_and_clearly_not_final(self) -> None:
        payload = build_realization_workbook(
            week_start="2026-08-03",
            week_end="2026-08-07",
            departments=[
                {
                    "name": "Development",
                    "status": "AKTUAL (SNAPSHOT DITOR)",
                    "report_mode": "LIVE_DAILY",
                    "people": [
                        {
                            "user_name": "Live Person",
                            "planned_count": 6,
                            "completed_on_time_count": 1,
                            "completed_late_count": 0,
                            "additional_count": 2,
                            "suggested_level": None,
                            "suggested_symbol": None,
                            "final_level": None,
                            "final_symbol": None,
                            "facts_json": {
                                "report_mode": "LIVE_DAILY",
                                "weekly_progress_percent": 16.7,
                                "questions": [
                                    {
                                        "key": "task_status",
                                        "label": "Statusi i detyrave",
                                        "auto_value": {"planned": 5, "completed": 0},
                                        "source_status": "AUTO",
                                        "evidence_ids": [],
                                    }
                                ],
                                "daily_timeline": [
                                    {
                                        "date": "2026-08-04",
                                        "daily_progress_percent": 0,
                                        "weekly_progress_percent": 16.7,
                                        "planned_count": 5,
                                        "completed_count": 0,
                                        "weekly_planned_count": 6,
                                        "weekly_completed_count": 1,
                                        "additional_count": 1,
                                        "attendance": [],
                                    }
                                ],
                            },
                        }
                    ],
                }
            ],
        )
        workbook = load_workbook(io.BytesIO(payload), data_only=False)
        summary = workbook["Përmbledhje"]
        development = workbook["Development"]

        self.assertEqual(summary["B5"].value, "Live Person")
        self.assertEqual(summary["C5"].value, 6)
        self.assertEqual(summary["D5"].value, 1)
        self.assertEqual(summary["H5"].value, "Në pritje të FINAL")
        self.assertIn("AKTUAL NGA SNAPSHOT-ET DITORE", summary["A2"].value)
        values = " ".join(
            str(cell.value or "")
            for row in development.iter_rows()
            for cell in row
        )
        self.assertIn("Sot: 0/5 (0%)", values)
        self.assertIn("Vlerësimi final", values)
        self.assertIn("PËR KONFIRMIM", values)
        for section_title, question_keys in REPORT_QUESTION_SECTIONS:
            self.assertIn(section_title, values)
            for key in question_keys:
                self.assertIn(QUESTION_LABELS[key], values)


if __name__ == "__main__":
    unittest.main()
