import io
import unittest
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook

from app.services.excel_standardizer import analyze_workbook, initials_from_user, standardize_workbook


class TestExcelStandardizer(unittest.TestCase):
    def test_initials_are_derived_from_primeflow_user_profile(self) -> None:
        self.assertEqual(initials_from_user("Arben Krasniqi", "arben.k", "arben@example.com"), "AK")
        self.assertEqual(initials_from_user("", "arben.krasniqi", "arben@example.com"), "AK")
        self.assertEqual(initials_from_user("", "", "arben.krasniqi@example.com"), "AK")

    def _source(self) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(["Name", "Amount", None])
        worksheet.append(["Primex", "1,370 EUR", "Shënim"])
        workbook.create_sheet("Empty")
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def test_analysis_detects_missing_header_without_asking_for_all_headers(self) -> None:
        analysis = analyze_workbook(self._source(), "stock_mi_gl.xlsx")

        self.assertEqual(len(analysis.sheets), 1)
        self.assertEqual(analysis.sheets[0].source_header_row, 1)
        self.assertEqual([item.column for item in analysis.sheets[0].missing_headers], ["C"])
        self.assertEqual(analysis.empty_sheets, ["Empty"])

    def test_standardization_applies_primex_structure_and_print_settings(self) -> None:
        generated_at = datetime(2026, 8, 3, 10, 15, tzinfo=ZoneInfo("Europe/Belgrade"))
        result, filename, report = standardize_workbook(
            self._source(),
            "stock_mi_gl.xlsx",
            initials="an",
            missing_headers={"Sheet1": {"C": "Notes"}},
            description="STOCK_MI_GL",
            generated_at=generated_at,
        )

        self.assertEqual(filename, "STOCK_MI_GL_03.08.2026_AN.xlsx")
        workbook = load_workbook(io.BytesIO(result), data_only=False)
        self.assertEqual(workbook.sheetnames, ["STOCK MI GL"])
        worksheet = workbook.active
        self.assertEqual(worksheet["A3"].value, "STOCK MI GL")
        self.assertEqual([worksheet.cell(6, column).value for column in range(1, 5)], ["NR", "NAME", "AMOUNT", "NOTES"])
        self.assertEqual(worksheet["A7"].value, 1)
        self.assertEqual(worksheet["B7"].value, "PX")
        self.assertEqual(worksheet["C7"].value, 1370)
        self.assertEqual(worksheet["C7"].number_format, "#,##0.00")
        self.assertEqual([worksheet["A8"].value, worksheet["A9"].value], [2, 3])
        self.assertEqual(worksheet.freeze_panes, "C7")
        self.assertEqual(worksheet.auto_filter.ref, "A6:D7")
        self.assertEqual(worksheet.print_title_rows, "$6:$6")
        self.assertEqual(worksheet.oddHeader.right.text, "03/08/2026 10:15")
        self.assertEqual(worksheet.oddFooter.center.text, "&P / &N")
        self.assertEqual(worksheet.oddFooter.right.text, "PUNOI:")
        self.assertTrue(any(item["category"] == "empty_sheets" for sheet in report["sheets"] for item in sheet["corrections"]))

    def test_existing_final_nr_rows_are_replaced_not_duplicated(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "STOCK"
        worksheet["A3"] = "STOCK"
        worksheet["A6"] = "NR"
        worksheet["B6"] = "ITEM"
        worksheet["A7"] = 1
        worksheet["B7"] = "Chair"
        worksheet["A8"] = 2
        worksheet["A9"] = 3
        source = io.BytesIO()
        workbook.save(source)

        result, _, _ = standardize_workbook(source.getvalue(), "stock.xlsx", "AN")
        standardized = load_workbook(io.BytesIO(result)).active

        self.assertEqual(standardized.max_row, 9)
        self.assertEqual([standardized[f"A{row}"].value for row in range(7, 10)], [1, 2, 3])
        self.assertIsNone(standardized["B8"].value)
        self.assertIsNone(standardized["B9"].value)


if __name__ == "__main__":
    unittest.main()
