from __future__ import annotations

import io
import asyncio
import unittest
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from types import SimpleNamespace
from zoneinfo import ZoneInfo
from unittest.mock import AsyncMock, patch

import httpx
from openpyxl import load_workbook

from app.services.weekly_planning_audit import (
    APPROVED_CHECK_DIMENSIONS,
    AUDIT_CHECK_REGISTRY,
    AuditError,
    AuditTaskOccurrence,
    PersonAudit,
    WeeklyPlanningAuditReport,
    audit_person_occurrences,
    _ai_title_introduces_unknown_abbreviation,
    clean_technical_markup,
    extract_px_abbreviations,
    is_reportable_person,
    load_px_abbreviations,
    monday_of_next_working_week,
    normalize_week_start,
    partition_users_by_full_week_leave,
    select_weekly_focus,
    suggested_concise_title,
    validate_report_integrity,
    validate_task_occurrence,
    validated_ai_errors_by_user,
)
from app.models.weekly_planning_audit import WeeklyPlanningAuditRun
from app.models.common_entry import CommonEntry
from app.models.enums import CommonCategory, CommonApprovalStatus
from app.services.common_leave import parse_common_view_annual_leave
from app.services.weekly_planning_audit_delivery import (
    build_delivery_record,
    record_delivery_failure,
    scheduled_idempotency_key,
    stable_smtp_message_id,
)
from app.services.weekly_planning_audit_excel import (
    SHEET_NAMES,
    build_weekly_planning_audit_workbook,
    report_subject,
    update_weekly_planning_audit_delivery_metadata,
)
from app.config import settings
from app.services.weekly_planning_audit_ai import analyze_weekly_planning_audit


WEEK_START = date(2026, 8, 3)


def occurrence(**overrides) -> AuditTaskOccurrence:
    values = {
        "user_id": uuid.UUID(int=1),
        "employee": "Elsa Ferati",
        "department": "Development",
        "task_id": uuid.UUID(int=2),
        "task_date": WEEK_START,
        "title": "Implementimi i modulit të auditimit",
        "description": None,
        "internal_notes": None,
        "status": "TODO",
        "priority": "NORMAL",
        "finish_period": "AM",
        "start_date": WEEK_START,
        "due_date": WEEK_START,
    }
    values.update(overrides)
    return AuditTaskOccurrence(**values)


class WeeklyPlanningAuditLogicTests(unittest.TestCase):
    def setUp(self) -> None:
        self.abbreviations, _ = load_px_abbreviations()

    def test_full_week_leave_user_is_excluded(self) -> None:
        user = SimpleNamespace(id=uuid.UUID(int=1))
        leave = {user.id: {WEEK_START.replace(day=WEEK_START.day + offset) for offset in range(5)}}
        included, excluded = partition_users_by_full_week_leave(
            [user], leave_dates_by_user=leave, week_start=WEEK_START
        )
        self.assertEqual(included, [])
        self.assertEqual(excluded, [user])

    def test_partial_week_leave_user_is_included(self) -> None:
        user = SimpleNamespace(id=uuid.UUID(int=1))
        included, excluded = partition_users_by_full_week_leave(
            [user], leave_dates_by_user={user.id: {date(2026, 8, 5)}}, week_start=WEEK_START
        )
        self.assertEqual(included, [user])
        self.assertEqual(excluded, [])

    def test_task_on_partial_leave_day_creates_one_error(self) -> None:
        task = occurrence(task_date=date(2026, 8, 5))
        errors = validate_task_occurrence(
            task,
            week_start=WEEK_START,
            leave_dates={date(2026, 8, 5)},
            official_abbreviations=self.abbreviations,
        )
        leave_errors = [error for error in errors if error.rule_code == "TASK_ON_ANNUAL_LEAVE"]
        self.assertEqual(len(leave_errors), 1)
        self.assertEqual(
            leave_errors[0].problem,
            "Detyra është planifikuar në ditën e Pushimit Vjetor.",
        )

    def test_task_outside_leave_days_is_not_flagged(self) -> None:
        errors = validate_task_occurrence(
            occurrence(task_date=date(2026, 8, 4)),
            week_start=WEEK_START,
            leave_dates={date(2026, 8, 5)},
            official_abbreviations=self.abbreviations,
        )
        self.assertNotIn("TASK_ON_ANNUAL_LEAVE", {error.rule_code for error in errors})

    def test_system_tasks_and_templates_are_never_focus(self) -> None:
        system = occurrence(
            title="Important system occurrence",
            is_system=True,
            system_template_origin_id=uuid.UUID(int=9),
            project_id=None,
        )
        real = occurrence(title="Real client launch", task_id=uuid.UUID(int=3), project_id=None)
        self.assertEqual(select_weekly_focus([system, real]).label, "Real client launch")

    def test_routine_tasks_are_not_focus(self) -> None:
        routine = occurrence(title="PLNF JAV routine", project_id=None)
        real = occurrence(title="Ndërtimi i portalit të klientit", task_id=uuid.UUID(int=3), project_id=None)
        self.assertEqual(select_weekly_focus([routine, real]).label, "Ndërtimi i portalit të klientit")

    def test_standard_meeting_project_is_not_focus(self) -> None:
        meeting = occurrence(
            title="Koordinim me ekipin",
            project_id=uuid.UUID(int=88),
            project_name="Takim standard me agjentët",
        )
        self.assertEqual(
            select_weekly_focus([meeting]).label,
            "Nuk është përcaktuar fokus jo-sistem",
        )

    def test_real_project_with_report_word_is_allowed_and_dominant_project_wins(self) -> None:
        project_id = uuid.UUID(int=10)
        tasks = [
            occurrence(
                task_id=uuid.UUID(int=20 + offset),
                task_date=WEEK_START.replace(day=WEEK_START.day + offset),
                title="Develop GDPR reporting product",
                project_id=project_id,
                project_name="GDPR Reporting Platform",
                priority="HIGH" if offset == 0 else "NORMAL",
            )
            for offset in range(3)
        ]
        tasks.append(occurrence(task_id=uuid.UUID(int=30), title="Single unrelated task", project_id=None))
        decision = select_weekly_focus(tasks)
        self.assertEqual(decision.label, "GDPR Reporting Platform")
        self.assertEqual(decision.source_project_id, str(project_id))

    def test_user_without_non_system_work_has_exact_focus_text(self) -> None:
        decision = select_weekly_focus([occurrence(is_system=True, project_id=None)])
        self.assertEqual(decision.label, "Nuk është përcaktuar fokus jo-sistem")

    def test_markup_is_removed_without_removing_content(self) -> None:
        self.assertEqual(
            clean_technical_markup("AT: P[[added]]X[[/added]] [[done]]WEB[[/done]]"),
            "AT: PX WEB",
        )

    def test_unofficial_abbreviation_is_not_invented(self) -> None:
        proposed = suggested_concise_title("RREG i faqes së klientit")
        self.assertIn("Rregullim", proposed)
        self.assertNotIn("RREG", proposed)
        self.assertNotIn("RREG", self.abbreviations)

    def test_long_title_proposes_shorter_title_and_preserves_clean_current_title(self) -> None:
        title = "Projekt klienti " + "udhëzim shumë i gjatë " * 12
        task = occurrence(title=title)
        errors = validate_task_occurrence(
            task, week_start=WEEK_START, leave_dates=set(), official_abbreviations=self.abbreviations
        )
        error = next(item for item in errors if item.rule_code == "TITLE_TOO_LONG")
        self.assertEqual(error.current_title, title.strip())
        self.assertTrue(error.move_to_notes)
        self.assertLess(len(error.proposed_title), len(title))

    def test_intrinsic_title_error_is_deduplicated_across_five_days(self) -> None:
        user = SimpleNamespace(
            id=uuid.UUID(int=80), full_name="Employee", username="employee", email="employee@example.com"
        )
        task_id = uuid.UUID(int=81)
        title = "Projekt klienti " + "udhëzim shumë i gjatë " * 12
        tasks = [
            occurrence(
                user_id=user.id,
                task_id=task_id,
                task_date=WEEK_START + timedelta(days=offset),
                title=title,
            )
            for offset in range(5)
        ]
        _, errors, cleanup = audit_person_occurrences(
            user,
            department="Development",
            occurrences=tasks,
            leave_dates=set(),
            week_start=WEEK_START,
            abbreviations=self.abbreviations,
        )
        self.assertEqual(sum(error.rule_code == "TITLE_TOO_LONG" for error in errors), 1)
        self.assertEqual(sum(row["task_id"] == str(task_id) for row in cleanup), 1)

    def test_leave_error_is_occurrence_specific_across_three_dates(self) -> None:
        user = SimpleNamespace(
            id=uuid.UUID(int=82), full_name="Employee", username="employee", email="employee@example.com"
        )
        task_id = uuid.UUID(int=83)
        leave_dates = {WEEK_START + timedelta(days=offset) for offset in range(3)}
        tasks = [
            occurrence(user_id=user.id, task_id=task_id, task_date=day)
            for day in sorted(leave_dates)
        ]
        _, errors, _ = audit_person_occurrences(
            user,
            department="Development",
            occurrences=tasks,
            leave_dates=leave_dates,
            week_start=WEEK_START,
            abbreviations=self.abbreviations,
        )
        self.assertEqual(sum(error.rule_code == "TASK_ON_ANNUAL_LEAVE" for error in errors), 3)

    def test_editor_metadata_is_cleaned_and_not_reported_as_an_error(self) -> None:
        task = occurrence(title="AT: P[[added]]X[[/added]] [[done]]WEB[[/done]]")
        errors = validate_task_occurrence(
            task, week_start=WEEK_START, leave_dates=set(), official_abbreviations=self.abbreviations
        )
        self.assertNotIn("TECHNICAL_MARKUP", {error.rule_code for error in errors})
        self.assertTrue(all("[[" not in error.current_title for error in errors))

    def test_ai_has_safe_deterministic_fallback_without_api_key(self) -> None:
        previous_key = settings.OPENAI_API_KEY
        previous_enabled = settings.WEEKLY_PLANNING_AUDIT_AI_ENABLED
        try:
            settings.OPENAI_API_KEY = None
            settings.WEEKLY_PLANNING_AUDIT_AI_ENABLED = True
            result, status = asyncio.run(analyze_weekly_planning_audit({
                "people": [{"user_id": "1", "tasks": [{"task_id": "task-1", "title": "Punë reale"}]}]
            }))
        finally:
            settings.OPENAI_API_KEY = previous_key
            settings.WEEKLY_PLANNING_AUDIT_AI_ENABLED = previous_enabled
        self.assertIsNone(result)
        self.assertEqual(status, "missing_api_key")

    def test_ai_timeout_returns_fallback(self) -> None:
        previous_key = settings.OPENAI_API_KEY
        previous_enabled = settings.WEEKLY_PLANNING_AUDIT_AI_ENABLED
        try:
            settings.OPENAI_API_KEY = "test-key"
            settings.WEEKLY_PLANNING_AUDIT_AI_ENABLED = True
            with patch(
                "app.services.weekly_planning_audit_ai.httpx.AsyncClient.post",
                new=AsyncMock(side_effect=httpx.ReadTimeout("timeout")),
            ):
                result, status = asyncio.run(analyze_weekly_planning_audit({
                    "people": [{"user_id": "1", "tasks": [{"task_id": "task-1", "title": "Punë"}]}]
                }))
        finally:
            settings.OPENAI_API_KEY = previous_key
            settings.WEEKLY_PLANNING_AUDIT_AI_ENABLED = previous_enabled
        self.assertIsNone(result)
        self.assertEqual(status, "fallback")

    def test_next_week_uses_local_tirana_time_and_next_monday(self) -> None:
        friday = datetime(2026, 7, 31, 9, 0, tzinfo=ZoneInfo("Europe/Tirane"))
        self.assertEqual(monday_of_next_working_week(friday, "Europe/Tirane"), WEEK_START)
        report_friday = datetime(2026, 8, 7, 10, 30, tzinfo=ZoneInfo("Europe/Tirane"))
        self.assertEqual(
            monday_of_next_working_week(report_friday, "Europe/Tirane"),
            date(2026, 8, 10),
        )

    def test_explicit_monday_is_accepted_and_non_monday_rejected(self) -> None:
        self.assertEqual(normalize_week_start(WEEK_START), WEEK_START)
        with self.assertRaisesRegex(ValueError, "Monday"):
            normalize_week_start(WEEK_START + timedelta(days=1))

    def test_active_admin_is_excluded_without_environment_configuration(self) -> None:
        admin = SimpleNamespace(
            id=uuid.UUID(int=70), username="admin", email="admin@example.com",
            full_name="Admin", role="ADMIN", is_active=True,
        )
        employee = SimpleNamespace(
            id=uuid.UUID(int=71), username="employee", email="employee@example.com",
            full_name="Normal Employee", role="STAFF", is_active=True,
        )
        inactive = SimpleNamespace(
            id=uuid.UUID(int=72), username="inactive", email="inactive@example.com",
            full_name="Inactive Employee", role="STAFF", is_active=False,
        )
        self.assertFalse(is_reportable_person(admin))
        self.assertTrue(is_reportable_person(employee))
        self.assertFalse(is_reportable_person(inactive))
        self.assertFalse(is_reportable_person(employee, ["employee@example.com"]))

    def test_px_matcher_uses_boundaries_longest_match_and_title_only(self) -> None:
        self.assertNotIn("T", extract_px_abbreviations("LH: WKF VS API", self.abbreviations))
        self.assertEqual(extract_px_abbreviations("LH:EF: PF", self.abbreviations), ["PF"])
        self.assertEqual(extract_px_abbreviations("TAK EXT: klienti", self.abbreviations), ["TAK EXT"])
        proposed = suggested_concise_title("LH:EF: PF\nJAV ekziston vetëm në notes")
        self.assertNotIn("JAV", extract_px_abbreviations(proposed, self.abbreviations))

    def test_scheduled_retry_uses_same_idempotency_key(self) -> None:
        first = scheduled_idempotency_key(week_start=WEEK_START, slot="09:00", recipient_config_version=4)
        retry = scheduled_idempotency_key(week_start=WEEK_START, slot="09:00", recipient_config_version=4)
        later_slot = scheduled_idempotency_key(week_start=WEEK_START, slot="09:30", recipient_config_version=4)
        self.assertEqual(first, retry)
        self.assertNotEqual(first, later_slot)

    def test_retry_reuses_same_initial_smtp_message_id(self) -> None:
        run_id = uuid.UUID(int=100)
        first = stable_smtp_message_id(
            run_id=run_id,
            delivery_id=uuid.UUID(int=101),
            resend=False,
            sender_domain="primexeu.com",
        )
        retry = stable_smtp_message_id(
            run_id=run_id,
            delivery_id=uuid.UUID(int=102),
            resend=False,
            sender_domain="primexeu.com",
        )
        self.assertEqual(first, retry)

    def test_manual_resend_builds_new_delivery_audit_record(self) -> None:
        run = WeeklyPlanningAuditRun(
            id=uuid.UUID(int=110),
            week_start=WEEK_START,
            week_end=date(2026, 8, 7),
            slot="10:30",
            recipients_snapshot={"to": ["ga@primexeu.com"], "cc": [], "bcc": []},
            filename="report.xlsx",
            file_checksum="a" * 64,
        )
        delivery = build_delivery_record(
            run,
            requested_by=uuid.UUID(int=111),
            resend=True,
            attempt=2,
        )
        self.assertEqual(delivery.delivery_type, "RESEND")
        self.assertEqual(delivery.attempt_number, 2)
        self.assertEqual(delivery.requested_by, uuid.UUID(int=111))

    def test_smtp_failure_is_recorded_on_run_and_delivery(self) -> None:
        run = WeeklyPlanningAuditRun(
            id=uuid.UUID(int=120),
            week_start=WEEK_START,
            week_end=date(2026, 8, 7),
            slot="09:00",
            recipients_snapshot={"to": ["ga@primexeu.com"], "cc": [], "bcc": []},
            filename="report.xlsx",
            file_checksum="b" * 64,
        )
        delivery = build_delivery_record(run, requested_by=None, resend=False, attempt=1)
        record_delivery_failure(delivery, run, RuntimeError("SMTP unavailable"))
        self.assertEqual(delivery.status, "FAILED")
        self.assertEqual(run.status, "FAILED")
        self.assertEqual(delivery.error_message, "SMTP unavailable")
        self.assertEqual(run.error_message, "SMTP unavailable")

    def test_common_view_leave_parser_handles_full_week_range(self) -> None:
        entry = CommonEntry(
            category=CommonCategory.annual_leave,
            title="PV",
            description="Date range: 2026-08-03 to 2026-08-07 (Full day)",
            entry_date=WEEK_START,
            created_by_user_id=uuid.UUID(int=1),
            approval_status=CommonApprovalStatus.approved,
            created_at=datetime(2026, 7, 20),
        )
        start, end, full_day, _, _, _, all_users = parse_common_view_annual_leave(entry)
        self.assertEqual((start, end), (WEEK_START, date(2026, 8, 7)))
        self.assertTrue(full_day)
        self.assertFalse(all_users)

    def test_common_view_leave_parser_preserves_partial_day_semantics(self) -> None:
        entry = CommonEntry(
            category=CommonCategory.annual_leave,
            title="PV",
            description="Date: 2026-08-05 (10:00 - 12:00)",
            entry_date=date(2026, 8, 5),
            created_by_user_id=uuid.UUID(int=1),
            approval_status=CommonApprovalStatus.approved,
            created_at=datetime(2026, 7, 20),
        )
        _, _, full_day, start_time, end_time, _, _ = parse_common_view_annual_leave(entry)
        self.assertFalse(full_day)
        self.assertEqual((start_time, end_time), ("10:00", "12:00"))

    def test_missing_am_pm_and_due_date_are_not_weekly_audit_errors(self) -> None:
        errors = validate_task_occurrence(
            occurrence(status=None, priority=None, finish_period=None, due_date=None),
            week_start=WEEK_START,
            leave_dates=set(),
            official_abbreviations=self.abbreviations,
        )
        codes = {error.rule_code for error in errors}
        self.assertTrue({"STATUS_MISSING", "PRIORITY_MISSING"} <= codes)
        self.assertNotIn("FINISH_PERIOD_MISSING", codes)
        self.assertNotIn("DUE_DATE_MISSING", codes)

    def test_finish_period_is_informational_and_never_modified(self) -> None:
        task = occurrence(finish_period="PM")
        errors = validate_task_occurrence(
            task, week_start=WEEK_START, leave_dates=set(), official_abbreviations=self.abbreviations
        )
        self.assertEqual(task.finish_period, "PM")
        self.assertNotIn("FINISH_PERIOD_MISSING", {error.rule_code for error in errors})

    def test_other_genuine_errors_remain_when_am_pm_is_missing(self) -> None:
        errors = validate_task_occurrence(
            occurrence(finish_period=None, status="INVALID"),
            week_start=WEEK_START,
            leave_dates=set(),
            official_abbreviations=self.abbreviations,
        )
        codes = {error.rule_code for error in errors}
        self.assertIn("STATUS_INVALID", codes)
        self.assertNotIn("FINISH_PERIOD_MISSING", codes)

    def test_missing_am_pm_does_not_increase_person_error_count(self) -> None:
        user = SimpleNamespace(
            id=uuid.UUID(int=90), full_name="Employee", username="employee",
            email="employee@example.com",
        )
        person, errors, _ = audit_person_occurrences(
            user,
            department="Development",
            occurrences=[occurrence(user_id=user.id, finish_period=None)],
            leave_dates=set(),
            week_start=WEEK_START,
            abbreviations=self.abbreviations,
        )
        self.assertEqual(errors, [])
        self.assertEqual(person.error_count, 0)

    def test_ko_word_in_title_is_not_a_ko_rule(self) -> None:
        errors = validate_task_occurrence(
            occurrence(title="Koordinim me klientin", ko_required=False),
            week_start=WEEK_START,
            leave_dates=set(),
            official_abbreviations=self.abbreviations,
        )
        self.assertFalse(any(error.rule_code.startswith("KO_") for error in errors))

    def test_no_planner_work_does_not_invent_missing_plan_error(self) -> None:
        user = SimpleNamespace(
            id=uuid.UUID(int=91), full_name="No Work", username="no.work", email="no.work@example.com"
        )
        person, errors, _ = audit_person_occurrences(
            user,
            department="Finance",
            occurrences=[],
            leave_dates=set(),
            week_start=WEEK_START,
            abbreviations=self.abbreviations,
        )
        self.assertEqual(person.focus, "Nuk është përcaktuar fokus jo-sistem")
        self.assertEqual(person.error_count, 0)
        self.assertNotIn("NO_MEANINGFUL_WEEKLY_PLAN", {error.rule_code for error in errors})

    def test_badged_task_format_validations_are_applicable_only(self) -> None:
        r1_errors = validate_task_occurrence(
            occurrence(title="R1 task without colon", is_r1=True),
            week_start=WEEK_START,
            leave_dates=set(),
            official_abbreviations=self.abbreviations,
        )
        normal_errors = validate_task_occurrence(
            occurrence(title="Normal task", is_r1=False),
            week_start=WEEK_START,
            leave_dates=set(),
            official_abbreviations=self.abbreviations,
        )
        self.assertIn("R1_FORMAT_INVALID", {error.rule_code for error in r1_errors})
        self.assertNotIn("R1_FORMAT_INVALID", {error.rule_code for error in normal_errors})
        applicable_cases = [
            (occurrence(title="Personal task", is_personal=True), "PERSONAL_FORMAT_INVALID"),
            (occurrence(title="BLL task", is_bllok=True), "BLL_FORMAT_INVALID"),
            (occurrence(title="WFC task"), "WFC_FORMAT_INVALID"),
            (occurrence(title="BKP task"), "BKP_FORMAT_INVALID"),
            (occurrence(title="1H report", is_1h_report=True, one_h_report_slot=None), "ONE_H_SLOT_MISSING"),
        ]
        for task, code in applicable_cases:
            with self.subTest(code=code):
                errors = validate_task_occurrence(
                    task,
                    week_start=WEEK_START,
                    leave_dates=set(),
                    official_abbreviations=self.abbreviations,
                )
                self.assertIn(code, {error.rule_code for error in errors})
        normal_codes = {error.rule_code for error in normal_errors}
        self.assertTrue({
            "PERSONAL_FORMAT_INVALID", "BLL_FORMAT_INVALID", "WFC_FORMAT_INVALID",
            "BKP_FORMAT_INVALID", "ONE_H_SLOT_MISSING",
        }.isdisjoint(normal_codes))

    def test_missing_priority_is_not_required_for_system_occurrence(self) -> None:
        errors = validate_task_occurrence(
            occurrence(priority=None, is_system=True),
            week_start=WEEK_START,
            leave_dates=set(),
            official_abbreviations=self.abbreviations,
        )
        self.assertNotIn("PRIORITY_MISSING", {error.rule_code for error in errors})

    def test_total_and_average_can_live_in_description(self) -> None:
        errors = validate_task_occurrence(
            occurrence(title="VS: Kontrolli i produkteve", description="Total: 120; Mesatare: 24"),
            week_start=WEEK_START,
            leave_dates=set(),
            official_abbreviations=self.abbreviations,
        )
        self.assertNotIn("TOTAL_AVERAGE_INVALID", {error.rule_code for error in errors})

    def test_total_and_average_are_checked_only_when_ko_rule_applies(self) -> None:
        ordinary = validate_task_occurrence(
            occurrence(title="Analizo Total të porosive", ko_required=False),
            week_start=WEEK_START,
            leave_dates=set(),
            official_abbreviations=self.abbreviations,
        )
        ko_task = validate_task_occurrence(
            occurrence(title="Kontroll produktesh", ko_required=True, ko_user_id_present=True),
            week_start=WEEK_START,
            leave_dates=set(),
            official_abbreviations=self.abbreviations,
        )
        self.assertNotIn("TOTAL_AVERAGE_INVALID", {error.rule_code for error in ordinary})
        self.assertIn("TOTAL_AVERAGE_INVALID", {error.rule_code for error in ko_task})

    def test_ai_proposed_title_cannot_introduce_unknown_abbreviation(self) -> None:
        self.assertTrue(
            _ai_title_introduces_unknown_abbreviation(
                "XYZ: Rregullimi i faqes", "Rregullimi i faqes", self.abbreviations
            )
        )
        self.assertFalse(
            _ai_title_introduces_unknown_abbreviation(
                "PF: Rregullimi i faqes", "Rregullimi i faqes", self.abbreviations
            )
        )
        self.assertFalse(
            _ai_title_introduces_unknown_abbreviation(
                "XYZ: Rregullimi i faqes", "XYZ: Faqja e klientit", self.abbreviations
            )
        )

    def test_ai_findings_are_rejected_when_ungrounded_or_invented(self) -> None:
        task = occurrence()
        result = {
            "errors": [
                {
                    "task_id": str(uuid.UUID(int=999)),
                    "problem": "Problem",
                    "proposed_title": "Titull",
                    "correction": "Korrigjim",
                    "severity": "LOW",
                },
                {
                    "task_id": str(task.task_id),
                    "problem": "Problem",
                    "proposed_title": "XYZ: Titull i ri",
                    "correction": "Korrigjim",
                    "severity": "LOW",
                },
            ]
        }
        self.assertEqual(validated_ai_errors_by_user(result, [task], self.abbreviations), {})

    def test_official_dictionary_contains_seeded_values_and_not_rreg(self) -> None:
        self.assertEqual(self.abbreviations["PF"], "PRIME FLOW/PLATFORMA")
        self.assertEqual(self.abbreviations["RIORG"], "RIORGANIZIM")
        self.assertGreaterEqual(len(self.abbreviations), 60)
        self.assertNotIn("RREG", self.abbreviations)


class WeeklyPlanningAuditAcceptanceSnapshotTests(unittest.TestCase):
    WEEK = date(2026, 8, 10)
    LAURENT_TASK_ID = uuid.UUID("ac583682-61c7-497c-93e6-fa3eacbe387c")

    def setUp(self) -> None:
        self.abbreviations, metadata = load_px_abbreviations()
        names = [
            "Anisa Ternava", "Endi Hyseni", "Florent Gara", "Rinesa Ahmedi",
            "Elsa Ferati", "Laurent Hoxha", "Enesa Sharku", "Haxhere Spahiu", "Gane Arifaj",
        ]
        self.users = [
            SimpleNamespace(
                id=uuid.UUID(int=100 + index), full_name=name,
                username=name.lower().replace(" ", "."), email=f"u{index}@example.com",
                department_id=None, is_active=True,
            )
            for index, name in enumerate(names)
        ]
        by_name = {user.full_name: user for user in self.users}
        week_dates = {self.WEEK + timedelta(days=offset) for offset in range(5)}
        self.leave = {
            by_name["Anisa Ternava"].id: set(week_dates),
            by_name["Endi Hyseni"].id: set(week_dates),
            by_name["Florent Gara"].id: set(week_dates),
            by_name["Rinesa Ahmedi"].id: {self.WEEK + timedelta(days=offset) for offset in range(4)},
        }
        included, excluded = partition_users_by_full_week_leave(
            self.users, leave_dates_by_user=self.leave, week_start=self.WEEK
        )
        self.excluded_names = {user.full_name for user in excluded}

        occurrences_by_user: dict[uuid.UUID, list[AuditTaskOccurrence]] = defaultdict(list)
        for offset in range(2):
            occurrences_by_user[by_name["Rinesa Ahmedi"].id].append(occurrence(
                user_id=by_name["Rinesa Ahmedi"].id,
                task_id=uuid.UUID(int=200 + offset),
                task_date=self.WEEK,
                start_date=self.WEEK,
                due_date=self.WEEK,
                title=f"RA: FRG: Detyra reale {offset + 1}",
            ))
        for offset, title in enumerate(("PLNF JAV", "BLL: Kontroll", "P: Email rutinë", "Raport javor")):
            occurrences_by_user[by_name["Elsa Ferati"].id].append(occurrence(
                user_id=by_name["Elsa Ferati"].id,
                task_id=uuid.UUID(int=300 + offset),
                task_date=self.WEEK + timedelta(days=offset),
                start_date=self.WEEK,
                due_date=self.WEEK + timedelta(days=offset),
                title=title,
                is_bllok=title.startswith("BLL"),
                is_personal=title.startswith("P:"),
            ))
        laurent_title = "LH: WKF VS API\n" + "1. Hapi i detajuar për implementim dhe verifikim.\n" * 8
        for offset in range(5):
            occurrences_by_user[by_name["Laurent Hoxha"].id].append(occurrence(
                user_id=by_name["Laurent Hoxha"].id,
                task_id=self.LAURENT_TASK_ID,
                task_date=self.WEEK + timedelta(days=offset),
                start_date=self.WEEK,
                due_date=self.WEEK + timedelta(days=4),
                title=laurent_title,
            ))
        for index, name in enumerate(("Enesa Sharku", "Haxhere Spahiu"), start=400):
            occurrences_by_user[by_name[name].id].append(occurrence(
                user_id=by_name[name].id, task_id=uuid.UUID(int=index),
                task_date=self.WEEK, start_date=self.WEEK, due_date=self.WEEK,
                title=f"{name}: Punë reale për klientin",
            ))
        occurrences_by_user[by_name["Gane Arifaj"].id].append(occurrence(
            user_id=by_name["Gane Arifaj"].id, task_id=uuid.UUID(int=500),
            task_date=self.WEEK, start_date=self.WEEK, due_date=self.WEEK,
            title="TAK EXT: Takim standard me agjentët",
            project_id=uuid.UUID(int=501), project_name="Takim standard me agjentët",
        ))

        people: list[PersonAudit] = []
        errors: list[AuditError] = []
        cleanup: list[dict[str, object]] = []
        executed = set(APPROVED_CHECK_DIMENSIONS)
        for user in included:
            person, person_errors, person_cleanup = audit_person_occurrences(
                user,
                department="Development",
                occurrences=occurrences_by_user.get(user.id, []),
                leave_dates=self.leave.get(user.id, set()),
                week_start=self.WEEK,
                abbreviations=self.abbreviations,
                executed_checks=executed,
            )
            people.append(person)
            errors.extend(person_errors)
            cleanup.extend(person_cleanup)
        self.report = WeeklyPlanningAuditReport(
            week_start=self.WEEK,
            week_end=self.WEEK + timedelta(days=4),
            generated_at=datetime(2026, 8, 7, 10, 30, tzinfo=ZoneInfo("Europe/Tirane")),
            timezone="Europe/Tirane",
            slot="10:30",
            people=people,
            errors=errors,
            title_cleanup=cleanup,
            excluded_full_leave=sorted(self.excluded_names),
            partial_leave_users=["Rinesa Ahmedi"],
            abbreviations=self.abbreviations,
            abbreviation_version=metadata["version"],
            abbreviation_source=metadata["source"],
            abbreviation_updated_at=metadata["updated_at"],
            executed_checks=list(APPROVED_CHECK_DIMENSIONS),
        )

    def test_snapshot_leave_plan_focus_and_dedup_acceptance(self) -> None:
        names = {person.employee for person in self.report.people}
        self.assertTrue({"Anisa Ternava", "Endi Hyseni", "Florent Gara"}.isdisjoint(names))
        rinesa = next(person for person in self.report.people if person.employee == "Rinesa Ahmedi")
        self.assertEqual(rinesa.leave_status, "Po, parcial: 10.08.2026–13.08.2026")
        leave_errors = [
            error for error in self.report.errors
            if error.employee == "Rinesa Ahmedi" and error.rule_code == "TASK_ON_ANNUAL_LEAVE"
        ]
        self.assertEqual(len(leave_errors), 2)
        for name in ("Elsa Ferati", "Laurent Hoxha", "Rinesa Ahmedi", "Enesa Sharku", "Haxhere Spahiu"):
            self.assertFalse(any(
                error.employee == name and error.rule_code == "NO_MEANINGFUL_WEEKLY_PLAN"
                for error in self.report.errors
            ))
        laurent_errors = [error for error in self.report.errors if error.task_id == str(self.LAURENT_TASK_ID)]
        self.assertLessEqual(sum(error.rule_code == "MULTIPLE_INSTRUCTIONS_IN_TITLE" for error in laurent_errors), 1)
        self.assertLessEqual(sum(error.rule_code == "TITLE_TOO_LONG" for error in laurent_errors), 1)
        self.assertEqual(sum(row["task_id"] == str(self.LAURENT_TASK_ID) for row in self.report.title_cleanup), 1)
        gane = next(person for person in self.report.people if person.employee == "Gane Arifaj")
        self.assertEqual(gane.focus, "Nuk është përcaktuar fokus jo-sistem")
        self.assertEqual(set(self.report.executed_checks), set(APPROVED_CHECK_DIMENSIONS))
        self.assertTrue(all(AUDIT_CHECK_REGISTRY[dimension] for dimension in self.report.executed_checks))
        validate_report_integrity(self.report)

    def test_snapshot_workbook_contract(self) -> None:
        recipients = ["130primex.eu@gmail.com", "info@primexeu.com", "ga@primexeu.com"]
        raw = build_weekly_planning_audit_workbook(
            self.report,
            recipients={"to": recipients, "cc": [], "bcc": []},
            run_id=str(uuid.UUID(int=999)),
        )
        workbook = load_workbook(io.BytesIO(raw), data_only=True)
        self.assertEqual(workbook.sheetnames, SHEET_NAMES)
        delivery = workbook["DËRGIMI AUTOMATIK"]
        self.assertEqual(delivery["C2"].value, "10:30")
        self.assertTrue(all(address in delivery["E2"].value for address in recipients))
        self.assertIsNone(delivery["G2"].value)
        workbook.close()


class WeeklyPlanningAuditWorkbookTests(unittest.TestCase):
    def _report(self) -> WeeklyPlanningAuditReport:
        abbreviations, metadata = load_px_abbreviations()
        return WeeklyPlanningAuditReport(
            week_start=WEEK_START,
            week_end=date(2026, 8, 7),
            generated_at=datetime(2026, 7, 31, 9, 0, tzinfo=ZoneInfo("Europe/Tirane")),
            timezone="Europe/Tirane",
            slot="10:30",
            people=[
                PersonAudit(
                    user_id=str(uuid.UUID(int=1)),
                    employee="Elsa Ferati",
                    department="Development",
                    leave_status="Jo",
                    focus="Audit Platform",
                    focus_source="Project 1",
                    focus_source_task_id=str(uuid.UUID(int=2)),
                    focus_source_project_id=str(uuid.UUID(int=3)),
                    task_count=2,
                    error_count=0,
                    critical_count=0,
                    high_count=0,
                    assessment="Në rregull",
                    required_action="Asnjë veprim",
                ),
                PersonAudit(
                    user_id=str(uuid.UUID(int=4)),
                    employee="No Tasks User",
                    department="Finance",
                    leave_status="Jo",
                    focus="Nuk është përcaktuar fokus jo-sistem",
                    focus_source="Nuk ka detyrë/projekt jo-sistem të vlefshëm",
                    focus_source_task_id=None,
                    focus_source_project_id=None,
                    task_count=0,
                    error_count=0,
                    critical_count=0,
                    high_count=0,
                    assessment="Në rregull",
                    required_action="Asnjë veprim",
                ),
            ],
            errors=[],
            abbreviations=abbreviations,
            abbreviation_version=metadata["version"],
            abbreviation_source=metadata["source"],
            abbreviation_updated_at=metadata["updated_at"],
            executed_checks=list(APPROVED_CHECK_DIMENSIONS),
        )

    def test_excel_has_exact_sheets_and_required_headers_and_opens(self) -> None:
        workbook_bytes = build_weekly_planning_audit_workbook(
            self._report(),
            recipients={"to": ["ga@primexeu.com"], "cc": [], "bcc": []},
            run_id=str(uuid.UUID(int=99)),
        )
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
        self.assertEqual(workbook.sheetnames, SHEET_NAMES)
        final_headers = [cell.value for cell in workbook["RAPORTI FINAL"][1]]
        detail_headers = [cell.value for cell in workbook["DETAJET E GABIMEVE"][1]]
        self.assertIn("Fokusi kryesor i javës", final_headers)
        self.assertIn("Kodi i rregullit", detail_headers)
        self.assertEqual(workbook["RAPORTI FINAL"].max_row, 3)
        delivery_headers = [cell.value for cell in workbook["DËRGIMI AUTOMATIK"][1]]
        self.assertIn("Versioni i fjalorit PX", delivery_headers)
        self.assertIn("Checksum i payload-it", delivery_headers)
        self.assertEqual(len(workbook["DËRGIMI AUTOMATIK"]["N2"].value), 64)
        workbook.close()

    def test_1030_subject_is_exact(self) -> None:
        self.assertEqual(
            report_subject(self._report()),
            "Kontrolli 10:30 | PLNF JAV 03.08.2026–07.08.2026 | Raporti 31.07.2026",
        )

    def test_delivery_metadata_never_invents_message_id(self) -> None:
        raw = build_weekly_planning_audit_workbook(
            self._report(),
            recipients={
                "to": ["130primex.eu@gmail.com", "info@primexeu.com", "ga@primexeu.com"],
                "cc": [],
                "bcc": [],
            },
            run_id=str(uuid.UUID(int=99)),
        )
        generated = load_workbook(io.BytesIO(raw), data_only=True)
        self.assertEqual(generated["DËRGIMI AUTOMATIK"]["F2"].value, "Generated, not sent")
        self.assertIsNone(generated["DËRGIMI AUTOMATIK"]["G2"].value)
        self.assertEqual(generated["DËRGIMI AUTOMATIK"]["K2"].value, "not_needed")
        generated.close()

        sent_raw = update_weekly_planning_audit_delivery_metadata(
            raw, delivery_status="Sent", message_id="provider-123", attempt_number=1
        )
        sent = load_workbook(io.BytesIO(sent_raw), data_only=True)
        self.assertEqual(sent["DËRGIMI AUTOMATIK"]["F2"].value, "Sent")
        self.assertEqual(sent["DËRGIMI AUTOMATIK"]["G2"].value, "provider-123")
        self.assertEqual(sent["DËRGIMI AUTOMATIK"]["H2"].value, 1)
        sent.close()


if __name__ == "__main__":
    unittest.main()
