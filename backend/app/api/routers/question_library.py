from __future__ import annotations

import io
import re
import uuid
from datetime import date, datetime, time, timedelta, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import APIRouter, Depends, HTTPException, Response, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import delete, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, require_admin, require_manager_or_admin
from app.config import settings
from app.db import get_db
from app.models.enums import NotificationType, ProjectPhaseStatus, TaskPriority, TaskStatus, UserRole
from app.models.question_library import (
    QuestionCategory,
    QuestionDailySignoff,
    QuestionDefinition,
    QuestionEditEvent,
    QuestionStatusEvent,
    QuestionUserStatus,
)
from app.models.task import Task
from app.models.task_assignee import TaskAssignee
from app.models.user import User
from app.schemas.question_library import (
    QuestionCategoryCreate,
    QuestionCategoryOut,
    QuestionCategoryUpdate,
    QuestionDailySignoffSummary,
    QuestionDailySignoffUpdate,
    QuestionDefinitionCreate,
    QuestionDefinitionOut,
    QuestionDefinitionUpdate,
    QuestionEditHistoryOut,
    QuestionStatusHistoryOut,
    QuestionStatusSummary,
    QuestionStatusUpdate,
)
from app.services.notifications import add_notification, publish_notification


router = APIRouter()
QUESTION_TASK_START_DATE = date(2026, 8, 3)


def can_manage_question_library(role: UserRole) -> bool:
    return role in (UserRole.ADMIN, UserRole.MANAGER)


def visible_status_owner_id(role: UserRole, current_user_id: uuid.UUID) -> uuid.UUID | None:
    return None if can_manage_question_library(role) else current_user_id


def _clean_required(value: str) -> str:
    cleaned = value.strip()
    if not cleaned:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Value cannot be empty")
    return cleaned


def _clean_optional(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


def _user_initials(full_name: str | None) -> str:
    parts = (full_name or "").strip().split()
    if not parts:
        return ""
    return "".join(part[0] for part in parts)[:2].upper()


def _is_question_participant(user: User) -> bool:
    return user.is_active and _user_initials(user.full_name) not in {"GA", "KA", "HV"}


def _question_task_title(text: str) -> str:
    return f"PYETJE E RE: {text}"


def _question_task_description(guidance: str | None) -> str:
    return guidance or "Përgjigju me ✓ ose X te faqja Pyetje për Barazim."


def _question_task_due_date(created_at: datetime) -> datetime:
    try:
        app_timezone = ZoneInfo(settings.APP_TIMEZONE)
    except ZoneInfoNotFoundError:
        app_timezone = timezone.utc
    local_created_at = created_at.astimezone(app_timezone)
    if local_created_at.hour >= 12:
        return (local_created_at + timedelta(days=1)).astimezone(timezone.utc)
    local_end_of_day = datetime.combine(
        local_created_at.date(),
        time.max,
        tzinfo=app_timezone,
    )
    return local_end_of_day.astimezone(timezone.utc)


def _daily_signoff_window() -> tuple[datetime, datetime]:
    try:
        app_timezone = ZoneInfo(settings.APP_TIMEZONE)
    except ZoneInfoNotFoundError:
        app_timezone = timezone.utc
    local_now = datetime.now(app_timezone)
    local_start = datetime.combine(local_now.date(), time.min, tzinfo=app_timezone)
    local_end = local_start + timedelta(days=1)
    return local_start.astimezone(timezone.utc), local_end.astimezone(timezone.utc)


def _question_tasks_enabled_at(moment: datetime) -> bool:
    try:
        app_timezone = ZoneInfo(settings.APP_TIMEZONE)
    except ZoneInfoNotFoundError:
        app_timezone = timezone.utc
    return moment.astimezone(app_timezone).date() >= QUESTION_TASK_START_DATE


def _question_task_batch_date(moment: datetime) -> date:
    try:
        app_timezone = ZoneInfo(settings.APP_TIMEZONE)
    except ZoneInfoNotFoundError:
        app_timezone = timezone.utc
    return moment.astimezone(app_timezone).date()


def _question_batch_content(
    questions: list[QuestionDefinition],
) -> tuple[str, str]:
    if len(questions) == 1:
        question = questions[0]
        return _question_task_title(question.text), _question_task_description(question.guidance)
    lines = ["Përgjigju për secilën pyetje te faqja Pyetje për Barazim.", ""]
    for index, question in enumerate(questions, start=1):
        lines.append(f"{index}. {question.text}")
        if question.guidance:
            lines.append(f"   Udhëzim: {question.guidance}")
    return f"PYETJE TË REJA ({len(questions)})", "\n".join(lines)


async def _question_task_members(
    db: AsyncSession,
    task_id: uuid.UUID,
) -> list[QuestionDefinition]:
    return list(
        (
            await db.execute(
                select(QuestionDefinition)
                .where(QuestionDefinition.task_id == task_id)
                .order_by(QuestionDefinition.created_at, QuestionDefinition.sort_order)
            )
        ).scalars().all()
    )


def _apply_question_task_content(
    task: Task,
    questions: list[QuestionDefinition],
) -> None:
    task.title, task.description = _question_batch_content(questions)


async def _refresh_question_task_status(
    db: AsyncSession,
    task: Task,
) -> None:
    question_ids = set(
        (
            await db.execute(
                select(QuestionDefinition.id).where(QuestionDefinition.task_id == task.id)
            )
        ).scalars().all()
    )
    if not question_ids and task.question_origin_id is not None:
        question_ids = {task.question_origin_id}
    assigned_user_ids = set(
        (
            await db.execute(
                select(TaskAssignee.user_id).where(TaskAssignee.task_id == task.id)
            )
        ).scalars().all()
    )
    responded_pairs = (
        {
            (row.question_id, row.user_id)
            for row in (
                await db.execute(
                    select(QuestionUserStatus.question_id, QuestionUserStatus.user_id).where(
                        QuestionUserStatus.question_id.in_(question_ids)
                    )
                )
            ).all()
        }
        if question_ids
        else set()
    )
    all_users_responded = bool(question_ids and assigned_user_ids) and all(
        (question_id, user_id) in responded_pairs
        for question_id in question_ids
        for user_id in assigned_user_ids
    )
    task.status = TaskStatus.DONE.value if all_users_responded else TaskStatus.TODO.value
    task.completed_at = datetime.now(timezone.utc) if all_users_responded else None


async def _split_or_update_edited_question_task(
    db: AsyncSession,
    question: QuestionDefinition,
) -> None:
    task = await db.get(Task, question.task_id) if question.task_id is not None else None
    if task is None:
        task = await db.scalar(
            select(Task).where(Task.question_origin_id == question.id)
        )
    if task is None:
        return

    members = await _question_task_members(db, task.id)
    if not members:
        members = [question]
    other_members = [item for item in members if item.id != question.id]
    if not other_members:
        question.task_id = task.id
        task.question_batch_date = None
        _apply_question_task_content(task, [question])
        task.status = TaskStatus.TODO.value
        task.completed_at = None
        return

    assignee_ids = list(
        (
            await db.execute(
                select(TaskAssignee.user_id).where(TaskAssignee.task_id == task.id)
            )
        ).scalars().all()
    )
    question.task_id = None
    if task.question_origin_id == question.id:
        task.question_origin_id = other_members[0].id
        task.fast_task_group_id = other_members[0].id
    _apply_question_task_content(task, other_members)
    await db.flush()
    await _refresh_question_task_status(db, task)

    separate_task = Task(
        title=_question_task_title(question.text),
        description=_question_task_description(question.guidance),
        assigned_to=None,
        created_by=task.created_by,
        question_origin_id=question.id,
        question_batch_date=None,
        fast_task_group_id=question.id,
        status=TaskStatus.TODO.value,
        priority=task.priority,
        phase=task.phase,
        progress_percentage=0,
        start_date=datetime.now(timezone.utc),
        due_date=task.due_date,
        is_deadline_important=task.is_deadline_important,
        is_r1=task.is_r1,
        is_active=task.is_active,
    )
    db.add(separate_task)
    await db.flush()
    question.task_id = separate_task.id
    for user_id in assignee_ids:
        db.add(TaskAssignee(task_id=separate_task.id, user_id=user_id))


async def _detach_question_from_shared_task(
    db: AsyncSession,
    question: QuestionDefinition,
) -> None:
    task = await db.get(Task, question.task_id) if question.task_id is not None else None
    if task is None:
        return
    members = await _question_task_members(db, task.id)
    other_members = [item for item in members if item.id != question.id]
    if not other_members:
        return
    question.task_id = None
    if task.question_origin_id == question.id:
        task.question_origin_id = other_members[0].id
        task.fast_task_group_id = other_members[0].id
    _apply_question_task_content(task, other_members)
    await db.flush()
    await _refresh_question_task_status(db, task)


async def _category_or_404(db: AsyncSession, category_id: uuid.UUID) -> QuestionCategory:
    category = await db.get(QuestionCategory, category_id)
    if category is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question category not found")
    return category


async def _question_or_404(db: AsyncSession, question_id: uuid.UUID) -> QuestionDefinition:
    question = await db.get(QuestionDefinition, question_id)
    if question is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question not found")
    return question


async def _question_out(
    db: AsyncSession,
    question: QuestionDefinition,
    current_user: User,
) -> QuestionDefinitionOut:
    owner_id = visible_status_owner_id(current_user.role, current_user.id)
    stmt = (
        select(QuestionUserStatus, User.full_name)
        .join(User, User.id == QuestionUserStatus.user_id)
        .where(QuestionUserStatus.question_id == question.id)
        .order_by(User.full_name, QuestionUserStatus.updated_at.desc())
    )
    if owner_id is not None:
        stmt = stmt.where(QuestionUserStatus.user_id == owner_id)
    status_rows = (await db.execute(stmt)).all()
    summaries = [
        QuestionStatusSummary(
            user_id=status_row.user_id,
            full_name=full_name,
            status=status_row.status,
            updated_at=status_row.updated_at,
        )
        for status_row, full_name in status_rows
    ]
    own = next((item.status for item in summaries if item.user_id == current_user.id), None)
    if own is None and owner_id is None:
        own_status = await db.scalar(
            select(QuestionUserStatus.status).where(
                QuestionUserStatus.question_id == question.id,
                QuestionUserStatus.user_id == current_user.id,
            )
        )
        own = own_status
    signoff_start, signoff_end = _daily_signoff_window()
    signoff_stmt = (
        select(QuestionDailySignoff, User.full_name)
        .join(User, User.id == QuestionDailySignoff.user_id)
        .where(
            QuestionDailySignoff.question_id == question.id,
            QuestionDailySignoff.signed_at >= signoff_start,
            QuestionDailySignoff.signed_at < signoff_end,
        )
        .order_by(User.full_name, QuestionDailySignoff.signed_at.desc())
    )
    if owner_id is not None:
        signoff_stmt = signoff_stmt.where(QuestionDailySignoff.user_id == owner_id)
    daily_signoffs = [
        QuestionDailySignoffSummary(
            user_id=signoff.user_id,
            full_name=full_name,
            signed_at=signoff.signed_at,
        )
        for signoff, full_name in (await db.execute(signoff_stmt)).all()
    ]
    task_status = None
    if question.task_id is not None:
        task_status = await db.scalar(select(Task.status).where(Task.id == question.task_id))
    if task_status is None:
        task_status = await db.scalar(
            select(Task.status).where(Task.question_origin_id == question.id)
        )
    return QuestionDefinitionOut(
        id=question.id,
        category_id=question.category_id,
        text=question.text,
        guidance=question.guidance,
        sort_order=question.sort_order,
        edit_count=question.edit_count,
        current_user_status=own,
        statuses=summaries,
        is_done=(task_status or "").upper() == TaskStatus.DONE.value,
        current_user_daily_signed=any(item.user_id == current_user.id for item in daily_signoffs),
        daily_signoffs=daily_signoffs,
        created_at=question.created_at,
        updated_at=question.updated_at,
    )


def _question_out_from_summaries(
    question: QuestionDefinition,
    summaries: list[QuestionStatusSummary],
    daily_signoffs: list[QuestionDailySignoffSummary],
    is_done: bool,
    current_user_id: uuid.UUID,
) -> QuestionDefinitionOut:
    return QuestionDefinitionOut(
        id=question.id,
        category_id=question.category_id,
        text=question.text,
        guidance=question.guidance,
        sort_order=question.sort_order,
        edit_count=question.edit_count,
        current_user_status=next(
            (item.status for item in summaries if item.user_id == current_user_id),
            None,
        ),
        statuses=summaries,
        is_done=is_done,
        current_user_daily_signed=any(item.user_id == current_user_id for item in daily_signoffs),
        daily_signoffs=daily_signoffs,
        created_at=question.created_at,
        updated_at=question.updated_at,
    )


@router.get("", response_model=list[QuestionCategoryOut])
async def list_question_library(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[QuestionCategoryOut]:
    categories = (
        await db.execute(select(QuestionCategory).order_by(QuestionCategory.sort_order, QuestionCategory.name))
    ).scalars().all()
    questions = (
        await db.execute(
            select(QuestionDefinition).order_by(
                QuestionDefinition.category_id,
                QuestionDefinition.sort_order,
                QuestionDefinition.created_at,
            )
        )
    ).scalars().all()
    by_category: dict[uuid.UUID, list[QuestionDefinition]] = {}
    for question in questions:
        by_category.setdefault(question.category_id, []).append(question)

    statuses_by_question: dict[uuid.UUID, list[QuestionStatusSummary]] = {}
    signoffs_by_question: dict[uuid.UUID, list[QuestionDailySignoffSummary]] = {}
    done_question_ids: set[uuid.UUID] = set()
    if questions:
        owner_id = visible_status_owner_id(current_user.role, current_user.id)
        status_stmt = (
            select(QuestionUserStatus, User.full_name)
            .join(User, User.id == QuestionUserStatus.user_id)
            .where(QuestionUserStatus.question_id.in_([item.id for item in questions]))
            .order_by(User.full_name, QuestionUserStatus.updated_at.desc())
        )
        if owner_id is not None:
            status_stmt = status_stmt.where(QuestionUserStatus.user_id == owner_id)
        for status_row, full_name in (await db.execute(status_stmt)).all():
            statuses_by_question.setdefault(status_row.question_id, []).append(
                QuestionStatusSummary(
                    user_id=status_row.user_id,
                    full_name=full_name,
                    status=status_row.status,
                    updated_at=status_row.updated_at,
                )
            )
        signoff_start, signoff_end = _daily_signoff_window()
        signoff_stmt = (
            select(QuestionDailySignoff, User.full_name)
            .join(User, User.id == QuestionDailySignoff.user_id)
            .where(
                QuestionDailySignoff.question_id.in_([item.id for item in questions]),
                QuestionDailySignoff.signed_at >= signoff_start,
                QuestionDailySignoff.signed_at < signoff_end,
            )
            .order_by(User.full_name, QuestionDailySignoff.signed_at.desc())
        )
        if owner_id is not None:
            signoff_stmt = signoff_stmt.where(QuestionDailySignoff.user_id == owner_id)
        for signoff, full_name in (await db.execute(signoff_stmt)).all():
            signoffs_by_question.setdefault(signoff.question_id, []).append(
                QuestionDailySignoffSummary(
                    user_id=signoff.user_id,
                    full_name=full_name,
                    signed_at=signoff.signed_at,
                )
            )
        done_question_ids = set(
            (
                await db.execute(
                    select(QuestionDefinition.id)
                    .join(Task, Task.id == QuestionDefinition.task_id)
                    .where(
                        QuestionDefinition.id.in_([item.id for item in questions]),
                        Task.status == TaskStatus.DONE.value,
                    )
                )
            ).scalars().all()
        )
        legacy_done_question_ids = set(
            (
                await db.execute(
                    select(Task.question_origin_id).where(
                        Task.question_origin_id.in_(
                            [item.id for item in questions if item.task_id is None]
                        ),
                        Task.status == TaskStatus.DONE.value,
                    )
                )
            ).scalars().all()
        )
        done_question_ids.update(legacy_done_question_ids)

    output: list[QuestionCategoryOut] = []
    for category in categories:
        question_output = [
            _question_out_from_summaries(
                item,
                statuses_by_question.get(item.id, []),
                signoffs_by_question.get(item.id, []),
                item.id in done_question_ids,
                current_user.id,
            )
            for item in by_category.get(category.id, [])
        ]
        output.append(
            QuestionCategoryOut(
                id=category.id,
                name=category.name,
                sort_order=category.sort_order,
                questions=question_output,
                created_at=category.created_at,
                updated_at=category.updated_at,
            )
        )
    return output


@router.get("/categories/{category_id}/export.xlsx")
async def export_question_category_excel(
    category_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    categories = await list_question_library(db=db, current_user=current_user)
    category = next((item for item in categories if item.id == category_id), None)
    if category is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question category not found")

    participants: list[User] = []
    if can_manage_question_library(current_user.role):
        active_users = (
            await db.execute(
                select(User)
                .where(User.is_active.is_(True))
                .order_by(User.full_name, User.email)
            )
        ).scalars().all()
        participants = [item for item in active_users if _is_question_participant(item)]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pyetje"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"

    sheet.merge_cells("A1:H1")
    title_cell = sheet["A1"]
    title_cell.value = "PYETJE PËR BARAZIM"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="183B68")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells("A2:H2")
    category_cell = sheet["A2"]
    category_cell.value = f"Kategoria: {category.name}"
    category_cell.font = Font(bold=True, size=12, color="071126")
    category_cell.fill = PatternFill("solid", fgColor="E7EDF5")
    category_cell.alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "NR",
        "PYETJA",
        "UDHËZIMI",
        "UNDERSTOOD / SIGNED",
        "NEEDS CLARIFICATION",
        "AWAITING RESPONSE",
        "CHECKED TODAY",
        "DONE",
    ]
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    border_side = Side(style="thin", color="183B68")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for column, label in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column, value=label)
        cell.font = Font(bold=True, size=9.5, color="071126")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 34

    participant_ids = {item.id for item in participants}
    participant_names = {
        item.id: item.full_name or item.username or item.email
        for item in participants
    }
    row_count = max(1, len(category.questions))
    base_row_height = min(42, max(22, 420 / row_count))
    for row_index, question in enumerate(category.questions, start=4):
        understood = [item.full_name for item in question.statuses if item.status == "DONE"]
        needs_clarification = [item.full_name for item in question.statuses if item.status == "X"]
        responded_ids = {item.user_id for item in question.statuses}
        awaiting = [
            participant_names[user_id]
            for user_id in participant_ids - responded_ids
        ] if participants else []
        checked_today = [item.full_name for item in question.daily_signoffs]
        values = [
            row_index - 3,
            question.text,
            question.guidance or "",
            ", ".join(sorted(understood)),
            ", ".join(sorted(needs_clarification)),
            ", ".join(sorted(awaiting)),
            ", ".join(sorted(checked_today)),
            "PO" if question.is_done else "JO",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.font = Font(size=10.5, color="071126")
            cell.border = cell_border
            cell.alignment = Alignment(
                horizontal="center" if column in {1, 8} else "left",
                vertical="top",
                wrap_text=True,
            )
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7FBFF")
        content_lines = max(
            1,
            len(question.text.splitlines()),
            len((question.guidance or "").splitlines()),
        )
        sheet.row_dimensions[row_index].height = max(base_row_height, 16 * content_lines + 8)

    # Every data column keeps its own fixed width and full border in print.
    widths = {"A": 6, "B": 42, "C": 34, "D": 23, "E": 23, "F": 27, "G": 21, "H": 9}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.auto_filter.ref = f"A3:H{max(3, len(category.questions) + 3)}"
    sheet.print_title_rows = "1:3"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.3
    sheet.page_margins.bottom = 0.3
    sheet.page_margins.header = 0.1
    sheet.page_margins.footer = 0.1
    sheet.print_area = f"A1:H{max(3, len(category.questions) + 3)}"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", category.name).strip("_") or "pyetje"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="pyetje_{safe_name}.xlsx"',
            "Cache-Control": "no-store",
        },
    )


@router.post("/categories", response_model=QuestionCategoryOut, status_code=status.HTTP_201_CREATED)
async def create_question_category(
    payload: QuestionCategoryCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> QuestionCategoryOut:
    name = _clean_required(payload.name)
    normalized_name = name.casefold()
    existing = await db.scalar(select(QuestionCategory.id).where(QuestionCategory.normalized_name == normalized_name))
    if existing is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A category with this name already exists")
    next_order = (await db.scalar(select(func.coalesce(func.max(QuestionCategory.sort_order), -1)))) + 1
    category = QuestionCategory(
        name=name,
        normalized_name=normalized_name,
        sort_order=next_order,
        created_by_user_id=current_user.id,
    )
    db.add(category)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A category with this name already exists")
    await db.refresh(category)
    return QuestionCategoryOut(
        id=category.id,
        name=category.name,
        sort_order=category.sort_order,
        questions=[],
        created_at=category.created_at,
        updated_at=category.updated_at,
    )


@router.patch("/categories/{category_id}", response_model=QuestionCategoryOut)
async def update_question_category(
    category_id: uuid.UUID,
    payload: QuestionCategoryUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> QuestionCategoryOut:
    category = await _category_or_404(db, category_id)
    name = _clean_required(payload.name)
    normalized_name = name.casefold()
    duplicate = await db.scalar(
        select(QuestionCategory.id).where(
            QuestionCategory.normalized_name == normalized_name,
            QuestionCategory.id != category_id,
        )
    )
    if duplicate is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A category with this name already exists")
    category.name = name
    category.normalized_name = normalized_name
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A category with this name already exists")
    await db.refresh(category)
    questions = (
        await db.execute(
            select(QuestionDefinition)
            .where(QuestionDefinition.category_id == category.id)
            .order_by(QuestionDefinition.sort_order, QuestionDefinition.created_at)
        )
    ).scalars().all()
    return QuestionCategoryOut(
        id=category.id,
        name=category.name,
        sort_order=category.sort_order,
        questions=[await _question_out(db, item, current_user) for item in questions],
        created_at=category.created_at,
        updated_at=category.updated_at,
    )


@router.delete("/categories/{category_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
async def delete_question_category(
    category_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
) -> Response:
    category = await _category_or_404(db, category_id)
    category_questions = (
        await db.execute(
            select(QuestionDefinition).where(QuestionDefinition.category_id == category_id)
        )
    ).scalars().all()
    for question in category_questions:
        await _detach_question_from_shared_task(db, question)
    await db.delete(category)
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post(
    "/categories/{category_id}/questions",
    response_model=QuestionDefinitionOut,
    status_code=status.HTTP_201_CREATED,
)
async def create_question_definition(
    category_id: uuid.UUID,
    payload: QuestionDefinitionCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> QuestionDefinitionOut:
    await _category_or_404(db, category_id)
    next_order = (
        await db.scalar(
            select(func.coalesce(func.max(QuestionDefinition.sort_order), -1)).where(
                QuestionDefinition.category_id == category_id
            )
        )
    ) + 1
    question = QuestionDefinition(
        category_id=category_id,
        text=_clean_required(payload.text),
        guidance=_clean_optional(payload.guidance),
        sort_order=next_order,
        created_by_user_id=current_user.id,
    )
    db.add(question)
    await db.flush()

    now = datetime.now(timezone.utc)
    if not _question_tasks_enabled_at(now):
        await db.commit()
        await db.refresh(question)
        return await _question_out(db, question, current_user)

    notifications = []
    batch_date = _question_task_batch_date(now)
    task = await db.scalar(
        select(Task).where(Task.question_batch_date == batch_date)
    )
    if task is None:
        participant_users = (
            await db.execute(select(User).where(User.is_active.is_(True)).order_by(User.created_at))
        ).scalars().all()
        participants = [user for user in participant_users if _is_question_participant(user)]
        task = Task(
            title=_question_task_title(question.text),
            description=_question_task_description(question.guidance),
            assigned_to=None,
            created_by=current_user.id,
            question_origin_id=question.id,
            question_batch_date=batch_date,
            fast_task_group_id=question.id,
            status=TaskStatus.TODO.value,
            priority=TaskPriority.NORMAL.value,
            phase=ProjectPhaseStatus.MEETINGS.value,
            progress_percentage=0,
            start_date=now,
            due_date=_question_task_due_date(now),
            is_deadline_important=True,
            is_r1=True,
            is_active=True,
        )
        db.add(task)
        await db.flush()
        question.task_id = task.id
        for participant in participants:
            db.add(TaskAssignee(task_id=task.id, user_id=participant.id))
            notifications.append(
                add_notification(
                    db=db,
                    user_id=participant.id,
                    type=NotificationType.assignment,
                    title="Detyrë e re",
                    body=question.text,
                    data={"task_id": str(task.id), "question_id": str(question.id)},
                )
            )
    else:
        question.task_id = task.id
        await db.flush()
        members = await _question_task_members(db, task.id)
        _apply_question_task_content(task, members)
        task.status = TaskStatus.TODO.value
        task.completed_at = None

    await db.commit()
    for notification in notifications:
        try:
            await publish_notification(user_id=notification.user_id, notification=notification)
        except Exception:
            pass
    await db.refresh(question)
    return await _question_out(db, question, current_user)


@router.patch("/questions/{question_id}", response_model=QuestionDefinitionOut)
async def update_question_definition(
    question_id: uuid.UUID,
    payload: QuestionDefinitionUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> QuestionDefinitionOut:
    question = await _question_or_404(db, question_id)
    siblings = (
        await db.execute(
            select(QuestionDefinition)
            .where(QuestionDefinition.category_id == question.category_id)
            .order_by(QuestionDefinition.sort_order, QuestionDefinition.created_at)
        )
    ).scalars().all()
    current_index = next((index for index, item in enumerate(siblings) if item.id == question.id), None)
    target_index = min(max(payload.sort_order, 0), len(siblings) - 1)
    if current_index is not None and current_index != target_index:
        siblings[current_index].sort_order, siblings[target_index].sort_order = (
            siblings[target_index].sort_order,
            siblings[current_index].sort_order,
        )
    updated_text = _clean_required(payload.text)
    updated_guidance = _clean_optional(payload.guidance)
    content_changed = question.text != updated_text or question.guidance != updated_guidance
    question.text = updated_text
    question.guidance = updated_guidance
    if content_changed:
        question.edit_count += 1
        db.add(
            QuestionEditEvent(
                question_id=question.id,
                user_id=current_user.id,
                user_full_name=current_user.full_name,
            )
        )
        await db.execute(
            delete(QuestionUserStatus).where(QuestionUserStatus.question_id == question.id)
        )
        await db.execute(
            delete(QuestionDailySignoff).where(QuestionDailySignoff.question_id == question.id)
        )
        await _split_or_update_edited_question_task(db, question)
    await db.commit()
    await db.refresh(question)
    return await _question_out(db, question, current_user)


@router.get("/questions/{question_id}/edit-history", response_model=list[QuestionEditHistoryOut])
async def list_question_edit_history(
    question_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
) -> list[QuestionEditHistoryOut]:
    await _question_or_404(db, question_id)
    events = (
        await db.execute(
            select(QuestionEditEvent)
            .where(QuestionEditEvent.question_id == question_id)
            .order_by(QuestionEditEvent.edited_at.desc(), QuestionEditEvent.id.desc())
        )
    ).scalars().all()
    return [
        QuestionEditHistoryOut(
            id=event.id,
            user_id=event.user_id,
            full_name=event.user_full_name,
            edited_at=event.edited_at,
        )
        for event in events
    ]


@router.delete("/questions/{question_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
async def delete_question_definition(
    question_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
) -> Response:
    question = await _question_or_404(db, question_id)
    category_id = question.category_id
    await _detach_question_from_shared_task(db, question)
    await db.delete(question)
    await db.flush()
    remaining = (
        await db.execute(
            select(QuestionDefinition)
            .where(QuestionDefinition.category_id == category_id)
            .order_by(QuestionDefinition.sort_order, QuestionDefinition.created_at)
        )
    ).scalars().all()
    for index, item in enumerate(remaining):
        item.sort_order = index
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.put("/questions/{question_id}/status", response_model=QuestionStatusSummary | None)
async def update_own_question_status(
    question_id: uuid.UUID,
    payload: QuestionStatusUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> QuestionStatusSummary | None:
    question = await _question_or_404(db, question_id)
    current = await db.scalar(
        select(QuestionUserStatus).where(
            QuestionUserStatus.question_id == question_id,
            QuestionUserStatus.user_id == current_user.id,
        )
    )
    if (current.status if current else None) == payload.status:
        if current is None:
            return None
        return QuestionStatusSummary(
            user_id=current_user.id,
            full_name=current_user.full_name,
            status=current.status,
            updated_at=current.updated_at,
        )

    if payload.status is None:
        if current is not None:
            await db.delete(current)
    elif current is None:
        current = QuestionUserStatus(
            question_id=question_id,
            user_id=current_user.id,
            status=payload.status,
        )
        db.add(current)
    else:
        current.status = payload.status
        current.updated_at = func.now()

    db.add(
        QuestionStatusEvent(
            question_id=question_id,
            user_id=current_user.id,
            user_full_name=current_user.full_name,
            status=payload.status,
        )
    )
    await db.flush()
    task = await db.get(Task, question.task_id) if question.task_id is not None else None
    if task is None:
        task = await db.scalar(select(Task).where(Task.question_origin_id == question_id))
    if task is not None:
        await _refresh_question_task_status(db, task)
    await db.commit()
    if payload.status is None:
        return None
    await db.refresh(current)
    return QuestionStatusSummary(
        user_id=current_user.id,
        full_name=current_user.full_name,
        status=current.status,
        updated_at=current.updated_at,
    )


@router.put(
    "/questions/{question_id}/daily-signoff",
    response_model=QuestionDailySignoffSummary | None,
)
async def update_own_question_daily_signoff(
    question_id: uuid.UUID,
    payload: QuestionDailySignoffUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> QuestionDailySignoffSummary | None:
    if current_user.role == UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admins can monitor daily sign-offs but cannot submit one",
        )
    await _question_or_404(db, question_id)
    current = await db.scalar(
        select(QuestionDailySignoff).where(
            QuestionDailySignoff.question_id == question_id,
            QuestionDailySignoff.user_id == current_user.id,
        )
    )
    if not payload.signed:
        if current is not None:
            await db.delete(current)
            await db.commit()
        return None

    if current is None:
        current = QuestionDailySignoff(
            question_id=question_id,
            user_id=current_user.id,
        )
        db.add(current)
    else:
        current.signed_at = func.now()
    await db.commit()
    await db.refresh(current)
    return QuestionDailySignoffSummary(
        user_id=current_user.id,
        full_name=current_user.full_name,
        signed_at=current.signed_at,
    )


@router.get("/questions/{question_id}/status-history", response_model=list[QuestionStatusHistoryOut])
async def list_question_status_history(
    question_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_manager_or_admin),
) -> list[QuestionStatusHistoryOut]:
    await _question_or_404(db, question_id)
    events = (
        await db.execute(
            select(QuestionStatusEvent)
            .where(QuestionStatusEvent.question_id == question_id)
            .order_by(QuestionStatusEvent.created_at.desc(), QuestionStatusEvent.id.desc())
        )
    ).scalars().all()
    return [
        QuestionStatusHistoryOut(
            id=item.id,
            user_id=item.user_id,
            full_name=item.user_full_name,
            status=item.status,
            created_at=item.created_at,
        )
        for item in events
    ]
