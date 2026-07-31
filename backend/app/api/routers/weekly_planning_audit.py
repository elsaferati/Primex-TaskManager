from __future__ import annotations

import io
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.access import ensure_manager_or_admin
from app.api.deps import get_current_user
from app.config import settings
from app.db import get_db
from app.models.user import User
from app.models.weekly_planning_audit import WeeklyPlanningAuditRun
from app.schemas.weekly_planning_audit import (
    WeeklyPlanningAuditAbbreviationImportOut,
    WeeklyPlanningAuditGenerateIn,
    WeeklyPlanningAuditHistoryOut,
    WeeklyPlanningAuditPreviewOut,
    WeeklyPlanningAuditRunOut,
    WeeklyPlanningAuditSendIn,
    WeeklyPlanningAuditSettingsOut,
    WeeklyPlanningAuditSettingsPatch,
)
from app.services.audit import add_audit_log
from app.services.weekly_planning_audit import build_weekly_planning_audit
from app.services.weekly_planning_audit_delivery import (
    MANUAL,
    WeeklyPlanningAuditEmailError,
    generate_report_run,
    get_or_create_settings,
    send_report_run,
)


router = APIRouter()


def _ensure_report_manager(user: User) -> None:
    ensure_manager_or_admin(user)


def _run_out(run: WeeklyPlanningAuditRun) -> WeeklyPlanningAuditRunOut:
    return WeeklyPlanningAuditRunOut(
        id=run.id,
        week_start=run.week_start,
        week_end=run.week_end,
        slot=run.slot,
        generated_at=run.generated_at,
        generated_by=run.generated_by,
        trigger_type=run.trigger_type,
        status=run.status,
        included_user_count=run.included_user_count,
        excluded_leave_count=run.excluded_leave_count,
        error_count=run.error_count,
        critical_count=run.critical_count,
        high_count=run.high_count,
        filename=run.filename,
        file_checksum=run.file_checksum,
        recipients_snapshot=run.recipients_snapshot,
        subject=run.subject,
        message_id=run.message_id,
        attempt_count=run.attempt_count,
        error_message=run.error_message,
        created_at=run.created_at,
        updated_at=run.updated_at,
        download_url=(
            f"/api/reports/weekly-planning-audit/{run.id}/download"
            if run.storage_path else None
        ),
    )


@router.get("/preview", response_model=WeeklyPlanningAuditPreviewOut)
async def preview_weekly_planning_audit(
    week_start: date | None = None,
    slot: str = Query(default="09:00", pattern=r"^(09:00|09:30|10:00|10:30|11:00)$"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> WeeklyPlanningAuditPreviewOut:
    _ensure_report_manager(user)
    config = await get_or_create_settings(db)
    report = await build_weekly_planning_audit(
        db,
        week_start=week_start,
        slot=slot,
        timezone_name=config.timezone,
        abbreviation_override=config.abbreviation_dictionary,
        abbreviation_version=config.abbreviation_version,
    )
    payload = report.to_dict()
    payload.pop("abbreviations", None)
    payload.pop("abbreviation_source", None)
    payload.pop("abbreviation_updated_at", None)
    return WeeklyPlanningAuditPreviewOut(**payload)


@router.post("/generate", response_model=WeeklyPlanningAuditRunOut)
async def generate_weekly_planning_audit(
    payload: WeeklyPlanningAuditGenerateIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> WeeklyPlanningAuditRunOut:
    _ensure_report_manager(user)
    try:
        run = await generate_report_run(
            db,
            week_start=payload.week_start,
            slot=payload.slot,
            trigger_type=MANUAL,
            generated_by=user.id,
        )
        await db.commit()
        await db.refresh(run)
        return _run_out(run)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.post("/send", response_model=WeeklyPlanningAuditRunOut)
async def send_weekly_planning_audit(
    payload: WeeklyPlanningAuditSendIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> WeeklyPlanningAuditRunOut:
    _ensure_report_manager(user)
    try:
        await send_report_run(db, run_id=payload.report_run_id, requested_by=user.id, resend=False)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc
    except WeeklyPlanningAuditEmailError as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(exc)) from exc
    run = await db.get(WeeklyPlanningAuditRun, payload.report_run_id)
    assert run is not None
    return _run_out(run)


@router.post("/generate-and-send", response_model=WeeklyPlanningAuditRunOut)
async def generate_and_send_weekly_planning_audit(
    payload: WeeklyPlanningAuditGenerateIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> WeeklyPlanningAuditRunOut:
    _ensure_report_manager(user)
    try:
        run = await generate_report_run(
            db,
            week_start=payload.week_start,
            slot=payload.slot,
            trigger_type=MANUAL,
            generated_by=user.id,
        )
        await db.commit()
        await send_report_run(db, run_id=run.id, requested_by=user.id, resend=False)
        refreshed = await db.get(WeeklyPlanningAuditRun, run.id)
        assert refreshed is not None
        return _run_out(refreshed)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    except WeeklyPlanningAuditEmailError as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(exc)) from exc


@router.post("/{run_id}/resend", response_model=WeeklyPlanningAuditRunOut)
async def resend_weekly_planning_audit(
    run_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> WeeklyPlanningAuditRunOut:
    _ensure_report_manager(user)
    try:
        await send_report_run(db, run_id=run_id, requested_by=user.id, resend=True)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc
    except WeeklyPlanningAuditEmailError as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(exc)) from exc
    run = await db.get(WeeklyPlanningAuditRun, run_id)
    assert run is not None
    add_audit_log(
        db=db,
        actor_user_id=user.id,
        entity_type="weekly_planning_audit_run",
        entity_id=run.id,
        action="RESEND",
        after={"status": run.status, "message_id": run.message_id},
    )
    await db.commit()
    return _run_out(run)


@router.get("/history", response_model=WeeklyPlanningAuditHistoryOut)
async def weekly_planning_audit_history(
    week_start: date | None = None,
    slot: str | None = None,
    run_status: str | None = Query(default=None, alias="status"),
    generated_from: date | None = None,
    generated_to: date | None = None,
    limit: int = Query(default=100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> WeeklyPlanningAuditHistoryOut:
    _ensure_report_manager(user)
    query = select(WeeklyPlanningAuditRun)
    if week_start:
        query = query.where(WeeklyPlanningAuditRun.week_start == week_start)
    if slot:
        query = query.where(WeeklyPlanningAuditRun.slot == slot)
    if run_status:
        query = query.where(WeeklyPlanningAuditRun.status == run_status)
    if generated_from:
        query = query.where(WeeklyPlanningAuditRun.created_at >= datetime.combine(generated_from, datetime.min.time(), tzinfo=timezone.utc))
    if generated_to:
        query = query.where(
            WeeklyPlanningAuditRun.created_at
            < datetime.combine(generated_to, datetime.min.time(), tzinfo=timezone.utc) + timedelta(days=1)
        )
    rows = (await db.execute(
        query.order_by(WeeklyPlanningAuditRun.created_at.desc()).limit(limit)
    )).scalars().all()
    return WeeklyPlanningAuditHistoryOut(items=[_run_out(row) for row in rows])


@router.get("/{run_id}/download")
async def download_weekly_planning_audit(
    run_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> FileResponse:
    _ensure_report_manager(user)
    run = await db.get(WeeklyPlanningAuditRun, run_id)
    if run is None or not run.storage_path or not run.filename:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report file not found")
    root = Path(settings.REPORT_STORAGE_DIR).expanduser().resolve()
    path = Path(run.storage_path).resolve()
    if root not in path.parents or not path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report file not found")
    return FileResponse(
        path,
        filename=run.filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/settings", response_model=WeeklyPlanningAuditSettingsOut)
async def get_weekly_planning_audit_settings(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> WeeklyPlanningAuditSettingsOut:
    _ensure_report_manager(user)
    config = await get_or_create_settings(db)
    return WeeklyPlanningAuditSettingsOut.model_validate(config, from_attributes=True)


@router.patch("/settings", response_model=WeeklyPlanningAuditSettingsOut)
async def patch_weekly_planning_audit_settings(
    payload: WeeklyPlanningAuditSettingsPatch,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> WeeklyPlanningAuditSettingsOut:
    _ensure_report_manager(user)
    config = await get_or_create_settings(db)
    before = {
        "enabled": config.enabled,
        "timezone": config.timezone,
        "recipients_to": config.recipients_to,
        "recipients_cc": config.recipients_cc,
        "recipients_bcc": config.recipients_bcc,
        "retention_days": config.retention_days,
        "recipient_config_version": config.recipient_config_version,
    }
    changes = {
        key: value
        for key, value in payload.model_dump(exclude_unset=True).items()
        if value is not None
    }
    recipient_changed = any(
        key.startswith("recipients_")
        and [str(item) for item in value] != list(getattr(config, key) or [])
        for key, value in changes.items()
    )
    for key, value in changes.items():
        if key.startswith("recipients_"):
            value = [str(item) for item in value]
        setattr(config, key, value)
    if recipient_changed:
        config.recipient_config_version += 1
    config.updated_by = user.id
    add_audit_log(
        db=db,
        actor_user_id=user.id,
        entity_type="weekly_planning_audit_settings",
        entity_id=config.id,
        action="UPDATE",
        before=before,
        after={**changes, "recipient_config_version": config.recipient_config_version},
    )
    await db.commit()
    await db.refresh(config)
    return WeeklyPlanningAuditSettingsOut.model_validate(config, from_attributes=True)


@router.post("/abbreviations/import", response_model=WeeklyPlanningAuditAbbreviationImportOut)
async def import_weekly_planning_audit_abbreviations(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> WeeklyPlanningAuditAbbreviationImportOut:
    _ensure_report_manager(user)
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="An XLSX file is required")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="XLSX file is too large")
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid XLSX file") from exc
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The XLSX file is empty")
    headers = [str(value or "").strip().casefold() for value in rows[0]]
    try:
        abbreviation_index = headers.index("shkurtesa")
        definition_index = headers.index("definicioni")
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Required columns: Shkurtesa, Definicioni",
        ) from exc
    imported: dict[str, str] = {}
    for row in rows[1:]:
        abbreviation = str(row[abbreviation_index] or "").strip()
        definition = str(row[definition_index] or "").strip()
        if abbreviation and definition:
            imported[abbreviation] = definition
    if not imported:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No abbreviations found")
    if "RREG" in {key.upper() for key in imported}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="RREG is not an official PX abbreviation")

    config = await get_or_create_settings(db)
    before = {
        "version": config.abbreviation_version,
        "entries": config.abbreviation_dictionary,
    }
    new_version = datetime.now(ZoneInfo(config.timezone)).strftime("%Y.%m.%d.%H%M%S")
    config.abbreviation_dictionary = imported
    config.abbreviation_version = new_version
    config.updated_by = user.id
    add_audit_log(
        db=db,
        actor_user_id=user.id,
        entity_type="weekly_planning_audit_abbreviations",
        entity_id=config.id,
        action="IMPORT_XLSX",
        before=before,
        after={"version": new_version, "entry_count": len(imported)},
    )
    await db.commit()
    return WeeklyPlanningAuditAbbreviationImportOut(version=new_version, entry_count=len(imported))
