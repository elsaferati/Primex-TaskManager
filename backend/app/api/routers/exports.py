from __future__ import annotations

import csv
import io
import re
import uuid
from datetime import date, datetime, time, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.encoders import jsonable_encoder
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from sqlalchemy import and_, case, func, or_, select
from sqlalchemy.orm import aliased
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel, Field

from app.api.access import ensure_department_access, ensure_manager_or_admin, ensure_reports_access
from app.api.deps import get_current_user
from app.db import get_db
from app.models.checklist import Checklist
from app.models.checklist_item import ChecklistItem, ChecklistItemAssignee
from app.models.common_entry import CommonEntry
from app.models.meeting import Meeting
from app.models.project import Project
from app.models.department import Department
from app.models.system_task_occurrence import SystemTaskOccurrence
from app.models.system_task_template import SystemTaskTemplate
from app.models.system_task_template_alignment_role import SystemTaskTemplateAlignmentRole
from app.models.system_task_template_alignment_user import SystemTaskTemplateAlignmentUser
from app.models.task import Task
from app.models.task_assignee import TaskAssignee
from app.models.task_assignee import TaskAssignee
from app.models.task_alignment_user import TaskAlignmentUser
from app.models.task_status import TaskStatus
from app.models.task_user_comment import TaskUserComment
from app.models.user import User
from app.models.weekly_planner_snapshot import WeeklyPlannerSnapshot
from app.models.ga_note import GaNote
from app.models.daily_report_ga_entry import DailyReportGaEntry
from app.models.enums import CommonCategory, TaskStatus as TaskStatusEnum, UserRole, ChecklistItemType, SystemTaskScope, GaNoteStatus
from app.api.routers import planners as planners_router
from app.api.routers.planners import weekly_table_planner
from app.schemas.planner import WeeklyTableDepartment
from app.schemas.weekly_planner_snapshot import WeeklySnapshotType
from app.services.system_task_occurrences import OPEN, ensure_occurrences_in_range
from app.services.daily_report_logic import (
    DailyReportTyoMode,
    _as_utc_date,
    completed_on_day,
    daily_report_tyo_label,
    planned_range_for_daily_report,
    task_is_visible_to_user,
)


router = APIRouter()


class AllTasksReportRowIn(BaseModel):
    typeLabel: str = Field(default="-")
    subtype: str = Field(default="-")
    period: str = Field(default="-")
    department: str = Field(default="-")
    title: str = Field(default="-")
    status: str = Field(default="-")
    bz: str = Field(default="-")
    kohaBz: str = Field(default="-")
    tyo: str = Field(default="-")
    comment: str | None = None
    userInitials: str | None = None


class AllTasksReportExportIn(BaseModel):
    title: str = Field(default="ALL TASKS REPORT")
    usersInitials: str | None = None
    rows: list[AllTasksReportRowIn]


class ExportPreviewOut(BaseModel):
    headers: list[str]
    rows: list[list[str]]
    total: int | None = None
    truncated: bool = False


DATE_LABEL_RE = re.compile(r"Date:\s*(\d{4}-\d{2}-\d{2})", re.IGNORECASE)
DATE_RANGE_RE = re.compile(r"Date range:\s*(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})", re.IGNORECASE)
DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")
START_RE = re.compile(r"Start:\s*(\d{1,2}:\d{2})", re.IGNORECASE)
UNTIL_RE = re.compile(r"Until:\s*(\d{1,2}:\d{2})", re.IGNORECASE)
FROM_TO_RE = re.compile(r"From:\s*(\d{1,2}:\d{2})\s*-\s*To:\s*(\d{1,2}:\d{2})", re.IGNORECASE)
TIME_RANGE_RE = re.compile(r"\((\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})\)")

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


def _initials_compact(label: str | None) -> str:
    if not label:
        return ""
    parts = [part for part in re.split(r"\s+", label.strip()) if part]
    return "".join(part[0].upper() for part in parts)


def _initials_filename(label: str | None) -> str:
    """
    Initials format for filenames, e.g. "R_A" instead of "RA".
    """
    compact = _initials_compact(label)
    if not compact:
        return ""
    return "_".join(list(compact))


def _format_excel_date(d: date) -> str:
    return f"{d.day:02d}-{d.month:02d}-{d.year}"


def _format_excel_datetime(value: date | datetime | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return _format_excel_date(value.date())
    return _format_excel_date(value)


def _day_code(d: date) -> str:
    codes = ["H", "M", "MR", "E", "P", "S", "D"]
    return codes[d.weekday()] if 0 <= d.weekday() < len(codes) else ""


def _strip_html(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"<[^>]+>", "", value).strip()


def _strip_html_keep_breaks(value: str | None) -> str:
    if not value:
        return ""
    text = re.sub(r"(?i)<br\s*/?>", "\n", value)
    text = re.sub(r"(?i)</p\s*>", "\n", text)
    text = re.sub(r"(?i)</div\s*>", "\n", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def _priority_label(value: str | None) -> str:
    normalized = value.upper() if value else "NORMAL"
    return "H" if normalized == "HIGH" else "N"


def _frequency_label(value: str | None) -> str:
    normalized = value.upper() if value else ""
    return {
        "DAILY": "D",
        "WEEKLY": "W",
        "MONTHLY": "M",
        "YEARLY": "Y",
        "3_MONTHS": "3M",
        "6_MONTHS": "6M",
    }.get(normalized, value or "")


_WEEKDAY_LABELS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
_MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _month_cycle_labels(start_month: int, interval: int) -> list[str]:
    if start_month < 1 or start_month > 12 or interval <= 0:
        return []
    steps = 12 // interval
    labels: list[str] = []
    for i in range(steps):
        month = ((start_month - 1 + (interval * i)) % 12) + 1
        labels.append(_MONTH_LABELS[month - 1])
    return labels


def _system_task_schedule_label(template: SystemTaskTemplate) -> str:
    frequency = (template.frequency or "").upper()

    if frequency == "WEEKLY":
        raw_days = list(template.days_of_week or [])
        if not raw_days and template.day_of_week is not None:
            raw_days = [template.day_of_week]
        day_values: set[int] = set()
        for raw in raw_days:
            try:
                day_value = int(raw)
            except (TypeError, ValueError):
                continue
            if 0 <= day_value <= 6:
                day_values.add(day_value)
        if not day_values:
            return ""
        return ", ".join(_WEEKDAY_LABELS[day] for day in sorted(day_values))

    if frequency == "DAILY":
        return "everyday"

    if frequency in {"MONTHLY", "3_MONTHS", "6_MONTHS"}:
        day_of_month = template.day_of_month
        if day_of_month is None:
            return ""
        if day_of_month == 0:
            day_label = "end_of_month"
        elif day_of_month == -1:
            day_label = "first_working_day"
        elif day_of_month > 0:
            day_label = f"{day_of_month:02d}"
        else:
            day_label = str(day_of_month)

        if frequency == "MONTHLY":
            return day_label

        interval = 3 if frequency == "3_MONTHS" else 6
        start_month = template.month_of_year
        month_labels = _month_cycle_labels(start_month or 0, interval) if start_month else []
        if month_labels:
            return f"{day_label} {', '.join(month_labels)}"
        return day_label

    if frequency == "YEARLY":
        month_label = ""
        month_of_year = template.month_of_year
        if month_of_year is not None and 1 <= month_of_year <= 12:
            month_label = _MONTH_LABELS[month_of_year - 1]

        day_label = ""
        day_of_month = template.day_of_month
        if day_of_month is not None:
            if day_of_month == 0:
                day_label = "end_of_month"
            elif day_of_month == -1:
                day_label = "first_working_day"
            elif day_of_month > 0:
                day_label = f"{day_of_month:02d}"
            else:
                day_label = str(day_of_month)

        if day_label and month_label:
            return f"{day_label} {month_label}"
        return day_label or month_label

    return ""


def _department_short(label: str) -> str:
    key = label.strip().upper()
    return {
        "DEVELOPMENT": "DEV",
        "GRAPHIC DESIGN": "GDS",
        "PRODUCT CONTENT": "PCM",
        "PROJECT CONTENT MANAGER": "PCM",
    }.get(key, label)


def _display_department_name(label: str | None) -> str:
    if not label:
        return ""
    return "Product Content" if label.strip() == "Project Content Manager" else label


def _safe_filename(label: str) -> str:
    cleaned = re.sub(r"[^\w\s\-]", "", label)
    cleaned = re.sub(r"\s+", "_", cleaned.strip())
    return cleaned.upper() or "EXPORT"


def _safe_filename_spaces(label: str) -> str:
    cleaned = re.sub(r"[^\w\s\-]", "", label)
    cleaned = re.sub(r"\s+", " ", cleaned.strip())
    return cleaned.upper() or "EXPORT"


def _fast_task_badge(task) -> str:
    if task.is_bllok:
        return "BLL"
    if task.is_r1:
        return "R1"
    if task.is_1h_report:
        return "1H"
    if task.ga_note_origin_id:
        return "GA"
    if task.is_personal:
        return "P:"
    if task.fast_task_type:
        return task.fast_task_type
    return ""


def _normalize_preview_value(value: object | None) -> str:
    if value is None:
        return "-"
    if isinstance(value, str):
        return value if value.strip() else "-"
    return str(value)


def _apply_preview_limit(rows: list[list[str]], limit: int | None) -> tuple[list[list[str]], int, bool]:
    total = len(rows)
    if not limit or limit <= 0 or total <= limit:
        return rows, total, False
    return rows[:limit], total, True


def _planner_cell_items(*, projects, system_tasks, fast_tasks, include_fast: bool, day_date: date) -> list[dict[str, str | None]]:
    """
    Returns list of task display items with status for coloring.
    Each item has: number, title, rest, status.
    """
    items: list[dict[str, str | None]] = []
    task_num = 1

    for project in projects:
        project_title = project.project_title or ""
        for task in project.tasks or []:
            task_title = task.task_title or ""
            task_name = task_title
            if task.daily_products is not None:
                task_name = f"{task_name} {task.daily_products} pcs"
            status_value = _effective_status(task.status, task.completed_at, day_date)
            if project_title:
                items.append(
                    {
                        "number": str(task_num),
                        "title": project_title,
                        "rest": task_name,
                        "status": status_value,
                    }
                )
            else:
                items.append(
                    {
                        "number": str(task_num),
                        "title": task_name,
                        "rest": "",
                        "status": status_value,
                    }
                )
            task_num += 1

    for task in system_tasks:
        title = task.title or ""
        if title:
            items.append(
                {
                    "number": str(task_num),
                    "title": title,
                    "rest": "",
                    "status": _effective_status(task.status, task.completed_at, day_date),
                }
            )
            task_num += 1

    if include_fast:
        for task in fast_tasks:
            label = task.title or ""
            badge = _fast_task_badge(task)
            if badge:
                label = f"{label} [{badge}]"
            if label:
                items.append(
                    {
                        "number": str(task_num),
                        "title": label,
                        "rest": "",
                        "status": _effective_status(task.status, task.completed_at, day_date),
                    }
                )
                task_num += 1

    return items


def _create_rich_text_cell(lines: list[tuple[str, str, str]]) -> CellRichText | str:
    """
    Creates a RichText cell with bold titles, or returns plain string if no lines.
    Format: "1. TITLE: REST" where TITLE is bold.
    """
    if not lines:
        return ""

    rich_text_parts: list[TextBlock] = []

    for idx, (number_part, title_part, rest_part) in enumerate(lines):
        prefix = "\n" if idx > 0 else ""
        # Add number and period
        rich_text_parts.append(TextBlock(text=f"{prefix}{number_part}. ", font=InlineFont()))

        # Add bold title
        rich_text_parts.append(TextBlock(text=title_part, font=InlineFont(b=True)))

        # Add colon and rest if there is rest
        if rest_part:
            rich_text_parts.append(TextBlock(text=f": {rest_part}", font=InlineFont()))

    return CellRichText(rich_text_parts)


def _planner_item_rich_text(item: dict[str, str | None]) -> CellRichText | str:
    number = item.get("number") or ""
    title = item.get("title") or ""
    rest = item.get("rest") or ""
    return _create_rich_text_cell([(number, title, rest)])


def _effective_status(status: str | None, completed_at: datetime | None, day_date: date) -> str:
    normalized = (status or "TODO").upper()
    if normalized != "DONE":
        return normalized
    if not completed_at:
        return "IN_PROGRESS"
    return "DONE" if completed_at.date() == day_date else "IN_PROGRESS"


def _scope_label(template: SystemTaskTemplate) -> str:
    if template.scope:
        return template.scope
    return "DEPARTMENT" if template.department_id else "ALL"


def _parse_internal_notes(notes: str | None) -> dict[str, str]:
    if not notes:
        return {}
    values: dict[str, str] = {}
    for raw in notes.splitlines():
        if ":" not in raw:
            continue
        key, value = raw.split(":", 1)
        key = key.strip().upper()
        value = value.strip()
        if key and value:
            values[key] = value
    return values


async def _assignees_for_tasks(db: AsyncSession, task_ids: list[uuid.UUID]) -> dict[uuid.UUID, list[str]]:
    if not task_ids:
        return {}
    rows = (
        await db.execute(
            select(TaskAssignee.task_id, User)
            .join(User, TaskAssignee.user_id == User.id)
            .where(TaskAssignee.task_id.in_(task_ids))
        )
    ).all()
    out: dict[uuid.UUID, list[str]] = {task_id: [] for task_id in task_ids}
    for task_id, user in rows:
        label = user.full_name or user.username or ""
        if not label:
            continue
        out.setdefault(task_id, []).append(label)
    return out


async def _query_tasks(
    *,
    db: AsyncSession,
    user,
    department_id: uuid.UUID | None,
    user_id: uuid.UUID | None,
    project_id: uuid.UUID | None,
    status_id: uuid.UUID | None,
    planned_from: date | None,
    planned_to: date | None,
) -> list[Task]:
    stmt = select(Task)

    role_value = getattr(user.role, "value", None)
    is_admin = user.role == UserRole.ADMIN or (isinstance(role_value, str) and role_value.upper() == "ADMIN")

    if not is_admin:
        if user.department_id is None:
            return []
        stmt = stmt.where(Task.department_id == user.department_id)

    if department_id:
        ensure_department_access(user, department_id)
        stmt = stmt.where(Task.department_id == department_id)
    if user_id:
        if hasattr(Task, "assigned_to_user_id"):
            stmt = stmt.where(Task.assigned_to_user_id == user_id)
        else:
            stmt = stmt.where(Task.assigned_to == user_id)
    if project_id:
        stmt = stmt.where(Task.project_id == project_id)
    if status_id:
        if hasattr(Task, "status_id"):
            stmt = stmt.where(Task.status_id == status_id)
        else:
            status_row = (await db.execute(select(TaskStatus).where(TaskStatus.id == status_id))).scalar_one_or_none()
            if status_row is not None:
                stmt = stmt.where(Task.status == status_row.name)
            else:
                return []

    planned_expr = None
    if hasattr(Task, "planned_for"):
        planned_expr = Task.planned_for
    start_expr = func.date(Task.start_date) if hasattr(Task, "start_date") else None
    due_expr = func.date(Task.due_date) if hasattr(Task, "due_date") else None
    exprs = [e for e in (planned_expr, start_expr, due_expr) if e is not None]
    if exprs:
        planned_expr = exprs[0] if len(exprs) == 1 else func.coalesce(*exprs)
        if planned_from:
            stmt = stmt.where(planned_expr >= planned_from)
        if planned_to:
            stmt = stmt.where(planned_expr <= planned_to)

    return (await db.execute(stmt.order_by(Task.created_at.desc()))).scalars().all()


async def _maps(db: AsyncSession, tasks: list[Task]) -> tuple[dict[uuid.UUID, str], dict[uuid.UUID, str]]:
    status_ids = {getattr(t, "status_id", None) for t in tasks}
    status_ids.discard(None)
    user_ids = {getattr(t, "assigned_to_user_id", None) for t in tasks}
    user_ids = {uid for uid in user_ids if uid is not None}
    if not user_ids:
        user_ids = {getattr(t, "assigned_to", None) for t in tasks}
        user_ids = {uid for uid in user_ids if uid is not None}
    status_map: dict[uuid.UUID, str] = {}
    user_map: dict[uuid.UUID, str] = {}

    if status_ids:
        statuses = (await db.execute(select(TaskStatus).where(TaskStatus.id.in_(status_ids)))).scalars().all()
        status_map = {s.id: s.name for s in statuses}
    if user_ids:
        users = (await db.execute(select(User).where(User.id.in_(user_ids)))).scalars().all()
        user_map = {u.id: (u.full_name or u.username) for u in users}
    return status_map, user_map


def _task_rows(tasks: list[Task], status_map: dict[uuid.UUID, str], user_map: dict[uuid.UUID, str]) -> list[list[str]]:
    rows: list[list[str]] = []
    for t in tasks:
        planned_value = getattr(t, "planned_for", None)
        if planned_value is None:
            planned_value = getattr(t, "start_date", None) or getattr(t, "due_date", None)
            if isinstance(planned_value, datetime):
                planned_value = planned_value.date()
        status_value = ""
        status_id = getattr(t, "status_id", None)
        if status_id:
            status_value = status_map.get(status_id, "")
        else:
            status_value = getattr(t, "status", "") or ""
        assigned_id = getattr(t, "assigned_to_user_id", None) or getattr(t, "assigned_to", None)
        is_carried_over = bool(getattr(t, "is_carried_over", False))
        carried_over_from = getattr(t, "carried_over_from", None)
        reminder_enabled = bool(getattr(t, "reminder_enabled", False))
        rows.append(
            [
                str(t.id),
                t.title,
                t.description or "",
                getattr(getattr(t, "task_type", None), "value", None) or getattr(t, "task_type", "") or "",
                status_value,
                str(t.department_id),
                str(t.project_id),
                user_map.get(assigned_id, "") if assigned_id else "",
                planned_value.isoformat() if planned_value else "",
                "yes" if is_carried_over else "no",
                carried_over_from.isoformat() if carried_over_from else "",
                "yes" if reminder_enabled else "no",
                t.created_at.isoformat() if t.created_at else "",
                t.completed_at.isoformat() if t.completed_at else "",
            ]
        )
    return rows


def _initials(label: str | None) -> str:
    # Backwards-compatible alias used throughout this module (compact initials like "DV").
    return _initials_compact(label)


def _resolve_period(finish_period: str | None, date_value: date | datetime | None) -> str:
    if finish_period in {"AM", "PM"}:
        return finish_period
    if date_value is None:
        return "AM"
    if isinstance(date_value, datetime):
        return "PM" if date_value.hour >= 12 else "AM"
    return "AM"


def _is_vs_vl_project_label(label: str | None) -> bool:
    if not label:
        return False
    upper = label.upper()
    return "VS" in upper or "VL" in upper


def _planned_range_for_task(task: Task) -> tuple[date | None, date | None]:
    if task.due_date is None:
        return None, None
    due = task.due_date.date()
    if task.start_date is not None:
        start = task.start_date.date()
        if start <= due:
            return start, due
    return due, due


async def _user_comments_for_tasks(
    db: AsyncSession, task_ids: list[uuid.UUID], user_id: uuid.UUID
) -> dict[uuid.UUID, str | None]:
    if not task_ids:
        return {}
    rows = (
        await db.execute(
            select(TaskUserComment.task_id, TaskUserComment.comment)
            .where(TaskUserComment.task_id.in_(task_ids))
            .where(TaskUserComment.user_id == user_id)
        )
    ).all()
    return {task_id: comment for task_id, comment in rows}


async def _alignment_maps_for_templates(
    db: AsyncSession, template_ids: list[uuid.UUID]
) -> tuple[dict[uuid.UUID, list[str]], dict[uuid.UUID, list[uuid.UUID]]]:
    if not template_ids:
        return {}, {}
    role_rows = (
        await db.execute(
            select(SystemTaskTemplateAlignmentRole.template_id, SystemTaskTemplateAlignmentRole.role)
            .where(SystemTaskTemplateAlignmentRole.template_id.in_(template_ids))
        )
    ).all()
    roles_map: dict[uuid.UUID, list[str]] = {}
    for tid, role in role_rows:
        roles_map.setdefault(tid, []).append(role)

    alignment_user_rows = (
        await db.execute(
            select(SystemTaskTemplateAlignmentUser.template_id, SystemTaskTemplateAlignmentUser.user_id)
            .where(SystemTaskTemplateAlignmentUser.template_id.in_(template_ids))
        )
    ).all()
    alignment_users_map: dict[uuid.UUID, list[uuid.UUID]] = {}
    for tid, uid in alignment_user_rows:
        alignment_users_map.setdefault(tid, []).append(uid)

    return roles_map, alignment_users_map


def _format_task_status(status: str | None) -> str:
    if not status:
        return "-"
    if status == "IN_PROGRESS":
        return "In Progress"
    if status == "TODO":
        return "To Do"
    if status == "DONE":
        return "Done"
    return status


def _format_system_status(status: str | None) -> str:
    if not status:
        return "-"
    if status == "NOT_DONE":
        return "Not Done"
    if status == "DONE":
        return "Done"
    if status == "OPEN":
        return "Open"
    if status == "SKIPPED":
        return "Skipped"
    return status


def _normalize_report_status(status: str | None) -> str:
    if not status:
        return ""
    normalized = status.strip().upper().replace(" ", "_")
    if normalized == "TO_DO":
        return "TODO"
    if normalized == "INPROGRESS":
        return "IN_PROGRESS"
    return normalized


def _format_alignment_time(value: time | None) -> str:
    if not value:
        return "-"
    return f"{value.hour:02d}:{value.minute:02d}"


def _tyo_label(base_date: date | None, completed_at: date | None, today: date) -> str:
    if completed_at and completed_at == today:
        return "T"
    if base_date is None:
        return "-"
    if base_date == today:
        return "T"
    delta = (today - base_date).days
    if delta == 1:
        return "Y"
    if delta > 1:
        return str(delta)
    return "-"


def _no_project_type_label(task: Task) -> str:
    if task.is_bllok:
        return "BLLOK"
    if task.is_1h_report:
        return "1H"
    if task.is_r1:
        return "R1"
    if task.is_personal:
        return "Personal"
    if task.ga_note_origin_id:
        return "GA"
    return "Normal"


def _fast_subtype_short(task: Task) -> str:
    base = _no_project_type_label(task)
    if base == "BLLOK":
        return "BLL"
    if base == "Personal":
        return "P:"
    if base == "Normal":
        return "N"
    return base


def _system_frequency_short_label(freq: str | None) -> str:
    if not freq:
        return "-"
    mapping = {
        "DAILY": "D",
        "WEEKLY": "W",
        "MONTHLY": "M",
        "YEARLY": "Y",
        "3_MONTHS": "3M",
        "6_MONTHS": "6M",
    }
    return mapping.get(freq, str(freq))


EXPORT_HEADERS = [
    "id",
    "title",
    "description",
    "task_type",
    "status",
    "department_id",
    "project_id",
    "assigned_to",
    "planned_for",
    "is_carried_over",
    "carried_over_from",
    "reminder_enabled",
    "created_at",
    "completed_at",
]


@router.get("/tasks.csv")
async def export_tasks_csv(
    department_id: uuid.UUID | None = None,
    user_id: uuid.UUID | None = None,
    project_id: uuid.UUID | None = None,
    status_id: uuid.UUID | None = None,
    planned_from: date | None = None,
    planned_to: date | None = None,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    ensure_reports_access(user)
    tasks = await _query_tasks(
        db=db,
        user=user,
        department_id=department_id,
        user_id=user_id,
        project_id=project_id,
        status_id=status_id,
        planned_from=planned_from,
        planned_to=planned_to,
    )
    status_map, user_map = await _maps(db, tasks)
    rows = _task_rows(tasks, status_map, user_map)

    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(EXPORT_HEADERS)
    writer.writerows(rows)
    stream.seek(0)
    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="tasks_export.csv"'},
    )


@router.get("/tasks.xlsx")
async def export_tasks_xlsx(
    department_id: uuid.UUID | None = None,
    user_id: uuid.UUID | None = None,
    project_id: uuid.UUID | None = None,
    status_id: uuid.UUID | None = None,
    planned_from: date | None = None,
    planned_to: date | None = None,
    title: str | None = None,
    standard: bool = False,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    ensure_reports_access(user)
    tasks = await _query_tasks(
        db=db,
        user=user,
        department_id=department_id,
        user_id=user_id,
        project_id=project_id,
        status_id=status_id,
        planned_from=planned_from,
        planned_to=planned_to,
    )
    status_map, user_map = await _maps(db, tasks)
    rows = _task_rows(tasks, status_map, user_map)

    wb = Workbook()
    ws = wb.active

    if standard:
        # Standard printable export (Admin Tasks requirement)
        std_title_raw = title or "TASKS"
        std_title = _safe_filename(std_title_raw)
        headers = ["NR", "TASK TITLE", "STATUS", "PRIORITY", "DUE DATE", "PUNOI"]

        ws.title = (std_title_raw[:31] if std_title_raw else "Tasks")[:31]

        last_col = len(headers)

        # Row 1: Title (merged across all columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        title_cell = ws.cell(row=1, column=1, value=std_title)
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=1)

        # Row 2: (blank / spacing)
        # Row 3: (blank / spacing)

        header_row = 4
        data_row = header_row + 1

        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header.upper())
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="left",
                vertical="bottom",
                wrap_text=True if header == "NR" else False,
                readingOrder=1,
            )

        status_fills = {
            "TODO": PatternFill(start_color="FFC4ED", end_color="FFC4ED", fill_type="solid"),
            "IN_PROGRESS": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            "DONE": PatternFill(start_color="C4FDC4", end_color="C4FDC4", fill_type="solid"),
        }

        # Data (simple table)
        # Use the same task list but present human-friendly columns
        # NOTE: This uses attributes already referenced elsewhere in this file, to avoid schema drift.
        for idx, t in enumerate(tasks, start=1):
            status_label = status_map.get(t.status_id, "") if hasattr(t, "status_id") else getattr(t, "status", "") or ""
            priority_value = getattr(t, "priority", "") or ""
            due_value = ""
            due_dt = getattr(t, "due_date", None)
            if due_dt:
                try:
                    due_value = due_dt.date().isoformat()
                except Exception:
                    due_value = str(due_dt)
            assignee_name = ""
            assignee_id = getattr(t, "assigned_to_user_id", None) or getattr(t, "assigned_to", None)
            if assignee_id:
                assignee_name = user_map.get(assignee_id, "") or ""
            row_values = [
                idx,
                getattr(t, "title", "") or "",
                status_label,
                str(priority_value).upper() if priority_value else "",
                due_value,
                _initials_compact(assignee_name),
            ]
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=data_row, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="bottom", readingOrder=1, wrap_text=False)
                if col_idx == 1:
                    cell.font = Font(bold=True)
                if col_idx == 3:
                    normalized_status = (str(value) if value is not None else "").upper().replace(" ", "_")
                    fill = status_fills.get(normalized_status)
                    if fill:
                        cell.fill = fill
            data_row += 1

        last_row = data_row - 1

        # Column widths
        widths = {
            1: 5,   # NR
            2: 46,  # TASK TITLE
            3: 14,  # STATUS
            4: 12,  # PRIORITY
            5: 14,  # DUE DATE
            6: 10,  # PUNOI
        }
        for col_idx in range(1, last_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 16)

        # Filters for all columns
        if last_row >= header_row:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"

        # Freeze: column A (NR) + header row (row 4)
        ws.freeze_panes = "B5"

        # Repeat header row on each printed page
        ws.print_title_rows = f"{header_row}:{header_row}"

        # Page setup + margins (exact values provided)
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.1
        ws.page_margins.right = 0.1
        ws.page_margins.top = 0.36
        ws.page_margins.bottom = 0.51
        ws.page_margins.header = 0.15
        ws.page_margins.footer = 0.2

        # Header/footer: date+time top-right, page x/y center bottom, initials bottom-right
        ws.oddHeader.right.text = "&D &T"
        ws.oddFooter.center.text = "Page &P / &N"
        ws.oddFooter.right.text = f"PUNOI: {_initials_compact(user.full_name or user.username or '') or '____'}"
        ws.evenHeader.right.text = ws.oddHeader.right.text
        ws.evenFooter.center.text = ws.oddFooter.center.text
        ws.evenFooter.right.text = ws.oddFooter.right.text
        ws.firstHeader.right.text = ws.oddHeader.right.text
        ws.firstFooter.center.text = ws.oddFooter.center.text
        ws.firstFooter.right.text = ws.oddFooter.right.text

        # Borders: thick outside, thin inside (entire table)
        thin = Side(style="thin", color="000000")
        thick = Side(style="medium", color="000000")
        if last_row >= header_row:
            for r in range(header_row, last_row + 1):
                for c in range(1, last_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = Border(
                        left=thick if c == 1 else thin,
                        right=thick if c == last_col else thin,
                        top=thick if r == header_row else thin,
                        bottom=thick if r == last_row else thin,
                    )

        # Ensure all cells bottom-aligned (including blanks within used range)
        for r in range(1, last_row + 1):
            for c in range(1, last_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.alignment is None:
                    cell.alignment = Alignment(horizontal="left", vertical="bottom", readingOrder=1)
                else:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal or "left",
                        vertical="bottom",
                        wrap_text=cell.alignment.wrap_text,
                        readingOrder=1,
                    )
    else:
        ws.title = "Tasks"
        ws.append(EXPORT_HEADERS)
        for row in rows:
            ws.append(row)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    today = datetime.now(timezone.utc).date()
    filename_date = f"{today.day:02d}_{today.month:02d}_{str(today.year)[-2:]}"
    user_initials = _initials_filename(user.full_name or user.username or "") or "U_S_E_R"
    if standard:
        file_title = _safe_filename(title or "TASKS")
        filename = f"{file_title} {filename_date}_{user_initials}.xlsx".replace("__", "_")
    else:
        filename = f"TASKS_{filename_date}_{_initials_compact(user.full_name or user.username or '') or 'USER'}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/project-tasks-preview", response_model=ExportPreviewOut)
async def preview_project_tasks(
    department_id: uuid.UUID | None = None,
    user_id: uuid.UUID | None = None,
    project_id: uuid.UUID | None = None,
    status_id: uuid.UUID | None = None,
    planned_from: date | None = None,
    planned_to: date | None = None,
    limit: int | None = 200,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    ensure_reports_access(user)
    tasks = await _query_tasks(
        db=db,
        user=user,
        department_id=department_id,
        user_id=user_id,
        project_id=project_id,
        status_id=status_id,
        planned_from=planned_from,
        planned_to=planned_to,
    )
    status_map, user_map = await _maps(db, tasks)
    rows: list[list[str]] = []
    for idx, t in enumerate(tasks, start=1):
        status_label = status_map.get(t.status_id, "") if hasattr(t, "status_id") else getattr(t, "status", "") or ""
        priority_value = getattr(t, "priority", "") or ""
        due_value = ""
        due_dt = getattr(t, "due_date", None)
        if due_dt:
            try:
                due_value = due_dt.date().isoformat()
            except Exception:
                due_value = str(due_dt)
        assignee_name = ""
        assignee_id = getattr(t, "assigned_to_user_id", None) or getattr(t, "assigned_to", None)
        if assignee_id:
            assignee_name = user_map.get(assignee_id, "") or ""
        row_values = [
            str(idx),
            _normalize_preview_value(getattr(t, "title", "") or ""),
            _normalize_preview_value(status_label),
            _normalize_preview_value(str(priority_value).upper() if priority_value else ""),
            _normalize_preview_value(due_value),
            _normalize_preview_value(_initials_compact(assignee_name)),
        ]
        rows.append(row_values)

    headers = ["NR", "TASK TITLE", "STATUS", "PRIORITY", "DUE DATE", "PUNOI"]
    limited_rows, total, truncated = _apply_preview_limit(rows, limit)
    return ExportPreviewOut(headers=headers, rows=limited_rows, total=total, truncated=truncated)


@router.get("/fast-tasks.xlsx")
async def export_fast_tasks_xlsx(
    department_id: uuid.UUID | None = None,
    user_id: uuid.UUID | None = None,
    status_id: uuid.UUID | None = None,
    planned_from: date | None = None,
    planned_to: date | None = None,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    ensure_reports_access(user)

    is_admin = user.role == UserRole.ADMIN or (getattr(user.role, "value", "").lower() == "admin")

    if not is_admin and department_id is None:
        department_id = user.department_id

    assigned_user = aliased(User)
    created_user = aliased(User)
    status_name: str | None = None
    if status_id is not None:
        status_row = (await db.execute(select(TaskStatus).where(TaskStatus.id == status_id))).scalar_one_or_none()
        if status_row is not None:
            status_name = status_row.name

    stmt = (
        select(Task, Department, assigned_user, created_user)
        .outerjoin(Department, Task.department_id == Department.id)
        .outerjoin(assigned_user, Task.assigned_to == assigned_user.id)
        .outerjoin(created_user, Task.created_by == created_user.id)
        .where(
            or_(
                Task.is_bllok.is_(True),
                Task.is_1h_report.is_(True),
                Task.is_personal.is_(True),
                Task.is_r1.is_(True),
                Task.fast_task_group_id.is_not(None),
            )
        )
        .where(Task.system_template_origin_id.is_(None))
        .where(Task.project_id.is_(None))
        .where(Task.internal_notes.is_(None))
        .where(Task.status != "DONE")
    )

    if not is_admin:
        if user.department_id is None:
            rows = []
        else:
            stmt = stmt.where(Task.department_id == user.department_id)
    else:
        if department_id:
            ensure_department_access(user, department_id)
            stmt = stmt.where(Task.department_id == department_id)

    if user_id:
        stmt = stmt.where(Task.assigned_to == user_id)
    if status_name:
        stmt = stmt.where(Task.status == status_name)
    if planned_from:
        stmt = stmt.where(func.date(Task.start_date) >= planned_from)
    if planned_to:
        stmt = stmt.where(func.date(Task.start_date) <= planned_to)

    sort_key = case(
        (Task.is_bllok.is_(True), 1),
        (Task.is_1h_report.is_(True), 2),
        (Task.is_r1.is_(True), 3),
        (Task.is_personal.is_(True), 4),
        else_=5,
    )

    if not (not is_admin and user.department_id is None):
        rows = (await db.execute(stmt.order_by(sort_key.asc(), Task.created_at.desc()))).all()

    headers = [
        "NR",
        "TITLE",
        "DESCRIPTION",
        "DEP",
        "PUNOI",
        "CREATED BY",
        "GA NOTE ORIGIN",
        "STATUS",
        "PRIORITY",
        "CREATION DATE",
        "START DATE",
        "DUE DATE",
    ]

    wb = Workbook()
    ws = wb.active

    today = datetime.now(timezone.utc).date()
    filename_date = f"{today.day:02d}.{today.month:02d}.{str(today.year)[-2:]}"
    initials_value = _initials_compact(user.full_name or user.username or "") or "USER"
    title_text = f"FAST TASKS_{filename_date}_{initials_value}"
    ws.title = "Fast Tasks"

    last_col = len(headers)

    # Row 1: Title (merged)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=1)

    header_row = 4
    data_row = header_row + 1

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header.upper())
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

    status_fills = {
        "TODO": PatternFill(start_color="FFC4ED", end_color="FFC4ED", fill_type="solid"),
        "IN_PROGRESS": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "DONE": PatternFill(start_color="C4FDC4", end_color="C4FDC4", fill_type="solid"),
    }

    for idx, (task, department, assigned, created) in enumerate(rows, start=1):
        values = [
            idx,
            task.title or "",
            _strip_html(task.description) or "",
            (department.code if department else ""),
            _initials_compact((assigned.full_name or assigned.username) if assigned else ""),
            _initials_compact((created.full_name or created.username) if created else ""),
            "Yes" if task.ga_note_origin_id else "No",
            task.status or "",
            (task.priority or "").upper(),
            _format_excel_datetime(task.created_at),
            _format_excel_datetime(task.start_date),
            _format_excel_datetime(task.due_date),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
            if col_idx == 1:
                cell.font = Font(bold=True)
            if col_idx == 8:
                normalized_status = (str(value) if value is not None else "").upper().replace(" ", "_")
                fill = status_fills.get(normalized_status)
                if fill:
                    cell.fill = fill
        data_row += 1

    last_row = data_row - 1

    widths = {
        1: 5,    # NR
        2: 36,   # TITLE
        3: 44,   # DESCRIPTION
        4: 8,    # DEP
        5: 6,    # PUNOI
        6: 8,    # CREATED BY
        7: 12,   # GA NOTE ORIGIN
        8: 12,   # STATUS
        9: 10,   # PRIORITY
        10: 16,  # CREATION DATE
        11: 14,  # START DATE
        12: 14,  # DUE DATE
    }
    for col_idx in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 16)

    if last_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"

    ws.freeze_panes = "B5"
    ws.print_title_rows = f"{header_row}:{header_row}"

    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2

    ws.oddHeader.right.text = "&D &T"
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.right.text = initials_value
    ws.evenHeader.right.text = ws.oddHeader.right.text
    ws.evenFooter.center.text = ws.oddFooter.center.text
    ws.evenFooter.right.text = ws.oddFooter.right.text
    ws.firstHeader.right.text = ws.oddHeader.right.text
    ws.firstFooter.center.text = ws.oddFooter.center.text
    ws.firstFooter.right.text = ws.oddFooter.right.text

    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    for r in range(header_row, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=thick if c == 1 else thin,
                right=thick if c == last_col else thin,
                top=thick if r == header_row else thin,
                bottom=thick if r == last_row else thin,
            )

    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.alignment is None:
                cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
            else:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "left",
                    vertical="bottom",
                    wrap_text=True,
                    readingOrder=1,
                )

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"{title_text}.xlsx".replace("__", "_")
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/fast-tasks-preview", response_model=ExportPreviewOut)
async def preview_fast_tasks(
    department_id: uuid.UUID | None = None,
    user_id: uuid.UUID | None = None,
    status_id: uuid.UUID | None = None,
    planned_from: date | None = None,
    planned_to: date | None = None,
    limit: int | None = 200,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    ensure_reports_access(user)

    is_admin = user.role == UserRole.ADMIN or (getattr(user.role, "value", "").lower() == "admin")

    if not is_admin and department_id is None:
        department_id = user.department_id

    assigned_user = aliased(User)
    created_user = aliased(User)
    status_name: str | None = None
    if status_id is not None:
        status_row = (await db.execute(select(TaskStatus).where(TaskStatus.id == status_id))).scalar_one_or_none()
        if status_row is not None:
            status_name = status_row.name

    stmt = (
        select(Task, Department, assigned_user, created_user)
        .outerjoin(Department, Task.department_id == Department.id)
        .outerjoin(assigned_user, Task.assigned_to == assigned_user.id)
        .outerjoin(created_user, Task.created_by == created_user.id)
        .where(
            or_(
                Task.is_bllok.is_(True),
                Task.is_1h_report.is_(True),
                Task.is_personal.is_(True),
                Task.is_r1.is_(True),
                Task.fast_task_group_id.is_not(None),
            )
        )
        .where(Task.system_template_origin_id.is_(None))
        .where(Task.project_id.is_(None))
        .where(Task.internal_notes.is_(None))
        .where(Task.status != "DONE")
    )

    if not is_admin:
        if user.department_id is None:
            rows: list[list[str]] = []
            headers = [
                "NR",
                "TITLE",
                "DESCRIPTION",
                "DEP",
                "PUNOI",
                "CREATED BY",
                "GA NOTE ORIGIN",
                "STATUS",
                "PRIORITY",
                "CREATION DATE",
                "START DATE",
                "DUE DATE",
            ]
            return ExportPreviewOut(headers=headers, rows=rows, total=0, truncated=False)
        stmt = stmt.where(Task.department_id == user.department_id)
    else:
        if department_id:
            ensure_department_access(user, department_id)
            stmt = stmt.where(Task.department_id == department_id)

    if user_id:
        stmt = stmt.where(Task.assigned_to == user_id)
    if status_name:
        stmt = stmt.where(Task.status == status_name)
    if planned_from:
        stmt = stmt.where(func.date(Task.start_date) >= planned_from)
    if planned_to:
        stmt = stmt.where(func.date(Task.start_date) <= planned_to)

    sort_key = case(
        (Task.is_bllok.is_(True), 1),
        (Task.is_1h_report.is_(True), 2),
        (Task.is_r1.is_(True), 3),
        (Task.is_personal.is_(True), 4),
        else_=5,
    )

    rows_raw = (await db.execute(stmt.order_by(sort_key.asc(), Task.created_at.desc()))).all()
    rows: list[list[str]] = []
    for idx, (task, department, assigned, created) in enumerate(rows_raw, start=1):
        values = [
            str(idx),
            _normalize_preview_value(task.title or ""),
            _normalize_preview_value(_strip_html(task.description) or ""),
            _normalize_preview_value(department.code if department else ""),
            _normalize_preview_value(_initials_compact((assigned.full_name or assigned.username) if assigned else "")),
            _normalize_preview_value(_initials_compact((created.full_name or created.username) if created else "")),
            _normalize_preview_value("Yes" if task.ga_note_origin_id else "No"),
            _normalize_preview_value(task.status or ""),
            _normalize_preview_value((task.priority or "").upper()),
            _normalize_preview_value(_format_excel_datetime(task.created_at)),
            _normalize_preview_value(_format_excel_datetime(task.start_date)),
            _normalize_preview_value(_format_excel_datetime(task.due_date)),
        ]
        rows.append(values)

    headers = [
        "NR",
        "TITLE",
        "DESCRIPTION",
        "DEP",
        "PUNOI",
        "CREATED BY",
        "GA NOTE ORIGIN",
        "STATUS",
        "PRIORITY",
        "CREATION DATE",
        "START DATE",
        "DUE DATE",
    ]
    limited_rows, total, truncated = _apply_preview_limit(rows, limit)
    return ExportPreviewOut(headers=headers, rows=limited_rows, total=total, truncated=truncated)

@router.post("/all-tasks-report.xlsx")
async def export_all_tasks_report_xlsx(
    payload: AllTasksReportExportIn,
    user=Depends(get_current_user),
):
    ensure_reports_access(user)

    headers = ["NR", "LL", "NLL", "AM/PM", "DEP", "TITULLI", "STS", "BZ", "KOHA BZ", "T/Y/O", "KOMENT", "PUNOI"]
    title_text = _safe_filename(payload.title or "ALL TASKS REPORT")

    wb = Workbook()
    ws = wb.active
    ws.title = (payload.title or "All Tasks Report")[:31]

    last_col = len(headers)

    # Row 1: merged title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=1)

    # Row 2-3: spacing (empty)

    header_row = 4
    data_row = header_row + 1

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        header_text = "AM/\nPM" if header == "AM/PM" else header.upper()
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
        cell.number_format = "@"

    # Data rows (exactly matching Print All columns)
    for idx, r in enumerate(payload.rows or [], start=1):
        values = [
            idx,
            (r.typeLabel or "-"),
            (r.subtype or "-"),
            (r.period or "-"),
            (r.department or "-"),
            (r.title or "-"),
            (r.status or "-"),
            (r.bz or "-"),
            (r.kohaBz or "-"),
            (r.tyo or "-"),
            (r.comment or ""),
            (r.userInitials or "-"),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
            if col_idx == 1:
                cell.font = Font(bold=True)
        data_row += 1

    last_row = data_row - 1

    # Column widths (close to print layout)
    widths = {
        1: 5,   # NR
        2: 6,   # LL
        3: 7,   # NLL
        4: 7,   # AM/PM
        5: 7,   # DEP
        6: 42,  # TITULLI
        7: 10,  # STS
        8: 10,  # BZ
        9: 12,  # KOHA BZ
        10: 6,  # T/Y/O
        11: 22, # KOMENT
        12: 10, # PUNOI
    }
    for col_idx in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 16)

    # Filters, freeze, repeated header row
    if last_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"

    ws.freeze_panes = "B5"
    ws.print_title_rows = f"{header_row}:{header_row}"

    # Page setup + margins
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2

    # Header/footer
    ws.oddHeader.right.text = "&D &T"
    ws.oddFooter.center.text = "Page &P / &N"
    user_initials_compact = _initials_compact(user.full_name or user.username or "") or "____"
    ws.oddFooter.right.text = f"PUNOI: {user_initials_compact}"

    # Borders: thick outside on header row, thin elsewhere with thick outside frame
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    if last_row >= header_row:
        for r_idx in range(header_row, last_row + 1):
            for c_idx in range(1, last_col + 1):
                is_header = r_idx == header_row
                is_first_col = c_idx == 1
                is_last_col = c_idx == last_col
                is_last_row = r_idx == last_row

                left = thick if is_first_col else thin
                right = thick if is_last_col else thin
                top = thick if is_header else thin
                bottom = thick if is_last_row else thin

                if is_header:
                    ws.cell(row=r_idx, column=c_idx).border = Border(left=thick, right=thick, top=thick, bottom=thick)
                else:
                    ws.cell(row=r_idx, column=c_idx).border = Border(left=left, right=right, top=top, bottom=bottom)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    today = datetime.now(timezone.utc).date()
    filename_date = f"{today.day:02d}_{today.month:02d}_{str(today.year)[-2:]}"

    # Use exported users initials (from the report rows) for the filename.
    filename_users = (payload.usersInitials or "").strip().upper()
    if not filename_users:
        unique_users = sorted(
            {
                (r.userInitials or "").strip().upper()
                for r in (payload.rows or [])
                if (r.userInitials or "").strip()
            }
        )
        filename_users = "_".join(unique_users) if unique_users else "USER"

    # Standard required filename:
    # ALL ADMIN TASK REPORT_DD_MM_YY_USERS_INITIALS.xlsx
    filename = f"ALL_ADMIN_TASK_REPORT_{filename_date}_{filename_users}.xlsx".replace("__", "_")

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/tasks.pdf")
async def export_tasks_pdf(
    department_id: uuid.UUID | None = None,
    user_id: uuid.UUID | None = None,
    project_id: uuid.UUID | None = None,
    status_id: uuid.UUID | None = None,
    planned_from: date | None = None,
    planned_to: date | None = None,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    ensure_reports_access(user)
    tasks = await _query_tasks(
        db=db,
        user=user,
        department_id=department_id,
        user_id=user_id,
        project_id=project_id,
        status_id=status_id,
        planned_from=planned_from,
        planned_to=planned_to,
    )
    status_map, user_map = await _maps(db, tasks)

    bio = io.BytesIO()
    c = canvas.Canvas(bio, pagesize=letter)
    width, height = letter
    y = height - inch

    c.setFont("Helvetica-Bold", 14)
    c.drawString(inch, y, "Tasks Export (Summary)")
    y -= 0.4 * inch

    c.setFont("Helvetica", 10)
    c.drawString(inch, y, f"Total tasks: {len(tasks)}")
    y -= 0.3 * inch

    counts: dict[str, int] = {}
    for t in tasks:
        name = status_map.get(t.status_id, "Unknown")
        counts[name] = counts.get(name, 0) + 1
    for name, count in sorted(counts.items(), key=lambda kv: kv[0].lower()):
        c.drawString(inch, y, f"{name}: {count}")
        y -= 0.22 * inch
        if y < inch:
            c.showPage()
            y = height - inch

    if tasks:
        if y < 1.5 * inch:
            c.showPage()
            y = height - inch
        c.setFont("Helvetica-Bold", 12)
        c.drawString(inch, y, "Tasks")
        y -= 0.3 * inch
        c.setFont("Helvetica", 9)
        for t in tasks[:200]:
            line = f"- {t.title} [{status_map.get(t.status_id, '')}]"
            assignee = user_map.get(t.assigned_to_user_id, "") if t.assigned_to_user_id else ""
            if assignee:
                line += f" @ {assignee}"
            c.drawString(inch, y, line[:120])
            y -= 0.2 * inch
            if y < inch:
                c.showPage()
                y = height - inch
                c.setFont("Helvetica", 9)

    c.save()
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="tasks_export.pdf"'},
    )


@router.get("/checklists.xlsx")
async def export_checklist_xlsx(
    checklist_id: uuid.UUID,
    include_ko2: bool = False,
    path: str | None = None,
    format: str | None = None,
    title: str | None = None,
    exclude_path: list[str] | None = Query(None),
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    checklist = (
        await db.execute(select(Checklist).where(Checklist.id == checklist_id))
    ).scalar_one_or_none()
    if checklist is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Checklist not found")

    if checklist.project_id is not None:
        project = (
            await db.execute(select(Project).where(Project.id == checklist.project_id))
        ).scalar_one_or_none()
        if project is not None and project.department_id is not None:
            ensure_department_access(user, project.department_id)
    elif checklist.task_id is not None:
        task = (
            await db.execute(select(Task).where(Task.id == checklist.task_id))
        ).scalar_one_or_none()
        if task is not None and task.department_id is not None:
            ensure_department_access(user, task.department_id)
    elif checklist.group_key and user.role != UserRole.ADMIN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin only")

    items_stmt = (
        select(ChecklistItem)
        .where(ChecklistItem.checklist_id == checklist.id)
        .where(ChecklistItem.item_type == ChecklistItemType.CHECKBOX)
    )
    if format == "mst":
        items_stmt = items_stmt.where(ChecklistItem.path.is_not(None), ChecklistItem.title.is_not(None))
    if path:
        items_stmt = items_stmt.where(ChecklistItem.path == path)
    if exclude_path:
        items_stmt = items_stmt.where(~ChecklistItem.path.in_(exclude_path))
    items = (
        await db.execute(items_stmt.order_by(ChecklistItem.position, ChecklistItem.id))
    ).scalars().all()

    if format == "mst":
        headers = ["NR", "PATH", "DETYRAT", "KEYWORDS", "PERSHKRIMI", "KATEGORIA", "CHECK", "INCL", "KOMENT"]
    else:
        headers = ["NR", "TASK", "COMMENT", "CHECK", "TIME", "KOMENT"]
        if include_ko2:
            headers.append("KO2")

    assignee_initials: dict[uuid.UUID, str] = {}
    if format == "mst" and items:
        item_ids = [item.id for item in items if item.id is not None]
        if item_ids:
            assignees = (
                await db.execute(
                    select(ChecklistItemAssignee.checklist_item_id, User.full_name, User.username)
                    .join(User, User.id == ChecklistItemAssignee.user_id)
                    .where(ChecklistItemAssignee.checklist_item_id.in_(item_ids))
                )
            ).all()
            initials_map: dict[uuid.UUID, list[str]] = {}
            for item_id, full_name, username in assignees:
                label = full_name or username or ""
                initials = "".join(part[0] for part in label.split() if part).upper()
                if not initials:
                    continue
                initials_map.setdefault(item_id, []).append(initials)
            assignee_initials = {
                item_id: ", ".join(sorted(set(values)))
                for item_id, values in initials_map.items()
                if values
            }

    wb = Workbook()
    ws = wb.active
    raw_title = title or checklist.title or "Checklist"
    title = raw_title.upper()
    ws.title = raw_title[:31]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=False if header == "NR" else True,
        )

    if format == "mst":
        column_widths = {
            "NR": 3,
            "PATH": 22,
            "DETYRAT": 28,
            "KEYWORDS": 20,
            "PERSHKRIMI": 36,
            "KATEGORIA": 16,
            "CHECK": 8,
            "INCL": 10,
            "KOMENT": 24,
        }
    else:
        column_widths = {
            "NR": 3,
            "TASK": 36,
            "COMMENT": 46,
            "CHECK": 8,
            "TIME": 10,
            "KOMENT": 24,
            "KO2": 6,
        }
    for col_idx, header in enumerate(headers, start=1):
        width = column_widths.get(header, 16)
        ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = width

    data_row = header_row + 1
    for idx, item in enumerate(items, start=1):
        if format == "mst":
            incl_value = assignee_initials.get(item.id) or (item.owner or "")
            row_values = [
                idx,
                item.path or "",
                item.title or "",
                item.keyword or "",
                item.description or "",
                item.category or "",
                "yes" if item.is_checked else "",
                incl_value,
                item.comment or "",
            ]
        else:
            row_values = [
                idx,
                item.title or "",
                item.description or "",
                "yes" if item.is_checked else "",
                item.keyword or "",
                item.comment or "",
            ]
            if include_ko2:
                row_values.append("yes" if str(item.time or "").lower() in {"1", "true", "yes"} else "")
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if col_idx == 1:
                cell.font = Font(bold=True)
        data_row += 1

    ws.freeze_panes = ws["B4"]
    ws.print_title_rows = "3:3"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToPage = True
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2

    last_row = data_row - 1
    last_col = len(headers)
    if last_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{ws.cell(row=header_row, column=last_col).column_letter}{last_row}"
        thin = Side(style="thin", color="000000")
        thick = Side(style="medium", color="000000")
        for r in range(header_row, last_row + 1):
            for c in range(1, last_col + 1):
                left = thick if c == 1 else thin
                right = thick if c == last_col else thin
                top = thick if r == header_row else thin
                bottom = thick if r == last_row else thin
                ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

    ws.oddHeader.right.text = ""
    ws.oddFooter.center.text = "Page &P / &N"
    user_initials = _initials(user.full_name or user.username or "")
    ws.oddFooter.right.text = f"PUNOI: {user_initials or '____'}"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = checklist.title or "checklist_export"
    safe_filename = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in filename)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{safe_filename}.xlsx\"'},
    )


async def _daily_report_rows_for_user(
    *,
    db: AsyncSession,
    day: date,
    department_id: uuid.UUID | None,
    user_id: uuid.UUID,
) -> list[list[str]]:
    dept_code: str | None = None
    if department_id is not None:
        dept = (
            await db.execute(select(Department).where(Department.id == department_id))
        ).scalar_one_or_none()
        if dept is not None:
            dept_code = (dept.code or "").strip().upper() or None
            if dept_code is None and dept.name:
                upper = dept.name.strip().upper()
                if "DEVELOPMENT" in upper:
                    dept_code = "DEV"
                elif "GRAPHIC" in upper and "DESIGN" in upper:
                    dept_code = "GDS"
                elif ("PRODUCT" in upper and "CONTENT" in upper) or "PROJECT CONTENT" in upper:
                    dept_code = "PCM"

    day_start = datetime.combine(day, time.min, tzinfo=timezone.utc)
    day_end = day_start + timedelta(days=1)

    task_stmt = (
        select(Task)
        .outerjoin(Project, Task.project_id == Project.id)
        .where(Task.is_active.is_(True))
        .where(Task.system_template_origin_id.is_(None))
        .where(Task.due_date.is_not(None))
        .where(or_(Task.completed_at.is_(None), and_(Task.completed_at >= day_start, Task.completed_at < day_end)))
        .outerjoin(TaskAssignee, TaskAssignee.task_id == Task.id)
        .where(
            or_(
                Task.assigned_to == user_id,
                TaskAssignee.user_id == user_id,
                Task.internal_notes.ilike(f"%ko_user_id={user_id}%"),
                Task.internal_notes.ilike(f"%ko_user_id:%{user_id}%"),
                Task.internal_notes.ilike(f"%ko_user_id%{user_id}%"),
            )
        )
        .distinct()
    )
    if department_id is not None:
        task_stmt = task_stmt.where(or_(Task.department_id == department_id, Project.department_id == department_id))

    tasks = (await db.execute(task_stmt.order_by(Task.due_date, Task.created_at))).scalars().all()

    task_ids = [t.id for t in tasks]
    assignee_ids_by_task: dict[uuid.UUID, set[uuid.UUID]] = {tid: set() for tid in task_ids}
    if task_ids:
        rows = (
            await db.execute(
                select(TaskAssignee.task_id, TaskAssignee.user_id).where(TaskAssignee.task_id.in_(task_ids))
            )
        ).all()
        for tid, uid in rows:
            assignee_ids_by_task.setdefault(tid, set()).add(uid)

    project_ids = {t.project_id for t in tasks if t.project_id is not None}
    projects: list[Project] = []
    if project_ids:
        projects = (
            await db.execute(select(Project).where(Project.id.in_(project_ids)))
        ).scalars().all()
    project_by_id = {p.id: p for p in projects}

    tasks_today: list[Task] = []
    tasks_overdue: list[Task] = []
    planned_range_by_task_id: dict[uuid.UUID, tuple[date, date]] = {}
    for task in tasks:
        project = project_by_id.get(task.project_id) if task.project_id else None
        if not task_is_visible_to_user(
            task,
            user_id=user_id,
            assignee_ids=assignee_ids_by_task.get(task.id),
            project=project,
            dept_code=dept_code,
        ):
            continue

        planned_start, planned_end = planned_range_for_daily_report(task, dept_code)
        if planned_start is None or planned_end is None:
            continue

        is_project_task = task.project_id is not None
        if is_project_task:
            # Match the UI daily report view:
            # project tasks show only when due date is today/past (or completed today).
            due_day = _as_utc_date(task.due_date)
            if due_day is None:
                continue
            planned_start, planned_end = due_day, due_day

        planned_range_by_task_id[task.id] = (planned_start, planned_end)

        if completed_on_day(task.completed_at, day):
            tasks_today.append(task)
            continue

        if is_project_task:
            if planned_end <= day:
                if planned_end == day:
                    tasks_today.append(task)
                else:
                    tasks_overdue.append(task)
            continue

        # Fast tasks: show once their range has started.
        if planned_start <= day <= planned_end:
            tasks_today.append(task)
        elif planned_end < day:
            tasks_overdue.append(task)

    daily_tasks = tasks_today + tasks_overdue
    task_comment_map = await _user_comments_for_tasks(db, [t.id for t in daily_tasks], user_id)

    project_ids = {t.project_id for t in daily_tasks if t.project_id is not None}
    project_map: dict[uuid.UUID, str] = {}
    if project_ids:
        projects = (
            await db.execute(select(Project).where(Project.id.in_(project_ids)))
        ).scalars().all()
        for project in projects:
            project_map[project.id] = project.title or project.name or "-"

    await ensure_occurrences_in_range(db=db, start=day - timedelta(days=60), end=day)
    await db.commit()

    occ_today_rows = (
        await db.execute(
            select(SystemTaskOccurrence, SystemTaskTemplate)
            .join(SystemTaskTemplate, SystemTaskOccurrence.template_id == SystemTaskTemplate.id)
            .where(SystemTaskOccurrence.user_id == user_id)
            .where(SystemTaskOccurrence.occurrence_date == day)
            .order_by(SystemTaskTemplate.title)
        )
    ).all()
    latest_occurrence = (
        select(
            SystemTaskOccurrence.template_id,
            func.max(SystemTaskOccurrence.occurrence_date).label("latest_date"),
        )
        .where(SystemTaskOccurrence.user_id == user_id)
        .where(SystemTaskOccurrence.occurrence_date <= day)
        .group_by(SystemTaskOccurrence.template_id)
    ).subquery()

    occ_overdue_rows = (
        await db.execute(
            select(SystemTaskOccurrence, SystemTaskTemplate)
            .join(latest_occurrence, SystemTaskOccurrence.template_id == latest_occurrence.c.template_id)
            .join(SystemTaskTemplate, SystemTaskOccurrence.template_id == SystemTaskTemplate.id)
            .where(SystemTaskOccurrence.occurrence_date == latest_occurrence.c.latest_date)
            .where(SystemTaskOccurrence.occurrence_date < day)
            .where(SystemTaskOccurrence.status == OPEN)
            .order_by(SystemTaskOccurrence.occurrence_date.desc(), SystemTaskTemplate.title)
        )
    ).all()

    overdue_rows: list[tuple[SystemTaskOccurrence, SystemTaskTemplate]] = []
    for occ, tmpl in occ_overdue_rows:
        overdue_rows.append((occ, tmpl))

    template_ids = list({tmpl.id for _, tmpl in occ_today_rows} | {tmpl.id for _, tmpl in overdue_rows})
    roles_map, alignment_users_map = await _alignment_maps_for_templates(db, template_ids)
    alignment_user_ids = {uid for ids in alignment_users_map.values() for uid in ids}
    alignment_user_map: dict[uuid.UUID, str] = {}
    if alignment_user_ids:
        users = (await db.execute(select(User).where(User.id.in_(alignment_user_ids)))).scalars().all()
        for u in users:
            label = u.full_name or u.username or ""
            alignment_user_map[u.id] = _initials(label)

    dept_ids = {t.department_id for t in daily_tasks if t.department_id}
    dept_ids |= {tmpl.department_id for _, tmpl in occ_today_rows if tmpl.department_id}
    dept_ids |= {tmpl.department_id for _, tmpl in overdue_rows if tmpl.department_id}
    department_map: dict[uuid.UUID, Department] = {}
    if dept_ids:
        departments = (
            await db.execute(select(Department).where(Department.id.in_(dept_ids)))
        ).scalars().all()
        department_map = {dept.id: dept for dept in departments}

    def department_label(
        department_id_value: uuid.UUID | None,
        scope_value: str | None = None,
        is_ga_note: bool = False,
    ) -> str:
        if scope_value == SystemTaskScope.GA.value:
            return "GA"
        if scope_value == SystemTaskScope.ALL.value:
            return "ALL"
        if department_id_value:
            dept = department_map.get(department_id_value)
            return (dept.code if dept and dept.code else dept.name) or "-"
        if is_ga_note:
            return "GA"
        return "-"

    def alignment_values(tmpl: SystemTaskTemplate) -> tuple[str, str]:
        roles = roles_map.get(tmpl.id, [])
        user_ids = alignment_users_map.get(tmpl.id, [])
        alignment_enabled = bool(
            tmpl.requires_alignment or tmpl.alignment_time or roles or user_ids
        )
        if not alignment_enabled:
            return "-", "-"
        bz = "-"
        if user_ids:
            initials = [alignment_user_map.get(uid, "") for uid in user_ids]
            initials = [value for value in initials if value]
            if initials:
                bz = "/".join(initials)
        elif roles:
            bz = ", ".join(roles)
        koha_bz = _format_alignment_time(tmpl.alignment_time) if tmpl.alignment_time else "-"
        return bz, koha_bz

    rows: list[list[str]] = []
    fast_rows: list[tuple[int, int, list[str]]] = []
    project_rows: list[list[str]] = []
    system_am_rows: list[list[str]] = []
    system_pm_rows: list[list[str]] = []
    fast_index = 0

    def fast_type_order(task: Task) -> int:
        label = _no_project_type_label(task)
        if label == "BLLOK":
            return 0
        if label == "1H":
            return 1
        if label == "Personal":
            return 2
        if label == "R1":
            return 3
        if label == "Normal":
            return 4
        return 5

    for task in daily_tasks:
        base_dt = task.due_date or task.start_date or task.created_at
        planned_start, planned_end = planned_range_by_task_id.get(task.id, (None, None))
        mode: DailyReportTyoMode = "range" if planned_start and planned_end and planned_start < planned_end else "dueOnly"
        tyo = daily_report_tyo_label(
            report_day=day,
            start_day=planned_start if mode == "range" else None,
            due_day=planned_end,
            mode=mode,
        )
        period = _resolve_period(task.finish_period, base_dt)
        status = _format_task_status("DONE" if task.completed_at else task.status)
        comment = task_comment_map.get(task.id) or ""
        if task.project_id is None:
            fast_rows.append(
                (
                    fast_type_order(task),
                    fast_index,
                    [
                        "",
                        "FT",
                        _fast_subtype_short(task),
                        period,
                        department_label(task.department_id, None, bool(getattr(task, "ga_note_origin_id", None))),
                        task.title or "-",
                        task.description or "",
                        (status or "").upper(),
                        "-",
                        "-",
                        tyo,
                        comment,
                    ],
                )
            )
            fast_index += 1
        else:
            project_label = project_map.get(task.project_id, "-")
            if task.finish_period in {"AM", "PM"}:
                period = task.finish_period
            elif _is_vs_vl_project_label(project_label):
                period = "AM/PM"
            else:
                period = _resolve_period(task.finish_period, base_dt)
            project_rows.append(
                [
                    "",
                    "PRJK",
                    "-",
                    period,
                    department_label(task.department_id, None, bool(getattr(task, "ga_note_origin_id", None))),
                    f"{project_label} - {task.title or '-'}",
                    task.description or "",
                    (status or "").upper(),
                    "-",
                    "-",
                    tyo,
                    comment,
                ]
            )

    def add_system_row(container: list[list[str]], occ: SystemTaskOccurrence, tmpl: SystemTaskTemplate) -> None:
        base_date = occ.occurrence_date
        acted_date = occ.acted_at.date() if occ.acted_at else None
        tyo = _tyo_label(base_date, acted_date, day)
        period = _resolve_period(tmpl.finish_period, occ.occurrence_date)
        bz, koha_bz = alignment_values(tmpl)
        container.append(
            [
                "",
                "SYS",
                _system_frequency_short_label(tmpl.frequency),
                period,
                department_label(tmpl.department_id, tmpl.scope, False),
                tmpl.title or "-",
                tmpl.description or "",
                _format_system_status(occ.status).upper(),
                bz,
                koha_bz,
                tyo,
                occ.comment or "",
            ]
        )

    for occ, tmpl in overdue_rows:
        target = system_pm_rows if _resolve_period(tmpl.finish_period, occ.occurrence_date) == "PM" else system_am_rows
        add_system_row(target, occ, tmpl)

    for occ, tmpl in occ_today_rows:
        target = system_pm_rows if _resolve_period(tmpl.finish_period, occ.occurrence_date) == "PM" else system_am_rows
        add_system_row(target, occ, tmpl)

    fast_rows.sort(key=lambda item: (item[0], item[1]))
    rows.extend([row for _, __, row in fast_rows])
    rows.extend(system_am_rows)
    rows.extend(project_rows)
    rows.extend(system_pm_rows)
    return rows


async def _daily_ga_table_for_user(
    *,
    db: AsyncSession,
    day: date,
    department_id: uuid.UUID | None,
    user_id: uuid.UUID,
) -> str:
    entry_content = ""
    if department_id is not None:
        entry = (
            await db.execute(
                select(DailyReportGaEntry).where(
                    DailyReportGaEntry.user_id == user_id,
                    DailyReportGaEntry.department_id == department_id,
                    DailyReportGaEntry.entry_date == day,
                )
            )
        ).scalar_one_or_none()
        if entry:
            entry_content = entry.content or ""
    return entry_content


@router.get("/daily-report.xlsx")
async def export_daily_report_xlsx(
    day: date,
    department_id: uuid.UUID | None = None,
    user_id: uuid.UUID | None = None,
    all_users: bool = False,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    if department_id is not None:
        ensure_department_access(user, department_id)
    elif user.role != UserRole.ADMIN:
        department_id = user.department_id

    target_user: User | None = None
    if not all_users:
        target_user_id = user_id or user.id
        target_user = (
            await db.execute(select(User).where(User.id == target_user_id))
        ).scalar_one_or_none()
        if target_user is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
        if department_id is not None and target_user.department_id != department_id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
        if user.role == UserRole.STAFF and user.department_id is None and target_user.id != user.id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")

    rows: list[list[str]] = []
    users_for_export: list[User] = []
    if all_users:
        if department_id is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="department_id is required")
        users_for_export = (
            await db.execute(
                select(User)
                .where(User.department_id == department_id)
                .order_by(User.full_name, User.username)
            )
        ).scalars().all()
    else:
        users_for_export = [target_user] if target_user is not None else []

    for member in users_for_export:
        member_rows = await _daily_report_rows_for_user(
            db=db,
            day=day,
            department_id=department_id,
            user_id=member.id,
        )
        member_label = _initials(member.full_name or member.username or "") or "-"
        for row in member_rows:
            rows.append(row + [member_label])

    # Sort by LL (index 1), NLL (index 2), and T/Y/O (index 9)
    if all_users:
        rows.sort(key=lambda r: (
            r[1] if len(r) > 1 else "",  # LL (typeLabel)
            r[2] if len(r) > 2 else "",  # NLL (subtype)
            r[9] if len(r) > 9 else "",  # T/Y/O (tyo)
        ))

    headers = ["NR", "LL", "NLL", "AM/PM", "DEP", "TITULLI", "PERSHKRIMI", "STS", "BZ", "KOHA BZ", "T/Y/O", "KOMENT", "USER"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    status_fills = {
        "TODO": PatternFill(start_color="FFC4ED", end_color="FFC4ED", fill_type="solid"),
        "IN_PROGRESS": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "DONE": PatternFill(start_color="C4FDC4", end_color="C4FDC4", fill_type="solid"),
    }

    title_row = 1
    dept_row = 2
    user_row = 3
    header_row = 5

    title_text = "ALL TODAY REPORT" if all_users else "DAILY TASK REPORT"
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(headers))
    title_cell = ws.cell(row=title_row, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    department_label = ""
    if department_id:
        department = (await db.execute(select(Department).where(Department.id == department_id))).scalar_one_or_none()
        department_label = department.name if department else ""
    ws.cell(row=dept_row, column=1, value=f"Department: {department_label or '-'}")
    if all_users:
        ws.cell(row=user_row, column=1, value="Users: All users")
    else:
        ws.cell(row=user_row, column=1, value=f"User: {users_for_export[0].full_name or users_for_export[0].username or '-'}")
    for col_idx, header in enumerate(headers, start=1):
        header_text = "AM/\nPM" if header == "AM/PM" else header.upper()
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(
            horizontal="left",
            vertical="bottom",
            wrap_text=True if header == "Nr" else True,
        )
        # Force uppercase display even if Excel auto-changes header text.
        cell.number_format = "@"

    column_widths = {
        "NR": 4,
        "LL": 5,
        "NLL": 6,
        "AM/PM": 7,
        "DEP": 6,
        "TITULLI": 32,
        "PERSHKRIMI": 26,
        "STS": 10,
        "BZ": 8,
        "KOHA BZ": 10,
        "T/Y/O": 6,
        "KOMENT": 22,
        "USER": 6,
    }
    for col_idx, header in enumerate(headers, start=1):
        width = column_widths.get(header, 16)
        ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 10
    ws.row_dimensions[header_row].height = 28

    data_row = header_row + 1
    for idx, row in enumerate(rows, start=1):
        row_values = row.copy()
        row_values[0] = idx
        status_value = row_values[7] if len(row_values) > 7 else ""
        status_key = _normalize_report_status(status_value)
        status_fill = status_fills.get(status_key)
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.alignment = Alignment(
                horizontal="left",
                vertical="bottom",
                wrap_text=col_idx in {6, 7, 13},
            )
            if col_idx == 1:
                cell.font = Font(bold=True)
            if status_fill and col_idx == 8:
                cell.fill = status_fill
        data_row += 1

    ws.freeze_panes = ws["B6"]
    ws.print_title_rows = "5:5"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToPage = True
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2

    last_row = data_row - 1
    last_col = len(headers)
    if last_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{ws.cell(row=header_row, column=last_col).column_letter}{last_row}"
        thin = Side(style="thin", color="000000")
        thick = Side(style="medium", color="000000")
        for r in range(header_row, last_row + 1):
            for c in range(1, last_col + 1):
                is_first_col = c == 1
                is_last_col = c == last_col
                is_header = r == header_row
                is_last_row = r == last_row

                left = thick if is_first_col else thin
                right = thick if is_last_col else thin
                top = thick if is_header else thin
                bottom = thick if is_last_row else thin
                # Header row: thick outside border (top/bottom/edges), thin inside separators.
                if is_header:
                    ws.cell(row=r, column=c).border = Border(
                        left=thick,
                        right=thick,
                        top=thick,
                        bottom=thick,
                    )
                else:
                    ws.cell(row=r, column=c).border = Border(
                        left=left,
                        right=right,
                        top=top,
                        bottom=bottom,
                    )

    if not all_users:
        ga_entry = await _daily_ga_table_for_user(
            db=db,
            day=day,
            department_id=department_id,
            user_id=users_for_export[0].id,
        )
        last_data_row = last_row
        ga_section_row = last_data_row + 2
        ga_label_start_col = 1
        ga_label_end_col = 4
        ga_value_col_start = 5
        ga_value_col_end = last_col

        if ga_label_end_col >= ga_label_start_col:
            ws.merge_cells(
                start_row=ga_section_row,
                start_column=ga_label_start_col,
                end_row=ga_section_row,
                end_column=ga_label_end_col,
            )
        ga_label_cell = ws.cell(row=ga_section_row, column=ga_label_start_col, value="GA/KUR/SI/KUJT/PRBL")
        ga_label_cell.font = Font(bold=True)
        ga_label_cell.alignment = Alignment(horizontal="left", vertical="bottom", readingOrder=1)
        if ga_value_col_end >= ga_value_col_start:
            ws.merge_cells(
                start_row=ga_section_row,
                start_column=ga_value_col_start,
                end_row=ga_section_row,
                end_column=ga_value_col_end,
            )
        ga_value_text = ga_entry or "-"
        ga_value_cell = ws.cell(row=ga_section_row, column=ga_value_col_start, value=ga_value_text)
        ga_value_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
        if ga_value_text:
            ga_lines = ga_value_text.count("\n") + 1
            ws.row_dimensions[ga_section_row].height = max(18, ga_lines * 15)

        # Outline border for GA section (single-row table).
        thin = Side(style="thin", color="000000")
        thick = Side(style="medium", color="000000")
        section_start = ga_section_row
        section_end = ga_section_row
        for r in range(section_start, section_end + 1):
            for c in range(1, last_col + 1):
                is_left = c == 1
                is_right = c == last_col
                is_divider_left = c == ga_value_col_start
                is_divider_right = c == ga_label_end_col

                left = thick if is_left else (thin if is_divider_left else None)
                right = thick if is_right else (thin if is_divider_right else None)
                ws.cell(row=r, column=c).border = Border(
                    left=left,
                    right=right,
                    top=thick,
                    bottom=thick,
                )

    ws.oddHeader.right.text = "&D &T"
    ws.oddFooter.center.text = "Page &P / &N"
    user_initials = _initials(user.full_name or user.username or "")
    ws.oddFooter.right.text = f"PUNOI: {user_initials or '____'}"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    date_label = day.strftime("%d_%m_%y")
    user_initials = _initials(user.full_name or user.username or "")
    initials_label = (user_initials or "USER").upper()
    filename = f"DAILY_REPORT_{date_label}_{initials_label}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )

@router.get("/system-tasks.xlsx")
async def export_system_tasks_xlsx(
    active_only: bool = True,
    department_id: uuid.UUID | None = None,
    user_id: uuid.UUID | None = None,
    status_id: uuid.UUID | None = None,
    planned_from: date | None = None,
    planned_to: date | None = None,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    ensure_reports_access(user)
    is_admin = user.role == UserRole.ADMIN or (getattr(user.role, "value", "").lower() == "admin")

    status_name: str | None = None
    if status_id is not None:
        status_row = (await db.execute(select(TaskStatus).where(TaskStatus.id == status_id))).scalar_one_or_none()
        if status_row is not None:
            status_name = status_row.name

    template_stmt = select(SystemTaskTemplate)
    if department_id is not None:
        ensure_department_access(user, department_id)
        template_stmt = template_stmt.where(
            or_(
                and_(
                    SystemTaskTemplate.scope == SystemTaskScope.DEPARTMENT.value,
                    SystemTaskTemplate.department_id == department_id,
                ),
                SystemTaskTemplate.scope == SystemTaskScope.ALL.value,
                SystemTaskTemplate.scope == SystemTaskScope.GA.value,
            )
        )
    elif not is_admin:
        if user.department_id is None:
            if user.role != UserRole.MANAGER:
                return []
        else:
            template_stmt = template_stmt.where(
                or_(
                    SystemTaskTemplate.scope == SystemTaskScope.ALL.value,
                    SystemTaskTemplate.scope == SystemTaskScope.GA.value,
                    and_(
                        SystemTaskTemplate.scope == SystemTaskScope.DEPARTMENT.value,
                        SystemTaskTemplate.department_id == user.department_id,
                    ),
                )
            )

    templates = (await db.execute(template_stmt.order_by(SystemTaskTemplate.title))).scalars().all()
    if not templates:
        return []

    template_assignee_ids_map: dict[uuid.UUID, list[uuid.UUID]] = {}
    template_assignee_user_ids: set[uuid.UUID] = set()
    for tmpl in templates:
        raw_ids = list(getattr(tmpl, "assignee_ids", None) or [])
        if tmpl.default_assignee_id and tmpl.default_assignee_id not in raw_ids:
            raw_ids.append(tmpl.default_assignee_id)
        if raw_ids:
            seen: set[uuid.UUID] = set()
            deduped: list[uuid.UUID] = []
            for uid in raw_ids:
                if uid in seen:
                    continue
                seen.add(uid)
                deduped.append(uid)
            template_assignee_ids_map[tmpl.id] = deduped
            template_assignee_user_ids.update(deduped)

    template_assignee_initials_map: dict[uuid.UUID, str] = {}
    if template_assignee_user_ids:
        template_assignee_users = (
            await db.execute(select(User).where(User.id.in_(template_assignee_user_ids)))
        ).scalars().all()
        template_assignee_initials_map = {
            u.id: _initials(u.full_name or u.username or "") for u in template_assignee_users
        }

    template_ids = [t.id for t in templates]
    task_stmt = (
        select(Task, SystemTaskTemplate)
        .join(SystemTaskTemplate, Task.system_template_origin_id == SystemTaskTemplate.id)
        .where(Task.system_template_origin_id.in_(template_ids))
    )
    if active_only:
        task_stmt = task_stmt.where(Task.is_active.is_(True))
    if department_id is not None:
        task_stmt = task_stmt.where(Task.department_id == department_id)
    if user_id is not None:
        assignee_subq = select(TaskAssignee.task_id).where(TaskAssignee.user_id == user_id)
        task_stmt = task_stmt.where(or_(Task.assigned_to == user_id, Task.id.in_(assignee_subq)))
    if status_name:
        task_stmt = task_stmt.where(Task.status == status_name)
    if planned_from:
        task_stmt = task_stmt.where(func.date(Task.start_date) >= planned_from)
    if planned_to:
        task_stmt = task_stmt.where(func.date(Task.start_date) <= planned_to)
    task_rows = (await db.execute(task_stmt.order_by(Task.created_at.desc()))).all()
    dedup: dict[uuid.UUID, tuple[Task, SystemTaskTemplate]] = {}
    for task, tmpl in task_rows:
        prev = dedup.get(tmpl.id)
        if prev is None or (task.created_at and prev[0].created_at and task.created_at > prev[0].created_at):
            dedup[tmpl.id] = (task, tmpl)
    task_rows = list(dedup.values())

    task_ids = [task.id for task, _ in task_rows]
    assignee_map: dict[uuid.UUID, list[str]] = {task_id: [] for task_id in task_ids}
    if task_ids:
        assignee_rows = (
            await db.execute(
                select(TaskAssignee.task_id, User)
                .join(User, TaskAssignee.user_id == User.id)
                .where(TaskAssignee.task_id.in_(task_ids))
            )
        ).all()
        for task_id, user_row in assignee_rows:
            label = user_row.full_name or user_row.username or ""
            if label:
                assignee_map.setdefault(task_id, []).append(label)

    fallback_ids = [
        task.assigned_to
        for task, _ in task_rows
        if task.assigned_to is not None and not assignee_map.get(task.id)
    ]
    fallback_map: dict[uuid.UUID, str] = {}
    if fallback_ids:
        fallback_users = (
            await db.execute(select(User).where(User.id.in_(fallback_ids)))
        ).scalars().all()
        fallback_map = {u.id: (u.full_name or u.username or "") for u in fallback_users}

    task_alignment_map: dict[uuid.UUID, list[uuid.UUID]] = {task_id: [] for task_id in task_ids}
    if task_ids:
        alignment_rows = (
            await db.execute(
                select(TaskAlignmentUser.task_id, TaskAlignmentUser.user_id)
                .where(TaskAlignmentUser.task_id.in_(task_ids))
            )
        ).all()
        for task_id, user_id in alignment_rows:
            task_alignment_map.setdefault(task_id, []).append(user_id)

    template_alignment_map: dict[uuid.UUID, list[uuid.UUID]] = {}
    if template_ids:
        template_alignment_rows = (
            await db.execute(
                select(SystemTaskTemplateAlignmentUser.template_id, SystemTaskTemplateAlignmentUser.user_id)
                .where(SystemTaskTemplateAlignmentUser.template_id.in_(template_ids))
            )
        ).all()
        for template_id, user_id in template_alignment_rows:
            template_alignment_map.setdefault(template_id, []).append(user_id)

    department_ids = {template.department_id for _, template in task_rows if template.department_id}
    department_map: dict[uuid.UUID, str] = {}
    if department_ids:
        departments = (
            await db.execute(select(Department).where(Department.id.in_(department_ids)))
        ).scalars().all()
        department_map = {d.id: d.name for d in departments}

    alignment_user_ids: set[uuid.UUID] = set()
    for task, template in task_rows:
        ids = task_alignment_map.get(task.id) or []
        if not ids:
            ids = template_alignment_map.get(template.id) or []
        for user_id in ids:
            alignment_user_ids.add(user_id)
    alignment_user_map: dict[uuid.UUID, str] = {}
    if alignment_user_ids:
        alignment_users = (
            await db.execute(select(User).where(User.id.in_(alignment_user_ids)))
        ).scalars().all()
        alignment_user_map = {
            u.id: _initials(u.full_name or u.username or "") for u in alignment_users
        }

    wb = Workbook()
    ws = wb.active
    ws.title = "System Tasks"

    headers = [
        "NR",
        "PRIO",
        "LL",
        "DITA",
        "DEP",
        "AM/\nPM",
        "TITULLI",
        "PERSHKRIMI",
        "USER",
        "REGJ/PATH/CHECKLISTA/TRAINING/BZ GROUP",
        "BZ ME",
        "KOHA BZ",
    ]
    last_col = len(headers)

    title_row = 3
    header_row = 5
    data_row = header_row + 1

    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=last_col)
    title_cell = ws.cell(row=title_row, column=1, value="SYSTEM TASKS")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

    col_widths = [len(header) for header in headers]

    for idx, (task, template) in enumerate(task_rows, start=1):
        template_assignee_ids = template_assignee_ids_map.get(template.id, [])
        assignee_initials = [
            template_assignee_initials_map.get(uid, "") for uid in template_assignee_ids
        ]
        assignee_label = ", ".join([initials for initials in assignee_initials if initials])
        if not assignee_label:
            assignees = assignee_map.get(task.id, [])
            if not assignees and task.assigned_to in fallback_map:
                assignees = [fallback_map[task.assigned_to]]
            fallback_initials = [(_initials(name) if name else "") for name in assignees]
            assignee_label = ", ".join([initials for initials in fallback_initials if initials])
        row_idx = data_row + idx - 1
        if template.department_id and template.department_id in department_map:
            department_label = _department_short(department_map[template.department_id])
        elif template.scope == "GA":
            department_label = "GA"
        else:
            department_label = "ALL"
        note_values = _parse_internal_notes(template.internal_notes)
        regj_value = note_values.get("REGJ", "") or "-"
        path_value = note_values.get("PATH", "") or "-"
        check_value = note_values.get("CHECKLISTA", "") or note_values.get("CHECK", "") or "-"
        training_value = note_values.get("TRAINING", "") or "-"
        bz_group_value = note_values.get("BZ GROUP", "") or "-"
        if all(value == "-" for value in [regj_value, path_value, check_value, training_value, bz_group_value]):
            details_value = ""
        else:
            details_value = "\n".join(
                [
                    f"1.REGJ: {regj_value}",
                    f"2.PATH: {path_value}",
                    f"3.CHECKLISTA: {check_value}",
                    f"4.TRAINING: {training_value}",
                    f"5.BZ GROUP: {bz_group_value}",
                ]
            )
        alignment_ids = task_alignment_map.get(task.id) or []
        if not alignment_ids:
            alignment_ids = template_alignment_map.get(template.id) or []
        bz_me_labels = [alignment_user_map.get(user_id, "") for user_id in alignment_ids]
        bz_me_value = ", ".join([label for label in bz_me_labels if label])
        alignment_time = getattr(template, "alignment_time", None)
        bz_time_value = ""
        if alignment_time:
            alignment_time_str = str(alignment_time)
            bz_time_value = alignment_time_str[:5]
        schedule_value = _system_task_schedule_label(template)
        values = [
            idx,
            _priority_label(template.priority),
            _frequency_label(template.frequency),
            schedule_value,
            department_label,
            task.finish_period or "",
            task.title,
            _strip_html_keep_breaks(task.description),
            assignee_label,
            details_value,
            bz_me_value,
            bz_time_value,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
            col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(value)))
        # Excel rich text caused repair prompts; keep plain text for compatibility.

    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 80)

    for row_idx in range(data_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        cell.font = Font(bold=True)
        cell.number_format = "#,##0"

    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    last_row = ws.max_row
    for r in range(header_row, last_row + 1):
        for c in range(1, last_col + 1):
            left = thick if c == 1 else thin
            right = thick if c == last_col else thin
            top = thick if r == header_row else thin
            if r == header_row or r == last_row:
                bottom = thick
            else:
                bottom = thin
            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

    ws.freeze_panes = "B6"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2
    ws.oddHeader.right.text = "&D &T"
    ws.oddFooter.center.text = "Page &P / &N"
    user_initials = _initials(user.full_name or user.username or "")
    ws.oddFooter.right.text = f"PUNOI: {user_initials or '____'}"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    today = datetime.now().date()
    filename = f"SYSTEM_TASKS_{today.day:02d}_{today.month:02d}_{str(today.year)[-2:]}_{user_initials or 'USER'}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/system-tasks-preview", response_model=ExportPreviewOut)
async def preview_system_tasks(
    active_only: bool = True,
    department_id: uuid.UUID | None = None,
    user_id: uuid.UUID | None = None,
    status_id: uuid.UUID | None = None,
    planned_from: date | None = None,
    planned_to: date | None = None,
    limit: int | None = 200,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    ensure_reports_access(user)
    is_admin = user.role == UserRole.ADMIN or (getattr(user.role, "value", "").lower() == "admin")

    status_name: str | None = None
    if status_id is not None:
        status_row = (await db.execute(select(TaskStatus).where(TaskStatus.id == status_id))).scalar_one_or_none()
        if status_row is not None:
            status_name = status_row.name

    template_stmt = select(SystemTaskTemplate)
    if department_id is not None:
        ensure_department_access(user, department_id)
        template_stmt = template_stmt.where(
            or_(
                and_(
                    SystemTaskTemplate.scope == SystemTaskScope.DEPARTMENT.value,
                    SystemTaskTemplate.department_id == department_id,
                ),
                SystemTaskTemplate.scope == SystemTaskScope.ALL.value,
                SystemTaskTemplate.scope == SystemTaskScope.GA.value,
            )
        )
    elif not is_admin:
        if user.department_id is None:
            if user.role != UserRole.MANAGER:
                return ExportPreviewOut(headers=[], rows=[], total=0, truncated=False)
        else:
            template_stmt = template_stmt.where(
                or_(
                    SystemTaskTemplate.scope == SystemTaskScope.ALL.value,
                    SystemTaskTemplate.scope == SystemTaskScope.GA.value,
                    and_(
                        SystemTaskTemplate.scope == SystemTaskScope.DEPARTMENT.value,
                        SystemTaskTemplate.department_id == user.department_id,
                    ),
                )
            )

    templates = (await db.execute(template_stmt.order_by(SystemTaskTemplate.title))).scalars().all()
    if not templates:
        headers = [
            "NR",
            "PRIO",
            "LL",
            "DEP",
            "AM/PM",
            "TITULLI",
            "PERSHKRIMI",
            "USER",
            "REGJ/PATH/CHECKLISTA/TRAINING/BZ GROUP",
            "BZ ME",
            "KOHA BZ",
        ]
        return ExportPreviewOut(headers=headers, rows=[], total=0, truncated=False)

    template_assignee_ids_map: dict[uuid.UUID, list[uuid.UUID]] = {}
    template_assignee_user_ids: set[uuid.UUID] = set()
    for tmpl in templates:
        raw_ids = list(getattr(tmpl, "assignee_ids", None) or [])
        if tmpl.default_assignee_id and tmpl.default_assignee_id not in raw_ids:
            raw_ids.append(tmpl.default_assignee_id)
        if raw_ids:
            seen: set[uuid.UUID] = set()
            deduped: list[uuid.UUID] = []
            for uid in raw_ids:
                if uid in seen:
                    continue
                seen.add(uid)
                deduped.append(uid)
            template_assignee_ids_map[tmpl.id] = deduped
            template_assignee_user_ids.update(deduped)

    template_assignee_initials_map: dict[uuid.UUID, str] = {}
    if template_assignee_user_ids:
        template_assignee_users = (
            await db.execute(select(User).where(User.id.in_(template_assignee_user_ids)))
        ).scalars().all()
        template_assignee_initials_map = {
            u.id: _initials(u.full_name or u.username or "") for u in template_assignee_users
        }

    template_ids = [t.id for t in templates]
    task_stmt = (
        select(Task, SystemTaskTemplate)
        .join(SystemTaskTemplate, Task.system_template_origin_id == SystemTaskTemplate.id)
        .where(Task.system_template_origin_id.in_(template_ids))
    )
    if active_only:
        task_stmt = task_stmt.where(Task.is_active.is_(True))
    if department_id is not None:
        task_stmt = task_stmt.where(Task.department_id == department_id)
    if user_id is not None:
        assignee_subq = select(TaskAssignee.task_id).where(TaskAssignee.user_id == user_id)
        task_stmt = task_stmt.where(or_(Task.assigned_to == user_id, Task.id.in_(assignee_subq)))
    if status_name:
        task_stmt = task_stmt.where(Task.status == status_name)
    if planned_from:
        task_stmt = task_stmt.where(func.date(Task.start_date) >= planned_from)
    if planned_to:
        task_stmt = task_stmt.where(func.date(Task.start_date) <= planned_to)
    task_rows = (await db.execute(task_stmt.order_by(Task.created_at.desc()))).all()
    dedup: dict[uuid.UUID, tuple[Task, SystemTaskTemplate]] = {}
    for task, tmpl in task_rows:
        prev = dedup.get(tmpl.id)
        if prev is None or (task.created_at and prev[0].created_at and task.created_at > prev[0].created_at):
            dedup[tmpl.id] = (task, tmpl)
    task_rows = list(dedup.values())

    task_ids = [task.id for task, _ in task_rows]
    assignee_map: dict[uuid.UUID, list[str]] = {task_id: [] for task_id in task_ids}
    if task_ids:
        assignee_rows = (
            await db.execute(
                select(TaskAssignee.task_id, User)
                .join(User, TaskAssignee.user_id == User.id)
                .where(TaskAssignee.task_id.in_(task_ids))
            )
        ).all()
        for task_id, user_row in assignee_rows:
            label = user_row.full_name or user_row.username or ""
            if label:
                assignee_map.setdefault(task_id, []).append(label)

    fallback_ids = [
        task.assigned_to
        for task, _ in task_rows
        if task.assigned_to is not None and not assignee_map.get(task.id)
    ]
    fallback_map: dict[uuid.UUID, str] = {}
    if fallback_ids:
        fallback_users = (
            await db.execute(select(User).where(User.id.in_(fallback_ids)))
        ).scalars().all()
        fallback_map = {u.id: (u.full_name or u.username or "") for u in fallback_users}

    task_alignment_map: dict[uuid.UUID, list[uuid.UUID]] = {task_id: [] for task_id in task_ids}
    if task_ids:
        alignment_rows = (
            await db.execute(
                select(TaskAlignmentUser.task_id, TaskAlignmentUser.user_id)
                .where(TaskAlignmentUser.task_id.in_(task_ids))
            )
        ).all()
        for task_id, alignment_user_id in alignment_rows:
            task_alignment_map.setdefault(task_id, []).append(alignment_user_id)

    template_alignment_map: dict[uuid.UUID, list[uuid.UUID]] = {}
    if template_ids:
        template_alignment_rows = (
            await db.execute(
                select(SystemTaskTemplateAlignmentUser.template_id, SystemTaskTemplateAlignmentUser.user_id)
                .where(SystemTaskTemplateAlignmentUser.template_id.in_(template_ids))
            )
        ).all()
        for template_id, alignment_user_id in template_alignment_rows:
            template_alignment_map.setdefault(template_id, []).append(alignment_user_id)

    department_ids = {template.department_id for _, template in task_rows if template.department_id}
    department_map: dict[uuid.UUID, str] = {}
    if department_ids:
        departments = (
            await db.execute(select(Department).where(Department.id.in_(department_ids)))
        ).scalars().all()
        department_map = {d.id: d.name for d in departments}

    alignment_user_ids: set[uuid.UUID] = set()
    for task, template in task_rows:
        ids = task_alignment_map.get(task.id) or []
        if not ids:
            ids = template_alignment_map.get(template.id) or []
        for alignment_user_id in ids:
            alignment_user_ids.add(alignment_user_id)
    alignment_user_map: dict[uuid.UUID, str] = {}
    if alignment_user_ids:
        alignment_users = (
            await db.execute(select(User).where(User.id.in_(alignment_user_ids)))
        ).scalars().all()
        alignment_user_map = {
            u.id: _initials(u.full_name or u.username or "") for u in alignment_users
        }

    headers = [
        "NR",
        "PRIO",
        "LL",
        "DEP",
        "AM/PM",
        "TITULLI",
        "PERSHKRIMI",
        "USER",
        "REGJ/PATH/CHECKLISTA/TRAINING/BZ GROUP",
        "BZ ME",
        "KOHA BZ",
    ]

    rows: list[list[str]] = []
    for idx, (task, template) in enumerate(task_rows, start=1):
        template_assignee_ids = template_assignee_ids_map.get(template.id, [])
        assignee_initials = [
            template_assignee_initials_map.get(uid, "") for uid in template_assignee_ids
        ]
        assignee_label = ", ".join([initials for initials in assignee_initials if initials])
        if not assignee_label:
            assignees = assignee_map.get(task.id, [])
            if not assignees and task.assigned_to in fallback_map:
                assignees = [fallback_map[task.assigned_to]]
            fallback_initials = [(_initials(name) if name else "") for name in assignees]
            assignee_label = ", ".join([initials for initials in fallback_initials if initials])
        if template.department_id and template.department_id in department_map:
            department_label = _department_short(department_map[template.department_id])
        elif template.scope == "GA":
            department_label = "GA"
        else:
            department_label = "ALL"
        note_values = _parse_internal_notes(template.internal_notes)
        regj_value = note_values.get("REGJ", "") or "-"
        path_value = note_values.get("PATH", "") or "-"
        check_value = note_values.get("CHECKLISTA", "") or note_values.get("CHECK", "") or "-"
        training_value = note_values.get("TRAINING", "") or "-"
        bz_group_value = note_values.get("BZ GROUP", "") or "-"
        if all(value == "-" for value in [regj_value, path_value, check_value, training_value, bz_group_value]):
            details_value = ""
        else:
            details_value = "\n".join(
                [
                    f"1.REGJ: {regj_value}",
                    f"2.PATH: {path_value}",
                    f"3.CHECKLISTA: {check_value}",
                    f"4.TRAINING: {training_value}",
                    f"5.BZ GROUP: {bz_group_value}",
                ]
            )
        alignment_ids = task_alignment_map.get(task.id) or []
        if not alignment_ids:
            alignment_ids = template_alignment_map.get(template.id) or []
        bz_me_labels = [alignment_user_map.get(alignment_user_id, "") for alignment_user_id in alignment_ids]
        bz_me_value = ", ".join([label for label in bz_me_labels if label])
        alignment_time = getattr(template, "alignment_time", None)
        bz_time_value = ""
        if alignment_time:
            alignment_time_str = str(alignment_time)
            bz_time_value = alignment_time_str[:5]
        values = [
            str(idx),
            _normalize_preview_value(_priority_label(template.priority)),
            _normalize_preview_value(_frequency_label(template.frequency)),
            _normalize_preview_value(department_label),
            _normalize_preview_value(task.finish_period or ""),
            _normalize_preview_value(task.title),
            _normalize_preview_value(_strip_html(task.description)),
            _normalize_preview_value(assignee_label),
            _normalize_preview_value(details_value),
            _normalize_preview_value(bz_me_value),
            _normalize_preview_value(bz_time_value),
        ]
        rows.append(values)

    limited_rows, total, truncated = _apply_preview_limit(rows, limit)
    return ExportPreviewOut(headers=headers, rows=limited_rows, total=total, truncated=truncated)


@router.get("/common.xlsx")
async def export_common_xlsx(
    week_start: date,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    week_dates = [week_start + timedelta(days=i) for i in range(5)]
    week_isos = [d.isoformat() for d in week_dates]
    start_dt = datetime.combine(week_dates[0], time.min, tzinfo=timezone.utc)
    end_dt = datetime.combine(week_dates[-1], time.max, tzinfo=timezone.utc)

    entries_stmt = select(CommonEntry)
    if user.department_id is not None:
        dept_user_ids = (
            await db.execute(select(User.id).where(User.department_id == user.department_id))
        ).scalars().all()
        if dept_user_ids:
            entries_stmt = entries_stmt.where(
                or_(
                    CommonEntry.created_by_user_id.in_(dept_user_ids),
                    CommonEntry.assigned_to_user_id.in_(dept_user_ids),
                )
            )
        else:
            entries_stmt = entries_stmt.where(CommonEntry.created_by_user_id.is_(None))
    entries_stmt = entries_stmt.where(
        or_(
            CommonEntry.entry_date.between(week_dates[0], week_dates[-1]),
            and_(
                CommonEntry.entry_date.is_(None),
                CommonEntry.created_at >= start_dt,
                CommonEntry.created_at <= end_dt,
            ),
        )
    )
    entries = (await db.execute(entries_stmt.order_by(CommonEntry.created_at.desc()))).scalars().all()

    tasks: list[Task] = []
    if user.department_id is not None or user.role != UserRole.STAFF:
        tasks_stmt = select(Task)
        if user.department_id is not None:
            tasks_stmt = tasks_stmt.where(Task.department_id == user.department_id)
        tasks_stmt = tasks_stmt.where(
            or_(
                Task.due_date.between(start_dt, end_dt),
                Task.start_date.between(start_dt, end_dt),
                Task.created_at.between(start_dt, end_dt),
            )
        )
        tasks = (await db.execute(tasks_stmt.order_by(Task.created_at.desc()))).scalars().all()

    meetings: list[Meeting] = []
    if user.department_id is not None or user.role != UserRole.STAFF:
        meetings_stmt = select(Meeting)
        if user.department_id is not None:
            meetings_stmt = meetings_stmt.where(Meeting.department_id == user.department_id)
        meetings_stmt = meetings_stmt.where(
            or_(
                Meeting.starts_at.between(start_dt, end_dt),
                Meeting.created_at.between(start_dt, end_dt),
            )
        )
        meetings = (await db.execute(meetings_stmt.order_by(Meeting.created_at.desc()))).scalars().all()

    task_ids = [t.id for t in tasks]
    assignees_by_task = await _assignees_for_tasks(db, task_ids)

    project_ids = {t.project_id for t in tasks if t.project_id is not None}
    project_name_map: dict[uuid.UUID, str] = {}
    if project_ids:
        projects = (
            await db.execute(select(Project).where(Project.id.in_(project_ids)))
        ).scalars().all()
        for project in projects:
            base_title = (project.title or "").strip()
            if not base_title:
                continue
            if project.project_type == "MST" and project.total_products is not None and project.total_products > 0:
                title = f"{base_title} - {project.total_products}"
            else:
                title = base_title
            project_name_map[project.id] = title

    user_ids: set[uuid.UUID] = set()
    for entry in entries:
        if entry.assigned_to_user_id:
            user_ids.add(entry.assigned_to_user_id)
        if entry.created_by_user_id:
            user_ids.add(entry.created_by_user_id)
    for task in tasks:
        if task.assigned_to:
            user_ids.add(task.assigned_to)
    for meeting in meetings:
        if meeting.created_by:
            user_ids.add(meeting.created_by)

    if user_ids:
        users = (await db.execute(select(User).where(User.id.in_(user_ids)))).scalars().all()
    else:
        users = []
    user_map = {u.id: (u.full_name or u.username or "") for u in users}

    data_by_day: dict[str, dict[str, list[str]]] = {
        iso: {
            "late": [],
            "absent": [],
            "leave": [],
            "blocked": [],
            "oneH": [],
            "personal": [],
            "external": [],
            "r1": [],
            "priority": [],
            "problem": [],
            "feedback": [],
        }
        for iso in week_isos
    }

    def add_for_day(iso: str, key: str, value: str) -> None:
        if iso not in data_by_day:
            return
        if value:
            data_by_day[iso][key].append(value)

    def entry_date_from(entry: CommonEntry) -> date | None:
        if entry.entry_date:
            return entry.entry_date
        if entry.description:
            date_match = DATE_LABEL_RE.search(entry.description)
            if date_match:
                return date.fromisoformat(date_match.group(1))
        return entry.created_at.date() if entry.created_at else None

    for entry in entries:
        person_id = entry.assigned_to_user_id or entry.created_by_user_id
        person_name = user_map.get(person_id, "") if person_id else ""
        person_label = person_name or entry.title or "Unknown"
        entry_date = entry_date_from(entry)
        if entry_date is None:
            continue
        entry_iso = entry_date.isoformat()

        if entry.category == CommonCategory.delays:
            note = entry.description or ""
            start = "08:00"
            until = "09:00"
            start_match = START_RE.search(note)
            if start_match:
                start = start_match.group(1)
            until_match = UNTIL_RE.search(note)
            if until_match:
                until = until_match.group(1)
            add_for_day(entry_iso, "late", f"{_initials(person_label)} {start}-{until}")
        elif entry.category == CommonCategory.absences:
            note = entry.description or ""
            from_time = "08:00"
            to_time = "23:00"
            from_to_match = FROM_TO_RE.search(note)
            if from_to_match:
                from_time = from_to_match.group(1)
                to_time = from_to_match.group(2)
            add_for_day(entry_iso, "absent", f"{_initials(person_label)} {from_time} - {to_time}")
        elif entry.category == CommonCategory.annual_leave:
            note = entry.description or ""
            start_date = entry_date
            end_date = entry_date
            range_match = DATE_RANGE_RE.search(note)
            if range_match:
                start_date = date.fromisoformat(range_match.group(1))
                end_date = date.fromisoformat(range_match.group(2))
            else:
                date_match = DATE_LABEL_RE.search(note)
                if date_match:
                    start_date = date.fromisoformat(date_match.group(1))
                    end_date = start_date
                else:
                    date_matches = DATE_RE.findall(note)
                    if date_matches:
                        start_date = date.fromisoformat(date_matches[0])
                        end_date = date.fromisoformat(date_matches[1]) if len(date_matches) > 1 else start_date

            full_day = "Full day" in note
            time_label = ""
            time_match = TIME_RANGE_RE.search(note)
            if not full_day and time_match:
                time_label = f"{time_match.group(1)}-{time_match.group(2)} "
            if full_day:
                time_label = "Full day "

            range_label = (
                f"{_format_excel_date(start_date)}-{_format_excel_date(end_date)}"
                if start_date != end_date
                else _format_excel_date(start_date)
            )
            leave_text = f"{_initials(person_label)} {time_label}{range_label}".strip()
            for day in week_dates:
                if start_date <= day <= end_date:
                    add_for_day(day.isoformat(), "leave", leave_text)
        elif entry.category == CommonCategory.blocks:
            note = entry.description or ""
            detail = f"{_initials(person_label)}: {note}" if note else _initials(person_label)
            add_for_day(entry_iso, "blocked", detail)
        elif entry.category == CommonCategory.external_tasks:
            add_for_day(entry_iso, "external", f"{entry.title} 14:00 ({_initials(person_label)})")
        elif entry.category == CommonCategory.problems:
            note = entry.description or ""
            detail = f"{_initials(person_label)}: {note}" if note else _initials(person_label)
            add_for_day(entry_iso, "problem", detail)
        elif entry.category in {CommonCategory.complaints, CommonCategory.requests, CommonCategory.proposals}:
            note = entry.description or ""
            detail = f"{_initials(person_label)}: {note}" if note else _initials(person_label)
            add_for_day(entry_iso, "feedback", detail)

    priority_map: dict[tuple[uuid.UUID, date], set[str]] = {}
    for task in tasks:
        if task.completed_at is not None or task.status == TaskStatusEnum.DONE.value:
            continue
        task_date = task.due_date or task.start_date or task.created_at
        if not task_date:
            continue
        day = task_date.date()
        if day.isoformat() not in data_by_day:
            continue
        assignee_names = assignees_by_task.get(task.id, [])
        if not assignee_names and task.assigned_to:
            fallback = user_map.get(task.assigned_to, "")
            if fallback:
                assignee_names = [fallback]

        if task.is_bllok:
            owner_label = assignee_names[0] if assignee_names else (user_map.get(task.assigned_to, "") if task.assigned_to else "Unknown")
            add_for_day(day.isoformat(), "blocked", f"{task.title} ({_initials(owner_label)})")
        if task.is_1h_report:
            owner_label = assignee_names[0] if assignee_names else (user_map.get(task.assigned_to, "") if task.assigned_to else "Unknown")
            add_for_day(day.isoformat(), "oneH", f"{task.title} ({_initials(owner_label)})")
        if task.is_personal:
            owner_label = assignee_names[0] if assignee_names else (user_map.get(task.assigned_to, "") if task.assigned_to else "Unknown")
            add_for_day(day.isoformat(), "personal", f"{task.title} ({_initials(owner_label)})")
        if task.is_r1:
            owner_label = assignee_names[0] if assignee_names else (user_map.get(task.assigned_to, "") if task.assigned_to else "Unknown")
            add_for_day(day.isoformat(), "r1", f"{task.title} ({_initials(owner_label)})")

        if task.project_id and task.project_id in project_name_map:
            key = (task.project_id, day)
            if key not in priority_map:
                priority_map[key] = set()
            for name in assignee_names:
                if name:
                    priority_map[key].add(name)

    for meeting in meetings:
        source = meeting.starts_at or meeting.created_at
        if not source:
            continue
        day = source.date()
        if day.isoformat() not in data_by_day:
            continue
        owner_label = user_map.get(meeting.created_by, "") if meeting.created_by else ""
        time_label = source.strftime("%H:%M") if meeting.starts_at else "TBD"
        owner_initials = _initials(owner_label) if owner_label else ""
        owner_suffix = f" ({owner_initials})" if owner_initials else ""
        add_for_day(day.isoformat(), "external", f"{meeting.title} {time_label}{owner_suffix}")

    for day_iso, day_data in data_by_day.items():
        priority_items = []
        for (project_id, day), assignees in sorted(priority_map.items(), key=lambda item: project_name_map.get(item[0][0], "")):
            if day.isoformat() != day_iso:
                continue
            assignee_initials = ", ".join(sorted({_initials(name) for name in assignees if name}))
            label = project_name_map.get(project_id, "")
            if assignee_initials:
                label = f"{label} [{assignee_initials}]"
            priority_items.append(label)
        day_data["priority"] = [item for item in priority_items if item]

    wb = Workbook()
    ws = wb.active
    ws.title = "Common View"

    last_col = 2 + len(week_dates)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value="COMMON VIEW")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_day_row = 3
    header_row = 4
    data_start_row = 5

    ws.cell(row=header_row, column=1, value="NR")
    ws.cell(row=header_row, column=2, value="LLOJI")
    for idx, day in enumerate(week_dates):
        col = 3 + idx
        ws.cell(row=header_day_row, column=col, value=f"{_day_code(day)} = {_format_excel_date(day)}".upper())
        ws.cell(row=header_row, column=col, value="KUSH/BZ ME/DET/SI/KUR/KUJT")

    row_specs = [
        ("late", "VONS"),
        ("absent", "MUNG"),
        ("leave", "PV/FEST"),
        ("external", "TAK EXT"),
        ("blocked", "BLL"),
        ("oneH", "1H"),
        ("personal", "P:"),
        ("r1", "R1"),
        ("priority", "PRJK"),
        ("problem", "PRBL"),
        ("feedback", "ANK/KRK/PRZ"),
    ]

    start_row = data_start_row
    for idx, (key, label) in enumerate(row_specs, start=1):
        row_idx = start_row + idx - 1
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=label.upper())
        for day_idx, iso in enumerate(week_isos):
            entries = data_by_day[iso][key]
            value = "\n".join(entries) if entries else ""
            ws.cell(row=row_idx, column=3 + day_idx, value=value)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    for col in range(3, 3 + len(week_dates)):
        ws.column_dimensions[get_column_letter(col)].width = 28

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 6
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 20

    header_font = Font(bold=True)
    label_font = Font(bold=True)
    for col in range(1, 3 + len(week_dates)):
        cell = ws.cell(row=header_day_row, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
        cell.number_format = "@"
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
        cell.number_format = "@"

    last_row = start_row + len(row_specs) - 1
    for row in range(start_row, last_row + 1):
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = label_font
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    table_top = header_day_row
    for r in range(table_top, last_row + 1):
        for c in range(1, last_col + 1):
            left = thick if c == 1 else thin
            right = thick if c == last_col else thin
            top = thick if r == table_top else thin
            bottom = thick if r == last_row else thin
            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

    for r in range(header_day_row, header_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            left = thick if c == 1 else thin
            right = thick if c == last_col else thin
            top = thick if r == header_day_row else thin
            bottom = thick if r == header_row else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"

    ws.freeze_panes = "B5"
    ws.print_title_rows = f"{header_day_row}:{header_row}"
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2
    ws.oddHeader.right.text = "&D &T"
    ws.oddFooter.center.text = "Page &P / &N"
    user_initials = _initials(user.full_name or user.username or "")
    ws.oddFooter.right.text = f"PUNOI: {user_initials or '____'}"
    ws.evenHeader.right.text = ws.oddHeader.right.text
    ws.evenFooter.center.text = ws.oddFooter.center.text
    ws.evenFooter.right.text = ws.oddFooter.right.text
    ws.firstHeader.right.text = ws.oddHeader.right.text
    ws.firstFooter.center.text = ws.oddFooter.center.text
    ws.firstFooter.right.text = ws.oddFooter.right.text

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename_date = f"{week_dates[0].day:02d}_{week_dates[0].month:02d}_{str(week_dates[0].year)[-2:]}"
    initials_value = user_initials or "USER"
    filename = f"COMMON VIEW {filename_date}_EF ({initials_value}).xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/weekly-planner.xlsx")
async def export_weekly_planner_xlsx(
    week_start: date | None = None,
    department_id: uuid.UUID | None = None,
    is_this_week: bool = False,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    if department_id is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="department_id is required")

    data = await weekly_table_planner(
        week_start=week_start,
        department_id=department_id,
        is_this_week=is_this_week,
        db=db,
        user=user,
    )

    if not data.departments:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No weekly planner data found")

    dept = data.departments[0]
    dept_label = _display_department_name(dept.department_name)
    week_start_date = data.week_start
    week_end_date = data.week_end

    # Add "THIS WEEK PLAN" or "NEXT WEEK PLAN" label
    week_label = "THIS WEEK PLAN" if is_this_week else "NEXT WEEK PLAN"
    title_label = f"{dept_label} {week_start_date.day:02d}-{week_start_date.month:02d}-{week_start_date.year} - {week_end_date.day:02d}-{week_end_date.month:02d}-{week_end_date.year} - {week_label}"
    title_upper = title_label.upper()

    # Collect users in stable order
    user_map: dict[uuid.UUID, str] = {}
    for day in dept.days:
        for user_day in day.users:
            user_map.setdefault(user_day.user_id, user_day.user_name or "")
    user_ids = list(user_map.keys())
    user_names = [user_map[user_id] for user_id in user_ids]

    wb = Workbook()
    ws = wb.active
    ws.title = dept_label[:31] if dept_label else "Weekly Planner"

    status_fills = {
        "TODO": PatternFill(start_color="FFC4ED", end_color="FFC4ED", fill_type="solid"),
        "IN_PROGRESS": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "DONE": PatternFill(start_color="C4FDC4", end_color="C4FDC4", fill_type="solid"),
    }

    last_col = 4 + len(user_names)  # NR, DAY, LL, TIME + users

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=title_upper)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10

    header_row = 4
    data_start_row = 5
    ws.row_dimensions[header_row].height = 15

    headers = ["NR", "DAY", "LLOJI", "TIME"] + [name.upper() for name in user_names]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 6
    for idx in range(5, last_col + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 28

    # Fill data rows (expand rows so each task can be colored by status)
    current_row = data_start_row
    for day_index, day in enumerate(dept.days, start=1):
        day_name = DAY_NAMES[day_index - 1].upper() if 0 <= day_index - 1 < len(DAY_NAMES) else ""
        day_date = f"{day.date.day:02d}-{day.date.month:02d}-{day.date.year}"
        day_label = f"{day_name}\n{day_date}".strip()

        slot_specs = [
            ("PRJK", "AM", False),
            ("FT", "AM", True),
            ("PRJK", "PM", False),
            ("FT", "PM", True),
        ]

        slot_rows: list[dict[str, object]] = []
        for ll_label, time_label, include_fast in slot_specs:
            per_user_items: list[list[dict[str, str | None]]] = []
            max_items = 0
            for user_id in user_ids:
                user_day = next((u for u in day.users if u.user_id == user_id), None)
                projects = []
                system_tasks = []
                fast_tasks = []
                if user_day:
                    if time_label == "AM":
                        projects = user_day.am_projects or []
                        system_tasks = user_day.am_system_tasks or []
                        fast_tasks = user_day.am_fast_tasks or []
                    else:
                        projects = user_day.pm_projects or []
                        system_tasks = user_day.pm_system_tasks or []
                        fast_tasks = user_day.pm_fast_tasks or []

                if include_fast:
                    projects = []
                    system_tasks = []

                items = _planner_cell_items(
                    projects=projects,
                    system_tasks=system_tasks,
                    fast_tasks=fast_tasks,
                    include_fast=include_fast,
                    day_date=day.date,
                )
                per_user_items.append(items)
                max_items = max(max_items, len(items))

            slot_rows.append(
                {
                    "ll_label": ll_label,
                    "time_label": time_label,
                    "include_fast": include_fast,
                    "row_count": max(1, max_items),
                    "per_user_items": per_user_items,
                }
            )

        total_day_rows = sum(slot["row_count"] for slot in slot_rows)
        base_row = current_row

        # Merge NR and DAY across all rows for this day
        ws.merge_cells(start_row=base_row, start_column=1, end_row=base_row + total_day_rows - 1, end_column=1)
        ws.merge_cells(start_row=base_row, start_column=2, end_row=base_row + total_day_rows - 1, end_column=2)
        nr_cell = ws.cell(row=base_row, column=1, value=day_index)
        nr_cell.font = Font(bold=True)
        nr_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
        day_cell = ws.cell(row=base_row, column=2, value=day_label)
        day_cell.font = Font(bold=True)
        day_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

        row_cursor = base_row
        for slot in slot_rows:
            slot_start = row_cursor
            row_count = slot["row_count"]
            ll_label = slot["ll_label"]
            time_label = slot["time_label"]
            per_user_items = slot["per_user_items"]

            if row_count > 1:
                ws.merge_cells(start_row=slot_start, start_column=3, end_row=slot_start + row_count - 1, end_column=3)
                ws.merge_cells(start_row=slot_start, start_column=4, end_row=slot_start + row_count - 1, end_column=4)

            ll_cell = ws.cell(row=slot_start, column=3, value=ll_label)
            ll_cell.font = Font(bold=True)
            ll_cell.alignment = Alignment(horizontal="left", vertical="bottom")
            time_cell = ws.cell(row=slot_start, column=4, value=time_label)
            time_cell.font = Font(bold=True)
            time_cell.alignment = Alignment(horizontal="left", vertical="bottom")

            for offset in range(row_count):
                row_idx = slot_start + offset
                for u_idx, _ in enumerate(user_ids):
                    items = per_user_items[u_idx]
                    item = items[offset] if offset < len(items) else None
                    cell_value = _planner_item_rich_text(item) if item else ""
                    cell = ws.cell(row=row_idx, column=5 + u_idx, value=cell_value)
                    cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
                    if item and item.get("status") in status_fills:
                        cell.fill = status_fills[item["status"]]

            row_cursor += row_count

        current_row = base_row + total_day_rows

    last_row = current_row - 1

    # Borders
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    for r in range(header_row, last_row + 1):
        for c in range(1, last_col + 1):
            left = thick if c == 1 else thin
            right = thick if c == last_col else thin
            top = thick if r == header_row else thin
            bottom = thick if r == last_row else thin
            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

    # Header row thicker outline
    for c in range(1, last_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.border = Border(
            left=thick if c == 1 else thin,
            right=thick if c == last_col else thin,
            top=thick,
            bottom=thick,
        )

    # Alignment and formats
    for r in range(header_row, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.alignment:
                cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
            else:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "left",
                    vertical="bottom",
                    wrap_text=True,
                )

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "B5"
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    user_initials = _initials(user.full_name or user.username or "")
    ws.oddHeader.right.text = "&D &T"
    ws.oddFooter.center.text = "Page &P / &N"
    ws.oddFooter.right.text = f"PUNOI: {user_initials or '____'}"
    ws.evenHeader.right.text = ws.oddHeader.right.text
    ws.evenFooter.center.text = ws.oddFooter.center.text
    ws.evenFooter.right.text = ws.oddFooter.right.text
    ws.firstHeader.right.text = ws.oddHeader.right.text
    ws.firstFooter.center.text = ws.oddFooter.center.text
    ws.firstFooter.right.text = ws.oddFooter.right.text

    # Number column formatting
    for r in range(data_start_row, last_row + 1):
        cell = ws.cell(row=r, column=1)
        cell.number_format = "#,##0"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    today = datetime.now(timezone.utc).date()
    initials_value = user_initials or "USER"
    filename_title = _safe_filename(title_upper)
    filename = f"{filename_title}_{today.day:02d}_{today.month:02d}_{str(today.year)[-2:]} ({initials_value}).xlsx"

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/weekly-snapshot.xlsx")
async def export_weekly_snapshot_xlsx(
    snapshot_id: uuid.UUID = Query(..., description="Weekly planner snapshot id"),
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    snapshot = (
        await db.execute(select(WeeklyPlannerSnapshot).where(WeeklyPlannerSnapshot.id == snapshot_id))
    ).scalar_one_or_none()
    if snapshot is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Snapshot not found")

    ensure_department_access(user, snapshot.department_id)

    payload = snapshot.payload or {}
    department_payload = payload.get("department")
    if not department_payload:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Snapshot payload missing department")

    try:
        dept = WeeklyTableDepartment.model_validate(department_payload)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Invalid snapshot payload") from exc

    def _parse_iso_date(value: object | None, fallback: date | None) -> date | None:
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            try:
                return date.fromisoformat(value[:10])
            except Exception:
                return fallback
        return fallback

    week_start_date = _parse_iso_date(payload.get("week_start"), snapshot.week_start_date)
    week_end_date = _parse_iso_date(payload.get("week_end"), snapshot.week_end_date)
    if week_start_date is None or week_end_date is None:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Snapshot payload missing week range")

    official_snapshot_id = (
        await db.execute(
            select(WeeklyPlannerSnapshot.id)
            .where(
                WeeklyPlannerSnapshot.department_id == snapshot.department_id,
                WeeklyPlannerSnapshot.week_start_date == snapshot.week_start_date,
                WeeklyPlannerSnapshot.snapshot_type == snapshot.snapshot_type,
            )
            .order_by(WeeklyPlannerSnapshot.created_at.asc())
            .limit(1)
        )
    ).scalar_one_or_none()
    is_official = official_snapshot_id == snapshot.id

    snapshot_type_label = (snapshot.snapshot_type or "").upper()
    if snapshot_type_label == "PLANNED":
        week_label = "PLANNED SNAPSHOT"
    elif snapshot_type_label == "FINAL":
        week_label = "FINAL SNAPSHOT"
    else:
        week_label = f"{snapshot_type_label} SNAPSHOT".strip()
    if is_official:
        week_label = f"OFFICIAL {week_label}".strip()

    dept_label = _display_department_name(dept.department_name)
    title_label = (
        f"{dept_label} "
        f"{week_start_date.day:02d}-{week_start_date.month:02d}-{week_start_date.year} - "
        f"{week_end_date.day:02d}-{week_end_date.month:02d}-{week_end_date.year} - "
        f"{week_label}"
    )
    title_upper = title_label.upper()

    # Collect users in stable order
    user_map: dict[uuid.UUID, str] = {}
    for day in dept.days:
        for user_day in day.users:
            user_map.setdefault(user_day.user_id, user_day.user_name or "")
    user_ids = list(user_map.keys())
    user_names = [user_map[user_id] for user_id in user_ids]

    wb = Workbook()
    ws = wb.active
    ws.title = dept_label[:31] if dept_label else "Weekly Snapshot"

    status_fills = {
        "TODO": PatternFill(start_color="FFC4ED", end_color="FFC4ED", fill_type="solid"),
        "IN_PROGRESS": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "DONE": PatternFill(start_color="C4FDC4", end_color="C4FDC4", fill_type="solid"),
    }

    last_col = 4 + len(user_names)  # NR, DAY, LL, TIME + users

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=title_upper)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10

    header_row = 4
    data_start_row = 5
    ws.row_dimensions[header_row].height = 15

    headers = ["NR", "DAY", "LLOJI", "TIME"] + [name.upper() for name in user_names]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 6
    for idx in range(5, last_col + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 28

    current_row = data_start_row
    for day_index, day in enumerate(dept.days, start=1):
        day_name = DAY_NAMES[day_index - 1].upper() if 0 <= day_index - 1 < len(DAY_NAMES) else ""
        day_date = f"{day.date.day:02d}-{day.date.month:02d}-{day.date.year}"
        day_label = f"{day_name}\n{day_date}".strip()

        slot_specs = [
            ("PRJK", "AM", False),
            ("FT", "AM", True),
            ("PRJK", "PM", False),
            ("FT", "PM", True),
        ]

        slot_rows: list[dict[str, object]] = []
        for ll_label, time_label, include_fast in slot_specs:
            per_user_items: list[list[dict[str, str | None]]] = []
            max_items = 0
            for user_id in user_ids:
                user_day = next((u for u in day.users if u.user_id == user_id), None)
                projects = []
                system_tasks = []
                fast_tasks = []
                if user_day:
                    if time_label == "AM":
                        projects = user_day.am_projects or []
                        system_tasks = user_day.am_system_tasks or []
                        fast_tasks = user_day.am_fast_tasks or []
                    else:
                        projects = user_day.pm_projects or []
                        system_tasks = user_day.pm_system_tasks or []
                        fast_tasks = user_day.pm_fast_tasks or []

                if include_fast:
                    projects = []
                    system_tasks = []

                items = _planner_cell_items(
                    projects=projects,
                    system_tasks=system_tasks,
                    fast_tasks=fast_tasks,
                    include_fast=include_fast,
                    day_date=day.date,
                )
                per_user_items.append(items)
                max_items = max(max_items, len(items))

            slot_rows.append(
                {
                    "ll_label": ll_label,
                    "time_label": time_label,
                    "include_fast": include_fast,
                    "row_count": max(1, max_items),
                    "per_user_items": per_user_items,
                }
            )

        total_day_rows = sum(slot["row_count"] for slot in slot_rows)
        base_row = current_row

        ws.merge_cells(start_row=base_row, start_column=1, end_row=base_row + total_day_rows - 1, end_column=1)
        ws.merge_cells(start_row=base_row, start_column=2, end_row=base_row + total_day_rows - 1, end_column=2)
        nr_cell = ws.cell(row=base_row, column=1, value=day_index)
        nr_cell.font = Font(bold=True)
        nr_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
        day_cell = ws.cell(row=base_row, column=2, value=day_label)
        day_cell.font = Font(bold=True)
        day_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

        row_cursor = base_row
        for slot in slot_rows:
            slot_start = row_cursor
            row_count = slot["row_count"]
            ll_label = slot["ll_label"]
            time_label = slot["time_label"]
            per_user_items = slot["per_user_items"]

            if row_count > 1:
                ws.merge_cells(start_row=slot_start, start_column=3, end_row=slot_start + row_count - 1, end_column=3)
                ws.merge_cells(start_row=slot_start, start_column=4, end_row=slot_start + row_count - 1, end_column=4)

            ll_cell = ws.cell(row=slot_start, column=3, value=ll_label)
            ll_cell.font = Font(bold=True)
            ll_cell.alignment = Alignment(horizontal="left", vertical="bottom")
            time_cell = ws.cell(row=slot_start, column=4, value=time_label)
            time_cell.font = Font(bold=True)
            time_cell.alignment = Alignment(horizontal="left", vertical="bottom")

            for offset in range(row_count):
                row_idx = slot_start + offset
                for u_idx, _ in enumerate(user_ids):
                    items = per_user_items[u_idx]
                    item = items[offset] if offset < len(items) else None
                    cell_value = _planner_item_rich_text(item) if item else ""
                    cell = ws.cell(row=row_idx, column=5 + u_idx, value=cell_value)
                    cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
                    if item and item.get("status") in status_fills:
                        cell.fill = status_fills[item["status"]]

            row_cursor += row_count

        current_row = base_row + total_day_rows

    last_row = current_row - 1

    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    for r in range(header_row, last_row + 1):
        for c in range(1, last_col + 1):
            left = thick if c == 1 else thin
            right = thick if c == last_col else thin
            top = thick if r == header_row else thin
            bottom = thick if r == last_row else thin
            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

    for c in range(1, last_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.border = Border(
            left=thick if c == 1 else thin,
            right=thick if c == last_col else thin,
            top=thick,
            bottom=thick,
        )

    for r in range(header_row, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.alignment:
                cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
            else:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "left",
                    vertical="bottom",
                    wrap_text=True,
                )

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "B5"
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    user_initials = _initials(user.full_name or user.username or "")
    ws.oddHeader.right.text = "&D &T"
    ws.oddFooter.center.text = "Page &P / &N"
    ws.oddFooter.right.text = f"PUNOI: {user_initials or '____'}"
    ws.evenHeader.right.text = ws.oddHeader.right.text
    ws.evenFooter.center.text = ws.oddFooter.center.text
    ws.evenFooter.right.text = ws.oddFooter.right.text
    ws.firstHeader.right.text = ws.oddHeader.right.text
    ws.firstFooter.center.text = ws.oddFooter.center.text
    ws.firstFooter.right.text = ws.oddFooter.right.text

    for r in range(data_start_row, last_row + 1):
        cell = ws.cell(row=r, column=1)
        cell.number_format = "#,##0"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    today = datetime.now(timezone.utc).date()
    initials_value = user_initials or "USER"
    filename_date = f"{today.day:02d}_{today.month:02d}_{str(today.year)[-2:]}"
    filename_title = _safe_filename_spaces(title_upper)
    filename = f"{filename_title} {filename_date}_EF ({initials_value}).xlsx"

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/weekly-plan-vs-final.xlsx")
async def export_weekly_plan_vs_final_xlsx(
    department_id: uuid.UUID,
    week_start: date,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    ensure_department_access(user, department_id)

    normalized_week_start = planners_router._week_start(week_start)
    week_dates = planners_router._get_next_5_working_days(normalized_week_start)
    week_end = week_dates[-1]

    department = (await db.execute(select(Department).where(Department.id == department_id))).scalar_one_or_none()
    dept_label = _display_department_name(department.name if department is not None else "")

    title_label = (
        f"{dept_label} "
        f"{_format_excel_date(normalized_week_start)} - {_format_excel_date(week_end)} - "
        "PLAN VS FINAL REPORT"
    ).strip()
    title_upper = title_label.upper() if title_label else "PLAN VS FINAL REPORT"

    planned_official_snapshot = (
        await db.execute(
            select(WeeklyPlannerSnapshot)
            .where(
                WeeklyPlannerSnapshot.department_id == department_id,
                WeeklyPlannerSnapshot.week_start_date == normalized_week_start,
                WeeklyPlannerSnapshot.snapshot_type == WeeklySnapshotType.PLANNED.value,
            )
            .order_by(WeeklyPlannerSnapshot.created_at.asc())
            .limit(1)
        )
    ).scalar_one_or_none()

    latest_final_snapshot = (
        await db.execute(
            select(WeeklyPlannerSnapshot)
            .where(
                WeeklyPlannerSnapshot.department_id == department_id,
                WeeklyPlannerSnapshot.week_start_date == normalized_week_start,
                WeeklyPlannerSnapshot.snapshot_type == WeeklySnapshotType.FINAL.value,
            )
            .order_by(WeeklyPlannerSnapshot.created_at.desc())
            .limit(1)
        )
    ).scalar_one_or_none()

    def _assignee_columns_from_snapshot(snapshot: WeeklyPlannerSnapshot | None) -> list[tuple[uuid.UUID | None, str]]:
        if snapshot is None:
            return []
        payload = snapshot.payload or {}
        department_payload = payload.get("department")
        if not department_payload:
            return []
        try:
            dept = WeeklyTableDepartment.model_validate(department_payload)
        except Exception:
            return []

        seen: set[str] = set()
        columns: list[tuple[uuid.UUID | None, str]] = []
        for day in dept.days:
            for user_day in day.users:
                name = (user_day.user_name or "").strip()
                if not name:
                    continue
                user_id = getattr(user_day, "user_id", None)
                key = str(user_id) if user_id is not None else f"name:{name.lower()}"
                if key in seen:
                    continue
                seen.add(key)
                columns.append((user_id, name))

        columns.sort(key=lambda item: item[1].lower())
        return columns

    # Prefer planned snapshot for column ordering (matches UI).
    assignee_columns = _assignee_columns_from_snapshot(planned_official_snapshot) or _assignee_columns_from_snapshot(latest_final_snapshot)

    message: str | None = None
    groups: list[dict[str, object]] = []
    if planned_official_snapshot is None:
        message = planners_router.NO_PLAN_SNAPSHOT_MESSAGE
    elif latest_final_snapshot is None:
        message = "No final snapshot found for this week. Save This Week (Final) first."
    else:
        planned_tasks_by_key = planners_router._tasks_by_key_from_snapshot_payload(planned_official_snapshot.payload or {})
        final_tasks_by_key = planners_router._tasks_by_key_from_snapshot_payload(latest_final_snapshot.payload or {})

        as_of_date = week_end + timedelta(days=1)
        buckets = planners_router._classify_weekly_plan_performance(
            planned_tasks=planned_tasks_by_key,
            actual_tasks=final_tasks_by_key,
            week_end=week_end,
            as_of_date=as_of_date,
        )

        completed = [planners_router._to_compare_task_out(task) for task in buckets.get("completed", [])]
        in_progress = [planners_router._to_compare_task_out(task) for task in buckets.get("in_progress", [])]
        pending = [planners_router._to_compare_task_out(task) for task in buckets.get("pending", [])]
        late = [planners_router._to_compare_task_out(task) for task in buckets.get("late", [])]
        additional = [planners_router._to_compare_task_out(task) for task in buckets.get("additional", [])]
        removed_or_canceled = [
            planners_router._to_compare_task_out(task) for task in buckets.get("removed_or_canceled", [])
        ]

        grouped = planners_router._group_compare_tasks_by_assignee(
            completed=completed,
            in_progress=in_progress,
            pending=pending,
            late=late,
            additional=additional,
            removed_or_canceled=removed_or_canceled,
        )

        if grouped:
            groups = [
                {
                    "assignee_id": g.assignee_id,
                    "assignee_name": g.assignee_name,
                    "completed": g.completed or [],
                    "in_progress": g.in_progress or [],
                    "pending": g.pending or [],
                    "late": g.late or [],
                    "additional": g.additional or [],
                    "removed_or_canceled": g.removed_or_canceled or [],
                }
                for g in grouped
            ]

    if not groups and assignee_columns:
        groups = [
            {
                "assignee_id": assignee_id,
                "assignee_name": assignee_name,
                "completed": [],
                "in_progress": [],
                "pending": [],
                "late": [],
                "additional": [],
                "removed_or_canceled": [],
            }
            for assignee_id, assignee_name in assignee_columns
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "PLAN VS FINAL"[:31]

    # Columns: NR, STATUS, then one column per user
    last_col = max(2, 2 + len(groups))

    # Row 1: merged title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=title_upper)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=1)
    ws.row_dimensions[1].height = 24

    # Row 2-3: spacing
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10

    user_initials = _initials(user.full_name or user.username or "")

    # Standard print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2
    ws.oddHeader.right.text = "&D &T"
    ws.oddFooter.center.text = "Page &P / &N"
    ws.oddFooter.right.text = f"PUNOI: {user_initials or '____'}"
    ws.evenHeader.right.text = ws.oddHeader.right.text
    ws.evenFooter.center.text = ws.oddFooter.center.text
    ws.evenFooter.right.text = ws.oddFooter.right.text
    ws.firstHeader.right.text = ws.oddHeader.right.text
    ws.firstFooter.center.text = ws.oddFooter.center.text
    ws.firstFooter.right.text = ws.oddFooter.right.text

    if message:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
        msg_cell = ws.cell(row=4, column=1, value=str(message))
        msg_cell.font = Font(bold=True)
        msg_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
        ws.row_dimensions[4].height = 30
        ws.print_area = f"A1:{get_column_letter(last_col)}4"
    else:
        header_row = 4
        data_start_row = 5

        # Column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 18
        for idx in range(3, last_col + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 44

        # Header row
        ws.row_dimensions[header_row].height = 22
        nr_header = ws.cell(row=header_row, column=1, value="NR")
        nr_header.font = Font(bold=True)
        nr_header.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

        status_header = ws.cell(row=header_row, column=2, value="STATUS")
        status_header.font = Font(bold=True)
        status_header.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

        for col_idx, group in enumerate(groups, start=3):
            label = str(group.get("assignee_name") or "").upper()
            cell = ws.cell(row=header_row, column=col_idx, value=label)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

        categories: list[tuple[str, str]] = [
            ("completed", "COMPLETED"),
            ("in_progress", "IN PROGRESS"),
            ("pending", "PENDING"),
            ("late", "LATE"),
            ("additional", "ADDITIONAL"),
            ("removed_or_canceled", "REMOVED/CANCELED"),
        ]

        def _cell_text(tasks: list[object]) -> str:
            if not tasks:
                return "-"
            lines: list[str] = [str(len(tasks))]
            for task in tasks:
                title = getattr(task, "title", None) or ""
                project_title = getattr(task, "project_title", None)
                suffix = f" ({project_title})" if project_title else ""
                lines.append(f"{title}{suffix}".strip())
            value = "\n".join(lines).strip()
            if len(value) > 32000:
                value = value[:31950].rstrip() + "\n...(TRUNCATED)"
            return value

        row_idx = data_start_row
        for idx, (key, label) in enumerate(categories, start=1):
            nr_cell = ws.cell(row=row_idx, column=1, value=idx)
            nr_cell.font = Font(bold=True)
            nr_cell.number_format = "#,##0"
            nr_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

            label_cell = ws.cell(row=row_idx, column=2, value=label)
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
            for col_idx, group in enumerate(groups, start=3):
                tasks = group.get(key) or []
                cell = ws.cell(row=row_idx, column=col_idx, value=_cell_text(list(tasks)))
                cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
            row_idx += 1

        last_row = row_idx - 1

        # Filters + freeze + print titles
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
        ws.freeze_panes = "C5" if last_col >= 3 else "B5"
        ws.print_title_rows = f"{header_row}:{header_row}"
        ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"

        # Borders (thick outside, thin inside; header thick outline)
        thin = Side(style="thin", color="000000")
        thick = Side(style="medium", color="000000")
        for r in range(header_row, last_row + 1):
            for c in range(1, last_col + 1):
                left = thick if c == 1 else thin
                right = thick if c == last_col else thin
                top = thick if r == header_row else thin
                bottom = thick if r == last_row else thin
                ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

        for c in range(1, last_col + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.border = Border(
                left=thick if c == 1 else thin,
                right=thick if c == last_col else thin,
                top=thick,
                bottom=thick,
            )

        # Row heights (based on max wrapped lines)
        for r in range(data_start_row, last_row + 1):
            max_lines = 1
            for c in range(1, last_col + 1):
                value = ws.cell(row=r, column=c).value
                lines = str(value).count("\n") + 1 if value is not None else 1
                max_lines = max(max_lines, lines)
            ws.row_dimensions[r].height = max(18, min(300, 14 * max_lines))

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    today = datetime.now(timezone.utc).date()
    initials_value = user_initials or "USER"
    filename_date = f"{today.day:02d}_{today.month:02d}_{str(today.year)[-2:]}"
    filename_title = _safe_filename_spaces(title_upper)
    filename = f"{filename_title} {filename_date}_EF ({initials_value}).xlsx"

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/ga-notes.xlsx")
async def export_ga_notes_xlsx(
    department_id: uuid.UUID | None = None,
    project_id: uuid.UUID | None = None,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """Export GA/KA notes to Excel"""
    from datetime import timedelta
    from app.models.enums import GaNoteStatus
    
    # Build query
    closed_cutoff = datetime.utcnow() - timedelta(days=30)

    stmt = select(GaNote).order_by(GaNote.created_at.desc())

    # Include all open notes; include closed notes only if recently closed
    stmt = stmt.where(
        or_(
            GaNote.status != GaNoteStatus.CLOSED,
            GaNote.completed_at.is_(None),
            GaNote.completed_at >= closed_cutoff,
        )
    )
    
    if project_id is not None:
        project = (await db.execute(select(Project).where(Project.id == project_id))).scalar_one_or_none()
        if project is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
        stmt = stmt.where(GaNote.project_id == project_id)
    elif department_id is not None:
        ensure_department_access(user, department_id)
        stmt = stmt.where(GaNote.department_id == department_id)
    elif user.role == UserRole.STAFF:
        # STAFF users can only see their department's notes
        if user.department_id:
            stmt = stmt.where(GaNote.department_id == user.department_id)
    
    notes = (await db.execute(stmt)).scalars().all()
    
    # Get related data
    user_ids = {n.created_by for n in notes if n.created_by}
    department_ids = {n.department_id for n in notes if n.department_id}
    project_ids = {n.project_id for n in notes if n.project_id}
    
    users_map = {}
    if user_ids:
        users = (await db.execute(select(User).where(User.id.in_(user_ids)))).scalars().all()
        users_map = {u.id: u for u in users}
    
    departments_map = {}
    if department_ids:
        departments = (await db.execute(select(Department).where(Department.id.in_(department_ids)))).scalars().all()
        departments_map = {d.id: d for d in departments}
    
    projects_map = {}
    if project_ids:
        projects = (await db.execute(select(Project).where(Project.id.in_(project_ids)))).scalars().all()
        projects_map = {p.id: p for p in projects}
    
    # Helper functions
    def get_user_initials(user_id: uuid.UUID | None) -> str:
        if not user_id or user_id not in users_map:
            return "-"
        u = users_map[user_id]
        return _initials(u.full_name or u.username or "")
    
    def get_department_abbrev(dept_id: uuid.UUID | None) -> str:
        if not dept_id or dept_id not in departments_map:
            return "-"
        dept_name = departments_map[dept_id].name
        key = dept_name.strip().upper()
        return {
            "DEVELOPMENT": "DEV",
            "GRAPHIC DESIGN": "GDS",
            "PRODUCT CONTENT": "PCM",
            "PROJECT CONTENT": "PCM",
        }.get(key, dept_name[:3].upper() if dept_name else "-")
    
    def format_note_date(dt: datetime | None) -> str:
        if not dt:
            return "-"
        return dt.strftime("%d.%m, %I:%M %p")
    
    # Build rows
    headers = ["NR", "SHENIMI", "DATA,ORA", "NGA", "DEP", "PRJK", "KRIJO DETYRE", "MBYLL SHENIM"]
    rows = []
    
    for idx, note in enumerate(notes, start=1):
        creator_initials = get_user_initials(note.created_by)
        dept_abbrev = get_department_abbrev(note.department_id) if note.department_id else "-"
        project_name = "-"
        if note.project_id and note.project_id in projects_map:
            p = projects_map[note.project_id]
            project_name = p.title or p.name or "Project"
        
        task_status = "Task Created" if note.is_converted_to_task else "No Task"
        note_status = "Closed" if note.status == GaNoteStatus.CLOSED else "Open"
        
        # Clean content (remove HTML if any)
        content = _strip_html(note.content) if note.content else ""
        
        rows.append([
            str(idx),
            content,
            format_note_date(note.created_at),
            creator_initials,
            dept_abbrev,
            project_name,
            task_status,
            note_status,
        ])
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "GA-KA Notes"
    
    # Define border styles
    thin = Side(style="thin", color="475569")
    thick = Side(style="thick", color="1e293b")
    
    # First two rows are empty
    # Title row (row 3)
    title_row = 3
    last_col = len(headers)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=last_col)
    title_cell = ws.cell(row=title_row, column=1, value="GA/KA NOTES")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Header row (row 4)
    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header.upper())
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
        # Thick outside borders for header
        left_border = thick if col_idx == 1 else thin
        right_border = thick if col_idx == last_col else thin
        top_border = thick
        bottom_border = thick
        cell.border = Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)
    
    # Data rows (starting from row 4)
    last_row = header_row + len(rows)
    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
            # Bold for NR column
            if col_idx == 1:
                cell.font = Font(bold=True)
            # Thick outside borders
            left_border = thick if col_idx == 1 else thin
            right_border = thick if col_idx == last_col else thin
            top_border = thin
            bottom_border = thick if row_idx == last_row else thin
            cell.border = Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)
    
    # Column widths
    ws.column_dimensions["A"].width = 5  # NR
    ws.column_dimensions["B"].width = 50  # SHENIMI
    ws.column_dimensions["C"].width = 18  # DATA,ORA
    ws.column_dimensions["D"].width = 8   # NGA
    ws.column_dimensions["E"].width = 8   # DEP
    ws.column_dimensions["F"].width = 20  # PRJK
    ws.column_dimensions["G"].width = 15  # KRIJO DETYRE
    ws.column_dimensions["H"].width = 15  # MBYLL SHENIM
    
    # Freeze panes: freeze column A (NR) and header row
    ws.freeze_panes = "B5"
    
    # Auto filter for all columns
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    
    # Page setup and margins
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2
    
    # Headers and footers
    now = datetime.now(timezone.utc)
    date_str = now.strftime("%d.%m.%Y")
    time_str = now.strftime("%I:%M %p")
    ws.oddHeader.right.text = f"{date_str}, {time_str}"
    ws.oddFooter.center.text = "Page &P / &N"
    user_initials = _initials(user.full_name or user.username or "") or "USER"
    ws.oddFooter.right.text = f"PUNOI: {user_initials}"
    
    # Apply to even and first pages as well
    ws.evenHeader.right.text = ws.oddHeader.right.text
    ws.evenFooter.center.text = ws.oddFooter.center.text
    ws.evenFooter.right.text = ws.oddFooter.right.text
    ws.firstHeader.right.text = ws.oddHeader.right.text
    ws.firstFooter.center.text = ws.oddFooter.center.text
    ws.firstFooter.right.text = ws.oddFooter.right.text
    
    # Print title rows (repeat header on each page)
    ws.print_title_rows = f"{header_row}:{header_row}"
    
    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    # Generate filename: GA/KA_NOTES DD_MM_YY_USER_INITIALS
    today = datetime.now(timezone.utc).date()
    filename_date = f"{today.day:02d}_{today.month:02d}_{str(today.year)[-2:]}"
    filename = f"GA/KA_NOTES {filename_date}_{user_initials}.xlsx"
    
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/weekly-plan-vs-actual.xlsx")
async def export_weekly_plan_vs_actual_xlsx(
    department_id: uuid.UUID,
    week_start: date,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    ensure_department_access(user, department_id)

    normalized_week_start = planners_router._week_start(week_start)
    week_dates = planners_router._get_next_5_working_days(normalized_week_start)
    week_end = week_dates[-1]

    department = (await db.execute(select(Department).where(Department.id == department_id))).scalar_one_or_none()
    dept_label = _display_department_name(department.name if department is not None else "")

    title_label = (
        f"{dept_label} "
        f"{_format_excel_date(normalized_week_start)} - {_format_excel_date(week_end)} - "
        "PLAN VS ACTUAL REPORT"
    ).strip()
    title_upper = title_label.upper() if title_label else "PLAN VS ACTUAL REPORT"

    planned_official_snapshot = (
        await db.execute(
            select(WeeklyPlannerSnapshot)
            .where(
                WeeklyPlannerSnapshot.department_id == department_id,
                WeeklyPlannerSnapshot.week_start_date == normalized_week_start,
                WeeklyPlannerSnapshot.snapshot_type == WeeklySnapshotType.PLANNED.value,
            )
            .order_by(WeeklyPlannerSnapshot.created_at.asc())
            .limit(1)
        )
    ).scalar_one_or_none()

    if planned_official_snapshot is None:
        wb = Workbook()
        ws = wb.active
        ws.title = "PLAN VS ACTUAL"[:31]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        title_cell = ws.cell(row=1, column=1, value=title_upper)
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=1)
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 10
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
        msg_cell = ws.cell(row=3, column=1, value=planners_router.NO_PLAN_SNAPSHOT_MESSAGE)
        msg_cell.font = Font(bold=True)
        msg_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
        ws.row_dimensions[3].height = 30

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        today = datetime.now(timezone.utc).date()
        initials_value = _initials(user.full_name or user.username or "") or "USER"
        filename_date = f"{today.day:02d}_{today.month:02d}_{str(today.year)[-2:]}"
        filename_title = _safe_filename_spaces(title_upper)
        filename = f"{filename_title} {filename_date}_EF ({initials_value}).xlsx"

        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
        )

    snapshot_tasks_by_key = planners_router._tasks_by_key_from_snapshot_payload(planned_official_snapshot.payload or {})

    current_weekly_table = await weekly_table_planner(
        week_start=normalized_week_start,
        department_id=department_id,
        is_this_week=False,
        db=db,
        user=user,
    )
    current_table_json = jsonable_encoder(current_weekly_table)
    current_departments = current_table_json.get("departments") or []
    current_department_payload = current_departments[0] if current_departments else None
    current_task_ids = planners_router._task_ids_from_department_payload(current_department_payload)
    current_task_priorities = await planners_router._load_task_priority_map(db, current_task_ids)
    current_task_items = planners_router._flatten_weekly_department_tasks(
        current_department_payload,
        task_priority_map=current_task_priorities,
    )
    current_tasks_by_key = {
        task["match_key"]: task
        for task in current_task_items
    }

    def _assignee_columns_from_current_table(department_payload: dict | None) -> list[tuple[uuid.UUID | None, str]]:
        if not department_payload:
            return []
        try:
            dept = WeeklyTableDepartment.model_validate(department_payload)
        except Exception:
            return []

        seen: set[str] = set()
        columns: list[tuple[uuid.UUID | None, str]] = []
        for day in dept.days:
            for user_day in day.users:
                name = (user_day.user_name or "").strip()
                if not name:
                    continue
                user_id = getattr(user_day, "user_id", None)
                key = str(user_id) if user_id is not None else f"name:{name.lower()}"
                if key in seen:
                    continue
                seen.add(key)
                columns.append((user_id, name))

        columns.sort(key=lambda item: item[1].lower())
        return columns

    assignee_columns = _assignee_columns_from_current_table(current_department_payload)

    today_local = planners_router._today_pristina_date()
    as_of_date = min(today_local, week_end)

    buckets = planners_router._classify_weekly_plan_performance(
        planned_tasks=snapshot_tasks_by_key,
        actual_tasks=current_tasks_by_key,
        week_end=week_end,
        as_of_date=as_of_date,
    )

    completed = [planners_router._to_compare_task_out(task) for task in buckets.get("completed", [])]
    in_progress = [planners_router._to_compare_task_out(task) for task in buckets.get("in_progress", [])]
    pending = [planners_router._to_compare_task_out(task) for task in buckets.get("pending", [])]
    late = [planners_router._to_compare_task_out(task) for task in buckets.get("late", [])]
    additional = [planners_router._to_compare_task_out(task) for task in buckets.get("additional", [])]
    removed_or_canceled = [
        planners_router._to_compare_task_out(task) for task in buckets.get("removed_or_canceled", [])
    ]

    grouped = planners_router._group_compare_tasks_by_assignee(
        completed=completed,
        in_progress=in_progress,
        pending=pending,
        late=late,
        additional=additional,
        removed_or_canceled=removed_or_canceled,
    )

    groups: list[dict[str, object]] = []
    if grouped:
        groups = [
            {
                "assignee_id": g.assignee_id,
                "assignee_name": g.assignee_name,
                "completed": g.completed or [],
                "in_progress": g.in_progress or [],
                "pending": g.pending or [],
                "late": g.late or [],
                "additional": g.additional or [],
                "removed_or_canceled": g.removed_or_canceled or [],
            }
            for g in grouped
        ]

    if not groups and assignee_columns:
        groups = [
            {
                "assignee_id": assignee_id,
                "assignee_name": assignee_name,
                "completed": [],
                "in_progress": [],
                "pending": [],
                "late": [],
                "additional": [],
                "removed_or_canceled": [],
            }
            for assignee_id, assignee_name in assignee_columns
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "PLAN VS ACTUAL"[:31]

    # Columns: NR, STATUS, then one column per user
    last_col = max(2, 2 + len(groups))

    # Row 1: merged title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=title_upper)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=1)
    ws.row_dimensions[1].height = 24

    # Row 2-3: spacing
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10

    user_initials = _initials(user.full_name or user.username or "")

    # Standard print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2
    ws.oddHeader.right.text = "&D &T"
    ws.oddFooter.center.text = "Page &P / &N"
    ws.oddFooter.right.text = f"PUNOI: {user_initials or '____'}"
    ws.evenHeader.right.text = ws.oddHeader.right.text
    ws.evenFooter.center.text = ws.oddFooter.center.text
    ws.evenFooter.right.text = ws.oddFooter.right.text
    ws.firstHeader.right.text = ws.oddHeader.right.text
    ws.firstFooter.center.text = ws.oddFooter.center.text
    ws.firstFooter.right.text = ws.oddFooter.right.text

    header_row = 4
    data_start_row = 5

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    for idx in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 44

    # Header row
    ws.row_dimensions[header_row].height = 22
    nr_header = ws.cell(row=header_row, column=1, value="NR")
    nr_header.font = Font(bold=True)
    nr_header.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

    status_header = ws.cell(row=header_row, column=2, value="STATUS")
    status_header.font = Font(bold=True)
    status_header.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

    for col_idx, group in enumerate(groups, start=3):
        label = str(group.get("assignee_name") or "").upper()
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

    categories: list[tuple[str, str]] = [
        ("completed", "COMPLETED"),
        ("in_progress", "IN PROGRESS"),
        ("pending", "PENDING"),
        ("late", "LATE"),
        ("additional", "ADDITIONAL"),
        ("removed_or_canceled", "REMOVED/CANCELED"),
    ]

    def _cell_text(tasks: list[object]) -> str:
        if not tasks:
            return "-"
        lines: list[str] = [str(len(tasks))]
        for task in tasks:
            title = getattr(task, "title", None) or ""
            project_title = getattr(task, "project_title", None)
            suffix = f" ({project_title})" if project_title else ""
            lines.append(f"{title}{suffix}".strip())
        value = "\n".join(lines).strip()
        if len(value) > 32000:
            value = value[:31950].rstrip() + "\n...(TRUNCATED)"
        return value

    row_idx = data_start_row
    for idx, (key, label) in enumerate(categories, start=1):
        nr_cell = ws.cell(row=row_idx, column=1, value=idx)
        nr_cell.font = Font(bold=True)
        nr_cell.number_format = "#,##0"
        nr_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

        label_cell = ws.cell(row=row_idx, column=2, value=label)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
        for col_idx, group in enumerate(groups, start=3):
            tasks = group.get(key) or []
            cell = ws.cell(row=row_idx, column=col_idx, value=_cell_text(list(tasks)))
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
        row_idx += 1

    last_row = row_idx - 1

    # Filters + freeze + print titles
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "C5" if last_col >= 3 else "B5"
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"

    # Borders (thick outside, thin inside; header thick outline)
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    for r in range(header_row, last_row + 1):
        for c in range(1, last_col + 1):
            left = thick if c == 1 else thin
            right = thick if c == last_col else thin
            top = thick if r == header_row else thin
            bottom = thick if r == last_row else thin
            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

    for c in range(1, last_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.border = Border(
            left=thick if c == 1 else thin,
            right=thick if c == last_col else thin,
            top=thick,
            bottom=thick,
        )

    # Row heights (based on max wrapped lines)
    for r in range(data_start_row, last_row + 1):
        max_lines = 1
        for c in range(1, last_col + 1):
            value = ws.cell(row=r, column=c).value
            lines = str(value).count("\n") + 1 if value is not None else 1
            max_lines = max(max_lines, lines)
        ws.row_dimensions[r].height = max(18, min(300, 14 * max_lines))

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    today = datetime.now(timezone.utc).date()
    initials_value = user_initials or "USER"
    filename_date = f"{today.day:02d}_{today.month:02d}_{str(today.year)[-2:]}"
    filename_title = _safe_filename_spaces(title_upper)
    filename = f"{filename_title} {filename_date}_EF ({initials_value}).xlsx"

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/weekly-snapshot-compare.xlsx")
async def export_weekly_snapshot_compare_xlsx(
    baseline_snapshot_id: uuid.UUID = Query(..., description="Baseline snapshot ID"),
    compare_snapshot_id: uuid.UUID = Query(..., description="Comparison snapshot ID"),
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    if baseline_snapshot_id == compare_snapshot_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Choose two different snapshots.")

    baseline_snapshot = (
        await db.execute(select(WeeklyPlannerSnapshot).where(WeeklyPlannerSnapshot.id == baseline_snapshot_id))
    ).scalar_one_or_none()
    if baseline_snapshot is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Baseline snapshot not found")

    compare_snapshot = (
        await db.execute(select(WeeklyPlannerSnapshot).where(WeeklyPlannerSnapshot.id == compare_snapshot_id))
    ).scalar_one_or_none()
    if compare_snapshot is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comparison snapshot not found")

    ensure_department_access(user, baseline_snapshot.department_id)
    ensure_department_access(user, compare_snapshot.department_id)

    if baseline_snapshot.department_id != compare_snapshot.department_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Snapshots must belong to the same department.")
    if baseline_snapshot.week_start_date != compare_snapshot.week_start_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Snapshots must be from the same week.")

    department = (
        await db.execute(select(Department).where(Department.id == baseline_snapshot.department_id))
    ).scalar_one_or_none()
    dept_label = _display_department_name(department.name if department is not None else "")

    week_start = baseline_snapshot.week_start_date
    week_end = baseline_snapshot.week_end_date or planners_router._get_next_5_working_days(week_start)[-1]

    title_label = (
        f"{dept_label} "
        f"{_format_excel_date(week_start)} - {_format_excel_date(week_end)} - "
        "SNAPSHOT COMPARISON"
    ).strip()
    title_upper = title_label.upper() if title_label else "SNAPSHOT COMPARISON"

    planned_tasks_by_key = planners_router._tasks_by_key_from_snapshot_payload(baseline_snapshot.payload or {})
    actual_tasks_by_key = planners_router._tasks_by_key_from_snapshot_payload(compare_snapshot.payload or {})

    def _assignee_columns_from_snapshot(snapshot: WeeklyPlannerSnapshot | None) -> list[tuple[uuid.UUID | None, str]]:
        if snapshot is None:
            return []
        payload = snapshot.payload or {}
        department_payload = payload.get("department")
        if not department_payload:
            return []
        try:
            dept = WeeklyTableDepartment.model_validate(department_payload)
        except Exception:
            return []

        seen: set[str] = set()
        columns: list[tuple[uuid.UUID | None, str]] = []
        for day in dept.days:
            for user_day in day.users:
                name = (user_day.user_name or "").strip()
                if not name:
                    continue
                user_id = getattr(user_day, "user_id", None)
                key = str(user_id) if user_id is not None else f"name:{name.lower()}"
                if key in seen:
                    continue
                seen.add(key)
                columns.append((user_id, name))

        columns.sort(key=lambda item: item[1].lower())
        return columns

    assignee_columns = _assignee_columns_from_snapshot(baseline_snapshot) or _assignee_columns_from_snapshot(compare_snapshot)

    as_of_date = week_end + timedelta(days=1)
    buckets = planners_router._classify_weekly_plan_performance(
        planned_tasks=planned_tasks_by_key,
        actual_tasks=actual_tasks_by_key,
        week_end=week_end,
        as_of_date=as_of_date,
    )

    completed = [planners_router._to_compare_task_out(task) for task in buckets.get("completed", [])]
    in_progress = [planners_router._to_compare_task_out(task) for task in buckets.get("in_progress", [])]
    pending = [planners_router._to_compare_task_out(task) for task in buckets.get("pending", [])]
    late = [planners_router._to_compare_task_out(task) for task in buckets.get("late", [])]
    additional = [planners_router._to_compare_task_out(task) for task in buckets.get("additional", [])]
    removed_or_canceled = [
        planners_router._to_compare_task_out(task) for task in buckets.get("removed_or_canceled", [])
    ]

    grouped = planners_router._group_compare_tasks_by_assignee(
        completed=completed,
        in_progress=in_progress,
        pending=pending,
        late=late,
        additional=additional,
        removed_or_canceled=removed_or_canceled,
    )

    groups: list[dict[str, object]] = []
    if grouped:
        groups = [
            {
                "assignee_id": g.assignee_id,
                "assignee_name": g.assignee_name,
                "completed": g.completed or [],
                "in_progress": g.in_progress or [],
                "pending": g.pending or [],
                "late": g.late or [],
                "additional": g.additional or [],
                "removed_or_canceled": g.removed_or_canceled or [],
            }
            for g in grouped
        ]

    if not groups and assignee_columns:
        groups = [
            {
                "assignee_id": assignee_id,
                "assignee_name": assignee_name,
                "completed": [],
                "in_progress": [],
                "pending": [],
                "late": [],
                "additional": [],
                "removed_or_canceled": [],
            }
            for assignee_id, assignee_name in assignee_columns
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "SNAPSHOT COMPARE"[:31]

    last_col = max(2, 2 + len(groups))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=title_upper)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=1)
    ws.row_dimensions[1].height = 24

    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10

    user_initials = _initials(user.full_name or user.username or "")

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.36
    ws.page_margins.bottom = 0.51
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.2
    ws.oddHeader.right.text = "&D &T"
    ws.oddFooter.center.text = "Page &P / &N"
    ws.oddFooter.right.text = f"PUNOI: {user_initials or '____'}"
    ws.evenHeader.right.text = ws.oddHeader.right.text
    ws.evenFooter.center.text = ws.oddFooter.center.text
    ws.evenFooter.right.text = ws.oddFooter.right.text
    ws.firstHeader.right.text = ws.oddHeader.right.text
    ws.firstFooter.center.text = ws.oddFooter.center.text
    ws.firstFooter.right.text = ws.oddFooter.right.text

    header_row = 4
    data_start_row = 5

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    for idx in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 44

    ws.row_dimensions[header_row].height = 22
    nr_header = ws.cell(row=header_row, column=1, value="NR")
    nr_header.font = Font(bold=True)
    nr_header.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

    status_header = ws.cell(row=header_row, column=2, value="STATUS")
    status_header.font = Font(bold=True)
    status_header.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

    for col_idx, group in enumerate(groups, start=3):
        label = str(group.get("assignee_name") or "").upper()
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

    categories: list[tuple[str, str]] = [
        ("completed", "COMPLETED"),
        ("in_progress", "IN PROGRESS"),
        ("pending", "PENDING"),
        ("late", "LATE"),
        ("additional", "ADDITIONAL"),
        ("removed_or_canceled", "REMOVED/CANCELED"),
    ]

    def _cell_text(tasks: list[object]) -> str:
        if not tasks:
            return "-"
        lines: list[str] = [str(len(tasks))]
        for task in tasks:
            title = getattr(task, "title", None) or ""
            project_title = getattr(task, "project_title", None)
            suffix = f" ({project_title})" if project_title else ""
            lines.append(f"{title}{suffix}".strip())
        value = "\n".join(lines).strip()
        if len(value) > 32000:
            value = value[:31950].rstrip() + "\n...(TRUNCATED)"
        return value

    row_idx = data_start_row
    for idx, (key, label) in enumerate(categories, start=1):
        nr_cell = ws.cell(row=row_idx, column=1, value=idx)
        nr_cell.font = Font(bold=True)
        nr_cell.number_format = "#,##0"
        nr_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)

        label_cell = ws.cell(row=row_idx, column=2, value=label)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
        for col_idx, group in enumerate(groups, start=3):
            tasks = group.get(key) or []
            cell = ws.cell(row=row_idx, column=col_idx, value=_cell_text(list(tasks)))
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True, readingOrder=1)
        row_idx += 1

    last_row = row_idx - 1

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "C5" if last_col >= 3 else "B5"
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"

    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    for r in range(header_row, last_row + 1):
        for c in range(1, last_col + 1):
            left = thick if c == 1 else thin
            right = thick if c == last_col else thin
            top = thick if r == header_row else thin
            bottom = thick if r == last_row else thin
            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

    for c in range(1, last_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.border = Border(
            left=thick if c == 1 else thin,
            right=thick if c == last_col else thin,
            top=thick,
            bottom=thick,
        )

    for r in range(data_start_row, last_row + 1):
        max_lines = 1
        for c in range(1, last_col + 1):
            value = ws.cell(row=r, column=c).value
            lines = str(value).count("\n") + 1 if value is not None else 1
            max_lines = max(max_lines, lines)
        ws.row_dimensions[r].height = max(18, min(300, 14 * max_lines))

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    today = datetime.now(timezone.utc).date()
    initials_value = user_initials or "USER"
    filename_date = f"{today.day:02d}_{today.month:02d}_{str(today.year)[-2:]}"
    filename_title = _safe_filename_spaces(title_upper)
    filename = f"{filename_title} {filename_date}_EF ({initials_value}).xlsx"

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )

