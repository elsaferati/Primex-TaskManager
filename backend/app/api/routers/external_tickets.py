from __future__ import annotations

import io
import json
import math
import re
import uuid
from datetime import date, datetime, time, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import String, asc, cast, desc, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.access import ensure_admin
from app.api.deps import get_current_user
from app.db import get_db
from app.models.enums import UserRole
from app.models.project import Project
from app.models.std_feedback_ticket import StdFeedbackSyncState, StdFeedbackTicket
from app.models.user import User
from app.schemas.std_feedback_ticket import (
    StdFeedbackSyncOut,
    StdFeedbackTicketDetailOut,
    StdFeedbackTicketListOut,
    StdFeedbackTicketOut,
    StdTicketCreateTaskOut,
    StdTicketCreateTaskRequest,
    StdTicketNoActionOut,
    StdTicketNoActionRequest,
    StdTicketProjectOption,
    StdTicketTaskOptionsOut,
    StdTicketUserOption,
)
from app.services.notifications import publish_notification
from app.services.std_feedback_task_creation import (
    create_ticket_task_bundle,
    is_std_project_title,
    mark_tickets_no_action,
)
from app.services.std_feedback_tickets import (
    StdFeedbackClient,
    _std_token,
    refresh_std_ticket_detail,
    sync_std_feedback_tickets,
    ticket_comments,
    ticket_files,
)


router = APIRouter()


def _ticket_out(ticket: StdFeedbackTicket) -> StdFeedbackTicketOut:
    return StdFeedbackTicketOut(
        id=ticket.id,
        external_id=ticket.external_id,
        issue_number=ticket.issue_number,
        order_ticket_number=ticket.order_ticket_number,
        title=ticket.title,
        description=ticket.description,
        affected_fields=list(ticket.affected_fields or []),
        category=ticket.category,
        priority=ticket.priority,
        status=ticket.status,
        dashboard_area=ticket.dashboard_area,
        reporter_username=ticket.reporter_username,
        reporter_email=ticket.reporter_email,
        comment_count=ticket.comment_count,
        file_count=ticket.file_count,
        reported_at=ticket.reported_at,
        source_updated_at=ticket.source_updated_at,
        closed_at=ticket.closed_at,
        synced_at=ticket.synced_at,
        review_status=ticket.review_status,
        review_note=ticket.review_note,
        reviewed_by=ticket.reviewed_by,
        reviewed_at=ticket.reviewed_at,
        ga_note_id=ticket.ga_note_id,
        task_id=ticket.task_id,
    )


def _ticket_detail_out(ticket: StdFeedbackTicket) -> StdFeedbackTicketDetailOut:
    base = _ticket_out(ticket).model_dump()
    return StdFeedbackTicketDetailOut(
        **base,
        creator_id=ticket.creator_id,
        assigned_admin=ticket.assigned_admin,
        closed_by=ticket.closed_by,
        related_order_id=ticket.related_order_id,
        order_snapshot_json=dict(ticket.order_snapshot_json or {}),
        comments=ticket_comments(ticket),
        files=ticket_files(ticket),
    )


def _search_condition(search: str):
    term = f"%{search.strip()}%"
    return or_(
        cast(StdFeedbackTicket.issue_number, String).ilike(term),
        StdFeedbackTicket.order_ticket_number.ilike(term),
        StdFeedbackTicket.title.ilike(term),
        StdFeedbackTicket.description.ilike(term),
        StdFeedbackTicket.reporter_username.ilike(term),
        StdFeedbackTicket.reporter_email.ilike(term),
    )


def _ticket_filters(
    *,
    search: str | None = None,
    ticket_status: str | None = None,
    category: str | None = None,
    priority: str | None = None,
    review_status: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> list:
    filters = [StdFeedbackTicket.is_external.is_(True)]
    if search and search.strip():
        filters.append(_search_condition(search))
    if ticket_status:
        filters.append(func.lower(StdFeedbackTicket.status) == ticket_status.casefold())
    if category:
        filters.append(func.lower(StdFeedbackTicket.category) == category.casefold())
    if priority:
        filters.append(func.lower(StdFeedbackTicket.priority) == priority.casefold())
    if review_status:
        filters.append(func.upper(StdFeedbackTicket.review_status) == review_status.upper())
    if date_from:
        filters.append(StdFeedbackTicket.reported_at >= datetime.combine(date_from, time.min, tzinfo=timezone.utc))
    if date_to:
        filters.append(
            StdFeedbackTicket.reported_at < datetime.combine(date_to + timedelta(days=1), time.min, tzinfo=timezone.utc)
        )
    return filters


def _ticket_ordering(sort_by: str, sort_dir: str):
    sort_columns = {
        "issue_number": StdFeedbackTicket.issue_number,
        "created_at": StdFeedbackTicket.reported_at,
        "updated_at": StdFeedbackTicket.source_updated_at,
        "priority": StdFeedbackTicket.priority,
        "status": StdFeedbackTicket.status,
    }
    sort_column = sort_columns.get(sort_by, StdFeedbackTicket.source_updated_at)
    return asc(sort_column) if sort_dir == "asc" else desc(sort_column)


def _excel_text(value, *, limit: int = 32767) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        rendered = json.dumps(value, ensure_ascii=False, default=str)
    else:
        rendered = str(value)
    return rendered[:limit]


def _excel_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is not None:
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value


def _external_tickets_workbook(tickets: list[StdFeedbackTicket]) -> io.BytesIO:
    headers = [
        "External ID",
        "Issue #",
        "Order Ticket #",
        "Title",
        "Problem / Description",
        "Affected Fields",
        "Status",
        "Category",
        "Priority",
        "Dashboard Area",
        "Reporter",
        "Reporter Email",
        "Assigned Admin",
        "Related Order ID",
        "Order Information",
        "Comments",
        "Attachments",
        "Comment Count",
        "File Count",
        "Created At (UTC)",
        "Updated At (UTC)",
        "Closed At (UTC)",
        "Review Decision",
        "Review Note",
        "Reviewed By ID",
        "Reviewed At (UTC)",
        "GA Note ID",
        "Task ID",
        "Source",
    ]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "STD Tickets"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.append(headers)

    for ticket in tickets:
        comments = ticket_comments(ticket)
        files = ticket_files(ticket)
        comment_lines = [
            " — ".join(part for part in (_excel_text(item.get("author") or item.get("creator") or item.get("user")), _excel_text(item.get("body") or item.get("comment") or item.get("content"))) if part)
            for item in comments
            if isinstance(item, dict)
        ]
        file_lines = [
            _excel_text(item.get("original_filename") or item.get("filename") or item.get("name") or item.get("id"))
            for item in files
            if isinstance(item, dict)
        ]
        worksheet.append(
            [
                ticket.external_id,
                ticket.issue_number,
                ticket.order_ticket_number or "",
                ticket.title or "",
                ticket.description or "",
                _excel_text(ticket.affected_fields or []),
                ticket.status or "",
                ticket.category or "",
                ticket.priority or "",
                ticket.dashboard_area or "",
                ticket.reporter_username or "",
                ticket.reporter_email or "",
                ticket.assigned_admin or "",
                ticket.related_order_id or "",
                _excel_text(ticket.order_snapshot_json or {}),
                _excel_text("\n".join(comment_lines)),
                _excel_text("\n".join(file_lines)),
                ticket.comment_count or 0,
                ticket.file_count or 0,
                _excel_datetime(ticket.reported_at),
                _excel_datetime(ticket.source_updated_at),
                _excel_datetime(ticket.closed_at),
                ticket.review_status or "",
                ticket.review_note or "",
                str(ticket.reviewed_by) if ticket.reviewed_by else "",
                _excel_datetime(ticket.reviewed_at),
                str(ticket.ga_note_id) if ticket.ga_note_id else "",
                str(ticket.task_id) if ticket.task_id else "",
                "STD External",
            ]
        )

    header_fill = PatternFill(fill_type="solid", fgColor="0F172A")
    alternating_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 28

    wrapped_columns = {4, 5, 6, 15, 16, 17, 24}
    datetime_columns = {20, 21, 22, 26}
    for row_index in range(2, worksheet.max_row + 1):
        if row_index % 2 == 0:
            for cell in worksheet[row_index]:
                cell.fill = alternating_fill
        for column_index in range(1, len(headers) + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(vertical="top", wrap_text=column_index in wrapped_columns)
            if column_index in datetime_columns and cell.value:
                cell.number_format = "dd.mm.yyyy hh:mm"

    widths = {
        1: 18, 2: 11, 3: 18, 4: 28, 5: 48, 6: 28, 7: 15, 8: 16, 9: 13, 10: 18,
        11: 22, 12: 30, 13: 22, 14: 18, 15: 36, 16: 42, 17: 32, 18: 14, 19: 11,
        20: 20, 21: 20, 22: 20, 23: 18, 24: 32, 25: 38, 26: 20, 27: 38, 28: 38, 29: 16,
    }
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, worksheet.max_row)}"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@router.get("", response_model=StdFeedbackTicketListOut)
async def list_external_tickets(
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=10, le=100),
    search: str | None = Query(default=None, max_length=200),
    ticket_status: str | None = Query(default=None, alias="status", max_length=50),
    category: str | None = Query(default=None, max_length=50),
    priority: str | None = Query(default=None, max_length=50),
    review_status: str | None = Query(default=None, max_length=30),
    date_from: date | None = None,
    date_to: date | None = None,
    sort_by: str = Query("updated_at"),
    sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> StdFeedbackTicketListOut:
    filters = _ticket_filters(
        search=search,
        ticket_status=ticket_status,
        category=category,
        priority=priority,
        review_status=review_status,
        date_from=date_from,
        date_to=date_to,
    )
    ordering = _ticket_ordering(sort_by, sort_dir)
    total = (await db.execute(select(func.count(StdFeedbackTicket.id)).where(*filters))).scalar_one()
    items = (
        await db.execute(
            select(StdFeedbackTicket)
            .where(*filters)
            .order_by(ordering.nullslast(), StdFeedbackTicket.id.asc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    ).scalars().all()

    async def distinct_values(column) -> list[str]:
        values = (
            await db.execute(
                select(column)
                .where(StdFeedbackTicket.is_external.is_(True), column.is_not(None))
                .distinct()
                .order_by(column)
            )
        ).scalars().all()
        return [str(value) for value in values if str(value).strip()]

    categories = await distinct_values(StdFeedbackTicket.category)
    priorities = await distinct_values(StdFeedbackTicket.priority)
    statuses = await distinct_values(StdFeedbackTicket.status)
    sync_state = (
        await db.execute(select(StdFeedbackSyncState).where(StdFeedbackSyncState.key == "default"))
    ).scalar_one_or_none()
    return StdFeedbackTicketListOut(
        items=[_ticket_out(item) for item in items],
        total=total,
        page=page,
        page_size=page_size,
        pages=max(1, math.ceil(total / page_size)),
        categories=categories,
        priorities=priorities,
        statuses=statuses,
        last_synchronized_at=sync_state.last_successful_sync_at if sync_state else None,
        last_sync_error=(sync_state.last_sync_error if sync_state and user.role == UserRole.ADMIN else None),
    )


@router.get("/export.xlsx")
async def export_external_tickets_xlsx(
    search: str | None = Query(default=None, max_length=200),
    ticket_status: str | None = Query(default=None, alias="status", max_length=50),
    category: str | None = Query(default=None, max_length=50),
    priority: str | None = Query(default=None, max_length=50),
    review_status: str | None = Query(default=None, max_length=30),
    date_from: date | None = None,
    date_to: date | None = None,
    sort_by: str = Query("updated_at"),
    sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    filters = _ticket_filters(
        search=search,
        ticket_status=ticket_status,
        category=category,
        priority=priority,
        review_status=review_status,
        date_from=date_from,
        date_to=date_to,
    )
    tickets = (
        await db.execute(
            select(StdFeedbackTicket)
            .where(*filters)
            .order_by(_ticket_ordering(sort_by, sort_dir).nullslast(), StdFeedbackTicket.id.asc())
        )
    ).scalars().all()
    filename = f"STD_Tickets_EXT_{datetime.now(timezone.utc).date().isoformat()}.xlsx"
    return StreamingResponse(
        _external_tickets_workbook(list(tickets)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/sync", response_model=StdFeedbackSyncOut)
async def sync_external_tickets_now(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> StdFeedbackSyncOut:
    ensure_admin(user)
    return StdFeedbackSyncOut(**(await sync_std_feedback_tickets(db)))


@router.get("/task-options", response_model=StdTicketTaskOptionsOut)
async def external_ticket_task_options(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> StdTicketTaskOptionsOut:
    projects = (
        await db.execute(
            select(Project)
            .where(Project.is_template.is_(False))
            .order_by(Project.title.asc())
        )
    ).scalars().all()
    projects = [project for project in projects if is_std_project_title(project.title)]
    users = (
        await db.execute(select(User).where(User.is_active.is_(True)).order_by(User.full_name.asc().nullslast(), User.username.asc()))
    ).scalars().all()
    return StdTicketTaskOptionsOut(
        projects=[
            StdTicketProjectOption(id=project.id, title=project.title, department_id=project.department_id)
            for project in projects
        ],
        users=[
            StdTicketUserOption(
                id=item.id,
                label=(item.full_name or item.username or item.email),
                department_id=item.department_id,
            )
            for item in users
        ],
    )


@router.post("/reviews/no-action", response_model=StdTicketNoActionOut)
async def review_external_tickets_no_action(
    payload: StdTicketNoActionRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> StdTicketNoActionOut:
    try:
        tickets = await mark_tickets_no_action(
            db,
            ticket_ids=payload.ticket_ids,
            actor_user_id=user.id,
            note=payload.note,
        )
    except ValueError as exc:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(exc)) from exc
    await db.commit()
    return StdTicketNoActionOut(updated=len(tickets))


@router.post("/tasks", response_model=StdTicketCreateTaskOut, status_code=status.HTTP_201_CREATED)
async def create_external_ticket_task(
    payload: StdTicketCreateTaskRequest,
    response: Response,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> StdTicketCreateTaskOut:
    project = (await db.execute(select(Project).where(Project.id == payload.project_id))).scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    try:
        bundle = await create_ticket_task_bundle(
            db,
            ticket_ids=payload.ticket_ids,
            project_id=payload.project_id,
            assignee_ids=payload.assignee_ids,
            actor_user_id=user.id,
            title=payload.title,
            description=payload.description,
            review_note=payload.review_note,
            priority=payload.priority,
            start_date=payload.start_date,
            due_date=payload.due_date,
        )
    except ValueError as exc:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(exc)) from exc
    await db.commit()
    if not bundle.created:
        response.status_code = status.HTTP_200_OK
    for notification in bundle.notifications:
        try:
            await publish_notification(user_id=notification.user_id, notification=notification)
        except Exception:
            pass
    return StdTicketCreateTaskOut(
        note_id=bundle.note.id,
        task_ids=[task.id for task in bundle.tasks],
        ticket_ids=[ticket.id for ticket in bundle.tickets],
        created=bundle.created,
    )


@router.get("/{ticket_id}", response_model=StdFeedbackTicketDetailOut)
async def get_external_ticket(
    ticket_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
) -> StdFeedbackTicketDetailOut:
    ticket = (
        await db.execute(
            select(StdFeedbackTicket).where(
                StdFeedbackTicket.id == ticket_id,
                StdFeedbackTicket.is_external.is_(True),
            )
        )
    ).scalar_one_or_none()
    if ticket is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="External ticket not found")
    ticket = await refresh_std_ticket_detail(db, ticket)
    return _ticket_detail_out(ticket)


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._() -]+", "_", value).strip(" .")
    return cleaned[:180] or "attachment"


@router.get("/{ticket_id}/files/{file_id}")
async def download_external_ticket_file(
    ticket_id: uuid.UUID,
    file_id: str,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
) -> Response:
    ticket = (
        await db.execute(
            select(StdFeedbackTicket).where(
                StdFeedbackTicket.id == ticket_id,
                StdFeedbackTicket.is_external.is_(True),
            )
        )
    ).scalar_one_or_none()
    if ticket is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="External ticket not found")
    if not _std_token():
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="STD integration is not configured")

    metadata = next(
        (item for item in ticket_files(ticket) if str(item.get("id")) == file_id),
        {},
    )
    async with StdFeedbackClient() as client:
        try:
            upstream = await client.get_file(ticket.external_id, file_id)
        except Exception as exc:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="STD attachment download failed") from exc
    filename = _safe_filename(
        str(metadata.get("original_filename") or metadata.get("filename") or metadata.get("name") or file_id)
    )
    media_type = upstream.headers.get("content-type") or str(metadata.get("content_type") or "application/octet-stream")
    return Response(
        content=upstream.content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
