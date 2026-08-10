from __future__ import annotations

import io
import re
from collections import Counter
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.realization_calculator import QUESTION_LABELS, REPORT_QUESTION_SECTIONS


NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
AMBER = "FFE699"
RED = "F4CCCC"
WHITE = "FFFFFF"
GRAY = "E7E6E6"
LEVEL_COLORS = {
    "A+": "00B050",
    "A": "70AD47",
    "B": "C6E0B4",
    "C": "FFD966",
    "M": "F4B183",
    "D": "F8696B",
    "E": "C00000",
}
LEVEL_BONUSES = {"A+": 50, "A": 40, "B": 30, "C": 20, "M": 15, "D": 10, "E": 0}
DEPARTMENT_COLORS = {
    "development": "2E7D32",
    "product content": "1565C0",
    "graphic design": "BF360C",
}


def _department_color(name: str) -> str:
    return DEPARTMENT_COLORS.get(name.strip().lower(), NAVY)
THIN = Side(style="thin", color="A6A6A6")
VALUE_LABELS = {
    "planned": "Planifikuar",
    "weekly_planned": "Në planin javor",
    "weekly_completed": "Të mbyllura deri sot",
    "weekly_remaining": "Të pambyllura",
    "today_planned": "Planifikuar sot",
    "today_completed": "Mbyllur sot",
    "today_in_progress": "Në progres sot",
    "today_no_progress": "Pa progres sot",
    "completed": "Mbyllur",
    "in_progress": "Në progres",
    "no_progress": "Pa progres",
    "count": "Gjithsej",
    "fast_tasks": "Fast tasks",
    "yes": "Ka detyra të reja",
    "approved": "Të konfirmuara",
    "unapproved": "Pa konfirmim",
    "needs_confirmation": "Kërkon konfirmim",
    "verified_categories": "Kategori të verifikuara",
    "additional_tasks_candidate": "Taska shtesë kandidate",
    "all_closed": "Të gjitha të mbyllura",
    "closed": "Të mbyllura",
    "remaining": "Të pambyllura",
    "attendance_tardiness": "Vonesa në prezencë",
    "frequent": "Të shpeshta",
    "threshold": "Pragu",
}
SOURCE_LABELS = {"system": "Sistem", "project": "Projekt", "fast": "Fast task"}
STATUS_LABELS = {
    "completed": "Mbyllur",
    "completed_on_time": "Mbyllur në kohë",
    "completed_late": "Mbyllur me vonesë",
    "in_progress": "Në progres",
    "no_progress": "Pa progres",
    "pending_confirmation": "Në pritje të konfirmimit",
}


def _value(value: Any) -> str | int | float | bool:
    if value is None:
        return "—"
    if isinstance(value, bool):
        return "Po" if value else "Jo"
    if isinstance(value, (str, int, float)):
        return value
    if isinstance(value, dict):
        return " · ".join(
            f"{VALUE_LABELS.get(str(key), str(key).replace('_', ' ').title())}: {_value(item)}"
            for key, item in value.items()
        ) or "—"
    if isinstance(value, (list, tuple, set)):
        return "; ".join(str(_value(item)) for item in value) or "—"
    return str(value)


def _clean_task_title(value: Any) -> str:
    return re.sub(r"\s+", " ", re.sub(r"\[\[/?[a-z_]+\]\]", "", str(value or ""), flags=re.I)).strip()


def _title(ws, title: str, subtitle: str, width: int, color: str = NAVY) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(color=WHITE, bold=True, size=16)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=color)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).font = Font(color="666666", italic=True)
    ws.cell(2, 1).alignment = Alignment(horizontal="center")


def _header(cell, fill: str = NAVY) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _sheet_name(value: str, used: set[str]) -> str:
    base = "".join("-" if char in "[]:*?/\\" else char for char in value)[:31] or "Department"
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f" {counter}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def build_realization_workbook(
    *,
    week_start: str,
    week_end: str,
    departments: list[dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Përmbledhje"
    has_live_data = any(
        department.get("report_mode") == "LIVE_DAILY" for department in departments
    )
    report_label = (
        "AKTUAL NGA SNAPSHOT-ET DITORE" if has_live_data else "PËRFUNDIMTAR"
    )
    _title(
        summary,
        "RAPORTI I REALIZIMIT JAVOR",
        f"{week_start} — {week_end} | {report_label}",
        10,
    )
    headers = [
        "Departamenti", "Personi", "Në planin javor", "Të mbyllura", "Me vonesë (FINAL)",
        "Shtesë gjatë javës", "Progresi javor", "Vlerësimi", "Simboli", "Statusi",
    ]
    for column, label in enumerate(headers, 1):
        summary.cell(4, column, label)
        _header(summary.cell(4, column))
    row = 5
    for department in departments:
        for person in department.get("people") or []:
            facts = person.get("facts_json") or {}
            is_live = facts.get("report_mode") == "LIVE_DAILY"
            grade = (
                "Në pritje të FINAL"
                if is_live
                else person.get("final_level") or person.get("suggested_level") or "—"
            )
            values = [
                department["name"], person["user_name"], person.get("planned_count", 0),
                person.get("completed_on_time_count", 0) + person.get("completed_late_count", 0),
                person.get("completed_late_count", 0),
                person.get("additional_count", 0), facts.get("weekly_progress_percent", 0) / 100,
                grade,
                "—" if is_live else person.get("final_symbol") or person.get("suggested_symbol") or "—",
                "AKTUAL" if is_live else department.get("status", "—"),
            ]
            for column, value in enumerate(values, 1):
                cell = summary.cell(row, column, value)
                cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            summary.cell(row, 7).number_format = "0.0%"
            level = str(values[7])
            if level in LEVEL_COLORS:
                summary.cell(row, 8).fill = PatternFill("solid", fgColor=LEVEL_COLORS[level])
                summary.cell(row, 8).font = Font(
                    bold=True, color=WHITE if level in {"A+", "A", "E"} else "000000"
                )
            summary.row_dimensions[row].height = 30
            row += 1
    if row == 5:
        summary.merge_cells(start_row=5, start_column=1, end_row=5, end_column=10)
        summary.cell(5, 1, "Nuk ka snapshot ditor ose rezultat FINAL për personat e përzgjedhur.")
        summary.cell(5, 1).alignment = Alignment(horizontal="center")
        summary.cell(5, 1).fill = PatternFill("solid", fgColor=AMBER)
    summary.freeze_panes = "A5"
    summary.auto_filter.ref = f"A4:J{max(4, row - 1)}"
    widths = [22, 24, 16, 14, 17, 16, 16, 22, 10, 16]
    for index, width in enumerate(widths, 1):
        summary.column_dimensions[get_column_letter(index)].width = width

    used_names = {summary.title}
    for department in departments:
        people = department.get("people") or []
        width = max(3, 1 + len(people) * 2)
        ws = workbook.create_sheet(_sheet_name(department["name"], used_names))
        department_color = _department_color(department["name"])
        manager_name = department.get("manager_name")
        subtitle = f"Java {week_start} — {week_end} | {department.get('status', '—')}"
        if manager_name:
            subtitle += f" | Menaxheri: {manager_name}"
        _title(
            ws,
            f"REALIZIMI — {department['name']}",
            subtitle,
            width,
            color=department_color,
        )
        ws.cell(4, 1, "Pyetja / Treguesi")
        _header(ws.cell(4, 1), department_color)
        for index, person in enumerate(people):
            col = 2 + index * 2
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
            ws.cell(4, col, person["user_name"])
            _header(ws.cell(4, col), department_color)
            ws.cell(5, col, "Përgjigjja")
            ws.cell(5, col + 1, "Argumenti / Evidenca")
            _header(ws.cell(5, col), "5B9BD5")
            _header(ws.cell(5, col + 1), "5B9BD5")
        ws.cell(5, 1, "Burimi")
        _header(ws.cell(5, 1), "5B9BD5")
        if not people:
            ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=width)
            ws.cell(6, 1, "Pa snapshot ditor dhe pa rezultat FINAL për këtë departament.")
            ws.cell(6, 1).alignment = Alignment(horizontal="center")
            ws.cell(6, 1).fill = PatternFill("solid", fgColor=AMBER)

        current_row = 7 if not people else 6
        source_labels = {
            "AUTO": "Automatik nga databaza",
            "AUTO_NEEDS_CONFIRMATION": "Kërkon konfirmim",
            "MISSING_EVIDENCE": "Kërkon evidencë",
            "MANAGER_CONFIRMED": "Konfirmuar nga menaxheri",
        }
        long_answer_keys = {
            "task_status",
            "week_positive",
            "week_problems",
            "repeated_after_clarification",
        }
        for section_title, question_keys in REPORT_QUESTION_SECTIONS:
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=width,
            )
            ws.cell(current_row, 1, section_title)
            _header(ws.cell(current_row, 1), "1F4E78")
            ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            for key in question_keys:
                ws.cell(current_row, 1, QUESTION_LABELS[key])
                ws.cell(current_row, 1).font = Font(bold=True)
                ws.cell(current_row, 1).fill = PatternFill("solid", fgColor=BLUE)
                ws.cell(current_row, 1).alignment = Alignment(vertical="top", wrap_text=True)
                ws.cell(current_row, 1).border = Border(
                    left=THIN, right=THIN, top=THIN, bottom=THIN
                )

                for index, person in enumerate(people):
                    col = 2 + index * 2
                    question = next(
                        (
                            item
                            for item in (person.get("facts_json") or {}).get("questions") or []
                            if item.get("key") == key
                        ),
                        {},
                    )
                    source_status = str(
                        question.get("source_status") or "AUTO_NEEDS_CONFIRMATION"
                    )
                    answer = question.get("final_value")
                    if answer is None:
                        answer = question.get("auto_value")
                    needs_confirmation = (
                        source_status in {"AUTO_NEEDS_CONFIRMATION", "MISSING_EVIDENCE"}
                        and answer is None
                    )
                    answer_cell = ws.cell(
                        current_row,
                        col,
                        "PËR KONFIRMIM" if needs_confirmation else _value(answer),
                    )
                    answer_cell.fill = PatternFill(
                        "solid", fgColor=AMBER if needs_confirmation else WHITE
                    )

                    evidence = question.get("evidence_ids") or []
                    explanation = str(question.get("explanation") or "").strip()
                    note_parts = [
                        source_labels.get(source_status, source_status.replace("_", " ").title()),
                        explanation,
                        f"{len(evidence)} evidencë/a — detajet në fletën Evidenca"
                        if evidence
                        else "",
                    ]
                    ws.cell(
                        current_row,
                        col + 1,
                        " | ".join(item for item in note_parts if item),
                    )
                    for target in (
                        ws.cell(current_row, col),
                        ws.cell(current_row, col + 1),
                    ):
                        target.alignment = Alignment(vertical="top", wrap_text=True)
                        target.border = Border(
                            left=THIN, right=THIN, top=THIN, bottom=THIN
                        )
                ws.row_dimensions[current_row].height = (
                    64 if key in long_answer_keys else 46
                )
                current_row += 1

        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=width,
        )
        ws.cell(current_row, 1, "5. SNAPSHOT-ET DITORE")
        _header(ws.cell(current_row, 1), "1F4E78")
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1
        daily_dates = sorted(
            {
                str(item.get("date"))
                for person in people
                for item in (person.get("facts_json") or {}).get("daily_timeline") or []
                if item.get("date")
            }
        )
        for daily_date in daily_dates:
            ws.cell(current_row, 1, f"Snapshot ditor — {daily_date}")
            ws.cell(current_row, 1).font = Font(bold=True)
            ws.cell(current_row, 1).fill = PatternFill("solid", fgColor=GREEN)
            for index, person in enumerate(people):
                col = 2 + index * 2
                snapshot = next(
                    (
                        item
                        for item in (person.get("facts_json") or {}).get("daily_timeline") or []
                        if str(item.get("date")) == daily_date
                    ),
                    None,
                )
                if snapshot:
                    ws.cell(current_row, col, float(snapshot.get("weekly_progress_percent") or 0) / 100)
                    ws.cell(current_row, col).number_format = "0.0%"
                    attendance = snapshot.get("attendance") or []
                    ws.cell(
                        current_row,
                        col + 1,
                        "Java deri atë ditë: "
                        f"{snapshot.get('weekly_completed_count', 0)}/"
                        f"{snapshot.get('weekly_planned_count', 0)} | "
                        "Sot: "
                        f"{snapshot.get('completed_count', 0)}/"
                        f"{snapshot.get('planned_count', 0)} "
                        f"({snapshot.get('daily_progress_percent', 0)}%) | "
                        f"Shtesë sot: {snapshot.get('additional_count', 0)} | "
                        f"Prezenca: {', '.join(str(item.get('type')) for item in attendance) or 'OK'}",
                    )
                else:
                    ws.cell(current_row, col, "—")
                    ws.cell(current_row, col + 1, "Pa snapshot")
                for target in (ws.cell(current_row, col), ws.cell(current_row, col + 1)):
                    target.alignment = Alignment(vertical="top", wrap_text=True)
                    target.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            ws.cell(current_row, 1).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            current_row += 1

        project_keys = sorted(
            {
                str(project.get("project_id") or project.get("project_title"))
                for person in people
                for project in (person.get("facts_json") or {}).get("project_progress") or []
            }
        )
        if project_keys:
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=width,
            )
            ws.cell(current_row, 1, "6. PROJEKTET MST / TT")
            _header(ws.cell(current_row, 1), "1F4E78")
            ws.cell(current_row, 1).alignment = Alignment(
                horizontal="left", vertical="center"
            )
            current_row += 1
        for project_key in project_keys:
            title = next(
                str(project.get("project_title") or "MST/TT")
                for person in people
                for project in (person.get("facts_json") or {}).get("project_progress") or []
                if str(project.get("project_id") or project.get("project_title")) == project_key
            )
            ws.cell(current_row, 1, f"Mesatarja e projektit — {title}")
            ws.cell(current_row, 1).font = Font(bold=True)
            ws.cell(current_row, 1).fill = PatternFill("solid", fgColor=AMBER)
            for index, person in enumerate(people):
                col = 2 + index * 2
                project = next(
                    (
                        item
                        for item in (person.get("facts_json") or {}).get("project_progress") or []
                        if str(item.get("project_id") or item.get("project_title")) == project_key
                    ),
                    None,
                )
                if project:
                    ws.cell(current_row, col, float(project.get("progress_percent") or 0) / 100)
                    ws.cell(current_row, col).number_format = "0.0%"
                    ws.cell(
                        current_row,
                        col + 1,
                        f"{project.get('task_count', 0)} detyra | {project.get('method', '')}",
                    )
                else:
                    ws.cell(current_row, col, "—")
                    ws.cell(current_row, col + 1, "Pa taska në këtë projekt")
                for target in (ws.cell(current_row, col), ws.cell(current_row, col + 1)):
                    target.alignment = Alignment(vertical="top", wrap_text=True)
                    target.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            ws.cell(current_row, 1).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            current_row += 1

        grade_row = current_row + 1
        department_is_live = department.get("report_mode") == "LIVE_DAILY"
        ws.merge_cells(
            start_row=grade_row,
            start_column=1,
            end_row=grade_row,
            end_column=width,
        )
        evaluation_section_number = 7 if project_keys else 6
        ws.cell(
            grade_row,
            1,
            f"{evaluation_section_number}. VLERËSIMI DHE KOMENTET",
        )
        _header(ws.cell(grade_row, 1), "1F4E78")
        ws.cell(grade_row, 1).alignment = Alignment(horizontal="left", vertical="center")

        def _final_level(person: dict[str, Any]) -> str | None:
            return person.get("final_level") or person.get("suggested_level")

        evaluation_rows = [
            (
                "Propozimi për nivelin e vlerësimit",
                lambda person: (
                    "Në pritje të FINAL"
                    if department_is_live
                    else person.get("suggested_level") or "—"
                ),
                lambda person: (
                    "Klasifikohet pas ruajtjes FINAL"
                    if department_is_live
                    else str(person.get("auto_narrative") or "")
                ),
                True,
                False,
            ),
            (
                "Vlerësimi final",
                lambda person: (
                    "Në pritje"
                    if department_is_live
                    else _final_level(person) or "—"
                ),
                lambda person: (
                    "Menaxheri e konfirmon pas FINAL"
                    if department_is_live
                    else str(person.get("override_reason") or "")
                ),
                True,
                False,
            ),
            (
                "Simboli",
                lambda person: (
                    "—"
                    if department_is_live
                    else person.get("final_symbol")
                    or person.get("suggested_symbol")
                    or "—"
                ),
                lambda person: "",
                False,
                False,
            ),
            (
                "Bonusi javor (€)",
                lambda person: (
                    "—"
                    if department_is_live
                    else LEVEL_BONUSES.get(str(_final_level(person)), "—")
                ),
                lambda person: (
                    "Sipas nivelit; D deri 10€, sipas gjykimit të menaxherit"
                    if not department_is_live and _final_level(person) == "D"
                    else ""
                ),
                False,
                False,
            ),
            (
                "Komente",
                lambda person: str(person.get("manager_comment") or "—"),
                lambda person: str(person.get("auto_narrative") or "—"),
                False,
                True,
            ),
        ]
        for row_offset, (label, value_builder, note_builder, is_level_row, is_tall) in enumerate(
            evaluation_rows, 1
        ):
            evaluation_row = grade_row + row_offset
            ws.cell(evaluation_row, 1, label)
            ws.cell(evaluation_row, 1).font = Font(bold=True)
            ws.cell(evaluation_row, 1).fill = PatternFill("solid", fgColor=BLUE)
            ws.cell(evaluation_row, 1).border = Border(
                left=THIN, right=THIN, top=THIN, bottom=THIN
            )
            for index, person in enumerate(people):
                col = 2 + index * 2
                value = value_builder(person)
                note = note_builder(person)
                fill = (
                    LEVEL_COLORS.get(str(value), GRAY)
                    if is_level_row
                    else WHITE
                )
                for target, target_value in (
                    (ws.cell(evaluation_row, col), value),
                    (ws.cell(evaluation_row, col + 1), note),
                ):
                    target.value = target_value
                    target.fill = PatternFill("solid", fgColor=fill)
                    target.font = Font(bold=is_level_row)
                    target.alignment = Alignment(
                        horizontal="center" if is_level_row else "left",
                        vertical="top",
                        wrap_text=True,
                    )
                    target.border = Border(
                        left=THIN, right=THIN, top=THIN, bottom=THIN
                    )
            ws.row_dimensions[evaluation_row].height = 46 if is_tall else 32
        last_row = grade_row + len(evaluation_rows)

        totals_row = last_row + 2
        ws.merge_cells(
            start_row=totals_row, start_column=1, end_row=totals_row, end_column=width
        )
        ws.cell(totals_row, 1, f"TOTALI I JAVËS — {department['name'].upper()}")
        _header(ws.cell(totals_row, 1), "1A1A1A")
        level_header_row = totals_row + 1
        levels = ["A+", "A", "B", "C", "M", "D", "E"]
        level_columns = [*levels, "TOTAL", "Komente / Vendim bordi"]
        for offset, label in enumerate(level_columns):
            cell = ws.cell(level_header_row, 1 + offset, label)
            _header(cell, department_color)
        level_counts = Counter(
            str(_final_level(person)) for person in people if _final_level(person)
        )
        counts_row = level_header_row + 1
        ws.cell(counts_row, 1, f"Nr. punëtorësh ({len(people)} gjithsej)")
        ws.cell(counts_row, 1).font = Font(bold=True)
        ws.cell(counts_row, 1).fill = PatternFill("solid", fgColor=GRAY)
        for offset, level in enumerate(levels, 1):
            cell = ws.cell(counts_row, 1 + offset, level_counts.get(level, 0))
            cell.fill = PatternFill("solid", fgColor=LEVEL_COLORS.get(level, WHITE))
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.cell(counts_row, 1 + len(levels) + 1, len(people)).font = Font(bold=True)
        ws.cell(counts_row, 1 + len(levels) + 1).alignment = Alignment(horizontal="center")
        for row in (level_header_row, counts_row):
            for col_index in range(1, len(level_columns) + 2):
                ws.cell(row, col_index).border = Border(
                    left=THIN, right=THIN, top=THIN, bottom=THIN
                )

        signature_row = counts_row + 2
        ws.cell(signature_row, 1, "Nënshkrimet:")
        ws.cell(signature_row, 1).font = Font(bold=True)
        ws.cell(signature_row, 1).fill = PatternFill("solid", fgColor=GRAY)
        ws.cell(signature_row + 1, 1, "Barazoi: ____________________")
        ws.cell(signature_row + 1, min(4, width), "Konfirmoi (GA): ____________________")
        last_row = signature_row + 1
        # The totals-by-level table (7 levels + TOTAL + Komente) can be wider
        # than the per-person columns for small departments — make sure the
        # print area and auto-filter cover it.
        width = max(width, len(level_columns) + 1)

        ws.freeze_panes = "B6"
        ws.column_dimensions["A"].width = 44
        for index in range(len(people)):
            ws.column_dimensions[get_column_letter(2 + index * 2)].width = 28
            ws.column_dimensions[get_column_letter(3 + index * 2)].width = 40
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = f"A5:{get_column_letter(width)}5"
        ws.print_title_rows = "1:5"
        ws.print_area = f"A1:{get_column_letter(width)}{max(5, last_row)}"

    evidence = workbook.create_sheet("Evidenca")
    _title(evidence, "REGJISTRI I EVIDENCËS", "Çdo përgjigje duhet të jetë e gjurmueshme", 8)
    evidence_headers = [
        "Departamenti", "Personi", "Lloji", "ID", "Kategoria", "Statusi",
        "Përshkrimi", "Verifikuar",
    ]
    for column, label in enumerate(evidence_headers, 1):
        evidence.cell(4, column, label)
        _header(evidence.cell(4, column))
    row = 5
    for department in departments:
        for person in department.get("people") or []:
            facts = person.get("facts_json") or {}
            for task in facts.get("tasks") or []:
                values = [
                    department["name"], person["user_name"], "TASK",
                    task.get("task_id") or task.get("match_key"),
                    SOURCE_LABELS.get(str(task.get("source_type")), task.get("source_type")),
                    STATUS_LABELS.get(str(task.get("classification")), task.get("classification")),
                    _clean_task_title(task.get("title")), "Po",
                ]
                for column, value in enumerate(values, 1):
                    evidence.cell(row, column, _value(value))
                row += 1
            for item in facts.get("observations") or []:
                values = [
                    department["name"], person["user_name"], "OBSERVATION", item.get("id"),
                    item.get("category"), item.get("marker"), item.get("comment"),
                    bool(item.get("verified")),
                ]
                for column, value in enumerate(values, 1):
                    evidence.cell(row, column, _value(value))
                row += 1
    evidence.freeze_panes = "A5"
    evidence.auto_filter.ref = f"A4:H{max(4, row - 1)}"
    for index, width in enumerate([22, 24, 14, 38, 20, 20, 55, 12], 1):
        evidence.column_dimensions[get_column_letter(index)].width = width

    guide = workbook.create_sheet("Udhëzuesi")
    _title(guide, "UDHËZUESI I VLERËSIMIT", "Rregulla deterministe, pa vlera monetare", 4)
    for column, label in enumerate(["Niveli", "Simboli", "Kushti", "Evidenca minimale"], 1):
        guide.cell(4, column, label)
        _header(guide.cell(4, column))
    rules = [
        ("A+", "+", "Plani i plotë + të paktën 2 angazhime ekstra të verifikuara", "Taska + observime të verifikuara"),
        ("A", "+", "Plani i plotë + 1 angazhim ekstra i verifikuar", "Taska + observim i verifikuar"),
        ("B", "+", "Plani normal i plotë", "Snapshot + prezenca"),
        ("C", "+/-", "3+ vonesa, përfundim me vonesë ose ndikim i vogël", "Prezenca/task/audit"),
        ("M", "+/-", "Mungesë personale e aprovuar dhe detyrat e mbuluara", "Aprovim + taska"),
        ("D", "-", "Shtyrje pa aprovim, takim i humbur ose ndikim i madh", "Evidencë e verifikuar"),
        ("E", "-", "Pa progres ose 2+ mungesa të papritura", "Progres + prezenca"),
    ]
    for row, rule in enumerate(rules, 5):
        for column, value in enumerate(rule, 1):
            cell = guide.cell(row, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        guide.cell(row, 1).fill = PatternFill("solid", fgColor=LEVEL_COLORS[rule[0]])
        guide.cell(row, 1).font = Font(bold=True)
    guide.merge_cells(start_row=13, start_column=1, end_row=13, end_column=4)
    guide.cell(13, 1, "SI LEXOHET RAPORTI")
    _header(guide.cell(13, 1))
    glossary = [
        ("Progresi javor", "Të mbyllura / në planin javor", "Akumulohet deri në snapshot-in e fundit", "Taskat e PLANNED"),
        ("Progresi ditor", "Të mbyllura sot / të planifikuara sot", "Nuk zëvendëson progresin javor", "Snapshot-i i orës 16:20"),
        ("Shtesë", "Taska pas PLANNED", "Raportohet veç; nuk fsheh planin e pambyllur", "Created_at + audit"),
        ("PV/FEST", "Pushim vjetor i plotë", "Personi përjashtohet nga realizimi", "Common View"),
    ]
    for row, item in enumerate(glossary, 14):
        for column, value in enumerate(item, 1):
            cell = guide.cell(row, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        guide.cell(row, 1).font = Font(bold=True)
        guide.row_dimensions[row].height = 34
    question_row = 20
    guide.merge_cells(
        start_row=question_row,
        start_column=1,
        end_row=question_row,
        end_column=4,
    )
    guide.cell(question_row, 1, "15 PYETJET E RAPORTIT")
    _header(guide.cell(question_row, 1))
    question_number = 1
    for section_title, question_keys in REPORT_QUESTION_SECTIONS:
        question_row += 1
        guide.merge_cells(
            start_row=question_row,
            start_column=1,
            end_row=question_row,
            end_column=4,
        )
        guide.cell(question_row, 1, section_title)
        _header(guide.cell(question_row, 1), "1F4E78")
        guide.cell(question_row, 1).alignment = Alignment(horizontal="left")
        for key in question_keys:
            question_row += 1
            values = (
                question_number,
                QUESTION_LABELS[key],
                "AUTO ose MANUAL",
                "Pa provë shënohet PËR KONFIRMIM; nuk hamendësohet.",
            )
            for column, value in enumerate(values, 1):
                cell = guide.cell(question_row, column, value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            guide.row_dimensions[question_row].height = 30
            question_number += 1
    for index, width in enumerate([18, 24, 65, 45], 1):
        guide.column_dimensions[get_column_letter(index)].width = width
    guide.freeze_panes = "A5"

    for ws in workbook.worksheets:
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
