from __future__ import annotations

import html
import os
import re
from io import BytesIO
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from app.config import settings
from app.services.meetings_report import common_view_item_sort_key, next_working_day
from app.services.primeflow_report import GmailService, PrimeFlowClient


TASK_ROWS = (
    ("oneH", "1H 10:00", "10:00"),
    ("oneH", "1H 11:00", "11:00"),
    ("oneH", "1H 11:50", "11:50"),
    ("oneH", "1H 14:20", "14:20"),
    ("blocked", "BLL\n14:30 - 15:30\nRAP 15:50", None),
    ("oneH", "1H 15:50", "15:50"),
    ("oneH", "1H NO SLOT", ""),
    ("important", "DEADLINE / 08:00", None),
    ("r1", "R1=1H", None),
    ("personal", "P:\nGA 08:15 / 13:15\n\nDV/LH 10:15 / 14:30", None),
)
MEETING_ROWS = (("external", "TAK EXT"), ("internal", "TAK INT"))
VALID_1H_SLOTS = {"10:00", "11:00", "11:50", "14:20", "15:50"}
ONE_H_BOARD_CHECKLIST = (
    ("Slotin paraprak/aktual", ""),
    ("A ke filluar me slotin aktual?", ""),
    ("Nese jo, kur?", ""),
    ("A kryhet sot?", ""),
    ("A kryhet kete jave?", ""),
    ("A arrihet RLZ javor?", ""),
    ("Done? / Strikes? / Notes te reja? Data? AM/PM? Kujt?", ""),
    ("BZ Notes", "Secili i lexon vet para BZ me GA"),
)
ONE_H_STAFF_CHECKLIST = (
    ("Hap doc dhe det", ""),
    ("Share screen side by side DET/REZULTATIN", ""),
    ("Sqaro slotin paraprak pastaj aktual", ""),
    ("BZ Det nga Stafi per GA", "Komunikimi GA temas Det nga Stafi/ KA email"),
)

# Gmail can remove style blocks from message bodies. Keep the styles that form
# the report grid inline so the received email matches the preview.
TABLE_STYLE = "width:100%;border-collapse:collapse;table-layout:fixed;margin:12px 0;font-family:Arial,sans-serif;font-size:12px;line-height:1.25;color:#000"
CELL_STYLE = "border:1px solid #000;padding:5px;vertical-align:top;text-align:left;overflow-wrap:anywhere;word-break:break-word"
HEADER_STYLE = f"{CELL_STYLE};text-align:center;font-weight:700"
SLOT_DIVIDER_STYLE = "border-top:2px solid #111827"
INTRA_SLOT_DIVIDER_STYLE = "border-top:1px solid #cbd5e1"
SLOT_LABEL_STYLE = f"{CELL_STYLE};font-weight:700"
PERSONAL_GA_COLOR = "#D8B4FE"
PERSONAL_GA_CELL_STYLE = f"{CELL_STYLE};background-color:{PERSONAL_GA_COLOR}"
PERSONAL_ROW_LABEL_STYLE = (
    f"{CELL_STYLE};font-size:10px;line-height:1.15;white-space:pre-line;"
    "overflow-wrap:normal;word-break:normal"
)
DEADLINE_COLOR = "#DC2626"
EIGHT_AM_BORDER_COLOR = "#DC2626"
NON_ROUTINE_MEETING_BORDER_COLOR = "#2563EB"
PERSONAL_TASK_INITIALS = re.compile(r"^[A-Z]{2,3}(?:\s*[:/]\s*[A-Z]{2,3})*(?=\s|:|/|$)", re.I)
NOTE_MARKERS_RE = re.compile(r"\[\[\s*/?\s*(?:added|done)\s*\]\]", re.I)
EIGHT_AM_MARKER_RE = re.compile(r"\b0?8:00\b")
STATUS_COLORS = {
    "TODO": "#FFC4ED",
    "IN_PROGRESS": "#FFFF00",
    "WAITING_CONFIRMATION": "#FFEDD5",
    "DONE": "#C4FDC4",
}
COMMENT_FIXED_INITIALS = ("AT", "RA", "EF", "EH", "LH", "FG")
COMMENT_ITEMS_PER_LINE = 4
COMMENT_WRITE_IN_LINE = "_" * 20
REQUIRED_SHTYPI_RECIPIENT = "130primex.eu@gmail.com"


def ensure_required_shtypi_recipient(
    recipients: dict[str, list[str]],
) -> dict[str, list[str]]:
    """Always include the required archive/print mailbox as a To recipient."""
    result = {key: [] for key in ("to", "cc", "bcc")}
    seen: set[str] = set()
    for key in ("to", "cc", "bcc"):
        for raw in recipients.get(key, []):
            email = str(raw or "").strip()
            normalized = email.casefold()
            if not email or normalized in seen:
                continue
            seen.add(normalized)
            result[key].append(email)
    required_key = REQUIRED_SHTYPI_RECIPIENT.casefold()
    if required_key not in seen:
        result["to"].append(REQUIRED_SHTYPI_RECIPIENT)
    elif all(
        email.casefold() != required_key for email in result["to"]
    ):
        for key in ("cc", "bcc"):
            result[key] = [email for email in result[key] if email.casefold() != required_key]
        result["to"].append(REQUIRED_SHTYPI_RECIPIENT)
    return result


def subject_for(target_date: date) -> str:
    return f"1H SHTYPI - {target_date:%d.%m.%Y}"


def _item_date(item: dict[str, Any]) -> date | None:
    raw = item.get("date") or item.get("entryDate") or item.get("entry_date")
    if not raw:
        return None
    try:
        return date.fromisoformat(str(raw)[:10])
    except ValueError:
        return None


def _slot(item: dict[str, Any]) -> str:
    raw = str(item.get("oneHReportSlot") or item.get("one_h_report_slot") or "").strip()
    return raw if raw in VALID_1H_SLOTS else ""


def _first_line(value: Any) -> str:
    return next((line.strip() for line in str(value or "").splitlines() if line.strip()), "")


def _report_text(value: Any) -> str:
    """Remove note-editor markup; recipients must only see the task text."""
    return re.sub(r"\s{2,}", " ", NOTE_MARKERS_RE.sub("", str(value or ""))).strip()


def _task_status(item: dict[str, Any]) -> str:
    raw = str(item.get("status") or item.get("task_status") or item.get("state") or "TODO")
    normalized = raw.strip().upper().replace(" ", "_").replace("-", "_")
    if normalized in {"COMPLETED", "COMPLETE"}:
        return "DONE"
    if normalized in {"TO_DO", "TODO"}:
        return "TODO"
    if normalized in {"INPROGRESS", "IN_PROGRESS"}:
        return "IN_PROGRESS"
    if normalized in {"WAITING", "PENDING_CONFIRMATION", "WAITING_CONFIRMATION"}:
        return "WAITING_CONFIRMATION"
    return normalized if normalized in STATUS_COLORS else "TODO"


def _task_period_label(item: dict[str, Any]) -> str:
    """Match the AM/PM indicator used by Common View task cards."""
    raw = str(item.get("finishPeriod") or item.get("finish_period") or "").strip().upper()
    return raw if raw in {"AM", "PM"} else "AM/PM"


def _task_cell_style(
    item: dict[str, Any], *, personal: bool, report_date: date | None = None
) -> tuple[str, str]:
    """Use red cards for deadlines outside the report day; today's keep status color."""
    deadline = bool(item.get("is_deadline_important") or item.get("isDeadlineImportant"))
    due_day = _task_due_day(item) if deadline else None
    if deadline and (report_date is None or due_day is None or due_day != report_date):
        color = DEADLINE_COLOR
    elif personal and _is_personal_task_for_ga(item):
        color = PERSONAL_GA_COLOR
    else:
        color = STATUS_COLORS[_task_status(item)]
    border = f";border:2px solid {EIGHT_AM_BORDER_COLOR}" if _is_eight_am_task(item) else ""
    text_color = ";color:#fff;font-weight:700" if color == DEADLINE_COLOR else ""
    return f"{CELL_STYLE};background-color:{color}{border}{text_color}", color


def _initials(value: str) -> str:
    cleaned = value.strip()
    if not cleaned:
        return ""
    if re.fullmatch(r"[A-Za-z]{1,4}", cleaned):
        return cleaned.upper()
    parts = re.findall(r"[^\W\d_]+", cleaned, flags=re.UNICODE)
    return "".join(part[0] for part in parts).upper()


def _comment_user_initials(payload: dict[str, Any]) -> list[str]:
    """Fixed report users followed by PCM users in Weekly Planner order."""
    departments = {
        str(row.get("id")): str(row.get("code") or row.get("name") or "").strip().upper()
        for row in (payload.get("departments") or [])
        if isinstance(row, dict) and row.get("id") is not None
    }
    pcm_aliases = {"PCM", "PRODUCT CONTENT", "PROJECT CONTENT", "PROJECT CONTENT MANAGER"}
    pcm_users: list[tuple[int, int, str, str]] = []
    for row in payload.get("users") or []:
        if not isinstance(row, dict) or row.get("is_active") is False:
            continue
        if departments.get(str(row.get("department_id"))) not in pcm_aliases:
            continue
        label = str(row.get("full_name") or row.get("username") or "").strip()
        initials = _initials(label)
        if not initials:
            continue
        order = row.get("weekly_planner_sort_order")
        pcm_users.append((1 if order is None else 0, int(order or 0), label.casefold(), initials))

    result = list(COMMENT_FIXED_INITIALS)
    for *_, initials in sorted(pcm_users):
        if initials not in result:
            result.append(initials)
    return result


def _comment_write_in_lines(initials: list[str]) -> list[str]:
    values = initials or list(COMMENT_FIXED_INITIALS)
    entries = [f"{value}: {COMMENT_WRITE_IN_LINE}" for value in values]
    return [
        ",    ".join(entries[index:index + COMMENT_ITEMS_PER_LINE])
        for index in range(0, len(entries), COMMENT_ITEMS_PER_LINE)
    ]


def _comments_table_html(initials: list[str]) -> str:
    """Render compact handwritten comment lines without table cells."""
    values = initials or list(COMMENT_FIXED_INITIALS)
    rows = []
    for index in range(0, len(values), COMMENT_ITEMS_PER_LINE):
        chunk = values[index:index + COMMENT_ITEMS_PER_LINE]
        entries = "".join(
            '<span data-user-comment="{initials}" style="display:inline-block;margin:0 20px 8px 0;'
            'font-family:Arial,sans-serif;font-size:13px;white-space:nowrap;">'
            '<strong>{initials}:</strong> {line}{suffix}</span>'.format(
                initials=html.escape(value),
                line=COMMENT_WRITE_IN_LINE,
                suffix="," if position < len(chunk) - 1 else "",
            )
            for position, value in enumerate(chunk)
        )
        rows.append(f'<div data-user-comment-line="true" style="white-space:nowrap;">{entries}</div>')
    return (
        '<div data-user-comments-lines="true" style="margin-top:18px;">'
        '<div style="font-family:Arial,sans-serif;font-size:16px;font-weight:800;margin:0 0 6px;">'
        "KOMENTE PER STAF</div>"
        f"{''.join(rows)}</div>"
    )


def _assignees(item: dict[str, Any]) -> list[str]:
    raw = item.get("assignees")
    if not isinstance(raw, list) or not raw:
        raw = str(item.get("person") or item.get("owner") or "").split(",")
    result: list[str] = []
    for value in raw:
        label = _initials(str(value or ""))
        if label and label not in result:
            result.append(label)
    return result


def _task_title(item: dict[str, Any], *, personal: bool) -> str:
    title = _report_text(_first_line(item.get("title")))
    if personal:
        return title
    title = re.sub(r"^[A-Z]{1,4}(?:/[A-Z]{1,4})*:\s*", "", title)
    owners = _assignees(item)
    return f"{'/'.join(owners)}: {title}" if owners else title


def _is_eight_am_task(item: dict[str, Any]) -> bool:
    title = " ".join(str(item.get(key) or "") for key in ("title", "task_title"))
    if EIGHT_AM_MARKER_RE.search(title):
        return True
    raw_due_date = item.get("due_date") or item.get("dueDate")
    if isinstance(raw_due_date, datetime):
        return raw_due_date.hour == 8 and raw_due_date.minute == 0
    match = re.search(r"T08:00(?::00)?", str(raw_due_date or ""))
    return bool(match)


def _task_due_day(item: dict[str, Any]) -> date | None:
    raw = (
        item.get("due_date")
        or item.get("dueDate")
        or item.get("deadline_date")
        or item.get("deadlineDate")
    )
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    try:
        return date.fromisoformat(str(raw or "")[:10])
    except ValueError:
        return None


def _task_badges_html(item: dict[str, Any], report_date: date | None) -> tuple[str, str]:
    top_badges: list[str] = []
    due_badge = ""
    badge_base = (
        "display:inline-block;float:right;margin:0 0 3px 4px;padding:3px 7px;border-radius:999px;"
        "font-family:Arial,sans-serif;font-size:10px;font-weight:800;line-height:1;white-space:nowrap;"
    )
    period = _task_period_label(item)
    period_style = (
        f"{badge_base}background-color:#E0F2FE;border:1px solid #BAE6FD;color:#0369A1;"
    )
    top_badges.append(
        f'<span data-task-badge="finish-period" style="{period_style}">{period}</span>'
    )
    if _is_eight_am_task(item):
        eight_am_style = (
            f"{badge_base}background-color:#DC2626;border:1px solid #B91C1C;color:#FFFFFF;"
        )
        top_badges.append(
            f'<span data-task-badge="08:00" style="{eight_am_style}">08:00</span>'
        )
    if bool(item.get("is_deadline_important") or item.get("isDeadlineImportant")):
        due_day = _task_due_day(item)
        if due_day:
            due_today = report_date is not None and due_day == report_date
            style = (
                "display:inline-block;padding:2px 5px;border:1px solid #93C5FD;border-radius:3px;"
                "background-color:#EFF6FF;color:#1D4ED8;font-family:Arial,sans-serif;"
                "font-size:11px;font-weight:900;line-height:1.1;white-space:nowrap;"
                if due_today else
                "display:inline-block;padding:1px 0;border:0;background-color:transparent;color:#FFFFFF;"
                "font-family:Arial,sans-serif;font-size:10px;font-weight:900;line-height:1.1;white-space:nowrap;"
            )
            due_badge = (
                f'<span data-task-badge="due-date" data-badge-position="bottom-right" '
                f'data-due-today="{str(due_day == report_date).lower()}" '
                f'style="{style}">{due_day:%d.%m.%Y}</span>'
            )
    return "".join(top_badges), due_badge


def _is_personal_task_for_ga(item: dict[str, Any]) -> bool:
    """Match the GA-assignee rule used by the P row in Common View printouts."""
    title = _report_text(_first_line(item.get("title")))
    match = PERSONAL_TASK_INITIALS.match(title.strip())
    if match is None:
        return False
    return "GA" in {value.strip().upper() for value in re.split(r"[:/]", match.group(0))}


def _dedupe(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for item in items:
        key = (
            _first_line(item.get("title")).casefold(),
            str(item.get("date") or ""),
            _slot(item),
            str(item.get("finishPeriod") or item.get("finish_period") or "").upper(),
        )
        if key not in seen:
            seen.add(key)
            result.append(item)
    return result


def _task_rows(items: dict[str, Any], target_date: date) -> list[tuple[str, list[dict[str, Any]], bool]]:
    by_bucket = {
        name: [item for item in values if isinstance(item, dict) and _item_date(item) == target_date]
        for name, values in items.items()
        if isinstance(values, list)
    }
    rows: list[tuple[str, list[dict[str, Any]], bool]] = []
    for bucket, label, requested_slot in TASK_ROWS:
        values = list(by_bucket.get(bucket, []))
        if bucket == "oneH":
            values = [item for item in values if _slot(item) == requested_slot]
        values = _dedupe(values)
        # Completed work belongs at the end of its slot so unfinished work is
        # immediately visible in the printed report.
        values.sort(key=lambda item: (_task_status(item) == "DONE", common_view_item_sort_key(item)))
        rows.append((label, values, bucket == "personal"))
    return rows


def _meeting_rows(items: dict[str, Any], target_date: date) -> list[tuple[str, list[dict[str, Any]]]]:
    rows: list[tuple[str, list[dict[str, Any]]]] = []
    for bucket, label in MEETING_ROWS:
        values = [item for item in items.get(bucket, []) if isinstance(item, dict) and _item_date(item) == target_date]
        values.sort(key=lambda item: (str(item.get("time") or ""), _first_line(item.get("title")).casefold()))
        rows.append((label, values))
    return rows


def _is_non_routine_meeting(item: dict[str, Any]) -> bool:
    recurrence = str(item.get("recurrence_type") or item.get("recurrenceType") or "").strip().lower()
    return recurrence not in {"daily", "weekly"}


def _one_h_checklists_html() -> str:
    """The two preparation checklists shown above every 1H Shtypi task grid."""

    def checklist(title: str, questions: tuple[tuple[str, str], ...], *, board: bool = False) -> str:
        separators = (
            '<span style="font-size:20px;font-weight:900;color:#111827;line-height:12px;"> / </span>'
        )
        question_text = separators.join(
            f'<strong>{index}. {html.escape(question)}</strong>'
            + (f' <span style="color:#475569;">({html.escape(description)})</span>' if description else "")
            for index, (question, description) in enumerate(questions, 1)
        )
        board_marker = ' data-board-checklist-columns="true"' if board else ""
        return (
            '<table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" '
            f'data-compact-checklist-row="true"{board_marker} style="width:100%;border-collapse:collapse;">'
            '<tr><th style="background-color:#eef2ff;border-left:5px solid #2563eb;padding:10px 12px;'
            f'font-family:Arial,sans-serif;font-size:14px;text-align:left;">{html.escape(title)}</th></tr>'
            '<tr><td style="border:1px solid #64748b;padding:8px 10px;font-family:Arial,sans-serif;'
            f'font-size:12px;line-height:1.45;">{question_text}</td></tr></table>'
        )

    return (
        '<table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" '
        'data-one-h-checklist-columns="true" style="width:100%;border-collapse:collapse;margin:0 0 14px;">'
        '<tr>'
        '<td width="50%" valign="top" style="width:50%;padding:0 6px 0 0;vertical-align:top;">'
        f"{checklist('STAFF - HAPAT PER 1H', ONE_H_STAFF_CHECKLIST)}"
        '</td>'
        '<td width="50%" valign="top" style="width:50%;padding:0 0 0 6px;vertical-align:top;">'
        f"{checklist('PYETJET PER 1H - BORD', ONE_H_BOARD_CHECKLIST, board=True)}"
        '</td>'
        '</tr></table>'
    )


def _html_table(
    rows: list[tuple[str, list[dict[str, Any]], bool]], *, meeting: bool = False, report_date: date | None = None
) -> str:
    header = "MEETING" if meeting else "TASK"
    label_header = "LLOJI" if meeting else "LLOJI DHE SLOTI"
    body: list[str] = []
    for number, (label, values, *rest) in enumerate(rows, 1):
        personal = bool(rest and rest[0])
        chunks = [values[index:index + 6] for index in range(0, len(values), 6)] or [[]]
        for chunk_index, chunk in enumerate(chunks):
            row_divider_style = INTRA_SLOT_DIVIDER_STYLE if chunk_index else SLOT_DIVIDER_STYLE
            cells: list[str] = []
            for item_index, item in enumerate(chunk):
                value = (
                    f"{_report_text(_first_line(item.get('title')))} {str(item.get('time') or '').strip()}".strip()
                    if meeting
                    else _task_title(item, personal=personal)
                )
                if meeting:
                    cell_style = (
                        f"{CELL_STYLE};border:2px solid {NON_ROUTINE_MEETING_BORDER_COLOR}"
                        if _is_non_routine_meeting(item) else CELL_STYLE
                    )
                    color = ""
                else:
                    cell_style, color = _task_cell_style(
                        item, personal=personal, report_date=report_date
                    )
                cell_style = f"{cell_style};{row_divider_style}"
                background = f' bgcolor="{color}"' if color else ""
                badges, due_badge = ("", "") if meeting else _task_badges_html(item, report_date)
                task_content = f'{badges}{item_index + (chunk_index * 6) + 1}. {html.escape(value)}'
                if due_badge:
                    task_content = (
                        '<table role="presentation" width="100%" height="100%" border="0" cellpadding="0" '
                        'cellspacing="0" style="width:100%;height:100%;border-collapse:collapse;">'
                        f'<tr><td valign="top" style="padding:0;vertical-align:top;">{task_content}</td></tr>'
                        '<tr><td valign="bottom" align="right" '
                        'style="padding:6px 0 0;text-align:right;vertical-align:bottom;">'
                        f'{due_badge}</td></tr></table>'
                    )
                cells.append(
                    f'<td{background} style="{cell_style}">{task_content}</td>'
                )
            cells.extend(f'<td style="{CELL_STYLE};{row_divider_style}"></td>' for _ in range(6 - len(cells)))
            row_header = (
                f'<th rowspan="{len(chunks)}" style="{SLOT_LABEL_STYLE};{row_divider_style}">{number}</th>'
                f'<th rowspan="{len(chunks)}" style="{PERSONAL_ROW_LABEL_STYLE if personal else SLOT_LABEL_STYLE};{row_divider_style}">{html.escape(label).replace(chr(10), "<br>")}</th>'
                if chunk_index == 0 else ""
            )
            body.append(f"<tr>{row_header}{''.join(cells)}</tr>")
    return (
        f'<table role="presentation" width="100%" border="1" cellpadding="0" cellspacing="0" style="{TABLE_STYLE}">'
        '<colgroup><col width="4%"><col width="9%"><col width="14.5%" span="6"></colgroup>'
        f'<thead><tr><th style="{HEADER_STYLE}">NR</th><th style="{HEADER_STYLE}">{label_header}</th>'
        + (
            "".join(f'<th style="{HEADER_STYLE}">{header} {index}</th>' for index in range(1, 7))
            if meeting else f'<th colspan="6" style="{HEADER_STYLE}">TASKS</th>'
        )
        + '</tr></thead>'
        f"<tbody>{''.join(body)}</tbody></table>"
    )


def _dated_meetings_html(
    sections: list[tuple[date, str, list[tuple[str, list[dict[str, Any]], bool]]]],
) -> str:
    if not sections:
        return ""

    visible_sections = sections[:2]
    while len(visible_sections) < 2:
        visible_sections.append((visible_sections[0][0], "", []))
    left, right = visible_sections
    divider_style = "border-left:4px solid #2563EB"

    def meeting_items(values: list[dict[str, Any]]) -> str:
        rendered: list[str] = []
        for index, item in enumerate(values, 1):
            value = (
                f"{_report_text(_first_line(item.get('title')))} "
                f"{str(item.get('time') or '').strip()}"
            ).strip()
            highlight = (
                f"border:2px solid {NON_ROUTINE_MEETING_BORDER_COLOR};padding:3px 5px;"
                if _is_non_routine_meeting(item)
                else "padding:3px 1px;"
            )
            rendered.append(
                f'<div data-meeting-card="true" style="{highlight}'
                'margin:0 0 3px;overflow-wrap:anywhere;word-break:break-word;">'
                f"{index}. {html.escape(value)}</div>"
            )
        return "".join(rendered) or "&nbsp;"

    left_rows = {label: values for label, values, *_ in left[2]}
    right_rows = {label: values for label, values, *_ in right[2]}
    labels = list(left_rows)
    labels.extend(label for label in right_rows if label not in left_rows)
    body = "".join(
        "<tr>"
        f'<th style="{SLOT_LABEL_STYLE};{SLOT_DIVIDER_STYLE}">{html.escape(label)}</th>'
        f'<td style="{CELL_STYLE};{SLOT_DIVIDER_STYLE}">{meeting_items(left_rows.get(label, []))}</td>'
        f'<th style="{SLOT_LABEL_STYLE};{SLOT_DIVIDER_STYLE};{divider_style}">{html.escape(label)}</th>'
        f'<td style="{CELL_STYLE};{SLOT_DIVIDER_STYLE}">{meeting_items(right_rows.get(label, []))}</td>'
        "</tr>"
        for label in labels
    )

    def day_header(section: tuple[date, str, list[tuple[str, list[dict[str, Any]], bool]]]) -> str:
        meeting_date, relative, _ = section
        return f"{html.escape(relative)} - {meeting_date:%d.%m.%Y}" if relative else "&nbsp;"

    return (
        '<table data-side-by-side-meetings="true" role="presentation" width="100%" border="1" '
        f'cellpadding="0" cellspacing="0" style="{TABLE_STYLE};margin-top:18px">'
        '<colgroup><col width="9%"><col width="41%"><col width="9%"><col width="41%"></colgroup>'
        '<thead><tr>'
        f'<th colspan="2" style="{HEADER_STYLE};background-color:#EEF2FF;border-left:5px solid #2563EB;'
        f'font-size:15px;">{day_header(left)}</th>'
        f'<th colspan="2" style="{HEADER_STYLE};background-color:#EEF2FF;{divider_style};font-size:15px;">'
        f'{day_header(right)}</th></tr>'
        '<tr>'
        f'<th style="{HEADER_STYLE}">LLOJI</th><th style="{HEADER_STYLE}">TAKIMET</th>'
        f'<th style="{HEADER_STYLE};{divider_style}">LLOJI</th>'
        f'<th style="{HEADER_STYLE}">TAKIMET</th>'
        f"</tr></thead><tbody>{body}</tbody></table>"
    )


def _excel_table_attachment(
    task_rows: list[tuple[str, list[dict[str, Any]], bool]],
    meeting_rows: list[tuple[str, list[dict[str, Any]], bool]],
    target_date: date,
    *,
    include_meetings: bool = True,
    comment_initials: list[str] | None = None,
    meeting_sections: list[tuple[date, str, list[tuple[str, list[dict[str, Any]], bool]]]] | None = None,
) -> tuple[str, bytes, str]:
    """Create the same printable grid as an XLSX attachment for email recipients."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "1H SHTYPI"
    sheet.merge_cells("A1:H1")
    title_cell = sheet["A1"]
    title_cell.value = f"1H SHTYPI - {target_date:%d.%m.%Y}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    border = Border(
        left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"),
    )
    slot_divider_border = Border(
        left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
        top=Side(style="medium", color="111827"), bottom=Side(style="thin", color="000000"),
    )
    intra_slot_divider_border = Border(
        left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="000000"),
    )
    header_fill = PatternFill("solid", fgColor="EAF0FF")
    fills = {
        status: PatternFill("solid", fgColor=color.removeprefix("#"))
        for status, color in STATUS_COLORS.items()
    }
    ga_fill = PatternFill("solid", fgColor=PERSONAL_GA_COLOR.removeprefix("#"))
    deadline_fill = PatternFill("solid", fgColor=DEADLINE_COLOR.removeprefix("#"))
    eight_am_border = Border(
        left=Side(style="medium", color=EIGHT_AM_BORDER_COLOR.removeprefix("#")),
        right=Side(style="medium", color=EIGHT_AM_BORDER_COLOR.removeprefix("#")),
        top=Side(style="medium", color=EIGHT_AM_BORDER_COLOR.removeprefix("#")),
        bottom=Side(style="medium", color=EIGHT_AM_BORDER_COLOR.removeprefix("#")),
    )
    non_routine_meeting_border = Border(
        left=Side(style="medium", color=NON_ROUTINE_MEETING_BORDER_COLOR.removeprefix("#")),
        right=Side(style="medium", color=NON_ROUTINE_MEETING_BORDER_COLOR.removeprefix("#")),
        top=Side(style="medium", color=NON_ROUTINE_MEETING_BORDER_COLOR.removeprefix("#")),
        bottom=Side(style="medium", color=NON_ROUTINE_MEETING_BORDER_COLOR.removeprefix("#")),
    )
    def write_checklists(row_number: int) -> int:
        """Write each preparation list as a title row plus one compact content row."""
        checklist_fill = PatternFill("solid", fgColor="EEF2FF")
        for start_column, end_column, title, questions in (
            (1, 4, "STAFF - HAPAT PER 1H", ONE_H_STAFF_CHECKLIST),
            (5, 8, "PYETJET PER 1H - BORD", ONE_H_BOARD_CHECKLIST),
        ):
            sheet.merge_cells(start_row=row_number, start_column=start_column, end_row=row_number, end_column=end_column)
            title_cell = sheet.cell(row_number, start_column, title)
            title_cell.fill = checklist_fill
            title_cell.font = Font(bold=True, size=11)
            title_cell.alignment = Alignment(vertical="center")
            title_cell.border = border
            sheet.merge_cells(
                start_row=row_number + 1,
                start_column=start_column,
                end_row=row_number + 1,
                end_column=end_column,
            )
            question_cell = sheet.cell(
                row_number + 1,
                start_column,
                " / ".join(
                    f"{index}. {question}" + (f" ({description})" if description else "")
                    for index, (question, description) in enumerate(questions, 1)
                ),
            )
            question_cell.font = Font(bold=True, size=10)
            question_cell.alignment = Alignment(vertical="center", wrap_text=True)
            question_cell.border = border
        sheet.row_dimensions[row_number + 1].height = 72
        return row_number + 2

    def write_section(
        rows: list[tuple[str, list[dict[str, Any]], bool]], *, meeting: bool, row_number: int
    ) -> int:
        section_headers = (
            ["NR", "LLOJI", "MEETING 1", "MEETING 2", "MEETING 3", "MEETING 4", "MEETING 5", "MEETING 6"]
            if meeting else ["NR", "LLOJI DHE SLOTI", "TASKS", None, None, None, None, None]
        )
        for column, value in enumerate(section_headers, 1):
            cell = sheet.cell(row_number, column, value)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if not meeting:
            sheet.merge_cells(start_row=row_number, start_column=3, end_row=row_number, end_column=8)
        row_number += 1
        for number, (label, values, personal) in enumerate(rows, 1):
            chunks = [values[index:index + 6] for index in range(0, len(values), 6)] or [[]]
            first_row = row_number
            for chunk_index, chunk in enumerate(chunks):
                row_border = intra_slot_divider_border if chunk_index else slot_divider_border
                if chunk_index == 0:
                    sheet.cell(row_number, 1, number)
                    label_cell = sheet.cell(row_number, 2, label)
                    label_cell.font = Font(bold=True, size=10)
                    if personal:
                        label_cell.alignment = Alignment(vertical="top", wrap_text=True)
                for item_index, item in enumerate(chunk, 3):
                    value = (
                        f"{_report_text(_first_line(item.get('title')))} {str(item.get('time') or '').strip()}".strip()
                        if meeting else _task_title(item, personal=personal)
                    )
                    if not meeting:
                        labels: list[str] = [f"[{_task_period_label(item)}]"]
                        if _is_eight_am_task(item):
                            labels.append("[08:00]")
                        if bool(item.get("is_deadline_important") or item.get("isDeadlineImportant")):
                            due_day = _task_due_day(item)
                            if due_day:
                                labels.append(f"[{due_day:%d.%m.%Y}]")
                        if labels:
                            value = f"{' '.join(labels)}\n{value}"
                    cell = sheet.cell(row_number, item_index, f"{item_index - 2 + chunk_index * 6}. {value}")
                    if not meeting:
                        deadline = bool(item.get("is_deadline_important") or item.get("isDeadlineImportant"))
                        due_day = _task_due_day(item) if deadline else None
                        if deadline and (due_day is None or due_day != target_date):
                            cell.fill = deadline_fill
                            cell.font = Font(color="FFFFFF", bold=True)
                        else:
                            cell.fill = ga_fill if personal and _is_personal_task_for_ga(item) else fills[_task_status(item)]
                        if _is_eight_am_task(item):
                            cell.border = eight_am_border
                    elif _is_non_routine_meeting(item):
                        cell.border = non_routine_meeting_border
                for column in range(1, 9):
                    cell = sheet.cell(row_number, column)
                    is_highlighted_meeting_cell = (
                        meeting
                        and column >= 3
                        and column - 3 < len(chunk)
                        and _is_non_routine_meeting(chunk[column - 3])
                    )
                    is_eight_am_task_cell = (
                        not meeting
                        and column >= 3
                        and column - 3 < len(chunk)
                        and _is_eight_am_task(chunk[column - 3])
                    )
                    if not is_highlighted_meeting_cell and not is_eight_am_task_cell:
                        cell.border = row_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                row_number += 1
            if len(chunks) > 1:
                sheet.merge_cells(start_row=first_row, start_column=1, end_row=row_number - 1, end_column=1)
                sheet.merge_cells(start_row=first_row, start_column=2, end_row=row_number - 1, end_column=2)
        return row_number

    task_header_row = write_checklists(3)
    next_row = write_section(task_rows, meeting=False, row_number=task_header_row)
    if include_meetings:
        if meeting_sections:
            for meeting_date, relative, dated_rows in meeting_sections:
                section_row = next_row + 1
                sheet.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=8)
                section_cell = sheet.cell(section_row, 1, f"TAKIMET - {relative} - {meeting_date:%d.%m.%Y}")
                section_cell.font = Font(bold=True, size=12)
                section_cell.fill = PatternFill("solid", fgColor="EEF2FF")
                section_cell.alignment = Alignment(horizontal="left", vertical="center")
                next_row = write_section(dated_rows, meeting=True, row_number=section_row + 1)
        else:
            next_row = write_section(meeting_rows, meeting=True, row_number=next_row + 1)

    comment_columns = comment_initials or list(COMMENT_FIXED_INITIALS)
    comment_start_row = next_row + 1
    sheet.merge_cells(
        start_row=comment_start_row,
        start_column=1,
        end_row=comment_start_row,
        end_column=8,
    )
    comment_title = sheet.cell(comment_start_row, 1, "KOMENTE PER STAF")
    comment_title.font = Font(bold=True, size=12)
    comment_title.alignment = Alignment(horizontal="left", vertical="center")
    for offset, line in enumerate(_comment_write_in_lines(comment_columns), 1):
        row = comment_start_row + offset
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        line_cell = sheet.cell(row, 1, line)
        line_cell.font = Font(size=11)
        line_cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[row].height = 22
    widths = [6, 22, 29, 29, 29, 29, 29, 29]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = f"C{task_header_row + 1}"

    output = BytesIO()
    workbook.save(output)
    return (
        f"1H_SHTYPI_{target_date:%Y-%m-%d}.xlsx",
        output.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _png_table_attachment(
    task_rows: list[tuple[str, list[dict[str, Any]], bool]], target_date: date,
    comment_initials: list[str] | None = None,
    meeting_sections: list[tuple[date, str, list[tuple[str, list[dict[str, Any]], bool]]]] | None = None,
) -> tuple[str, bytes, str]:
    """Render the Today SHTYPI task grid with the same task-state colours."""
    margin = 28
    column_widths = [58, 210, *([267] * 6)]
    width = sum(column_widths) + (margin * 2)
    try:
        regular = ImageFont.truetype(os.getenv("PRIMEFLOW_REPORT_FONT_PATH", r"C:\Windows\Fonts\segoeui.ttf"), 16)
        bold = ImageFont.truetype(r"C:\Windows\Fonts\arialbd.ttf", 16)
        small_bold = ImageFont.truetype(r"C:\Windows\Fonts\arialbd.ttf", 14)
        heading = ImageFont.truetype(r"C:\Windows\Fonts\arialbd.ttf", 27)
    except OSError:
        regular = bold = small_bold = heading = ImageFont.load_default()
    measure = ImageDraw.Draw(Image.new("RGB", (1, 1), "white"))

    def wrap(value: str, font: Any, max_width: int) -> list[str]:
        result: list[str] = []
        for source in str(value or "").splitlines() or [""]:
            words, current = source.split() or [""], ""
            for word in words:
                candidate = word if not current else f"{current} {word}"
                if current and measure.textlength(candidate, font=font) > max_width:
                    result.append(current)
                    current = word
                else:
                    current = candidate
            result.append(current)
        return result or [""]

    layout: list[tuple[str, list[dict[str, Any]], bool, int, int]] = []
    for label, values, personal in task_rows:
        chunks = [values[index:index + 6] for index in range(0, len(values), 6)] or [[]]
        for chunk_index, chunk in enumerate(chunks):
            line_counts = [len(wrap(label, bold, column_widths[1] - 12))]
            for item in chunk:
                # AM/PM and 08:00 share the top badge row. A deadline date
                # uses its own row at the bottom of the task card.
                badges = 1 + int(
                    bool(item.get("is_deadline_important") or item.get("isDeadlineImportant"))
                    and _task_due_day(item) is not None
                )
                line_counts.append(
                    len(wrap(_task_title(item, personal=personal), regular, column_widths[2] - 12)) + badges
                )
            layout.append((label, chunk, personal, chunk_index, max(44, 12 + max(line_counts) * 21)))

    header_top, header_height = 92, 40
    comment_columns = comment_initials or list(COMMENT_FIXED_INITIALS)
    comment_lines = _comment_write_in_lines(comment_columns)
    comment_title_height, comment_line_height = 30, 28
    comment_block_height = 20 + comment_title_height + len(comment_lines) * comment_line_height
    meeting_pair = list((meeting_sections or [])[:2])
    if meeting_pair:
        while len(meeting_pair) < 2:
            meeting_pair.append((meeting_pair[0][0], "", []))
    meeting_half_width = (width - (margin * 2)) // 2
    meeting_label_width = 150
    meeting_content_width = meeting_half_width - meeting_label_width
    meeting_layout: list[tuple[str, list[dict[str, Any]], list[dict[str, Any]], int]] = []
    if meeting_pair:
        left_rows = {label: values for label, values, *_ in meeting_pair[0][2]}
        right_rows = {label: values for label, values, *_ in meeting_pair[1][2]}
        meeting_labels = list(left_rows)
        meeting_labels.extend(label for label in right_rows if label not in left_rows)

        def meeting_values_height(values: list[dict[str, Any]]) -> int:
            height = 0
            for index, item in enumerate(values, 1):
                value = (
                    f"{index}. {_report_text(_first_line(item.get('title')))} "
                    f"{str(item.get('time') or '').strip()}"
                ).strip()
                height += max(26, len(wrap(value, regular, meeting_content_width - 18)) * 20 + 8)
            return height

        for label in meeting_labels:
            left_values = left_rows.get(label, [])
            right_values = right_rows.get(label, [])
            row_height = max(
                44,
                meeting_values_height(left_values) + 10,
                meeting_values_height(right_values) + 10,
            )
            meeting_layout.append((label, left_values, right_values, row_height))
    meeting_block_height = (
        20 + 42 + 38 + sum(row[3] for row in meeting_layout)
        if meeting_pair else 0
    )
    height = (
        header_top + header_height + sum(row[4] for row in layout)
        + meeting_block_height + comment_block_height + margin
    )
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    draw.text((margin, 22), f"1H SHTYPI TODAY - {target_date:%d.%m.%Y}", fill="#111827", font=heading)
    draw.text((margin, 59), "Current Common View state used by the 1H report", fill="#475569", font=regular)

    y, x = header_top, margin
    for column, label in enumerate(["NR", "LLOJI DHE SLOTI"]):
        right = x + column_widths[column]
        draw.rectangle((x, y, right, y + header_height), fill="#F8FAFC", outline="#111827")
        draw.text((x + 6, y + 10), label, fill="#111827", font=bold)
        x = right
    tasks_right = width - margin
    draw.rectangle((x, y, tasks_right, y + header_height), fill="#F8FAFC", outline="#111827")
    tasks_bounds = draw.textbbox((0, 0), "TASKS", font=bold)
    tasks_width = tasks_bounds[2] - tasks_bounds[0]
    draw.text((x + (tasks_right - x - tasks_width) / 2, y + 10), "TASKS", fill="#111827", font=bold)
    y += header_height

    number = 0
    for label, chunk, personal, chunk_index, row_height in layout:
        if chunk_index == 0:
            number += 1
        x = margin
        for column, value in ((0, str(number) if chunk_index == 0 else ""), (1, label if chunk_index == 0 else "")):
            right = x + column_widths[column]
            draw.rectangle((x, y, right, y + row_height), fill="#FFFFFF", outline="#111827", width=2 if chunk_index == 0 else 1)
            for line_index, line in enumerate(wrap(value, bold, column_widths[column] - 12)):
                draw.text((x + 6, y + 6 + line_index * 20), line, fill="#111827", font=bold)
            x = right
        for item_index in range(6):
            right = x + column_widths[2 + item_index]
            item = chunk[item_index] if item_index < len(chunk) else None
            fill, text_color, outline, outline_width = "#FFFFFF", "#111827", "#111827", 1
            if item is not None:
                deadline = bool(item.get("is_deadline_important") or item.get("isDeadlineImportant"))
                due_day = _task_due_day(item) if deadline else None
                if deadline and (due_day is None or due_day != target_date):
                    fill, text_color = DEADLINE_COLOR, "#FFFFFF"
                elif personal and _is_personal_task_for_ga(item):
                    fill = PERSONAL_GA_COLOR
                else:
                    fill = STATUS_COLORS[_task_status(item)]
                if _is_eight_am_task(item):
                    outline, outline_width = EIGHT_AM_BORDER_COLOR, 4
            draw.rectangle((x, y, right, y + row_height), fill=fill, outline=outline, width=outline_width)
            if item is not None:
                text_y = y + 6
                deadline = bool(item.get("is_deadline_important") or item.get("isDeadlineImportant"))
                due_day = _task_due_day(item) if deadline else None
                badge_right = right - 5
                if due_day:
                    date_label = due_day.strftime("%d.%m.%Y")
                    badge_width = int(measure.textlength(date_label, font=small_bold)) + 12
                    badge_left = badge_right - badge_width
                    badge_bottom = y + row_height - 5
                    badge_top = badge_bottom - 23
                    if due_day == target_date:
                        draw.rounded_rectangle(
                            (badge_left, badge_top, badge_right, badge_bottom),
                            radius=3,
                            fill="#EFF6FF",
                            outline="#93C5FD",
                            width=1,
                        )
                        draw.text((badge_left + 6, badge_top + 3), date_label, fill="#1D4ED8", font=small_bold)
                    else:
                        draw.text((badge_left + 6, badge_top + 3), date_label, fill="#FFFFFF", font=small_bold)
                if _is_eight_am_task(item):
                    badge_width = int(measure.textlength("08:00", font=small_bold)) + 12
                    badge_left = badge_right - badge_width
                    draw.rounded_rectangle(
                        (badge_left, text_y, badge_right, text_y + 23),
                        radius=10,
                        fill="#DC2626",
                        outline="#B91C1C",
                        width=1,
                    )
                    draw.text((badge_left + 6, text_y + 3), "08:00", fill="#FFFFFF", font=small_bold)
                    badge_right = badge_left - 5
                period_label = _task_period_label(item)
                badge_width = int(measure.textlength(period_label, font=small_bold)) + 12
                badge_left = badge_right - badge_width
                draw.rounded_rectangle(
                    (badge_left, text_y, badge_right, text_y + 23),
                    radius=10,
                    fill="#E0F2FE",
                    outline="#BAE6FD",
                    width=1,
                )
                draw.text((badge_left + 6, text_y + 3), period_label, fill="#0369A1", font=small_bold)
                text_y += 28
                value = f"{item_index + 1 + chunk_index * 6}. {_task_title(item, personal=personal)}"
                task_font = bold if fill == DEADLINE_COLOR else regular
                for line_index, line in enumerate(wrap(value, task_font, column_widths[2 + item_index] - 12)):
                    draw.text((x + 6, text_y + line_index * 20), line, fill=text_color, font=task_font)
            x = right
        y += row_height

    if meeting_pair:
        y += 20
        table_right = width - margin
        center = margin + meeting_half_width
        left_date, left_relative, _ = meeting_pair[0]
        right_date, right_relative, _ = meeting_pair[1]
        group_bottom = y + 42
        draw.rectangle((margin, y, center, group_bottom), fill="#EEF2FF", outline="#111827")
        draw.rectangle((center, y, table_right, group_bottom), fill="#EEF2FF", outline="#111827")
        left_header = f"{left_relative} - {left_date:%d.%m.%Y}"
        left_bounds = draw.textbbox((0, 0), left_header, font=bold)
        left_text_width = left_bounds[2] - left_bounds[0]
        draw.text(
            (margin + (meeting_half_width - left_text_width) / 2, y + 10),
            left_header,
            fill="#111827",
            font=bold,
        )
        right_header = f"{right_relative} - {right_date:%d.%m.%Y}" if right_relative else ""
        right_bounds = draw.textbbox((0, 0), right_header, font=bold)
        right_text_width = right_bounds[2] - right_bounds[0]
        draw.text(
            (center + (meeting_half_width - right_text_width) / 2, y + 10),
            right_header,
            fill="#111827",
            font=bold,
        )
        draw.line((center, y, center, group_bottom), fill=NON_ROUTINE_MEETING_BORDER_COLOR, width=5)
        y = group_bottom

        x = margin
        meeting_column_widths = [
            meeting_label_width,
            meeting_content_width,
            meeting_label_width,
            meeting_content_width,
        ]
        for column, label in enumerate(["LLOJI", "TAKIMET", "LLOJI", "TAKIMET"]):
            right = x + meeting_column_widths[column]
            draw.rectangle((x, y, right, y + 38), fill="#F8FAFC", outline="#111827")
            draw.text((x + 6, y + 9), label, fill="#111827", font=bold)
            x = right
        draw.line((center, y, center, y + 38), fill=NON_ROUTINE_MEETING_BORDER_COLOR, width=5)
        y += 38
        for label, left_values, right_values, row_height in meeting_layout:
            x = margin
            row_bottom = y + row_height
            for side_values in (left_values, right_values):
                label_right = x + meeting_label_width
                draw.rectangle((x, y, label_right, row_bottom), fill="#FFFFFF", outline="#111827")
                draw.text((x + 6, y + 8), label, fill="#111827", font=bold)
                content_right = label_right + meeting_content_width
                draw.rectangle((label_right, y, content_right, row_bottom), fill="#FFFFFF", outline="#111827")
                item_y = y + 5
                for index, item in enumerate(side_values, 1):
                    value = (
                        f"{index}. {_report_text(_first_line(item.get('title')))} "
                        f"{str(item.get('time') or '').strip()}"
                    ).strip()
                    lines = wrap(value, regular, meeting_content_width - 18)
                    item_height = max(26, len(lines) * 20 + 8)
                    if _is_non_routine_meeting(item):
                        draw.rectangle(
                            (label_right + 4, item_y, content_right - 4, item_y + item_height - 3),
                            fill="#FFFFFF",
                            outline=NON_ROUTINE_MEETING_BORDER_COLOR,
                            width=3,
                        )
                    for line_index, line in enumerate(lines):
                        draw.text(
                            (label_right + 9, item_y + 4 + line_index * 20),
                            line,
                            fill="#111827",
                            font=regular,
                        )
                    item_y += item_height
                x = content_right
            draw.line((center, y, center, row_bottom), fill=NON_ROUTINE_MEETING_BORDER_COLOR, width=5)
            y = row_bottom

    y += 20
    draw.text((margin, y + 3), "KOMENTE PER STAF", fill="#111827", font=bold)
    y += comment_title_height
    for line in comment_lines:
        draw.text((margin, y + 3), line, fill="#111827", font=regular)
        y += comment_line_height

    output = BytesIO()
    image.crop((0, 0, width, y + margin)).save(output, format="PNG", optimize=True)
    return f"1H-SHTYPI-Today-{target_date:%Y-%m-%d}.png", output.getvalue(), "image/png"


async def _build_print_report(
    target_date: date, *, include_attachment: bool = False, include_meetings: bool = True,
    include_png: bool = False, first_meeting_day_label: str = "NESER",
    payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if payload is None:
        base_url = settings.PRIMEFLOW_API_BASE_URL
        if not base_url:
            raise RuntimeError("PRIMEFLOW_API_BASE_URL is required to generate 1H SHTYPI")
        client = PrimeFlowClient(
            base_url.rstrip("/"),
            settings.PRIMEFLOW_EMAIL or settings.ADMIN_EMAIL,
            settings.PRIMEFLOW_PASSWORD or settings.ADMIN_PASSWORD,
            settings.PRIMEFLOW_ACCESS_TOKEN,
        )
        payload = await client.common_view(target_date)
    items = payload.get("items") or {}
    comment_initials = _comment_user_initials(payload)
    task_rows = _task_rows(items, target_date)
    meeting_dates = [target_date, next_working_day(target_date)] if include_meetings else []
    next_meeting_payload = payload
    if len(meeting_dates) > 1 and payload.get("week_end"):
        try:
            payload_week_end = date.fromisoformat(str(payload["week_end"])[:10])
        except ValueError:
            payload_week_end = meeting_dates[-1]
        if meeting_dates[-1] > payload_week_end:
            base_url = settings.PRIMEFLOW_API_BASE_URL
            if not base_url:
                raise RuntimeError("PRIMEFLOW_API_BASE_URL is required to load next-day meetings")
            next_meeting_payload = await PrimeFlowClient(
                base_url.rstrip("/"),
                settings.PRIMEFLOW_EMAIL or settings.ADMIN_EMAIL,
                settings.PRIMEFLOW_PASSWORD or settings.ADMIN_PASSWORD,
                settings.PRIMEFLOW_ACCESS_TOKEN,
            ).common_view(meeting_dates[-1])
    meeting_sections = [
        (
            meeting_date,
            (
                first_meeting_day_label
                if index == 0
                else ("NESER" if first_meeting_day_label == "SOT" else "DITA PAS NESER")
            ),
            [
                (label, values, False)
                for label, values in _meeting_rows(
                    (payload if meeting_date == target_date else next_meeting_payload).get("items") or {},
                    meeting_date,
                )
            ],
        )
        for index, meeting_date in enumerate(meeting_dates)
    ]
    meeting_rows = meeting_sections[0][2] if meeting_sections else []
    report_date = target_date.strftime("%d.%m.%Y")
    html_body = f"""<!doctype html><html><body style=\"margin:0;color:#000;font-family:Arial,sans-serif\">
<div style=\"text-align:center;font-size:20px;font-weight:700;margin:0 0 12px\">1H SHTYPI — {report_date}</div>
{_one_h_checklists_html()}{_html_table(task_rows, report_date=target_date)}{_dated_meetings_html(meeting_sections)}{_comments_table_html(comment_initials)}</body></html>"""
    content_html = (
        '<div data-today-print-report="true" style="margin:18px 0 14px">'
        + re.sub(r"^.*?<body[^>]*>|</body>.*$", "", html_body, flags=re.S)
        + "</div>"
    )
    plain_rows = [
        f"1H SHTYPI - {report_date}",
        "",
        "PYETJET PER 1H - BORD",
        *(
            f"{index}. {question}" + (f" ({description})" if description else "")
            for index, (question, description) in enumerate(ONE_H_BOARD_CHECKLIST, 1)
        ),
        "",
        "STAFF - HAPAT PER 1H",
        *(
            f"{index}. {question}" + (f" ({description})" if description else "")
            for index, (question, description) in enumerate(ONE_H_STAFF_CHECKLIST, 1)
        ),
        "",
        "TASKS",
    ]
    for label, values, personal in task_rows:
        plain_rows.append(
            f"{label}: "
            + "; ".join(
                f"[{_task_period_label(item)}] {_task_title(item, personal=personal)}"
                for item in values
            )
        )
    if include_meetings:
        for meeting_date, relative, dated_rows in meeting_sections:
            plain_rows.append("")
            plain_rows.append(f"MEETINGS - {relative} - {meeting_date:%d.%m.%Y}")
            for label, values, _ in dated_rows:
                plain_rows.append(
                    f"{label}: " + "; ".join(
                        f"{_report_text(_first_line(item.get('title')))} {item.get('time') or ''}".strip()
                        for item in values
                    )
                )
    plain_rows.extend([
        "",
        "KOMENTE PER STAF",
        *_comment_write_in_lines(comment_initials),
    ])
    report: dict[str, Any] = {
        "subject": subject_for(target_date),
        "target_date": target_date.isoformat(),
        "html": html_body,
        "content_html": content_html,
        "plain_text": "\n".join(plain_rows),
    }
    if include_attachment:
        attachments = [
            _excel_table_attachment(
                task_rows, meeting_rows, target_date, include_meetings=include_meetings,
                comment_initials=comment_initials, meeting_sections=meeting_sections,
            )
        ]
        if include_png:
            attachments.append(
                _png_table_attachment(task_rows, target_date, comment_initials, meeting_sections)
            )
        report["attachments"] = attachments
    return report


async def build_tomorrow_print_report(
    delivery_date: date, *, include_attachment: bool = False
) -> dict[str, Any]:
    return await _build_print_report(
        next_working_day(delivery_date), include_attachment=include_attachment, include_meetings=True
    )


async def build_today_print_report(
    report_date: date, *, include_attachment: bool = False, payload: dict[str, Any] | None = None
) -> dict[str, Any]:
    """Build today's task grid plus today/next-working-day meeting sections."""
    return await _build_print_report(
        report_date, include_attachment=include_attachment, include_meetings=True,
        include_png=True, first_meeting_day_label="SOT", payload=payload
    )


async def send_tomorrow_print_report(report: dict[str, Any], recipients: dict[str, list[str]]) -> dict[str, Any]:
    recipients = ensure_required_shtypi_recipient(recipients)
    return await GmailService().send_verified(
        report["subject"], recipients, report["plain_text"], report["html"], attachments=report.get("attachments")
    )
