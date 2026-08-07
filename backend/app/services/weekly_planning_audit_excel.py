from __future__ import annotations

import io
import hashlib
import json
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.weekly_planning_audit import (
    REPORT_VERSION,
    WeeklyPlanningAuditReport,
    validate_report_integrity,
)


SHEET_NAMES = [
    "RAPORTI FINAL",
    "DETAJET E GABIMEVE",
    "TITUJT - SHKURTESAT PX",
    "SHKURTESAT PX",
    "DËRGIMI AUTOMATIK",
]
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SEVERITY_FILLS = {
    "CRITICAL": PatternFill("solid", fgColor="C00000"),
    "HIGH": PatternFill("solid", fgColor="F4B183"),
    "MEDIUM": PatternFill("solid", fgColor="FFE699"),
    "LOW": PatternFill("solid", fgColor="DDEBF7"),
}
TECHNICAL_MARKUP_TEXT = re.compile(r"\[\[\s*/?\s*(?:added|done)\s*\]\]", re.IGNORECASE)
FINAL_HEADERS = [
    "Nr.", "Personi", "Departamenti", "Statusi PV", "Fokusi kryesor i javës",
    "Numri i detyrave", "Numri i gabimeve", "Gabime kritike", "Gabime të larta",
    "Vlerësimi", "Veprimi i kërkuar",
]
DETAIL_HEADERS = [
    "Nr.", "Personi", "Departamenti", "Data", "Task ID", "Titulli aktual",
    "Problemi konkret", "Titulli i propozuar", "Korrigjimi", "Kodi i rregullit",
    "Çfarë kalon në Description/Notes", "Ashpërsia", "Fokusi i javës", "Burimi",
]
TITLE_HEADERS = [
    "Personi", "Task ID", "Titulli aktual", "Problemi i titullit",
    "Titulli i propozuar", "Çfarë kalon në Description/Notes",
    "Shkurtesat e përdorura", "Burimi i rregullit",
]
ABBREVIATION_HEADERS = ["Shkurtesa", "Definicioni", "Versioni", "Burimi", "Data e përditësimit"]
DELIVERY_HEADERS = [
    "Java e raportuar", "Data dhe ora e gjenerimit", "Kontrolli", "Timezone",
    "Marrësit", "Statusi i dërgimit", "Message/provider ID", "Numri i tentimit",
    "Report run ID", "Versioni i raportit", "AI status", "Modeli AI",
    "Versioni i fjalorit PX", "Checksum i payload-it",
]


def report_filename(report: WeeklyPlanningAuditReport) -> str:
    return (
        f"Raporti_PF_PLNF_JAV_{report.week_start:%d-%m-%Y}_{report.week_end:%d-%m-%Y}_"
        f"{report.slot.replace(':', '-')}_{report.generated_at:%d-%m-%Y}.xlsx"
    )


def report_subject(report: WeeklyPlanningAuditReport) -> str:
    week_range = f"{report.week_start:%d.%m.%Y}–{report.week_end:%d.%m.%Y}"
    return f"Kontrolli {report.slot} | PLNF JAV {week_range} | Raporti {report.generated_at:%d.%m.%Y}"


def report_email_body(report: WeeklyPlanningAuditReport) -> str:
    critical = sum(item.severity == "CRITICAL" for item in report.errors)
    high = sum(item.severity == "HIGH" for item in report.errors)
    return (
        "Përshëndetje,\n\n"
        "Bashkëngjitur është raporti aktual i kontrollit të planifikimit javor në PrimeFlow "
        f"për javën {report.week_start:%d.%m.%Y}–{report.week_end:%d.%m.%Y}.\n\n"
        f"Raporti është gjeneruar nga gjendja aktuale në PrimeFlow në orën {report.slot}.\n\n"
        "Përmbledhje:\n"
        f"- Persona të përfshirë: {len(report.people)}\n"
        f"- Persona të përjashtuar për PV të plotë: {len(report.excluded_full_leave)}\n"
        f"- Gabime gjithsej: {len(report.errors)}\n"
        f"- Gabime kritike: {critical}\n"
        f"- Gabime të larta: {high}\n\n"
        "Me respekt,\nPrimeFlow"
    )


def _style_sheet(ws, *, severity_column: int | None = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if severity_column:
            severity = str(row[severity_column - 1].value or "").upper()
            if severity in SEVERITY_FILLS:
                row[severity_column - 1].fill = SEVERITY_FILLS[severity]
                if severity == "CRITICAL":
                    row[severity_column - 1].font = Font(color="FFFFFF", bold=True)
    for column_cells in ws.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max(12, max((len(value) for value in values), default=0) + 2), 55)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def build_weekly_planning_audit_workbook(
    report: WeeklyPlanningAuditReport,
    *,
    recipients: dict[str, list[str]],
    run_id: str,
    delivery_status: str = "Generated, not sent",
    message_id: str | None = None,
    attempt_number: int = 0,
) -> bytes:
    validate_report_integrity(report)
    if any(error.rule_code == "FINISH_PERIOD_MISSING" for error in report.errors):
        raise ValueError("AM/PM must not be audited by the Friday weekly planning report")
    payload_checksum = hashlib.sha256(
        json.dumps(report.to_dict(), ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEET_NAMES:
        workbook.create_sheet(name)

    final = workbook["RAPORTI FINAL"]
    final.append(FINAL_HEADERS)
    for index, person in enumerate(report.people, 1):
        final.append([
            index, person.employee, person.department, person.leave_status, person.focus,
            person.task_count, person.error_count, person.critical_count, person.high_count,
            person.assessment, person.required_action,
        ])
    _style_sheet(final)

    details = workbook["DETAJET E GABIMEVE"]
    details.append(DETAIL_HEADERS)
    for index, error in enumerate(report.errors, 1):
        details.append([
            index, error.employee, error.department, error.task_date, error.task_id,
            error.current_title, error.problem, error.proposed_title, error.correction,
            error.rule_code, error.move_to_notes, error.severity, error.weekly_focus, error.source,
        ])
        if error.task_date:
            details.cell(details.max_row, 4).number_format = "dd.mm.yyyy"
    _style_sheet(details, severity_column=12)

    cleanup = workbook["TITUJT - SHKURTESAT PX"]
    cleanup.append(TITLE_HEADERS)
    for item in report.title_cleanup:
        cleanup.append([
            item["employee"], item["task_id"], item["current_title"], item["title_problem"],
            item["proposed_title"], item["move_to_notes"], item["used_abbreviations"],
            item["rule_source"],
        ])
    _style_sheet(cleanup)

    abbreviations = workbook["SHKURTESAT PX"]
    abbreviations.append(ABBREVIATION_HEADERS)
    for abbreviation, definition in report.abbreviations.items():
        abbreviations.append([
            abbreviation, definition, report.abbreviation_version,
            report.abbreviation_source, report.abbreviation_updated_at,
        ])
    _style_sheet(abbreviations)

    delivery = workbook["DËRGIMI AUTOMATIK"]
    delivery.append(DELIVERY_HEADERS)
    recipient_text = "; ".join(
        f"{kind.upper()}: {', '.join(values)}" for kind, values in recipients.items() if values
    )
    delivery.append([
        f"{report.week_start:%d.%m.%Y}–{report.week_end:%d.%m.%Y}",
        report.generated_at.replace(tzinfo=None),
        report.slot,
        report.timezone,
        recipient_text,
        delivery_status,
        message_id or "",
        attempt_number,
        run_id,
        REPORT_VERSION,
        report.ai_status,
        report.ai_model or "",
        report.abbreviation_version,
        payload_checksum,
    ])
    delivery.cell(2, 2).number_format = "dd.mm.yyyy hh:mm"
    _style_sheet(delivery)

    output = io.BytesIO()
    workbook.save(output)
    # Re-open before returning so corrupt workbooks fail generation rather than download.
    output.seek(0)
    verified = load_workbook(output, read_only=True, data_only=False)
    if verified.sheetnames != SHEET_NAMES:
        raise ValueError("Weekly planning audit workbook has an invalid sheet layout")
    expected_headers = {
        "RAPORTI FINAL": FINAL_HEADERS,
        "DETAJET E GABIMEVE": DETAIL_HEADERS,
        "TITUJT - SHKURTESAT PX": TITLE_HEADERS,
        "SHKURTESAT PX": ABBREVIATION_HEADERS,
        "DËRGIMI AUTOMATIK": DELIVERY_HEADERS,
    }
    for sheet_name, headers in expected_headers.items():
        actual = [cell.value for cell in next(verified[sheet_name].iter_rows(min_row=1, max_row=1))]
        if actual != headers:
            raise ValueError(f"Weekly planning audit workbook has invalid headers: {sheet_name}")
    if verified["RAPORTI FINAL"].max_row != len(report.people) + 1:
        raise ValueError("Weekly planning audit summary row count is invalid")
    if verified["DETAJET E GABIMEVE"].max_row != len(report.errors) + 1:
        raise ValueError("Weekly planning audit detail row count is invalid")
    for sheet in verified.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and TECHNICAL_MARKUP_TEXT.search(cell.value):
                    raise ValueError("Technical diff markup leaked into the workbook")
    verified.close()
    return output.getvalue()


def update_weekly_planning_audit_delivery_metadata(
    workbook_bytes: bytes,
    *,
    delivery_status: str,
    message_id: str | None,
    attempt_number: int,
) -> bytes:
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    delivery = workbook["DËRGIMI AUTOMATIK"]
    delivery.cell(2, 6, delivery_status)
    delivery.cell(2, 7, message_id or "")
    delivery.cell(2, 8, attempt_number)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
