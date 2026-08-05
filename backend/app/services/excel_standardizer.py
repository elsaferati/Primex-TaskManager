from __future__ import annotations

import csv
import io
import math
import re
import unicodedata
from copy import copy
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException


KOSOVO_TIMEZONE = ZoneInfo("Europe/Belgrade")
MAX_UPLOAD_BYTES = 20 * 1024 * 1024
GENERIC_SHEET_NAMES = {"sheet", "sheet1", "sheet2", "data", "table", "final"}
INVALID_SHEET_TITLE = re.compile(r"[\\/*?:\[\]]")
FILENAME_DATE = re.compile(r"(?<!\d)\d{1,2}[._-]\d{1,2}[._-]\d{2,4}(?!\d)")


@dataclass(frozen=True)
class MissingHeader:
    column: str
    column_index: int


@dataclass(frozen=True)
class SheetAnalysis:
    name: str
    source_header_row: int
    source_first_column: int
    source_last_column: int
    source_last_row: int
    headers: list[str]
    missing_headers: list[MissingHeader]
    suggested_title: str


@dataclass(frozen=True)
class WorkbookAnalysis:
    filename: str
    suggested_description: str
    sheets: list[SheetAnalysis]
    empty_sheets: list[str]

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["has_missing_headers"] = any(sheet.missing_headers for sheet in self.sheets)
        return payload


class ExcelStandardizationError(ValueError):
    pass


def initials_from_user(
    full_name: str | None,
    username: str | None = None,
    email: str | None = None,
) -> str:
    label = _text(full_name) or _text(username) or _text(email).split("@", 1)[0]
    parts = [part for part in re.split(r"[^0-9A-Za-zÀ-ž]+", label) if part]
    return "".join(part[0].upper() for part in parts)[:10]


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _normalize(value: Any) -> str:
    raw = unicodedata.normalize("NFKD", _text(value))
    ascii_value = "".join(char for char in raw if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", ascii_value.upper()).strip(" .,:;_-/")


ABBREVIATIONS: dict[str, str] = {
    _normalize(full): short
    for short, full in {
        "PX": "PRIMEX",
        "PV": "PUSHIM VJETOR",
        "MUNG": "MUNGESE",
        "VONS": "VONESE",
        "TAK EXT": "TAKIM EXTERN",
        "TAK INT": "TAKIM INTERN",
        "BZ GA": "BARAZIM ME GANEN",
        "BLL": "BLLOK",
        "R1": "RAST I PARE",
        "P:": "PERSONALISHT",
        "PRJK": "PROJEKTE",
        "ANK": "ANKESA",
        "KRK": "KERKESA",
        "PRZ": "PROPOZIME",
        "MBL": "MBLEDHJA",
        "BZ": "BARAZIM",
        "MOSBZ": "MOSBARAZIM",
        "TAK": "TAKIM",
        "JAV": "JAVOR",
        "CV": "COMMON VIEW",
        "PLNF": "PLANIFIKIMI",
        "RLZ": "REALIZIMI",
        "EM": "EMAIL",
        "DEP": "DEPARTAMENTI",
        "PCM": "PROJEKTE / PRODUCT CONTENT MANAGMENT",
        "ZHVLL": "ZHVILLIMI",
        "GD": "GRAPHIC DESIGN",
        "ADM": "ADMINISTRATA",
        "RAP": "RAPORTIMI",
        "PRBL": "PROBLEME",
        "DET": "DETYRA",
        "DET SYS": "DETYRA TE SISTEMIT",
        "SYS": "SISTEM",
        "FT": "FAST TASK",
        "KO": "KONTROLLA",
        "T": "TODAY",
        "Y": "YESTERDAY",
        "O": "OVERVIEW",
        "PAG": "PAGESE",
        "INV": "INVOICE",
        "FAT": "FATURE",
        "KONF": "KONFIRMIM",
        "PF": "PRIME FLOW/PLATFORMA",
        "DOC": "DOKUMENTE",
        "TRANS": "TRANSAKSIONE",
        "FINC": "FINANCA",
        "VERIF": "VERIFIKU",
        "GR": "GRUPI",
        "RJ": "RAPORTI JAVOR",
        "RN": "RAPORTI NESER",
        "RS": "RAPORTI SOT",
        "DG": "DETYRA GRUPORE",
        "STAND": "STANDARDET",
        "J.T": "JAVEN TJETER",
        "PROD": "PRODUKTE",
        "ZGJ": "ZGJIDHJE",
        "DRZ": "DOREZOHET",
        "APL": "APLIKANTA",
        "KOMPL": "KOMPLET",
        "REGJ": "REGJISTRATOR",
        "CHL": "CHECKLIST",
        "P/P": "PYETJE/PERGJIGJE",
        "RIORG": "RIORGANIZIM",
    }.items()
}

ABBREVIATIONS.update(
    {
        _normalize(variant): short
        for short, variants in {
            "DRZ": ["DORZOHET", "DOREZOHET", "DORËZOHET", "DORZIM", "DOREZIM", "DORËZIM"],
            "RIORG": ["RIORGANZIMI", "RIORGANIZIMI", "RIORGANIZIM", "RIORGANIZIMI I PUNES"],
            "CHL": ["CHECKLISTA", "CHECKLIST-A", "CHECK LIST", "LISTA E KONTROLLIT"],
            "KOMPL": ["KOMPLETE", "KOMPLOTO", "KOMPLETIMI"],
            "REGJ": ["REGJISTRIM", "REGJISTRIMI", "REGJISTRATORI"],
            "PROD": ["PRODUKTI", "PRODUKTET", "PRODUCT", "PRODUCTS"],
            "ZGJ": ["ZGJIDHJA", "ZGJIDHET"],
            "DET": ["DETYRE", "DETYRAT", "TASK", "TASKS"],
            "PRBL": ["PROBLEM", "PROBLEMI", "PROBLEMET"],
            "KRK": ["KERKESE", "KERKESA", "REQUEST", "REQUESTS"],
            "ANK": ["ANKESE", "COMPLAINT", "COMPLAINTS"],
            "PAG": ["PAGESE", "PAGESA", "PAYMENT", "PAYMENTS"],
        }.items()
        for variant in variants
    }
)

NUMERIC_HEADER_TERMS = {
    "NR", "NUMBER", "QTY", "QUANTITY", "SASI", "CMIM", "PRICE", "AMOUNT", "SHUMA",
    "TOTAL", "TOTALI", "COST", "KOSTO", "HOURS", "ORE", "KOHA", "PERCENT", "PERCENTAGE",
    "VALUE", "VLERA", "EUR", "USD",
}
CURRENCY_HEADER_TERMS = {"CMIM", "PRICE", "AMOUNT", "SHUMA", "TOTAL", "TOTALI", "COST", "KOSTO", "EUR", "USD"}


def _load_workbook(content: bytes, filename: str):
    if not content:
        raise ExcelStandardizationError("Skedari është bosh.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ExcelStandardizationError("Skedari tejkalon kufirin prej 20 MB.")

    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        try:
            return load_workbook(io.BytesIO(content), data_only=False)
        except (InvalidFileException, BadZipFile, KeyError, OSError) as exc:
            raise ExcelStandardizationError("Skedari .xlsx nuk mund të lexohet ose është i dëmtuar.") from exc
    if suffix == ".csv":
        decoded: str | None = None
        for encoding in ("utf-8-sig", "cp1252"):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise ExcelStandardizationError("CSV-ja nuk ka encoding të mbështetur.")
        try:
            dialect = csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "DATA"
        for row in csv.reader(io.StringIO(decoded), dialect):
            worksheet.append(row)
        return workbook
    raise ExcelStandardizationError("Lejohen vetëm skedarët .xlsx ose .csv.")


def _used_bounds(worksheet) -> tuple[int, int, int, int] | None:
    populated: list[tuple[int, int]] = []
    for row in worksheet.iter_rows():
        for cell in row:
            if _text(cell.value):
                populated.append((cell.row, cell.column))
    if not populated:
        return None
    rows = [item[0] for item in populated]
    columns = [item[1] for item in populated]
    return min(rows), max(rows), min(columns), max(columns)


def _merged_singleton_row(worksheet, row: int) -> bool:
    return any(cell_range.min_row <= row <= cell_range.max_row and cell_range.max_col > cell_range.min_col for cell_range in worksheet.merged_cells.ranges)


def _find_header_row(worksheet, bounds: tuple[int, int, int, int]) -> int:
    first_row, last_row, first_col, last_col = bounds
    candidates: list[tuple[float, int]] = []
    for row in range(first_row, min(last_row, first_row + 24) + 1):
        values = [worksheet.cell(row, column).value for column in range(first_col, last_col + 1)]
        nonempty_columns = [first_col + index for index, value in enumerate(values) if _text(value)]
        if not nonempty_columns:
            continue
        text_count = sum(isinstance(value, str) and not str(value).startswith("=") for value in values if _text(value))
        below_support = 0
        for column in nonempty_columns:
            if any(_text(worksheet.cell(next_row, column).value) for next_row in range(row + 1, min(last_row, row + 3) + 1)):
                below_support += 1
        score = len(nonempty_columns) * 10 + text_count * 3 + below_support * 2
        score -= (row - first_row) * 12
        if len(nonempty_columns) == 1:
            score -= 35
        if _merged_singleton_row(worksheet, row) and len(nonempty_columns) == 1:
            score -= 15
        candidates.append((score, row))
    if not candidates:
        return first_row
    best_score = max(score for score, _ in candidates)
    return min(row for score, row in candidates if score == best_score)


def _suggested_title(worksheet, header_row: int, filename: str) -> str:
    for row in range(header_row - 1, 0, -1):
        values = [_text(worksheet.cell(row, column).value) for column in range(1, worksheet.max_column + 1)]
        visible = [value for value in values if value]
        if len(visible) == 1 and not visible[0].startswith("="):
            return visible[0].upper()
    if worksheet.title.casefold() not in GENERIC_SHEET_NAMES:
        return worksheet.title.upper()
    return _description_from_filename(filename).replace("_", " ")


def _description_from_filename(filename: str) -> str:
    stem = Path(filename).stem
    had_date = FILENAME_DATE.search(stem) is not None
    stem = FILENAME_DATE.sub(" ", stem)
    if had_date:
        stem = re.sub(r"(?:^|[_\s-])[A-Z]{1,5}$", " ", stem.strip(), flags=re.IGNORECASE)
    cleaned = _sanitize_description(stem)
    return cleaned or "EXCEL_STANDARD"


def _sanitize_description(value: str) -> str:
    return re.sub(r"[^0-9A-Za-zÀ-ž]+", "_", value.upper()).strip("_")


def analyze_workbook(content: bytes, filename: str) -> WorkbookAnalysis:
    workbook = _load_workbook(content, filename)
    sheets: list[SheetAnalysis] = []
    empty_sheets: list[str] = []
    for worksheet in workbook.worksheets:
        bounds = _used_bounds(worksheet)
        if bounds is None:
            empty_sheets.append(worksheet.title)
            continue
        _, last_row, first_col, last_col = bounds
        header_row = _find_header_row(worksheet, bounds)
        missing: list[MissingHeader] = []
        headers: list[str] = []
        for column in range(first_col, last_col + 1):
            value = _text(worksheet.cell(header_row, column).value)
            headers.append(value)
            has_data_below = any(_text(worksheet.cell(row, column).value) for row in range(header_row + 1, last_row + 1))
            if not value and has_data_below:
                missing.append(MissingHeader(column=get_column_letter(column), column_index=column))
        sheets.append(
            SheetAnalysis(
                name=worksheet.title,
                source_header_row=header_row,
                source_first_column=first_col,
                source_last_column=last_col,
                source_last_row=last_row,
                headers=headers,
                missing_headers=missing,
                suggested_title=_suggested_title(worksheet, header_row, filename),
            )
        )
    if not sheets:
        raise ExcelStandardizationError("Skedari nuk përmban asnjë sheet me të dhëna.")
    return WorkbookAnalysis(
        filename=filename,
        suggested_description=_description_from_filename(filename),
        sheets=sheets,
        empty_sheets=empty_sheets,
    )


def _parse_number(value: Any) -> int | float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    text = str(value).strip().replace("\u00a0", " ")
    if not re.search(r"\d", text):
        return None
    cleaned = re.sub(r"[^0-9,.-]", "", text)
    if not cleaned or cleaned.count("-") > 1 or ("-" in cleaned and not cleaned.startswith("-")):
        return None
    if re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", cleaned):
        cleaned = cleaned.replace(",", "")
    elif re.fullmatch(r"-?\d+,\d{1,2}", cleaned):
        cleaned = cleaned.replace(",", ".")
    elif "," in cleaned and "." not in cleaned:
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    try:
        number = float(cleaned)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return int(number) if number.is_integer() else number


def _is_numeric_column(worksheet, analysis: SheetAnalysis, source_col: int) -> tuple[bool, bool]:
    header = _normalize(worksheet.cell(analysis.source_header_row, source_col).value)
    header_tokens = set(re.findall(r"[A-Z0-9]+", header))
    keyword_match = bool(header_tokens & NUMERIC_HEADER_TERMS) or "%" in header
    currency = bool(header_tokens & CURRENCY_HEADER_TERMS) or any(symbol in _text(worksheet.cell(analysis.source_header_row, source_col).value) for symbol in ("€", "$"))
    values = [worksheet.cell(row, source_col).value for row in range(analysis.source_header_row + 1, analysis.source_last_row + 1)]
    values = [value for value in values if _text(value) and not (isinstance(value, str) and value.startswith("="))]
    if not values:
        return False, currency
    numeric_count = sum(_parse_number(value) is not None for value in values)
    real_numeric_count = sum(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values)
    ratio = numeric_count / len(values)
    real_ratio = real_numeric_count / len(values)
    return (keyword_match and ratio >= 0.6) or real_ratio >= 0.85, currency


def _safe_sheet_name(value: str, existing: set[str]) -> str:
    base = INVALID_SHEET_TITLE.sub(" ", value).strip().upper() or "EXCEL"
    base = re.sub(r"\s+", " ", base)[:31]
    candidate = base
    suffix = 2
    while candidate.casefold() in existing:
        ending = f" {suffix}"
        candidate = f"{base[:31 - len(ending)]}{ending}"
        suffix += 1
    existing.add(candidate.casefold())
    return candidate


def _copy_value(source: Cell, target: Cell) -> None:
    value = source.value
    if isinstance(value, str) and value.startswith("="):
        try:
            value = Translator(value, origin=source.coordinate).translate_formula(target.coordinate)
        except Exception:
            value = source.value
    target.value = value
    if source.comment is not None:
        target.comment = copy(source.comment)
    if source.hyperlink is not None:
        target._hyperlink = copy(source.hyperlink)


def _add_correction(corrections: list[dict[str, Any]], category: str, detail: str, count: int | None = None) -> None:
    item: dict[str, Any] = {"category": category, "detail": detail}
    if count is not None:
        item["count"] = count
    corrections.append(item)


def _estimate_row_height(worksheet, row: int, first_col: int, last_col: int) -> float:
    lines = 1
    for column in range(first_col, last_col + 1):
        value = _text(worksheet.cell(row, column).value)
        width = worksheet.column_dimensions[get_column_letter(column)].width or 12
        explicit = value.splitlines() or [""]
        cell_lines = sum(max(1, math.ceil(len(line) / max(1, int(width) - 2))) for line in explicit)
        lines = max(lines, cell_lines)
    return min(300, max(18, 15 * lines))


def _headers_match_freeze(value: Any) -> bool:
    return str(value or "").upper() == "C7"


def _source_last_data_row(worksheet, analysis: SheetAnalysis) -> int:
    last_row = analysis.source_last_row
    first_header = _normalize(worksheet.cell(analysis.source_header_row, analysis.source_first_column).value)
    if first_header != "NR":
        return last_row
    removed = 0
    while last_row > analysis.source_header_row and removed < 2:
        first_value = worksheet.cell(last_row, analysis.source_first_column).value
        other_values = [
            worksheet.cell(last_row, column).value
            for column in range(analysis.source_first_column + 1, analysis.source_last_column + 1)
        ]
        if _parse_number(first_value) is None or any(_text(value) for value in other_values):
            break
        last_row -= 1
        removed += 1
    return last_row


def standardize_workbook(
    content: bytes,
    filename: str,
    initials: str,
    missing_headers: dict[str, dict[str, str]] | None = None,
    description: str | None = None,
    generated_at: datetime | None = None,
) -> tuple[bytes, str, dict[str, Any]]:
    clean_initials = re.sub(r"[^0-9A-Za-zÀ-ž]", "", initials).upper()[:10]
    if not clean_initials:
        raise ExcelStandardizationError("Inicialet janë të detyrueshme për skedarin aktual.")
    analysis = analyze_workbook(content, filename)
    missing_headers = missing_headers or {}
    for sheet in analysis.sheets:
        provided = missing_headers.get(sheet.name, {})
        unresolved = [item.column for item in sheet.missing_headers if not _text(provided.get(item.column))]
        if unresolved:
            raise ExcelStandardizationError(f"Plotëso header-at që mungojnë në {sheet.name}: {', '.join(unresolved)}.")

    source_workbook = _load_workbook(content, filename)
    target_workbook = Workbook()
    target_workbook.remove(target_workbook.active)
    now = generated_at.astimezone(KOSOVO_TIMEZONE) if generated_at else datetime.now(KOSOVO_TIMEZONE)
    description_value = _sanitize_description(description) if description else _description_from_filename(filename)
    description_value = description_value or "EXCEL_STANDARD"
    output_filename = f"{description_value}_{now.strftime('%d.%m.%Y')}_{clean_initials}.xlsx"
    existing_names: set[str] = set()
    report_sheets: list[dict[str, Any]] = []
    workbook_corrections: list[dict[str, Any]] = []
    if filename != output_filename:
        _add_correction(workbook_corrections, "filename", f"Emri i skedarit u standardizua nga '{filename}' në '{output_filename}'.")
    if Path(filename).suffix.lower() == ".csv":
        _add_correction(workbook_corrections, "file_format", "CSV-ja u konvertua në formatin final .xlsx.")
    if workbook_corrections:
        report_sheets.append({"name": "Workbook", "source_name": filename, "corrections": workbook_corrections})
    sheet_name_map: dict[str, str] = {}

    analysis_by_name = {sheet.name: sheet for sheet in analysis.sheets}
    for source_sheet in source_workbook.worksheets:
        sheet_analysis = analysis_by_name.get(source_sheet.title)
        if sheet_analysis is None:
            continue
        title = sheet_analysis.suggested_title.upper()
        generic_name = source_sheet.title.casefold() in GENERIC_SHEET_NAMES
        target_name = _safe_sheet_name(title if generic_name else source_sheet.title, existing_names)
        sheet_name_map[source_sheet.title] = target_name
        target_sheet = target_workbook.create_sheet(target_name)
        corrections: list[dict[str, Any]] = []

        if generic_name or target_name != source_sheet.title:
            _add_correction(corrections, "sheet_name", f"Sheet-i '{source_sheet.title}' u emërtua '{target_name}'.")
        if sheet_analysis.source_header_row != 6:
            _add_correction(corrections, "row_structure", f"Header-i u zhvendos nga rreshti {sheet_analysis.source_header_row} në rreshtin 6.")

        first_col = sheet_analysis.source_first_column
        last_col = sheet_analysis.source_last_column
        source_last_data_row = _source_last_data_row(source_sheet, sheet_analysis)
        source_headers: list[str] = []
        for source_col in range(first_col, last_col + 1):
            column_letter = get_column_letter(source_col)
            value = _text(source_sheet.cell(sheet_analysis.source_header_row, source_col).value)
            if not value:
                value = _text(missing_headers.get(sheet_analysis.name, {}).get(column_letter))
                _add_correction(corrections, "header", f"Header-i bosh në kolonën {column_letter} u plotësua me '{value}'.")
            source_headers.append(value)

        add_nr = _normalize(source_headers[0] if source_headers else "") != "NR"
        if add_nr:
            _add_correction(corrections, "nr", "U shtua kolona e parë NR dhe numerimi sekuencial.")
        target_last_col = len(source_headers) + (1 if add_nr else 0)
        numeric_columns: dict[int, tuple[bool, bool]] = {
            source_col: _is_numeric_column(source_sheet, sheet_analysis, source_col)
            for source_col in range(first_col, last_col + 1)
        }

        target_sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=target_last_col)
        title_cell = target_sheet.cell(3, 1, title)
        title_cell.font = Font(name="Calibri", size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        target_sheet.row_dimensions[3].height = 24
        source_title = source_sheet.cell(3, first_col)
        source_title_merged = any(
            cell_range.min_row == 3
            and cell_range.max_row == 3
            and cell_range.min_col <= first_col
            and cell_range.max_col >= last_col
            for cell_range in source_sheet.merged_cells.ranges
        )
        source_title_valid = (
            _text(source_title.value) == _text(source_title.value).upper()
            and bool(source_title.font.bold)
            and (source_title.font.name or "Calibri") == "Calibri"
            and float(source_title.font.sz or 0) == 16
            and source_title.alignment.horizontal == "center"
            and source_title_merged
        )
        if not source_title_valid:
            _add_correction(corrections, "title", "Titulli u krijua/standardizua në rreshtin 3, me Calibri 16 bold dhe i centruar mbi të gjitha kolonat.")

        if add_nr:
            target_sheet.cell(6, 1, "NR")
        abbreviation_changes: list[str] = []
        numeric_changes: list[str] = []
        for source_row in range(sheet_analysis.source_header_row, source_last_data_row + 1):
            target_row = 6 + (source_row - sheet_analysis.source_header_row)
            if add_nr and source_row > sheet_analysis.source_header_row:
                target_sheet.cell(target_row, 1, target_row - 6)
            for source_col in range(first_col, last_col + 1):
                target_col = (source_col - first_col + 1) + (1 if add_nr else 0)
                source_cell = source_sheet.cell(source_row, source_col)
                target_cell = target_sheet.cell(target_row, target_col)
                _copy_value(source_cell, target_cell)
                if source_row == sheet_analysis.source_header_row:
                    header_value = source_headers[source_col - first_col]
                    normalized = _normalize(header_value)
                    target_cell.value = ABBREVIATIONS.get(normalized, header_value.upper())
                    if _text(target_cell.value) != _text(source_cell.value):
                        abbreviation_changes.append(target_cell.coordinate)
                    continue
                is_numeric, _ = numeric_columns[source_col]
                if is_numeric and not (isinstance(target_cell.value, str) and target_cell.value.startswith("=")):
                    parsed = _parse_number(target_cell.value)
                    if parsed is not None and parsed != target_cell.value:
                        target_cell.value = parsed
                        numeric_changes.append(target_cell.coordinate)
                elif isinstance(target_cell.value, str):
                    normalized = _normalize(target_cell.value)
                    replacement = ABBREVIATIONS.get(normalized)
                    if replacement and replacement != target_cell.value:
                        target_cell.value = replacement
                        abbreviation_changes.append(target_cell.coordinate)

        if not add_nr:
            target_sheet.cell(6, 1, "NR")
            for row in range(7, 7 + max(0, source_last_data_row - sheet_analysis.source_header_row)):
                target_sheet.cell(row, 1, row - 6)
            _add_correction(corrections, "nr", "Kolona NR u rinumerua në mënyrë sekuenciale.")

        data_row_count = max(0, source_last_data_row - sheet_analysis.source_header_row)
        last_data_row = 6 + data_row_count
        final_row = last_data_row + 2
        for row in range(last_data_row + 1, final_row + 1):
            target_sheet.cell(row, 1, row - 6)
            for column in range(2, target_last_col + 1):
                target_sheet.cell(row, column, None)
        _add_correction(corrections, "final_rows", "U vendosën saktësisht 2 rreshta bosh brenda kufirit të tabelës.")

        header_range = target_sheet.iter_rows(min_row=6, max_row=6, min_col=1, max_col=target_last_col)
        for row in header_range:
            for cell in row:
                cell.font = copy(cell.font)
                cell.font = Font(name="Calibri", size=11, bold=True)

        thin = Side(style="thin", color="000000")
        thick = Side(style="medium", color="000000")
        misaligned = 0
        not_wrapped = 0
        missing_border = 0
        for source_row in range(sheet_analysis.source_header_row, source_last_data_row + 1):
            for source_col in range(first_col, last_col + 1):
                source_cell = source_sheet.cell(source_row, source_col)
                try:
                    source_alignment = source_cell.alignment
                except (AttributeError, IndexError, KeyError, TypeError):
                    # Excel can still open workbooks whose style table contains a
                    # dangling reference. Treat that source style as non-standard
                    # instead of aborting the whole conversion.
                    source_alignment = None
                if (
                    source_alignment is None
                    or source_alignment.horizontal != "left"
                    or source_alignment.vertical != "bottom"
                ):
                    misaligned += 1
                if isinstance(source_cell.value, str) and (source_alignment is None or not source_alignment.wrap_text):
                    not_wrapped += 1
                try:
                    source_border = source_cell.border
                    border_missing = any(
                        side.style is None
                        for side in (
                            source_border.left,
                            source_border.right,
                            source_border.top,
                            source_border.bottom,
                        )
                    )
                except (AttributeError, IndexError, KeyError, TypeError):
                    border_missing = True
                if border_missing:
                    missing_border += 1

        for row in range(6, final_row + 1):
            for column in range(1, target_last_col + 1):
                cell = target_sheet.cell(row, column)
                cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
                cell.font = Font(name="Calibri", size=11, bold=row == 6)
                cell.border = Border(
                    left=thick if column == 1 else thin,
                    right=thick if column == target_last_col else thin,
                    top=thick if row == 6 else thin,
                    bottom=thick if row in (6, final_row) else thin,
                )
        if misaligned:
            _add_correction(corrections, "alignment", "Qelizat jo-titull u vendosën LEFT + BOTTOM.", misaligned)
        if not_wrapped:
            _add_correction(corrections, "wrap_text", "U aktivizua Wrap Text për qelizat me tekst.", not_wrapped)
        if missing_border:
            _add_correction(corrections, "borders", "U plotësuan kufijtë e tabelës dhe kufiri i jashtëm i trashë.", missing_border)
        if numeric_changes:
            _add_correction(corrections, "numeric_cleanup", f"U kthyen në numra realë qelizat: {', '.join(numeric_changes[:12])}{'…' if len(numeric_changes) > 12 else ''}.", len(numeric_changes))
        if abbreviation_changes:
            _add_correction(corrections, "abbreviations", f"U standardizuan header/etiketa në: {', '.join(abbreviation_changes[:12])}{'…' if len(abbreviation_changes) > 12 else ''}.", len(abbreviation_changes))

        for target_col in range(1, target_last_col + 1):
            if target_col == 1:
                target_sheet.column_dimensions[get_column_letter(target_col)].width = 6
                continue
            values = [_text(target_sheet.cell(row, target_col).value) for row in range(6, final_row + 1)]
            longest_line = max((len(line) for value in values for line in (value.splitlines() or [""])), default=0)
            target_sheet.column_dimensions[get_column_letter(target_col)].width = min(45, max(10, longest_line + 2))

        for source_col, (is_numeric, currency) in numeric_columns.items():
            if not is_numeric:
                continue
            target_col = (source_col - first_col + 1) + (1 if add_nr else 0)
            numeric_values = [target_sheet.cell(row, target_col).value for row in range(7, last_data_row + 1)]
            has_decimal = any(isinstance(value, float) and not value.is_integer() for value in numeric_values)
            number_format = "#,##0.00" if currency else "#,##0.##" if has_decimal else "#,##0"
            for row in range(7, final_row + 1):
                target_sheet.cell(row, target_col).number_format = number_format
        for row in range(7, final_row + 1):
            target_sheet.cell(row, 1).number_format = "#,##0"

        target_sheet.row_dimensions[6].height = _estimate_row_height(target_sheet, 6, 1, target_last_col)
        for row in range(7, last_data_row + 1):
            target_sheet.row_dimensions[row].height = _estimate_row_height(target_sheet, row, 1, target_last_col)
        target_sheet.row_dimensions[last_data_row + 1].height = 18
        target_sheet.row_dimensions[last_data_row + 2].height = 18

        if source_sheet.auto_filter.ref != f"{get_column_letter(first_col)}{sheet_analysis.source_header_row}:{get_column_letter(last_col)}{sheet_analysis.source_last_row}":
            _add_correction(corrections, "filters", "AutoFilter u vendos në të gjithë header-in e rreshtit 6.")
        if not _headers_match_freeze(source_sheet.freeze_panes):
            _add_correction(corrections, "freeze_panes", "Freeze Panes u vendos në C7.")
        target_sheet.auto_filter.ref = f"A6:{get_column_letter(target_last_col)}{last_data_row}"
        target_sheet.freeze_panes = "C7"
        target_sheet.print_title_rows = "$6:$6"
        target_sheet.print_area = f"$A$1:${get_column_letter(target_last_col)}${final_row}"
        target_sheet.sheet_view.showGridLines = False
        target_sheet.page_setup.orientation = "landscape" if target_last_col > 7 else "portrait"
        target_sheet.page_setup.paperSize = target_sheet.PAPERSIZE_A4
        target_sheet.page_setup.fitToPage = True
        target_sheet.page_setup.fitToWidth = 1
        target_sheet.page_setup.fitToHeight = 0
        target_sheet.sheet_properties.pageSetUpPr.fitToPage = True
        target_sheet.page_margins.left = 0.25
        target_sheet.page_margins.right = 0.25
        target_sheet.page_margins.top = 0.5
        target_sheet.page_margins.bottom = 0.5
        target_sheet.page_margins.header = 0.2
        target_sheet.page_margins.footer = 0.2
        target_sheet.oddHeader.right.text = now.strftime("%d/%m/%Y %H:%M")
        target_sheet.evenHeader.right.text = target_sheet.oddHeader.right.text
        target_sheet.firstHeader.right.text = target_sheet.oddHeader.right.text
        target_sheet.oddFooter.center.text = "&P / &N"
        target_sheet.oddFooter.right.text = "PUNOI:"
        target_sheet.evenFooter.center.text = target_sheet.oddFooter.center.text
        target_sheet.evenFooter.right.text = target_sheet.oddFooter.right.text
        target_sheet.firstFooter.center.text = target_sheet.oddFooter.center.text
        target_sheet.firstFooter.right.text = target_sheet.oddFooter.right.text
        _add_correction(corrections, "print_setup", "U standardizuan print area, margjinat, orientimi, rreshti përsëritës, header-i dhe footer-i.")

        report_sheets.append({
            "name": target_name,
            "source_name": source_sheet.title,
            "corrections": corrections,
        })

    if analysis.empty_sheets:
        report_sheets.append({
            "name": "Sheets removed",
            "source_name": None,
            "corrections": [{
                "category": "empty_sheets",
                "detail": f"U hoqën sheet-et bosh: {', '.join(analysis.empty_sheets)}.",
                "count": len(analysis.empty_sheets),
            }],
        })

    for worksheet in target_workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                formula = cell.value
                for source_name, target_name in sheet_name_map.items():
                    if source_name == target_name:
                        continue
                    formula = formula.replace(f"'{source_name}'!", f"'{target_name}'!")
                    formula = re.sub(
                        rf"(?<![A-Za-z0-9_']){re.escape(source_name)}!",
                        f"'{target_name}'!",
                        formula,
                    )
                cell.value = formula
    output = io.BytesIO()
    target_workbook.save(output)
    report = {
        "filename": output_filename,
        "generated_at": now.isoformat(),
        "summary": "Excel-i u standardizua sipas standardeve Primex dhe është gati për shkarkim e printim.",
        "sheets": report_sheets,
    }
    return output.getvalue(), output_filename, report
